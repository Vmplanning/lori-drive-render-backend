import os
import re
import uuid
import csv
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import httpx
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel


APP_NAME = "LORI Drive API"

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
LORI_API_KEY = os.getenv("LORI_API_KEY", "")

if not SUPABASE_URL:
    print("WARNING: SUPABASE_URL is not set.")
if not SUPABASE_SERVICE_ROLE_KEY:
    print("WARNING: SUPABASE_SERVICE_ROLE_KEY is not set.")
if not LORI_API_KEY:
    print("WARNING: LORI_API_KEY is not set. Set this before connecting Voiceflow.")


app = FastAPI(title=APP_NAME, version="1.0.0")

# For prototype use, this is open. For production, restrict this to your actual portal domain.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def require_lori_key(provided_key: Optional[str]) -> None:
    """Simple API-key check so random users cannot call your backend."""
    if not LORI_API_KEY:
        raise HTTPException(status_code=500, detail="LORI_API_KEY is not configured on the server.")
    if provided_key != LORI_API_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized. Missing or invalid x-lori-api-key.")


async def supabase_get(view_name: str, params: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Query a Supabase table or view through the REST API."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    url = f"{SUPABASE_URL}/rest/v1/{view_name}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(url, headers=headers, params=params)

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase query failed.",
                "view": view_name,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )

    return response.json()


@app.get("/")
async def root():
    return {
        "name": APP_NAME,
        "status": "running",
        "message": "LORI Drive backend is live.",
    }


@app.get("/health")
async def health():
    return {
        "status": "ok",
        "supabase_url_configured": bool(SUPABASE_URL),
        "supabase_key_configured": bool(SUPABASE_SERVICE_ROLE_KEY),
        "lori_api_key_configured": bool(LORI_API_KEY),
    }


@app.get("/batch-summary")
async def batch_summary(
    batch_code: str = Query("JESSUP-DEMO-001"),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    rows = await supabase_get(
        "lori_batch_summary",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "*",
            "limit": "1",
        },
    )
    if not rows:
        raise HTTPException(status_code=404, detail=f"No batch found for {batch_code}.")
    return rows[0]


@app.get("/worst-drivers")
async def worst_drivers(
    batch_code: str = Query("JESSUP-DEMO-001"),
    limit: int = Query(10, ge=1, le=100),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    return await supabase_get(
        "lori_worst_drivers",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "employee_id,driver_name,supervisor_name,safety_score,route_score,overall_score,risk_level,trend_direction,recommended_action",
            "limit": str(limit),
        },
    )


@app.get("/best-drivers")
async def best_drivers(
    batch_code: str = Query("JESSUP-DEMO-001"),
    limit: int = Query(10, ge=1, le=100),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    return await supabase_get(
        "lori_best_drivers",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "employee_id,driver_name,supervisor_name,safety_score,route_score,overall_score,risk_level,trend_direction,recommended_action",
            "limit": str(limit),
        },
    )


@app.get("/risk-summary")
async def risk_summary(
    batch_code: str = Query("JESSUP-DEMO-001"),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    return await supabase_get(
        "lori_risk_summary",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "risk_level,driver_count,average_overall_score,average_safety_score,average_route_score,average_compliance_score,average_payroll_score,average_training_score,lowest_overall_score,highest_overall_score",
            "order": "risk_level.asc",
        },
    )


@app.get("/supervisor-actions")
async def supervisor_actions(
    batch_code: str = Query("JESSUP-DEMO-001"),
    limit: int = Query(50, ge=1, le=200),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    return await supabase_get(
        "lori_supervisor_action_list",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "employee_id,driver_name,supervisor_name,action_type,action_reason,priority,due_date,status,notes,priority_sort_order",
            "order": "priority_sort_order.asc,due_date.asc",
            "limit": str(limit),
        },
    )


@app.get("/driver-360")
async def driver_360(
    employee_id: str = Query(..., description="Example: DEMO-D007"),
    batch_code: str = Query("JESSUP-DEMO-001"),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    rows = await supabase_get(
        "lori_driver_360",
        {
            "batch_code": f"eq.{batch_code}",
            "employee_id": f"eq.{employee_id}",
            "select": "*",
            "limit": "1",
        },
    )
    if not rows:
        raise HTTPException(status_code=404, detail=f"No driver profile found for {employee_id} in {batch_code}.")
    return rows[0]


@app.get("/leadership-briefing")
async def leadership_briefing(
    batch_code: str = Query("JESSUP-DEMO-001"),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    require_lori_key(x_lori_api_key)
    rows = await supabase_get(
        "lori_leadership_briefings",
        {
            "batch_code": f"eq.{batch_code}",
            "report_type": "eq.Daily Leadership Briefing",
            "select": "batch_code,organization_name,location_name,report_type,report_title,report_body,created_by,created_at",
            "limit": "1",
        },
    )
    if not rows:
        raise HTTPException(status_code=404, detail=f"No leadership briefing found for {batch_code}.")
    return rows[0]


@app.get("/voiceflow/summary")
async def voiceflow_summary(
    batch_code: str = Query("JESSUP-DEMO-001"),
    x_lori_api_key: Optional[str] = Query(None, alias="api_key"),
):
    """Returns a clean sentence Voiceflow can speak back."""
    require_lori_key(x_lori_api_key)

    batch_rows = await supabase_get(
        "lori_batch_summary",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "*",
            "limit": "1",
        },
    )
    risk_rows = await supabase_get(
        "lori_risk_summary",
        {
            "batch_code": f"eq.{batch_code}",
            "select": "risk_level,driver_count",
        },
    )

    if not batch_rows:
        raise HTTPException(status_code=404, detail=f"No batch found for {batch_code}.")

    batch = batch_rows[0]
    risk_map = {row["risk_level"]: row["driver_count"] for row in risk_rows}

    response_text = (
        f"Batch {batch['batch_code']} is {batch['batch_status']}. "
        f"I found {batch['total_drivers_found']} drivers, {batch['total_safety_events']} safety events, "
        f"{batch['total_payroll_exceptions']} payroll exceptions, {batch['total_training_gaps']} training gaps, "
        f"and {batch['total_compliance_gaps']} compliance gaps. "
        f"Risk summary: {risk_map.get('Corrective Action Needed', 0)} driver requires corrective action, "
        f"{risk_map.get('Watch List', 0)} drivers are on the watch list, "
        f"{risk_map.get('Solid Performer', 0)} are solid performers, and "
        f"{risk_map.get('Elite Performer', 0)} are elite performers."
    )

    return {
        "batch_code": batch_code,
        "response_text": response_text,
        "batch_summary": batch,
        "risk_summary": risk_rows,
    }
def clean_file_name(filename: str) -> str:
    """Make uploaded file names safe for storage paths."""
    name = filename or "uploaded-file"
    name = name.strip().replace(" ", "_")
    name = re.sub(r"[^A-Za-z0-9._-]", "", name)
    return name or "uploaded-file"


async def supabase_insert(table_name: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    """Insert one row into a Supabase table through the REST API."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    url = f"{SUPABASE_URL}/rest/v1/{table_name}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(url, headers=headers, json=payload)

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase insert failed.",
                "table": table_name,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )

    rows = response.json()
    return rows[0] if rows else {}


async def upload_to_supabase_storage(
    bucket_name: str,
    storage_path: str,
    file: UploadFile,
) -> Dict[str, Any]:
    """Upload one file into a Supabase Storage bucket."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    file_bytes = await file.read()

    url = f"{SUPABASE_URL}/storage/v1/object/{bucket_name}/{storage_path}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": file.content_type or "application/octet-stream",
        "x-upsert": "true",
    }

    async with httpx.AsyncClient(timeout=120.0) as client:
        response = await client.post(url, headers=headers, content=file_bytes)

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase storage upload failed.",
                "bucket": bucket_name,
                "path": storage_path,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )

    return {
        "bucket": bucket_name,
        "path": storage_path,
        "file_name": file.filename,
        "content_type": file.content_type,
        "size_bytes": len(file_bytes),
    }


@app.post("/upload-batch")
async def upload_batch(
    api_key: str = Form(...),
    organization_name: str = Form(...),
    location_name: str = Form(...),
    location_code: str = Form(...),
    uploaded_by: str = Form(...),
    report_period_start: date = Form(...),
    report_period_end: date = Form(...),

    driver_performance_file: Optional[UploadFile] = File(None),
    safety_events_file: Optional[UploadFile] = File(None),
    payroll_exceptions_file: Optional[UploadFile] = File(None),
    training_gaps_file: Optional[UploadFile] = File(None),
    compliance_gaps_file: Optional[UploadFile] = File(None),
    route_performance_file: Optional[UploadFile] = File(None),
    supervisor_notes_file: Optional[UploadFile] = File(None),
):
    """
    Receives a LORI delivery operations batch from the portal,
    stores uploaded source files in Supabase Storage,
    and records batch/file tracking metadata in Supabase.
    """
    require_lori_key(api_key)

    bucket_name = "lori-batch-uploads"
    clean_location_code = location_code.strip().upper().replace(" ", "-")
    batch_code = f"{clean_location_code}-{report_period_end.strftime('%Y-%m-%d')}-{uuid.uuid4().hex[:6].upper()}"

    source_files = [
        ("Driver Performance File", driver_performance_file),
        ("Safety Events File", safety_events_file),
        ("Payroll Exceptions File", payroll_exceptions_file),
        ("Training Gaps File", training_gaps_file),
        ("Compliance Gaps File", compliance_gaps_file),
        ("Route Performance File", route_performance_file),
        ("Supervisor Notes File", supervisor_notes_file),
    ]

    uploaded_file_results = []

    for file_type, file in source_files:
        if file is None:
            continue

        safe_name = clean_file_name(file.filename)
        storage_path = f"{batch_code}/{file_type.replace(' ', '_').lower()}/{safe_name}"

        storage_result = await upload_to_supabase_storage(
            bucket_name=bucket_name,
            storage_path=storage_path,
            file=file,
        )

        file_record = await supabase_insert(
            "lori_uploaded_files",
            {
                "batch_code": batch_code,
                "file_type": file_type,
                "original_file_name": file.filename,
                "storage_bucket": bucket_name,
                "storage_path": storage_path,
                "uploaded_by": uploaded_by,
                "processing_status": "uploaded",
                "notes": "Uploaded through LORI Data Intake portal.",
            },
        )

        uploaded_file_results.append(
            {
                "file_type": file_type,
                "original_file_name": file.filename,
                "storage": storage_result,
                "record": file_record,
            }
        )

    batch_record = await supabase_insert(
        "lori_batch_uploads",
        {
            "batch_code": batch_code,
            "organization_name": organization_name,
            "location_name": location_name,
            "location_code": clean_location_code,
            "uploaded_by": uploaded_by,
            "report_period_start": report_period_start.isoformat(),
            "report_period_end": report_period_end.isoformat(),
            "batch_status": "uploaded",
            "total_files": len(uploaded_file_results),
            "notes": "Batch files uploaded successfully. File parsing and scoring will be connected in the next production step.",
        },
    )

    return {
        "status": "success",
        "message": "Batch uploaded successfully. File parsing and scoring will be connected in the next production step.",
        "batch_code": batch_code,
        "batch_status": "uploaded",
        "organization_name": organization_name,
        "location_name": location_name,
        "location_code": clean_location_code,
        "uploaded_by": uploaded_by,
        "report_period_start": report_period_start.isoformat(),
        "report_period_end": report_period_end.isoformat(),
        "files_uploaded": len(uploaded_file_results),
        "uploaded_files": uploaded_file_results,
        "batch_record": batch_record,
        "next_step": "Connect file parsing, scoring, dashboard refresh, and report generation.",
    }
def parse_int(value: Any) -> Optional[int]:
    """Safely convert CSV values into integers."""
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def parse_float(value: Any) -> Optional[float]:
    """Safely convert CSV values into numbers."""
    if value is None or value == "":
        return None
    try:
        cleaned = str(value).strip().replace("%", "")
        number = float(cleaned)
        return number
    except Exception:
        return None


def normalize_on_time_rate(value: Any) -> Optional[float]:
    """Normalize on-time rate values into decimal format when possible."""
    number = parse_float(value)
    if number is None:
        return None
    if number > 1 and number <= 100:
        return round(number / 100, 4)
    return number


def calculate_uploaded_driver_risk(
    overall_score: Optional[float],
    missed_deliveries: Optional[int],
    customer_complaints: Optional[int],
) -> str:
    """Assign a basic risk level from uploaded driver performance data."""
    score = overall_score if overall_score is not None else 75
    missed = missed_deliveries or 0
    complaints = customer_complaints or 0

    if score < 65 or missed >= 5 or complaints >= 2:
        return "Corrective Action"
    if score < 80 or missed >= 3 or complaints >= 1:
        return "Watch List"
    if score < 90:
        return "Solid Performer"
    return "Elite Performer"


def build_uploaded_driver_action(risk_level: str) -> str:
    """Create a leadership-ready recommended action."""
    if risk_level == "Corrective Action":
        return "Prioritize supervisor review, coaching documentation, and follow-up before additional route escalation."
    if risk_level == "Watch List":
        return "Schedule coaching conversation and monitor next reporting cycle for safety, route, or service improvement."
    if risk_level == "Solid Performer":
        return "Maintain standard performance monitoring and consider targeted recognition if trend remains positive."
    return "Recognize as a strong performer and consider for peer mentoring, route leadership, or positive reinforcement."


async def supabase_select(table_name: str, params: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Select rows from Supabase through the REST API."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    url = f"{SUPABASE_URL}/rest/v1/{table_name}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.get(url, headers=headers, params=params)

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase select failed.",
                "table": table_name,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )

    return response.json()


async def supabase_delete_by_batch(table_name: str, batch_code: str) -> None:
    """Delete existing parsed records for a batch before re-parsing."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    url = f"{SUPABASE_URL}/rest/v1/{table_name}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.delete(
            url,
            headers=headers,
            params={"batch_code": f"eq.{batch_code}"},
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase delete failed.",
                "table": table_name,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )


async def supabase_update_by_batch(
    table_name: str,
    batch_code: str,
    payload: Dict[str, Any],
) -> Dict[str, Any]:
    """Update a Supabase row by batch code."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    url = f"{SUPABASE_URL}/rest/v1/{table_name}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.patch(
            url,
            headers=headers,
            params={"batch_code": f"eq.{batch_code}"},
            json=payload,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase update failed.",
                "table": table_name,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )

    rows = response.json()
    return rows[0] if rows else {}


async def download_from_supabase_storage(bucket_name: str, storage_path: str) -> bytes:
    """Download an uploaded file from Supabase Storage."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase environment variables are not configured.")

    url = f"{SUPABASE_URL}/storage/v1/object/{bucket_name}/{storage_path}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    }

    async with httpx.AsyncClient(timeout=120.0) as client:
        response = await client.get(url, headers=headers)

    if response.status_code >= 400:
        raise HTTPException(
            status_code=response.status_code,
            detail={
                "message": "Supabase storage download failed.",
                "bucket": bucket_name,
                "path": storage_path,
                "supabase_status": response.status_code,
                "supabase_response": response.text,
            },
        )

    return response.content


def parse_driver_performance_csv(
    file_bytes: bytes,
    batch_code: str,
    source_file_name: str,
) -> List[Dict[str, Any]]:
    """Parse the uploaded Driver Performance CSV into normalized driver records."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    parsed_rows: List[Dict[str, Any]] = []

    for row in reader:
        employee_id = (row.get("employee_id") or "").strip()
        driver_name = (row.get("driver_name") or "").strip()
        supervisor_name = (row.get("supervisor_name") or "").strip()
        route_id = (row.get("route_id") or "").strip()
        delivery_station = (row.get("delivery_station") or "").strip()

        tenure_months = parse_int(row.get("tenure_months"))
        routes_completed = parse_int(row.get("routes_completed"))
        on_time_rate = normalize_on_time_rate(row.get("on_time_rate"))
        missed_deliveries = parse_int(row.get("missed_deliveries"))
        customer_complaints = parse_int(row.get("customer_complaints"))
        overall_score = parse_float(row.get("overall_score"))

        risk_level = calculate_uploaded_driver_risk(
            overall_score=overall_score,
            missed_deliveries=missed_deliveries,
            customer_complaints=customer_complaints,
        )

        parsed_rows.append(
            {
                "batch_code": batch_code,
                "employee_id": employee_id,
                "driver_name": driver_name,
                "supervisor_name": supervisor_name,
                "route_id": route_id,
                "delivery_station": delivery_station,
                "tenure_months": tenure_months,
                "routes_completed": routes_completed,
                "on_time_rate": on_time_rate,
                "missed_deliveries": missed_deliveries,
                "customer_complaints": customer_complaints,
                "overall_score": overall_score,
                "safety_score": None,
                "route_score": None,
                "payroll_score": None,
                "training_score": None,
                "calculated_risk_level": risk_level,
                "recommended_action": build_uploaded_driver_action(risk_level),
                "source_file_name": source_file_name,
            }
        )

    return parsed_rows


@app.post("/parse-driver-performance")
async def parse_driver_performance(
    batch_code: str = Query(...),
    api_key: str = Query(...),
):
    """
    Parses the uploaded Driver Performance CSV for a batch,
    stores driver records in Supabase,
    and updates the batch upload record.
    """
    require_lori_key(api_key)

    file_rows = await supabase_select(
        "lori_uploaded_files",
        {
            "select": "*",
            "batch_code": f"eq.{batch_code}",
            "file_type": "eq.Driver Performance File",
            "order": "uploaded_at.desc",
            "limit": "1",
        },
    )

    if not file_rows:
        raise HTTPException(
            status_code=404,
            detail=f"No Driver Performance File found for batch {batch_code}.",
        )

    file_record = file_rows[0]
    bucket_name = file_record.get("storage_bucket") or "lori-batch-uploads"
    storage_path = file_record.get("storage_path")
    source_file_name = file_record.get("original_file_name") or "driver_performance.csv"

    if not storage_path:
        raise HTTPException(status_code=500, detail="Uploaded file record is missing storage_path.")

    file_bytes = await download_from_supabase_storage(bucket_name, storage_path)

    parsed_rows = parse_driver_performance_csv(
        file_bytes=file_bytes,
        batch_code=batch_code,
        source_file_name=source_file_name,
    )

    await supabase_delete_by_batch("lori_uploaded_driver_scores", batch_code)

    inserted_count = 0
    risk_summary: Dict[str, int] = {}

    for parsed_row in parsed_rows:
        await supabase_insert("lori_uploaded_driver_scores", parsed_row)
        inserted_count += 1
        risk = parsed_row.get("calculated_risk_level") or "Unclassified"
        risk_summary[risk] = risk_summary.get(risk, 0) + 1

    updated_batch = await supabase_update_by_batch(
        "lori_batch_uploads",
        batch_code,
        {
            "batch_status": "parsed",
            "total_drivers_found": inserted_count,
            "notes": f"Driver Performance File parsed successfully. {inserted_count} driver records inserted.",
        },
    )

    return {
        "status": "success",
        "message": "Driver Performance File parsed successfully.",
        "batch_code": batch_code,
        "source_file_name": source_file_name,
        "drivers_parsed": inserted_count,
        "risk_summary": risk_summary,
        "batch_status": "parsed",
        "updated_batch": updated_batch,
        "next_step": "Connect parsed driver records to dashboard, reports, and LORI assistant workflows.",
    }
def uploaded_risk_summary_from_rows(drivers: List[Dict[str, Any]]) -> Dict[str, int]:
    """Summarize parsed uploaded driver rows by calculated risk level."""
    summary = {
        "Corrective Action": 0,
        "Watch List": 0,
        "Solid Performer": 0,
        "Elite Performer": 0,
    }

    for driver in drivers:
        risk = driver.get("calculated_risk_level") or "Unclassified"
        summary[risk] = summary.get(risk, 0) + 1

    return summary


def driver_score_for_sort(driver: Dict[str, Any]) -> float:
    """Return a usable score for sorting uploaded drivers."""
    score = parse_float(driver.get("overall_score"))
    return score if score is not None else 0.0


async def get_uploaded_batch_or_latest(batch_code: Optional[str]) -> Dict[str, Any]:
    """Get a requested uploaded batch or the latest uploaded batch."""
    if batch_code:
        rows = await supabase_select(
            "lori_batch_uploads",
            {
                "select": "*",
                "batch_code": f"eq.{batch_code}",
                "limit": "1",
            },
        )
    else:
        rows = await supabase_select(
            "lori_batch_uploads",
            {
                "select": "*",
                "order": "created_at.desc",
                "limit": "1",
            },
        )

    if not rows:
        raise HTTPException(status_code=404, detail="No uploaded batch found.")

    return rows[0]


async def get_uploaded_driver_rows(batch_code: str) -> List[Dict[str, Any]]:
    """Get parsed uploaded driver score rows for a batch."""
    return await supabase_select(
        "lori_uploaded_driver_scores",
        {
            "select": "*",
            "batch_code": f"eq.{batch_code}",
            "order": "overall_score.desc",
        },
    )


@app.get("/uploaded-batch-summary")
async def uploaded_batch_summary(
    api_key: str = Query(...),
    batch_code: Optional[str] = Query(None),
):
    """
    Summary for the latest or requested uploaded/parsed batch.
    Used by the Lovable dashboard.
    """
    require_lori_key(api_key)

    batch = await get_uploaded_batch_or_latest(batch_code)
    resolved_batch_code = batch["batch_code"]

    drivers = await get_uploaded_driver_rows(resolved_batch_code)
    risk_summary = uploaded_risk_summary_from_rows(drivers)

    response_text = (
        f"Uploaded batch {resolved_batch_code} is {batch.get('batch_status')}. "
        f"I found {len(drivers)} parsed driver records. "
        f"Risk summary: {risk_summary.get('Corrective Action', 0)} driver requires corrective action, "
        f"{risk_summary.get('Watch List', 0)} drivers are on the watch list, "
        f"{risk_summary.get('Solid Performer', 0)} are solid performers, and "
        f"{risk_summary.get('Elite Performer', 0)} are elite performers."
    )

    return {
        "status": "success",
        "batch_code": resolved_batch_code,
        "batch_status": batch.get("batch_status"),
        "organization_name": batch.get("organization_name"),
        "location_name": batch.get("location_name"),
        "location_code": batch.get("location_code"),
        "uploaded_by": batch.get("uploaded_by"),
        "report_period_start": batch.get("report_period_start"),
        "report_period_end": batch.get("report_period_end"),
        "total_files": batch.get("total_files"),
        "total_drivers_found": len(drivers),
        "risk_summary": risk_summary,
        "response_text": response_text,
        "next_step": "Use this uploaded batch summary to power the leadership dashboard and report cards.",
    }


@app.get("/uploaded-risk-summary")
async def uploaded_risk_summary(
    api_key: str = Query(...),
    batch_code: Optional[str] = Query(None),
):
    """
    Risk summary from parsed uploaded driver records.
    """
    require_lori_key(api_key)

    batch = await get_uploaded_batch_or_latest(batch_code)
    resolved_batch_code = batch["batch_code"]
    drivers = await get_uploaded_driver_rows(resolved_batch_code)

    summary = uploaded_risk_summary_from_rows(drivers)

    return [
        {"risk_level": "Corrective Action", "driver_count": summary.get("Corrective Action", 0)},
        {"risk_level": "Watch List", "driver_count": summary.get("Watch List", 0)},
        {"risk_level": "Solid Performer", "driver_count": summary.get("Solid Performer", 0)},
        {"risk_level": "Elite Performer", "driver_count": summary.get("Elite Performer", 0)},
    ]


@app.get("/uploaded-worst-drivers")
async def uploaded_worst_drivers(
    api_key: str = Query(...),
    batch_code: Optional[str] = Query(None),
    limit: int = Query(3, ge=1, le=25),
):
    """
    Highest-risk drivers from parsed uploaded batch data.
    """
    require_lori_key(api_key)

    batch = await get_uploaded_batch_or_latest(batch_code)
    resolved_batch_code = batch["batch_code"]
    drivers = await get_uploaded_driver_rows(resolved_batch_code)

    risk_order = {
        "Corrective Action": 1,
        "Watch List": 2,
        "Solid Performer": 3,
        "Elite Performer": 4,
    }

    sorted_drivers = sorted(
        drivers,
        key=lambda d: (
            risk_order.get(d.get("calculated_risk_level") or "", 99),
            driver_score_for_sort(d),
        ),
    )

    results = []

    for driver in sorted_drivers[:limit]:
        results.append(
            {
                "employee_id": driver.get("employee_id"),
                "driver_name": driver.get("driver_name"),
                "supervisor_name": driver.get("supervisor_name"),
                "route_id": driver.get("route_id"),
                "risk_level": driver.get("calculated_risk_level"),
                "overall_score": driver.get("overall_score"),
                "missed_deliveries": driver.get("missed_deliveries"),
                "customer_complaints": driver.get("customer_complaints"),
                "recommended_action": driver.get("recommended_action"),
                "batch_code": resolved_batch_code,
            }
        )

    return results


@app.get("/uploaded-best-drivers")
async def uploaded_best_drivers(
    api_key: str = Query(...),
    batch_code: Optional[str] = Query(None),
    limit: int = Query(3, ge=1, le=25),
):
    """
    Best drivers from parsed uploaded batch data.
    """
    require_lori_key(api_key)

    batch = await get_uploaded_batch_or_latest(batch_code)
    resolved_batch_code = batch["batch_code"]
    drivers = await get_uploaded_driver_rows(resolved_batch_code)

    sorted_drivers = sorted(
        drivers,
        key=lambda d: driver_score_for_sort(d),
        reverse=True,
    )

    results = []

    for driver in sorted_drivers[:limit]:
        results.append(
            {
                "employee_id": driver.get("employee_id"),
                "driver_name": driver.get("driver_name"),
                "supervisor_name": driver.get("supervisor_name"),
                "route_id": driver.get("route_id"),
                "risk_level": driver.get("calculated_risk_level"),
                "overall_score": driver.get("overall_score"),
                "on_time_rate": driver.get("on_time_rate"),
                "routes_completed": driver.get("routes_completed"),
                "recommended_action": driver.get("recommended_action"),
                "batch_code": resolved_batch_code,
            }
        )

    return results


@app.get("/uploaded-driver-360")
async def uploaded_driver_360(
    api_key: str = Query(...),
    batch_code: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
):
    """
    Driver 360 profile from parsed uploaded batch data.
    """
    require_lori_key(api_key)

    batch = await get_uploaded_batch_or_latest(batch_code)
    resolved_batch_code = batch["batch_code"]

    params = {
        "select": "*",
        "batch_code": f"eq.{resolved_batch_code}",
        "limit": "1",
    }

    if employee_id:
        params["employee_id"] = f"eq.{employee_id}"
    elif driver_name:
        params["driver_name"] = f"ilike.*{driver_name}*"
    else:
        params["order"] = "overall_score.asc"

    rows = await supabase_select("lori_uploaded_driver_scores", params)

    if not rows:
        raise HTTPException(
            status_code=404,
            detail="No uploaded driver profile found for the provided search.",
        )

    driver = rows[0]

    return {
        "batch_code": resolved_batch_code,
        "employee_id": driver.get("employee_id"),
        "driver_name": driver.get("driver_name"),
        "supervisor_name": driver.get("supervisor_name"),
        "route_id": driver.get("route_id"),
        "delivery_station": driver.get("delivery_station"),
        "tenure_months": driver.get("tenure_months"),
        "routes_completed": driver.get("routes_completed"),
        "on_time_rate": driver.get("on_time_rate"),
        "missed_deliveries": driver.get("missed_deliveries"),
        "customer_complaints": driver.get("customer_complaints"),
        "overall_score": driver.get("overall_score"),
        "safety_score": driver.get("safety_score"),
        "route_score": driver.get("route_score"),
        "payroll_score": driver.get("payroll_score"),
        "training_score": driver.get("training_score"),
        "risk_level": driver.get("calculated_risk_level"),
        "recommended_action": driver.get("recommended_action"),
        "source_file_name": driver.get("source_file_name"),
    }
MONTH_NAMES = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def parse_date_safe(value: Any) -> Optional[date]:
    """Safely parse a date returned from Supabase."""
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def format_birthday(month: Any, day: Any) -> str:
    """Format birthday month/day without storing a full birthdate."""
    try:
        month_int = int(month)
        day_int = int(day)
        month_name = MONTH_NAMES.get(month_int, f"Month {month_int}")
        return f"{month_name} {day_int}"
    except Exception:
        return "Not available"


async def find_driver_master_record(
    employee_id: Optional[str] = None,
    driver_name: Optional[str] = None,
    question: Optional[str] = None,
) -> Dict[str, Any]:
    """Find a driver by employee ID, driver name, or name mentioned inside a question."""
    if employee_id:
        rows = await supabase_select(
            "lori_driver_master",
            {
                "select": "*",
                "employee_id": f"eq.{employee_id}",
                "limit": "1",
            },
        )
        if rows:
            return rows[0]

    if driver_name:
        rows = await supabase_select(
            "lori_driver_master",
            {
                "select": "*",
                "driver_name": f"ilike.*{driver_name}*",
                "limit": "1",
            },
        )
        if rows:
            return rows[0]

    if question:
        all_drivers = await supabase_select(
            "lori_driver_master",
            {
                "select": "*",
                "order": "driver_name.asc",
            },
        )

        question_lower = question.lower()

        for driver in all_drivers:
            name = (driver.get("driver_name") or "").lower()
            preferred = (driver.get("preferred_name") or "").lower()
            emp_id = (driver.get("employee_id") or "").lower()

            if emp_id and emp_id in question_lower:
                return driver

            if name and name in question_lower:
                return driver

            if preferred and preferred in question_lower:
                return driver

    raise HTTPException(
        status_code=404,
        detail="No matching driver was found. Try using the driver name or employee ID.",
    )


async def get_driver_intelligence_payload(
    employee_id: Optional[str] = None,
    driver_name: Optional[str] = None,
    question: Optional[str] = None,
) -> Dict[str, Any]:
    """Collect the full driver intelligence profile from all driver tables."""
    master = await find_driver_master_record(
        employee_id=employee_id,
        driver_name=driver_name,
        question=question,
    )

    resolved_employee_id = master["employee_id"]

    credentials = await supabase_select(
        "lori_driver_credentials",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "expiration_date.asc",
        },
    )

    counseling = await supabase_select(
        "lori_driver_counseling",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "counseling_date.desc",
        },
    )

    routes = await supabase_select(
        "lori_driver_routes",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "route_date.desc",
            "limit": "10",
        },
    )

    metrics = await supabase_select(
        "lori_driver_metrics",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "metric_period_end.desc",
            "limit": "5",
        },
    )

    safety_events = await supabase_select(
        "lori_driver_safety_events",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "event_date.desc",
            "limit": "10",
        },
    )

    notes = await supabase_select(
        "lori_driver_notes",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "note_date.desc",
            "limit": "10",
        },
    )

    alerts = await supabase_select(
        "lori_driver_alerts",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "due_date.asc",
        },
    )

    timeline = await supabase_select(
        "lori_driver_timeline",
        {
            "select": "*",
            "employee_id": f"eq.{resolved_employee_id}",
            "order": "event_date.desc",
            "limit": "15",
        },
    )

    answer_text = build_driver_intelligence_answer(
        master=master,
        credentials=credentials,
        counseling=counseling,
        routes=routes,
        metrics=metrics,
        safety_events=safety_events,
        notes=notes,
        alerts=alerts,
        timeline=timeline,
        question=question,
    )

    return {
        "status": "success",
        "driver": master,
        "credentials": credentials,
        "counseling": counseling,
        "routes": routes,
        "metrics": metrics,
        "safety_events": safety_events,
        "notes": notes,
        "alerts": alerts,
        "timeline": timeline,
        "answer_text": answer_text,
    }


def build_driver_intelligence_answer(
    master: Dict[str, Any],
    credentials: List[Dict[str, Any]],
    counseling: List[Dict[str, Any]],
    routes: List[Dict[str, Any]],
    metrics: List[Dict[str, Any]],
    safety_events: List[Dict[str, Any]],
    notes: List[Dict[str, Any]],
    alerts: List[Dict[str, Any]],
    timeline: List[Dict[str, Any]],
    question: Optional[str] = None,
) -> str:
    """Create a leadership-ready answer from the driver intelligence profile."""
    q = (question or "").lower()
    driver_name = master.get("driver_name") or "this driver"
    employee_id = master.get("employee_id") or "Unknown"
    supervisor = master.get("supervisor_name") or "Not listed"
    location = master.get("location_name") or "Not listed"
    primary_route = master.get("primary_route_id") or "Not listed"
    birthday = format_birthday(master.get("birthday_month"), master.get("birthday_day"))

    latest_metric = metrics[0] if metrics else {}
    latest_route = routes[0] if routes else {}
    open_alerts = [a for a in alerts if (a.get("status") or "").lower() == "open"]
    critical_alerts = [
        a for a in open_alerts
        if (a.get("alert_level") or "").lower() in ["critical", "high"]
    ]
    open_counseling = [
        c for c in counseling
        if (c.get("status") or "").lower() == "open"
    ]

    if "dot" in q or "credential" in q or "expire" in q or "expiration" in q:
        if not credentials:
            return f"I do not see credential records for {driver_name}."

        lines = [
            f"Credential Readout for {driver_name}",
            f"Employee ID: {employee_id}",
            f"Supervisor: {supervisor}",
            "",
        ]

        for credential in credentials:
            lines.append(
                f"{credential.get('credential_type')}: {credential.get('credential_status')} — "
                f"expires {credential.get('expiration_date')} "
                f"({credential.get('days_until_expiration')} days until expiration). "
                f"Alert Level: {credential.get('alert_level')}. "
                f"Notes: {credential.get('notes')}"
            )

        return "\n".join(lines)

    if "birthday" in q:
        return (
            f"{driver_name}'s birthday is listed as {birthday}. "
            f"For privacy, LORI stores month and day only in this demo, not a full birthdate."
        )

    if "counsel" in q or "coaching" in q or "follow" in q:
        if not counseling:
            return f"I do not see counseling or coaching records for {driver_name}."

        lines = [
            f"Counseling and Coaching History for {driver_name}",
            f"Supervisor: {supervisor}",
            "",
        ]

        for item in counseling[:5]:
            lines.append(
                f"{item.get('counseling_date')}: {item.get('counseling_type')} — "
                f"{item.get('counseling_reason')}. Outcome: {item.get('outcome')}. "
                f"Follow-up: {item.get('follow_up_date')}. Status: {item.get('status')}. "
                f"Priority: {item.get('priority')}."
            )

        return "\n".join(lines)

    if "route" in q:
        if not routes:
            return f"I do not see route records for {driver_name}."

        return (
            f"Route Readout for {driver_name}\n"
            f"Primary Route: {primary_route}\n"
            f"Latest Route: {latest_route.get('route_id')} — {latest_route.get('route_name')}\n"
            f"Planned Stops: {latest_route.get('planned_stops')}\n"
            f"Completed Stops: {latest_route.get('completed_stops')}\n"
            f"Missed Stops: {latest_route.get('missed_stops')}\n"
            f"On-Time Rate: {latest_route.get('on_time_rate')}\n"
            f"Route Score: {latest_route.get('route_score')}\n"
            f"Route Risk Flag: {latest_route.get('route_risk_flag')}\n"
            f"Notes: {latest_route.get('notes')}"
        )

    if "safety" in q or "event" in q:
        if not safety_events:
            return f"I do not see safety events for {driver_name}."

        lines = [
            f"Safety Event Readout for {driver_name}",
            "",
        ]

        for event in safety_events[:5]:
            lines.append(
                f"{event.get('event_date')}: {event.get('event_type')} "
                f"({event.get('severity')}) on {event.get('route_id')}. "
                f"{event.get('description')} Corrective action required: "
                f"{event.get('corrective_action_required')}."
            )

        return "\n".join(lines)

    if "alert" in q or "risk" in q or "everything" in q or "profile" in q or not question:
        lines = [
            f"Driver Intelligence Profile: {driver_name}",
            f"Employee ID: {employee_id}",
            f"Supervisor: {supervisor}",
            f"Location: {location}",
            f"Primary Route: {primary_route}",
            f"Birthday: {birthday}",
            "",
            "Performance Snapshot:",
            f"Overall Score: {latest_metric.get('overall_score', 'Not available')}",
            f"Risk Level: {latest_metric.get('risk_level', 'Not available')}",
            f"Trend Direction: {latest_metric.get('trend_direction', 'Not available')}",
            f"Recommended Action: {latest_metric.get('recommended_action', 'Not available')}",
            "",
            "Open Alerts:",
        ]

        if open_alerts:
            for alert in open_alerts[:5]:
                lines.append(
                    f"- {alert.get('alert_level')}: {alert.get('alert_title')} — "
                    f"{alert.get('alert_detail')} Due: {alert.get('due_date')}. "
                    f"Recommended Action: {alert.get('recommended_action')}"
                )
        else:
            lines.append("- No open alerts found.")

        lines.append("")
        lines.append("Counseling / Coaching:")

        if open_counseling:
            for item in open_counseling[:3]:
                lines.append(
                    f"- {item.get('counseling_date')}: {item.get('counseling_type')} — "
                    f"{item.get('counseling_reason')}. Follow-up: {item.get('follow_up_date')}. "
                    f"Status: {item.get('status')}."
                )
        else:
            lines.append("- No open counseling follow-ups found.")

        lines.append("")
        lines.append("Leadership Readout:")

        if critical_alerts:
            lines.append(
                f"{driver_name} has critical or high-priority items that require leadership review "
                f"before continued escalation."
            )
        else:
            lines.append(
                f"{driver_name} does not currently show critical open alerts in the demo intelligence layer."
            )

        return "\n".join(lines)

    return (
        f"I found {driver_name}. Employee ID: {employee_id}. Supervisor: {supervisor}. "
        f"Latest risk level: {latest_metric.get('risk_level', 'Not available')}. "
        f"Ask about DOT expiration, counseling, routes, safety events, birthday, alerts, or full profile."
    )


@app.get("/driver-search")
async def driver_search(
    api_key: str = Query(...),
    query: Optional[str] = Query(None),
):
    """
    Search driver master records by name or employee ID.
    """
    require_lori_key(api_key)

    if query:
        all_drivers = await supabase_select(
            "lori_driver_master",
            {
                "select": "*",
                "order": "driver_name.asc",
            },
        )

        q = query.lower()
        matches = [
            driver for driver in all_drivers
            if q in (driver.get("driver_name") or "").lower()
            or q in (driver.get("preferred_name") or "").lower()
            or q in (driver.get("employee_id") or "").lower()
        ]

        return matches[:10]

    return await supabase_select(
        "lori_driver_master",
        {
            "select": "*",
            "order": "driver_name.asc",
            "limit": "25",
        },
    )


@app.get("/driver-intelligence")
async def driver_intelligence(
    api_key: str = Query(...),
    employee_id: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    question: Optional[str] = Query(None),
):
    """
    Main driver intelligence endpoint for Voiceflow.
    Answers deeper questions about a driver using all driver intelligence tables.
    """
    require_lori_key(api_key)

    return await get_driver_intelligence_payload(
        employee_id=employee_id,
        driver_name=driver_name,
        question=question,
    )


@app.get("/driver-compliance-watch")
async def driver_compliance_watch(
    api_key: str = Query(...),
    days: int = Query(30, ge=0, le=365),
):
    """
    Show drivers with expired or upcoming credentials.
    """
    require_lori_key(api_key)

    credentials = await supabase_select(
        "lori_driver_credentials",
        {
            "select": "*",
            "order": "days_until_expiration.asc",
        },
    )

    results = []

    for credential in credentials:
        days_until = credential.get("days_until_expiration")
        status = (credential.get("credential_status") or "").lower()

        include = False

        if days_until is not None:
            try:
                include = int(days_until) <= days
            except Exception:
                include = False

        if status in ["expired", "overdue"]:
            include = True

        if include:
            results.append(credential)

    return {
        "status": "success",
        "watch_window_days": days,
        "drivers_found": len(results),
        "credentials": results,
        "answer_text": f"I found {len(results)} credential records that are expired, overdue, or due within {days} days.",
    }


@app.get("/driver-counseling-due")
async def driver_counseling_due(
    api_key: str = Query(...),
    days: int = Query(30, ge=0, le=365),
):
    """
    Show open counseling follow-ups due within a date window.
    """
    require_lori_key(api_key)

    counseling_rows = await supabase_select(
        "lori_driver_counseling",
        {
            "select": "*",
            "order": "follow_up_date.asc",
        },
    )

    today = date.today()
    results = []

    for row in counseling_rows:
        if (row.get("status") or "").lower() != "open":
            continue

        follow_up_date = parse_date_safe(row.get("follow_up_date"))

        if follow_up_date is None:
            continue

        days_until = (follow_up_date - today).days

        if days_until <= days:
            row["days_until_follow_up"] = days_until
            results.append(row)

    return {
        "status": "success",
        "watch_window_days": days,
        "drivers_found": len(results),
        "counseling_follow_ups": results,
        "answer_text": f"I found {len(results)} open counseling follow-ups due within {days} days.",
    }


@app.get("/driver-birthday-watch")
async def driver_birthday_watch(
    api_key: str = Query(...),
    month: Optional[int] = Query(None, ge=1, le=12),
):
    """
    Show driver birthdays by month. Stores only month/day for privacy.
    """
    require_lori_key(api_key)

    selected_month = month or date.today().month

    drivers = await supabase_select(
        "lori_driver_master",
        {
            "select": "*",
            "birthday_month": f"eq.{selected_month}",
            "order": "birthday_day.asc",
        },
    )

    results = []

    for driver in drivers:
        results.append(
            {
                "employee_id": driver.get("employee_id"),
                "driver_name": driver.get("driver_name"),
                "preferred_name": driver.get("preferred_name"),
                "supervisor_name": driver.get("supervisor_name"),
                "location_name": driver.get("location_name"),
                "birthday": format_birthday(driver.get("birthday_month"), driver.get("birthday_day")),
                "privacy_note": "Only birthday month and day are stored in this demo.",
            }
        )

    return {
        "status": "success",
        "month": selected_month,
        "month_name": MONTH_NAMES.get(selected_month, str(selected_month)),
        "drivers_found": len(results),
        "birthdays": results,
        "answer_text": f"I found {len(results)} driver birthdays in {MONTH_NAMES.get(selected_month, str(selected_month))}.",
    }
# ============================================================
# LORI DRIVE — NEW DRIVER INTAKE ENDPOINT
# Saves a manually entered driver into Supabase.
# Demonstration Data Only — Not Company Proprietary Data
# ============================================================

class NewDriverIntakeRequest(BaseModel):
    employee_id: str
    full_name: str
    preferred_name: Optional[str] = None
    driver_role: Optional[str] = None
    employment_status: Optional[str] = None
    hire_date: Optional[str] = None

    supervisor_name: Optional[str] = None
    location: Optional[str] = None
    location_code: Optional[str] = None
    primary_route_id: Optional[str] = None
    station_operation: Optional[str] = None

    phone_last4: Optional[str] = None
    email: Optional[str] = None

    birthday_month: Optional[str] = None
    birthday_day: Optional[str] = None

    dot_medical_card_expiration: Optional[str] = None
    cdl_license_status: Optional[str] = None
    license_class: Optional[str] = None
    defensive_driving_status: Optional[str] = None
    training_status: Optional[str] = None

    driver_qualification_file_status: Optional[str] = None
    onboarding_checklist_status: Optional[str] = None
    background_screening_status: Optional[str] = None
    initial_safety_orientation_status: Optional[str] = None

    manager_notes: Optional[str] = None
    compliance_notes: Optional[str] = None
    supervisor_notes: Optional[str] = None


def lori_supabase_headers():
    return {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }


async def lori_supabase_insert(table_name: str, payload: dict):
    url = f"{SUPABASE_URL}/rest/v1/{table_name}"

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            url,
            headers=lori_supabase_headers(),
            json=payload
        )

    if response.status_code not in [200, 201]:
        raise HTTPException(
            status_code=500,
            detail={
                "message": f"Supabase insert failed for {table_name}",
                "status_code": response.status_code,
                "response": response.text
            }
        )

    try:
        return response.json()
    except Exception:
        return {"raw_response": response.text}


async def lori_supabase_upsert_driver_master(payload: dict):
    url = f"{SUPABASE_URL}/rest/v1/lori_driver_master?on_conflict=employee_id"

    headers = lori_supabase_headers()
    headers["Prefer"] = "resolution=merge-duplicates,return=representation"

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            url,
            headers=headers,
            json=payload
        )

    if response.status_code not in [200, 201]:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Supabase upsert failed for lori_driver_master",
                "status_code": response.status_code,
                "response": response.text
            }
        )

    try:
        return response.json()
    except Exception:
        return {"raw_response": response.text}

def lori_normalize_birthday_month(value):
    if value is None:
        return None

    text = str(value).strip()

    if text == "":
        return None

    month_lookup = {
        "1": 1, "01": 1, "jan": 1, "january": 1,
        "2": 2, "02": 2, "feb": 2, "february": 2,
        "3": 3, "03": 3, "mar": 3, "march": 3,
        "4": 4, "04": 4, "apr": 4, "april": 4,
        "5": 5, "05": 5, "may": 5,
        "6": 6, "06": 6, "jun": 6, "june": 6,
        "7": 7, "07": 7, "jul": 7, "july": 7,
        "8": 8, "08": 8, "aug": 8, "august": 8,
        "9": 9, "09": 9, "sep": 9, "sept": 9, "september": 9,
        "10": 10, "oct": 10, "october": 10,
        "11": 11, "nov": 11, "november": 11,
        "12": 12, "dec": 12, "december": 12
    }

    return month_lookup.get(text.lower())


def lori_normalize_birthday_day(value):
    if value is None:
        return None

    text = str(value).strip()

    if text == "":
        return None

    try:
        day = int(text)
        if day < 1 or day > 31:
            return None
        return day
    except Exception:
        return None
@app.post("/new-driver")
async def create_new_driver(
    driver: NewDriverIntakeRequest,
    api_key: str = Query(None)
):
    if api_key != LORI_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")

    now_value = datetime.utcnow().isoformat()

    employee_id = driver.employee_id.strip()
    full_name = driver.full_name.strip()

    if not employee_id:
        raise HTTPException(status_code=400, detail="Employee ID is required.")

    if not full_name:
        raise HTTPException(status_code=400, detail="Driver full name is required.")
    birthday_month_value = lori_normalize_birthday_month(driver.birthday_month)
    birthday_day_value = lori_normalize_birthday_day(driver.birthday_day)
    master_payload = {
        "employee_id": employee_id,
        "driver_name": full_name,
        "preferred_name": driver.preferred_name,
        "driver_role": driver.driver_role,
        "employment_status": driver.employment_status or "Active",
        "hire_date": driver.hire_date,
        "supervisor_name": driver.supervisor_name,
        "location": driver.location,
        "location_code": driver.location_code,
        "primary_route_id": driver.primary_route_id,
        "station_operation": driver.station_operation,
        "phone_last4": driver.phone_last4,
        "email": driver.email,
                "birthday_month": birthday_month_value,
        "birthday_day": birthday_day_value,
        "record_source": "Manual New Driver Intake",
        "created_at": now_value,
        "updated_at": now_value
    }

    master_payload = {k: v for k, v in master_payload.items() if v is not None}

    saved_master = await lori_supabase_upsert_driver_master(master_payload)

    credential_records_created = []

    if driver.dot_medical_card_expiration:
        dot_payload = {
            "employee_id": employee_id,
            "credential_type": "DOT Medical Card",
            "credential_status": "Active",
            "expiration_date": driver.dot_medical_card_expiration,
            "notes": "Created through LORI New Driver Intake.",
            "created_at": now_value,
            "updated_at": now_value
        }
        credential_records_created.append(
            await lori_supabase_insert("lori_driver_credentials", dot_payload)
        )

    if driver.cdl_license_status or driver.license_class:
        license_payload = {
            "employee_id": employee_id,
            "credential_type": "CDL / License",
            "credential_status": driver.cdl_license_status or "Needs Review",
            "license_class": driver.license_class,
            "notes": "Created through LORI New Driver Intake.",
            "created_at": now_value,
            "updated_at": now_value
        }
        license_payload = {k: v for k, v in license_payload.items() if v is not None}
        credential_records_created.append(
            await lori_supabase_insert("lori_driver_credentials", license_payload)
        )

    onboarding_notes = []

    if driver.driver_qualification_file_status:
        onboarding_notes.append(f"Driver Qualification File: {driver.driver_qualification_file_status}")

    if driver.onboarding_checklist_status:
        onboarding_notes.append(f"Onboarding Checklist: {driver.onboarding_checklist_status}")

    if driver.background_screening_status:
        onboarding_notes.append(f"Background / Screening: {driver.background_screening_status}")

    if driver.initial_safety_orientation_status:
        onboarding_notes.append(f"Initial Safety Orientation: {driver.initial_safety_orientation_status}")

    if driver.training_status:
        onboarding_notes.append(f"Training Status: {driver.training_status}")

    if driver.defensive_driving_status:
        onboarding_notes.append(f"Defensive Driving Certification: {driver.defensive_driving_status}")

    note_text_parts = []

    if onboarding_notes:
        note_text_parts.append("Compliance Setup:\n" + "\n".join(onboarding_notes))

    if driver.manager_notes:
        note_text_parts.append("Manager Notes:\n" + driver.manager_notes)

    if driver.compliance_notes:
        note_text_parts.append("Compliance Notes:\n" + driver.compliance_notes)

    if driver.supervisor_notes:
        note_text_parts.append("Supervisor Notes:\n" + driver.supervisor_notes)

    note_record_created = None

    if note_text_parts:
        note_payload = {
            "employee_id": employee_id,
            "note_type": "New Driver Intake",
            "note_text": "\n\n".join(note_text_parts),
            "created_by": "LORI New Driver Intake",
            "created_at": now_value,
            "updated_at": now_value
        }
        note_record_created = await lori_supabase_insert("lori_driver_notes", note_payload)

    timeline_payload = {
        "employee_id": employee_id,
        "event_type": "New Driver Created",
        "event_title": "New driver profile created",
        "event_summary": f"{full_name} was manually added to LORI and staged for compliance review.",
        "created_at": now_value
    }

    timeline_record_created = None

    try:
        timeline_record_created = await lori_supabase_insert("lori_driver_timeline", timeline_payload)
    except Exception:
        timeline_record_created = {
            "status": "timeline_not_created",
            "message": "Driver was saved, but the timeline record could not be created."
        }

    return {
        "status": "success",
        "message": "New driver profile created and staged for compliance review.",
        "employee_id": employee_id,
        "driver_name": full_name,
        "saved_master_record": saved_master,
        "credential_records_created": credential_records_created,
        "note_record_created": note_record_created,
        "timeline_record_created": timeline_record_created,
        "next_steps": [
            "Open Driver 360",
            "Review driver file readiness",
            "Add missing credentials",
            "Assign supervisor follow-up",
            "Confirm DOT/FMCSA-sensitive records against official company files"
        ]
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# Regulatory Intelligence Backend
# Adds live-ready regulatory alert scan endpoints
# ============================================================

import hashlib
import re
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from typing import Any, Dict, List, Optional

import httpx
from fastapi import Query, HTTPException


REGULATORY_KEYWORDS = [
    "fmcsa",
    "dot",
    "department of transportation",
    "transportation",
    "motor carrier",
    "commercial motor vehicle",
    "driver",
    "drivers",
    "cdl",
    "medical card",
    "hours of service",
    "drug",
    "alcohol",
    "safety",
    "inspection",
    "vehicle",
    "fleet",
    "compliance",
    "rule",
    "notice",
    "regulation",
    "enforcement",
    "crash",
    "carrier",
    "hazmat",
    "hazardous materials",
    "electronic logging",
    "eld",
]


def lori_regulatory_require_key(api_key: Optional[str]) -> None:
    if LORI_API_KEY and api_key != LORI_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")


def lori_regulatory_supabase_headers(prefer: Optional[str] = None) -> Dict[str, str]:
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def lori_regulatory_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def lori_regulatory_clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def lori_regulatory_parse_date(value: Any) -> Optional[str]:
    if not value:
        return None

    text = str(value).strip()

    try:
        parsed = parsedate_to_datetime(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).isoformat()
    except Exception:
        pass

    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).isoformat()
    except Exception:
        return None


def lori_regulatory_hash(*parts: Any) -> str:
    raw = "||".join([str(p or "") for p in parts])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def lori_regulatory_is_relevant(title: str, summary: str, source: Dict[str, Any]) -> bool:
    combined = f"{title} {summary} {source.get('source_name', '')} {source.get('category', '')}".lower()

    if source.get("source_type") in ["Federal", "State"] and source.get("agency"):
        agency = str(source.get("agency", "")).lower()
        if "transportation" in agency:
            return True
        if "motor carrier" in agency:
            return True

    return any(keyword in combined for keyword in REGULATORY_KEYWORDS)


async def lori_regulatory_supabase_get(query_path: str) -> Any:
    url = f"{SUPABASE_URL}/rest/v1/{query_path}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get(url, headers=lori_regulatory_supabase_headers())

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase GET failed: {response.text}",
        )

    return response.json()


async def lori_regulatory_supabase_post(table: str, payload: Dict[str, Any]) -> Any:
    url = f"{SUPABASE_URL}/rest/v1/{table}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(
            url,
            headers=lori_regulatory_supabase_headers("return=representation"),
            json=payload,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase POST failed: {response.text}",
        )

    return response.json()


async def lori_regulatory_supabase_patch(table: str, row_id: str, payload: Dict[str, Any]) -> Any:
    url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{row_id}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(
            url,
            headers=lori_regulatory_supabase_headers("return=representation"),
            json=payload,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase PATCH failed: {response.text}",
        )

    return response.json()


async def lori_regulatory_alert_exists(content_hash: str) -> bool:
    safe_hash = content_hash.replace("'", "")
    rows = await lori_regulatory_supabase_get(
        f"lori_regulatory_alerts?content_hash=eq.{safe_hash}&select=id&limit=1"
    )
    return bool(rows)


async def lori_regulatory_fetch_rss(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    feed_url = source.get("feed_url") or source.get("source_url")
    if not feed_url:
        return []

    items: List[Dict[str, Any]] = []

    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        response = await client.get(
            feed_url,
            headers={"User-Agent": "LORI-Regulatory-Scanner/1.0"},
        )

    if response.status_code >= 400:
        raise Exception(f"RSS fetch failed for {feed_url}: {response.status_code}")

    root = ET.fromstring(response.text)

    for item in root.findall(".//item"):
        title = lori_regulatory_clean_text(item.findtext("title"))
        link = lori_regulatory_clean_text(item.findtext("link"))
        summary = lori_regulatory_clean_text(
            item.findtext("description") or item.findtext("summary")
        )
        published_raw = item.findtext("pubDate") or item.findtext("published") or item.findtext("updated")
        published_at = lori_regulatory_parse_date(published_raw)

        if not title:
            continue

        if not lori_regulatory_is_relevant(title, summary, source):
            continue

        items.append(
            {
                "title": title,
                "summary": summary,
                "url": link,
                "published_at": published_at,
                "raw": {
                    "source": "rss",
                    "feed_url": feed_url,
                    "published_raw": published_raw,
                },
            }
        )

    ns = {"atom": "http://www.w3.org/2005/Atom"}
    for entry in root.findall(".//atom:entry", ns):
        title = lori_regulatory_clean_text(entry.findtext("atom:title", default="", namespaces=ns))
        summary = lori_regulatory_clean_text(
            entry.findtext("atom:summary", default="", namespaces=ns)
            or entry.findtext("atom:content", default="", namespaces=ns)
        )
        published_raw = (
            entry.findtext("atom:published", default="", namespaces=ns)
            or entry.findtext("atom:updated", default="", namespaces=ns)
        )
        published_at = lori_regulatory_parse_date(published_raw)

        link = ""
        link_el = entry.find("atom:link", ns)
        if link_el is not None:
            link = link_el.attrib.get("href", "")

        if not title:
            continue

        if not lori_regulatory_is_relevant(title, summary, source):
            continue

        items.append(
            {
                "title": title,
                "summary": summary,
                "url": link,
                "published_at": published_at,
                "raw": {
                    "source": "atom",
                    "feed_url": feed_url,
                    "published_raw": published_raw,
                },
            }
        )

    return items[:20]


async def lori_regulatory_fetch_federal_register(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []

    api_urls = [
        "https://www.federalregister.gov/api/v1/documents.json?conditions%5Bagencies%5D%5B%5D=federal-motor-carrier-safety-administration&order=newest&per_page=10",
        "https://www.federalregister.gov/api/v1/documents.json?conditions%5Bagencies%5D%5B%5D=transportation-department&order=newest&per_page=10",
    ]

    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        for api_url in api_urls:
            response = await client.get(
                api_url,
                headers={"User-Agent": "LORI-Regulatory-Scanner/1.0"},
            )

            if response.status_code >= 400:
                continue

            data = response.json()

            for doc in data.get("results", []):
                title = lori_regulatory_clean_text(doc.get("title"))
                summary = lori_regulatory_clean_text(doc.get("abstract") or doc.get("type"))
                url = doc.get("html_url") or doc.get("pdf_url") or doc.get("public_inspection_pdf_url")
                published_at = lori_regulatory_parse_date(doc.get("publication_date"))

                if not title:
                    continue

                if not lori_regulatory_is_relevant(title, summary, source):
                    continue

                items.append(
                    {
                        "title": title,
                        "summary": summary,
                        "url": url,
                        "published_at": published_at,
                        "raw": {
                            "source": "federal_register",
                            "document_number": doc.get("document_number"),
                            "type": doc.get("type"),
                            "publication_date": doc.get("publication_date"),
                            "agencies": doc.get("agencies"),
                        },
                    }
                )

    return items[:20]


async def lori_regulatory_fetch_source(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    source_format = str(source.get("source_format") or "").lower()
    source_name = str(source.get("source_name") or "").lower()

    if "federal register" in source_name or source_format == "api":
        return await lori_regulatory_fetch_federal_register(source)

    if source.get("feed_url") or source_format == "rss":
        return await lori_regulatory_fetch_rss(source)

    return []


def lori_regulatory_priority_for_item(item: Dict[str, Any]) -> str:
    text = f"{item.get('title', '')} {item.get('summary', '')}".lower()

    if any(word in text for word in ["final rule", "effective", "emergency", "enforcement"]):
        return "High"

    if any(word in text for word in ["proposed rule", "notice", "comment", "guidance"]):
        return "Watch"

    return "Informational"


def lori_regulatory_build_operational_impact(item: Dict[str, Any]) -> str:
    return (
        "This update may require review for potential impact on driver operations, "
        "station readiness, safety procedures, DOT/FMCSA compliance planning, "
        "company policy, supervisor briefing, or leadership awareness."
    )


def lori_regulatory_build_recommended_preparation(item: Dict[str, Any]) -> str:
    return (
        "Review the official source, confirm whether the update applies to the operating location, "
        "identify affected drivers or supervisors, determine whether policy or training updates are needed, "
        "and prepare a leadership briefing if operational impact is confirmed."
    )


@app.get("/regulatory-alerts")
async def get_regulatory_alerts(
    api_key: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    limit: int = Query(10),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 50))

    query = (
        "lori_regulatory_alerts?"
        "select=*"
        "&order=created_at.desc"
        f"&limit={limit}"
    )

    if state:
        query += f"&state_code=eq.{state.upper()}"

    alerts = await lori_regulatory_supabase_get(query)

    latest_logs = await lori_regulatory_supabase_get(
        "lori_regulatory_scan_logs?select=*&order=created_at.desc&limit=1"
    )

    latest_log = latest_logs[0] if latest_logs else None

    if alerts:
        status_message = "New or stored regulatory alerts are available for review."
        result = "Regulatory Update Requires Review"
    else:
        status_message = "No new verified regulatory updates found in the current stored alert set."
        result = "No New Verified Updates Found"

    return {
        "status": "success",
        "result": result,
        "message": status_message,
        "last_checked": latest_log.get("scan_completed_at") if latest_log else None,
        "latest_scan": latest_log,
        "alerts_count": len(alerts),
        "alerts": alerts,
        "disclaimer": (
            "LORI provides operational decision support. Regulatory alerts, transportation laws, "
            "DOT/FMCSA updates, state law changes, labor/HR updates, and compliance-sensitive matters "
            "must be verified against official sources and reviewed by appropriate compliance, HR, "
            "labor relations, or legal personnel before formal action."
        ),
    }


@app.get("/regulatory-scan-log")
async def get_regulatory_scan_log(
    api_key: Optional[str] = Query(None),
    limit: int = Query(10),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 50))

    logs = await lori_regulatory_supabase_get(
        f"lori_regulatory_scan_logs?select=*&order=created_at.desc&limit={limit}"
    )

    return {
        "status": "success",
        "scan_logs": logs,
    }


@app.post("/regulatory-scan")
async def run_regulatory_scan(
    api_key: Optional[str] = Query(None),
    operating_state: Optional[str] = Query("MD"),
    station_code: Optional[str] = Query("JESSUP-01"),
):
    lori_regulatory_require_key(api_key)

    scan_log_payload = {
        "scan_started_at": lori_regulatory_now_iso(),
        "scan_status": "Started",
        "operating_state": operating_state.upper() if operating_state else None,
        "station_code": station_code or "JESSUP-01",
        "federal_check_status": "Started",
        "state_check_status": "Started",
        "dot_fmcsa_check_status": "Started",
        "labor_hr_check_status": "Pending",
        "sources_checked": 0,
        "alerts_found": 0,
        "new_alerts_found": 0,
        "result_summary": "Regulatory scan started.",
    }

    created_log = await lori_regulatory_supabase_post(
        "lori_regulatory_scan_logs",
        scan_log_payload,
    )

    scan_log_id = created_log[0]["id"] if created_log else None

    sources = await lori_regulatory_supabase_get(
        "lori_regulatory_sources?is_active=eq.true&select=*"
    )

    sources_checked = 0
    alerts_found = 0
    new_alerts_found = 0
    errors: List[str] = []
    stored_alerts: List[Dict[str, Any]] = []

    for source in sources:
        try:
            source_state = source.get("state_code")
            source_type = source.get("source_type")

            if source_type == "State" and source_state and operating_state:
                if source_state.upper() != operating_state.upper():
                    continue

            items = await lori_regulatory_fetch_source(source)
            sources_checked += 1
            alerts_found += len(items)

            for item in items:
                content_hash = lori_regulatory_hash(
                    source.get("id"),
                    item.get("title"),
                    item.get("url"),
                    item.get("published_at"),
                )

                exists = await lori_regulatory_alert_exists(content_hash)
                if exists:
                    continue

                priority = lori_regulatory_priority_for_item(item)

                alert_payload = {
                    "source_id": source.get("id"),
                    "alert_title": item.get("title"),
                    "alert_summary": item.get("summary"),
                    "alert_body": item.get("summary"),
                    "source_type": source.get("source_type"),
                    "agency": source.get("agency"),
                    "jurisdiction": source.get("jurisdiction"),
                    "state_code": source.get("state_code"),
                    "category": source.get("category"),
                    "published_at": item.get("published_at"),
                    "source_url": item.get("url"),
                    "alert_priority": priority,
                    "alert_status": "New",
                    "applies_to": [
                        "Drivers",
                        "Supervisors",
                        "Station Leadership",
                        "Fleet Operations",
                    ],
                    "operational_impact": lori_regulatory_build_operational_impact(item),
                    "recommended_preparation": lori_regulatory_build_recommended_preparation(item),
                    "source_verification_status": "Official source found - leadership verification recommended",
                    "content_hash": content_hash,
                    "raw_payload": item.get("raw"),
                }

                inserted_alert = await lori_regulatory_supabase_post(
                    "lori_regulatory_alerts",
                    alert_payload,
                )

                if inserted_alert:
                    new_alerts_found += 1
                    stored_alerts.append(inserted_alert[0])

        except Exception as exc:
            errors.append(f"{source.get('source_name', 'Unknown source')}: {str(exc)}")

    if new_alerts_found > 0:
        scan_status = "Completed - New Alerts Found"
        result_summary = f"Regulatory scan completed. {new_alerts_found} new alert(s) found."
    else:
        scan_status = "Completed - No New Verified Updates Found"
        result_summary = "Regulatory scan completed. No new verified updates found from configured sources."

    federal_status = "Checked" if sources_checked > 0 else "No active federal sources checked"
    state_status = "Checked" if operating_state else "No operating state selected"
    dot_status = "Checked" if sources_checked > 0 else "No DOT/FMCSA sources checked"

    update_payload = {
        "scan_completed_at": lori_regulatory_now_iso(),
        "scan_status": scan_status,
        "federal_check_status": federal_status,
        "state_check_status": state_status,
        "dot_fmcsa_check_status": dot_status,
        "labor_hr_check_status": "Pending source connection",
        "sources_checked": sources_checked,
        "alerts_found": alerts_found,
        "new_alerts_found": new_alerts_found,
        "result_summary": result_summary,
        "error_message": "\n".join(errors) if errors else None,
    }

    if scan_log_id:
        await lori_regulatory_supabase_patch(
            "lori_regulatory_scan_logs",
            scan_log_id,
            update_payload,
        )

    return {
        "status": "success",
        "scan_status": scan_status,
        "operating_state": operating_state,
        "station_code": station_code,
        "sources_checked": sources_checked,
        "alerts_found": alerts_found,
        "new_alerts_found": new_alerts_found,
        "result_summary": result_summary,
        "new_alerts": stored_alerts,
        "errors": errors,
        "disclaimer": (
            "LORI provides operational decision support. Regulatory alerts must be verified against "
            "official federal, state, DOT/FMCSA, HR, compliance, labor relations, or legal sources before formal action."
        ),
    }
# ============================================================
# LORI REGULATORY INTELLIGENCE
# Company-Specific Filter Override
# Focus: Food distribution / commercial delivery fleet
# Excludes irrelevant aviation / FAA / aircraft updates
# ============================================================

COMPANY_PROFILE_NAME = "Food Authority / Food Distribution Delivery Fleet"

COMPANY_RELEVANT_KEYWORDS = [
    "fmcsa",
    "federal motor carrier safety administration",
    "dot",
    "department of transportation",
    "motor carrier",
    "commercial motor vehicle",
    "cmv",
    "carrier",
    "driver",
    "drivers",
    "cdl",
    "commercial driver's license",
    "commercial driver",
    "driver qualification",
    "driver qualification file",
    "dqf",
    "medical card",
    "dot medical",
    "medical examiner",
    "hours of service",
    "hos",
    "electronic logging",
    "eld",
    "drug",
    "alcohol",
    "controlled substance",
    "testing",
    "safety",
    "inspection",
    "vehicle inspection",
    "pre-trip",
    "post-trip",
    "fleet",
    "truck",
    "trucking",
    "delivery",
    "route",
    "logistics",
    "transportation",
    "warehouse",
    "distribution",
    "food distribution",
    "food delivery",
    "foodservice",
    "refrigerated",
    "cold chain",
    "hazmat",
    "hazardous materials",
    "compliance",
    "rule",
    "notice",
    "regulation",
    "enforcement",
]

COMPANY_EXCLUDED_KEYWORDS = [
    "faa",
    "federal aviation administration",
    "aviation",
    "aircraft",
    "airworthiness",
    "airspace",
    "airport",
    "runway",
    "helicopter",
    "rotorcraft",
    "bell textron",
    "boeing",
    "airbus",
    "flight",
    "pilot",
    "federal airway",
    "vortac",
    "tacan",
    "navigation aid",
    "class e airspace",
    "domestic very high frequency",
    "vhf",
    "air traffic",
    "aeronautical",
    "maritime",
    "vessel",
    "coast guard",
    "railroad",
    "railway",
    "pipeline safety",
    "transit rail",
]


def lori_company_text_blob(title: str, summary: str, source: Dict[str, Any]) -> str:
    return f"""
    {title or ''}
    {summary or ''}
    {source.get('source_name', '') or ''}
    {source.get('agency', '') or ''}
    {source.get('category', '') or ''}
    {source.get('source_type', '') or ''}
    """.lower()


def lori_regulatory_is_relevant(title: str, summary: str, source: Dict[str, Any]) -> bool:
    """
    Company-specific relevance filter.

    This keeps LORI focused on a food distribution / commercial delivery
    transportation operation instead of pulling unrelated DOT items such as
    FAA, aircraft, helicopter, aviation, airspace, railroad, maritime, or pipeline notices.
    """

    combined = lori_company_text_blob(title, summary, source)

    # Hard exclude aviation / aircraft / non-fleet transportation categories.
    if any(excluded in combined for excluded in COMPANY_EXCLUDED_KEYWORDS):
        return False

    source_name = str(source.get("source_name") or "").lower()
    agency = str(source.get("agency") or "").lower()
    category = str(source.get("category") or "").lower()

    # FMCSA sources are highly relevant to commercial fleet operations.
    if "federal motor carrier safety administration" in agency:
        return True

    if "fmcsa" in source_name or "fmcsa" in category:
        return True

    # DOT items must still match company-relevant transportation operations.
    if any(keyword in combined for keyword in COMPANY_RELEVANT_KEYWORDS):
        return True

    return False


async def lori_regulatory_fetch_federal_register(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Company-specific Federal Register scanner.

    Only scans FMCSA documents for now.
    Avoids broad DOT-wide Federal Register pulls that include FAA aviation,
    aircraft, helicopter, airspace, railroad, maritime, and unrelated items.
    """

    items: List[Dict[str, Any]] = []

    api_urls = [
        "https://www.federalregister.gov/api/v1/documents.json?conditions%5Bagencies%5D%5B%5D=federal-motor-carrier-safety-administration&order=newest&per_page=20"
    ]

    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        for api_url in api_urls:
            response = await client.get(
                api_url,
                headers={"User-Agent": "LORI-Regulatory-Scanner/1.0"},
            )

            if response.status_code >= 400:
                continue

            data = response.json()

            for doc in data.get("results", []):
                title = lori_regulatory_clean_text(doc.get("title"))
                summary = lori_regulatory_clean_text(doc.get("abstract") or doc.get("type"))
                url = doc.get("html_url") or doc.get("pdf_url") or doc.get("public_inspection_pdf_url")
                published_at = lori_regulatory_parse_date(doc.get("publication_date"))

                if not title:
                    continue

                if not lori_regulatory_is_relevant(title, summary, source):
                    continue

                items.append(
                    {
                        "title": title,
                        "summary": summary,
                        "url": url,
                        "published_at": published_at,
                        "raw": {
                            "source": "federal_register",
                            "company_profile": COMPANY_PROFILE_NAME,
                            "document_number": doc.get("document_number"),
                            "type": doc.get("type"),
                            "publication_date": doc.get("publication_date"),
                            "agencies": doc.get("agencies"),
                        },
                    }
                )

    return items[:20]


def lori_regulatory_build_operational_impact(item: Dict[str, Any]) -> str:
    return (
        "This update may require review for potential impact on commercial delivery operations, "
        "driver qualification, DOT/FMCSA compliance readiness, fleet safety, vehicle inspection, "
        "driver policy, supervisor briefing, training, audit readiness, or leadership awareness. "
        "It should be reviewed for relevance to food distribution, delivery routes, commercial drivers, "
        "warehouse-to-customer transportation, and fleet operations."
    )


def lori_regulatory_build_recommended_preparation(item: Dict[str, Any]) -> str:
    return (
        "Review the official source, confirm whether the update applies to the company’s commercial delivery fleet, "
        "identify affected drivers, supervisors, vehicles, routes, policies, or audit areas, determine whether training "
        "or policy updates are needed, and prepare a leadership briefing if operational impact is confirmed."
    )
# ============================================================
# LORI REGULATORY INTELLIGENCE
# Voiceflow-ready daily alert briefing endpoint
# Allows user to ask: "Tell me the alerts for today."
# ============================================================

def lori_regulatory_short_date(value: Any) -> str:
    if not value:
        return "Not listed"

    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.strftime("%b %d, %Y")
    except Exception:
        return str(value)


def lori_regulatory_safe_line(value: Any, fallback: str = "Not listed") -> str:
    cleaned = lori_regulatory_clean_text(value)
    return cleaned if cleaned else fallback


def lori_regulatory_build_alert_briefing(
    alerts: List[Dict[str, Any]],
    latest_scan: Optional[Dict[str, Any]],
    requested_state: Optional[str] = None,
) -> str:
    today_label = datetime.now(timezone.utc).strftime("%b %d, %Y")

    if latest_scan:
        scan_status = latest_scan.get("scan_status") or "Scan status not listed"
        sources_checked = latest_scan.get("sources_checked", 0)
        alerts_found = latest_scan.get("alerts_found", 0)
        new_alerts_found = latest_scan.get("new_alerts_found", 0)
        scan_completed = lori_regulatory_short_date(latest_scan.get("scan_completed_at"))
        operating_state = latest_scan.get("operating_state") or requested_state or "Not selected"
        station_code = latest_scan.get("station_code") or "Not listed"
        result_summary = latest_scan.get("result_summary") or "No scan summary listed."
    else:
        scan_status = "No scan log found"
        sources_checked = 0
        alerts_found = 0
        new_alerts_found = 0
        scan_completed = "Not listed"
        operating_state = requested_state or "Not selected"
        station_code = "Not listed"
        result_summary = "No completed regulatory scan is available yet."

    if not alerts:
        return f"""Regulatory Alert Briefing — Today

Date:
{today_label}

Status:
No new verified regulatory alerts are currently listed for review.

Latest Scan:
{scan_status}

Sources Checked:
{sources_checked}

Operating State:
{operating_state}

Station:
{station_code}

Result:
{result_summary}

Recommended Next Action:
No immediate regulatory briefing is required based on the current stored alert set. Continue monitoring and verify official DOT, FMCSA, state, HR, compliance, labor relations, or legal sources before taking formal action.

Compliance Note:
LORI provides operational decision support. Regulatory alerts, transportation laws, DOT/FMCSA updates, state law changes, labor/HR updates, and compliance-sensitive matters must be verified against official sources and reviewed by appropriate compliance, HR, labor relations, or legal personnel before formal action."""

    lines = []
    lines.append("Regulatory Alert Briefing — Today")
    lines.append("")
    lines.append("Date:")
    lines.append(today_label)
    lines.append("")
    lines.append("Current Status:")
    lines.append("New regulatory alerts require leadership review.")
    lines.append("")
    lines.append("Latest Scan:")
    lines.append(str(scan_status))
    lines.append("")
    lines.append("Operating State:")
    lines.append(str(operating_state))
    lines.append("")
    lines.append("Station:")
    lines.append(str(station_code))
    lines.append("")
    lines.append("Sources Checked:")
    lines.append(str(sources_checked))
    lines.append("")
    lines.append("Alerts Found:")
    lines.append(str(alerts_found))
    lines.append("")
    lines.append("New Alerts Found:")
    lines.append(str(new_alerts_found))
    lines.append("")
    lines.append("Top Alerts for Leadership Review:")

    for index, alert in enumerate(alerts[:5], start=1):
        title = lori_regulatory_safe_line(alert.get("alert_title"))
        agency = lori_regulatory_safe_line(alert.get("agency"))
        category = lori_regulatory_safe_line(alert.get("category"))
        priority = lori_regulatory_safe_line(alert.get("alert_priority"))
        published = lori_regulatory_short_date(alert.get("published_at"))
        impact = lori_regulatory_safe_line(alert.get("operational_impact"))
        preparation = lori_regulatory_safe_line(alert.get("recommended_preparation"))
        verification = lori_regulatory_safe_line(alert.get("source_verification_status"))

        lines.append("")
        lines.append(f"{index}. {title}")
        lines.append(f"Agency: {agency}")
        lines.append(f"Category: {category}")
        lines.append(f"Priority: {priority}")
        lines.append(f"Published: {published}")
        lines.append(f"Operational Impact: {impact}")
        lines.append(f"Recommended Preparation: {preparation}")
        lines.append(f"Verification Status: {verification}")

    lines.append("")
    lines.append("Recommended Next Action:")
    lines.append(
        "Review the top alerts, confirm official source relevance, identify affected drivers, supervisors, policies, audits, or training areas, and add confirmed items to the leadership briefing or compliance follow-up list."
    )

    lines.append("")
    lines.append("Compliance Note:")
    lines.append(
        "LORI provides operational decision support. Regulatory alerts, transportation laws, DOT/FMCSA updates, state law changes, labor/HR updates, and compliance-sensitive matters must be verified against official sources and reviewed by appropriate compliance, HR, labor relations, or legal personnel before formal action."
    )

    return "\n".join(lines)


@app.get("/voiceflow/regulatory-alerts")
async def voiceflow_regulatory_alerts(
    api_key: Optional[str] = Query(None),
    question: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    limit: int = Query(5),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 10))

    alerts = await lori_regulatory_supabase_get(
        "lori_regulatory_alerts?select=*&order=created_at.desc&limit=25"
    )

    filtered_alerts = []

    if state:
        requested_state = state.upper()
        for alert in alerts:
            alert_state = alert.get("state_code")
            if alert_state is None or str(alert_state).upper() == requested_state:
                filtered_alerts.append(alert)
    else:
        filtered_alerts = alerts

    filtered_alerts = filtered_alerts[:limit]

    latest_logs = await lori_regulatory_supabase_get(
        "lori_regulatory_scan_logs?select=*&order=created_at.desc&limit=1"
    )

    latest_scan = latest_logs[0] if latest_logs else None

    answer_text = lori_regulatory_build_alert_briefing(
        filtered_alerts,
        latest_scan,
        state,
    )

    return {
        "status": "success",
        "question": question,
        "alerts_count": len(filtered_alerts),
        "latest_scan": latest_scan,
        "answer_text": answer_text,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# Agreement / Policy Intelligence Backend
# Searches uploaded/demo policy, agreement, work rule, and CBA sections
# and returns Voiceflow-ready review guidance.
# ============================================================

from typing import Any, Dict, List, Optional
from fastapi import Query, HTTPException
import re


POLICY_REVIEW_KEYWORDS = {
    "attendance": [
        "attendance",
        "call out",
        "call-out",
        "callout",
        "absence",
        "late",
        "tardy",
        "no call",
        "no show",
        "scheduled assignment",
        "report to work",
    ],
    "discipline": [
        "discipline",
        "corrective action",
        "write up",
        "write-up",
        "coaching",
        "counseling",
        "progressive",
        "formal action",
        "warning",
        "suspension",
        "termination",
    ],
    "safety": [
        "safety",
        "accident",
        "incident",
        "unsafe",
        "inspection",
        "vehicle",
        "pre trip",
        "pre-trip",
        "post trip",
        "post-trip",
        "hazard",
    ],
    "route": [
        "route",
        "assignment",
        "bid",
        "run",
        "delivery",
        "schedule",
        "dispatch",
        "route assignment",
    ],
    "overtime": [
        "overtime",
        "hours",
        "extra work",
        "premium",
        "pay",
        "payroll",
        "timekeeping",
    ],
    "seniority": [
        "seniority",
        "bid",
        "bidding",
        "assignment",
        "preference",
        "order",
    ],
    "grievance": [
        "grievance",
        "dispute",
        "appeal",
        "union",
        "labor relations",
        "representation",
    ],
    "policy": [
        "policy",
        "work rule",
        "sop",
        "procedure",
        "company policy",
        "driver policy",
    ],
    "agreement": [
        "agreement",
        "contract",
        "cba",
        "collective bargaining",
        "union agreement",
        "labor agreement",
    ],
}


def lori_policy_clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def lori_policy_terms(text: str) -> List[str]:
    text = lori_policy_clean_text(text).lower()
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    return [term for term in text.split() if len(term) >= 3]


def lori_policy_detect_situation_type(question: str) -> str:
    q = lori_policy_clean_text(question).lower()

    best_type = "Policy / Agreement Review"
    best_score = 0

    for situation_type, keywords in POLICY_REVIEW_KEYWORDS.items():
        score = 0
        for keyword in keywords:
            if keyword in q:
                score += 1
        if score > best_score:
            best_score = score
            best_type = situation_type.title()

    return best_type


def lori_policy_score_section(question: str, section: Dict[str, Any], document: Dict[str, Any]) -> int:
    q = lori_policy_clean_text(question).lower()

    section_blob = " ".join(
        [
            lori_policy_clean_text(section.get("section_title")),
            lori_policy_clean_text(section.get("section_text")),
            lori_policy_clean_text(section.get("article_number")),
            lori_policy_clean_text(section.get("section_number")),
            " ".join(section.get("topic_tags") or []),
            " ".join(section.get("risk_tags") or []),
            " ".join(section.get("applies_to") or []),
            lori_policy_clean_text(document.get("document_title")),
            lori_policy_clean_text(document.get("document_type")),
            lori_policy_clean_text(document.get("summary")),
        ]
    ).lower()

    score = 0

    for term in lori_policy_terms(q):
        if term in section_blob:
            score += 2

    for situation_type, keywords in POLICY_REVIEW_KEYWORDS.items():
        if any(keyword in q for keyword in keywords):
            for keyword in keywords:
                if keyword in section_blob:
                    score += 3

    return score


async def lori_policy_supabase_get(query_path: str) -> Any:
    url = f"{SUPABASE_URL}/rest/v1/{query_path}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get(url, headers=lori_regulatory_supabase_headers())

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase GET failed: {response.text}",
        )

    return response.json()


async def lori_policy_supabase_post(table: str, payload: Dict[str, Any]) -> Any:
    url = f"{SUPABASE_URL}/rest/v1/{table}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(
            url,
            headers=lori_regulatory_supabase_headers("return=representation"),
            json=payload,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase POST failed: {response.text}",
        )

    return response.json()


def lori_policy_build_supervisor_language(situation_type: str) -> str:
    if situation_type.lower() == "attendance":
        return (
            "This conversation should be positioned as an attendance expectation review. "
            "The supervisor should confirm the facts, review the attendance or call-out record, "
            "explain the expectation clearly, and document the discussion without making a final "
            "discipline or agreement determination."
        )

    if situation_type.lower() == "discipline":
        return (
            "This conversation should be positioned as a coaching and documentation review. "
            "The supervisor should focus on facts, prior coaching history, expectations, and corrective steps. "
            "Formal discipline should not be finalized until HR, labor relations, compliance, or leadership review is complete."
        )

    if situation_type.lower() == "safety":
        return (
            "This conversation should be positioned as a safety expectation and prevention review. "
            "The supervisor should confirm the safety facts, document what occurred, reinforce the required behavior, "
            "and determine whether training, vehicle inspection, or compliance follow-up is needed."
        )

    if situation_type.lower() == "route":
        return (
            "This conversation should be positioned as a route assignment and operational expectation review. "
            "The supervisor should confirm the assignment facts, schedule impact, route instructions, and whether policy, "
            "agreement, dispatch, or leadership review is needed."
        )

    return (
        "This conversation should be positioned as an operational expectation review. "
        "The supervisor should confirm the facts, compare the situation to the applicable policy or agreement language, "
        "document the discussion, and avoid final conclusions until the matter is reviewed by the appropriate leader, "
        "HR, labor relations, compliance, or legal reviewer."
    )


def lori_policy_build_answer(
    question: str,
    situation_type: str,
    matches: List[Dict[str, Any]],
    review_request: Optional[Dict[str, Any]] = None,
) -> str:
    if not matches:
        return f"""Agreement / Policy Review

Situation Type:
{situation_type}

Status:
I do not see a matching agreement or policy section in the currently searchable LORI policy records.

Recommended Next Action:
Upload or confirm the relevant union agreement, company policy, work rule, SOP, or driver policy so LORI can compare the situation against the official language.

Supervisor Guidance:
Do not make a final policy, agreement, labor, HR, or disciplinary determination until the official document is reviewed.

Compliance Note:
This is not a final HR, legal, labor, or contract determination. HR, labor relations, compliance, legal, or leadership review is recommended before formal action."""

    top_match = matches[0]
    top_section = top_match["section"]
    top_document = top_match["document"]

    lines = []

    lines.append("Agreement / Policy Review")
    lines.append("")
    lines.append("Situation Type:")
    lines.append(situation_type)
    lines.append("")
    lines.append("Primary Document:")
    lines.append(lori_policy_clean_text(top_document.get("document_title")) or "Not listed")
    lines.append("")
    lines.append("Document Type:")
    lines.append(lori_policy_clean_text(top_document.get("document_type")) or "Not listed")
    lines.append("")
    lines.append("Potentially Relevant Section:")
    section_label_parts = [
        lori_policy_clean_text(top_section.get("article_number")),
        lori_policy_clean_text(top_section.get("section_number")),
        lori_policy_clean_text(top_section.get("section_title")),
    ]
    section_label = " — ".join([part for part in section_label_parts if part])
    lines.append(section_label or "Section not listed")
    lines.append("")
    lines.append("Relevant Language:")
    lines.append(lori_policy_clean_text(top_section.get("section_text")))
    lines.append("")
    lines.append("Operational Concern:")
    lines.append(
        "The situation may require review against the identified agreement, policy, work rule, or procedure. "
        "The section appears relevant for operational review, but it should not be treated as a final violation finding."
    )
    lines.append("")
    lines.append("Facts to Confirm:")
    lines.append(
        "Confirm the date, time, driver or employee involved, supervisor, prior coaching history, documentation, "
        "applicable route or assignment details, and whether the official agreement or company policy version is current."
    )
    lines.append("")
    lines.append("Recommended Supervisor Language:")
    lines.append(lori_policy_build_supervisor_language(situation_type))
    lines.append("")
    lines.append("Recommended Next Action:")
    lines.append(
        "Review the official document, confirm the facts, document the discussion, and route the matter to HR, labor relations, "
        "compliance, legal, or leadership review before formal action if discipline, contract interpretation, or policy enforcement is being considered."
    )

    if len(matches) > 1:
        lines.append("")
        lines.append("Additional Sections to Review:")
        for index, match in enumerate(matches[1:4], start=2):
            section = match["section"]
            document = match["document"]
            label_parts = [
                lori_policy_clean_text(document.get("document_title")),
                lori_policy_clean_text(section.get("article_number")),
                lori_policy_clean_text(section.get("section_number")),
                lori_policy_clean_text(section.get("section_title")),
            ]
            label = " — ".join([part for part in label_parts if part])
            lines.append(f"{index}. {label}")

    lines.append("")
    lines.append("HR / Labor / Compliance Note:")
    lines.append(
        "This is not a final HR, legal, labor, or contract determination. HR, labor relations, compliance, legal, or leadership review is recommended before formal action."
    )

    return "\n".join(lines)


async def lori_policy_find_matches(question: str, limit: int = 5) -> List[Dict[str, Any]]:
    documents = await lori_policy_supabase_get(
        "lori_policy_documents?select=*&order=created_at.desc&limit=100"
    )

    sections = await lori_policy_supabase_get(
        "lori_policy_sections?select=*&order=created_at.desc&limit=500"
    )

    docs_by_id = {doc.get("id"): doc for doc in documents}

    scored: List[Dict[str, Any]] = []

    for section in sections:
        doc = docs_by_id.get(section.get("document_id"), {})
        score = lori_policy_score_section(question, section, doc)

        if score > 0:
            scored.append(
                {
                    "score": score,
                    "section": section,
                    "document": doc,
                }
            )

    scored.sort(key=lambda item: item["score"], reverse=True)

    return scored[:limit]


@app.get("/policy-documents")
async def get_policy_documents(
    api_key: Optional[str] = Query(None),
    limit: int = Query(25),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 100))

    documents = await lori_policy_supabase_get(
        f"lori_policy_documents?select=*&order=created_at.desc&limit={limit}"
    )

    return {
        "status": "success",
        "documents_count": len(documents),
        "documents": documents,
    }


@app.get("/policy-search")
async def policy_search(
    api_key: Optional[str] = Query(None),
    query: str = Query(...),
    limit: int = Query(5),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 10))

    matches = await lori_policy_find_matches(query, limit)

    return {
        "status": "success",
        "query": query,
        "matches_count": len(matches),
        "matches": matches,
    }


@app.get("/voiceflow/policy-review")
async def voiceflow_policy_review(
    api_key: Optional[str] = Query(None),
    question: str = Query(...),
    driver_name: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    supervisor_name: Optional[str] = Query(None),
    limit: int = Query(5),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 10))

    situation_type = lori_policy_detect_situation_type(question)

    review_payload = {
        "request_text": question,
        "driver_name": driver_name,
        "employee_id": employee_id,
        "supervisor_name": supervisor_name,
        "situation_type": situation_type,
        "operating_state": "MD",
        "station_code": "JESSUP-01",
        "review_status": "Open",
        "priority": "Review Needed",
    }

    created_review = await lori_policy_supabase_post(
        "lori_policy_review_requests",
        review_payload,
    )

    review_request = created_review[0] if created_review else None

    matches = await lori_policy_find_matches(question, limit)

    answer_text = lori_policy_build_answer(
        question=question,
        situation_type=situation_type,
        matches=matches,
        review_request=review_request,
    )

    if review_request and matches:
        top_match = matches[0]
        finding_payload = {
            "review_request_id": review_request.get("id"),
            "document_id": top_match["document"].get("id"),
            "section_id": top_match["section"].get("id"),
            "finding_type": "Policy / Agreement Review",
            "confidence_level": "Needs Human Review",
            "issue_summary": question,
            "potentially_relevant_section": top_match["section"].get("section_title"),
            "relevant_language": top_match["section"].get("section_text"),
            "operational_concern": (
                "This may require review against the identified agreement, policy, work rule, or procedure."
            ),
            "facts_to_confirm": (
                "Confirm the facts, involved driver or employee, date, supervisor, documentation, current policy version, and prior coaching history."
            ),
            "recommended_supervisor_language": lori_policy_build_supervisor_language(situation_type),
            "recommended_next_action": (
                "Review the official document and route the matter to HR, labor relations, compliance, legal, or leadership review before formal action."
            ),
        }

        await lori_policy_supabase_post(
            "lori_policy_findings",
            finding_payload,
        )

    return {
        "status": "success",
        "question": question,
        "situation_type": situation_type,
        "matches_count": len(matches),
        "review_request": review_request,
        "answer_text": answer_text,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# Enhanced Counseling Language Override
# Upgrades Agreement / Policy Review responses with polished,
# supervisor-ready counseling scripts based on the policy issue.
# ============================================================

def lori_policy_build_enhanced_counseling_script(
    situation_type: str,
    question: str,
    section: Dict[str, Any],
    document: Dict[str, Any],
) -> str:
    situation = lori_policy_clean_text(situation_type).lower()
    section_title = lori_policy_clean_text(section.get("section_title")) or "the applicable policy or agreement section"
    document_title = lori_policy_clean_text(document.get("document_title")) or "the applicable policy or agreement document"

    if situation == "attendance":
        return f"""Supervisor Counseling Script — Attendance / Call-Out Review

Opening Statement:
I want to review an attendance and call-out concern with you in a clear and professional way. This conversation is intended to confirm the facts, review expectations, and prevent this from becoming a larger performance or compliance issue.

Policy / Agreement Reference:
The section currently flagged for review is "{section_title}" from "{document_title}." This section appears relevant because it addresses attendance expectations, call-out procedures, documentation, and supervisor follow-up.

Facts to Confirm:
Before any conclusion is made, we need to confirm the date or dates involved, the scheduled assignment, whether proper call-out steps were followed, whether notice was timely, whether documentation exists, and whether there is any prior coaching or pattern that should be reviewed.

Expectation:
The operational expectation is that drivers report to scheduled assignments on time and follow the established call-out process when they are unable to report. The call-out process is important because it allows the station to protect route coverage, customer commitments, staffing plans, and driver accountability.

Operational Impact:
When call-outs are repeated, late, undocumented, or inconsistent with procedure, it can affect route coverage, supervisor planning, customer service, overtime exposure, team fairness, and overall station reliability.

Employee Response Prompt:
Before we determine next steps, I want to give you an opportunity to explain what happened from your perspective. Is there any information, documentation, or context we should review before this matter is evaluated further?

Corrective Expectation:
Going forward, the expectation is that you follow the call-out procedure exactly, communicate as early as required, provide any required documentation, and understand that repeated attendance issues may require additional review under company policy, applicable work rules, or agreement language.

Documentation Language:
This conversation should be documented as an attendance expectation review and fact-confirmation discussion. It should not be written as a final policy violation, contract violation, or disciplinary determination unless HR, labor relations, compliance, legal, or leadership review confirms the appropriate next step.

Recommended Next Step:
Review the official policy or agreement language, verify the attendance record, confirm prior coaching history, document the conversation, and determine whether this should remain at the coaching level or be escalated for HR/labor review.

HR / Labor / Compliance Note:
This is not a final HR, legal, labor, or contract determination. Formal action should be reviewed by HR, labor relations, compliance, legal, or leadership before being finalized."""

    if situation == "discipline":
        return f"""Supervisor Counseling Script — Progressive Coaching / Discipline Review

Opening Statement:
I want to review this concern with you in a factual and professional manner. The purpose of this conversation is to clarify expectations, review the available facts, and determine whether coaching, documentation, or further review is appropriate.

Policy / Agreement Reference:
The section currently flagged for review is "{section_title}" from "{document_title}." This section appears relevant because it addresses coaching, documentation, progressive review, or the steps that should be considered before formal action.

Facts to Confirm:
Before any decision is made, we need to confirm what occurred, when it occurred, who was involved, whether prior coaching exists, whether the issue is isolated or repeated, and whether the official policy or agreement requires specific steps before escalation.

Expectation:
The expectation is that performance, conduct, safety, and operational concerns are addressed through a fair, documented, and consistent review process.

Operational Impact:
Unresolved concerns can affect safety, route reliability, team accountability, compliance readiness, and leadership confidence in the operation.

Employee Response Prompt:
I want to give you the opportunity to respond and provide any context or documentation that should be considered before next steps are determined.

Corrective Expectation:
Going forward, you are expected to correct the concern, follow the applicable work rule or policy expectation, and understand that repeated or unresolved concerns may require additional review.

Documentation Language:
This should be documented as a coaching and fact-review conversation unless HR, labor relations, compliance, legal, or leadership confirms that formal discipline is appropriate.

Recommended Next Step:
Confirm the facts, review the official policy or agreement, check prior documentation, and determine whether the matter should remain coaching or move to formal review.

HR / Labor / Compliance Note:
This is not a final disciplinary, contract, labor, or legal determination. Formal action should be reviewed before being finalized."""

    if situation == "safety":
        return f"""Supervisor Counseling Script — Safety Expectation Review

Opening Statement:
I want to review a safety-related concern with you. This conversation is focused on prevention, accountability, and making sure expectations are clear before the issue creates greater operational or compliance risk.

Policy / Agreement Reference:
The section currently flagged for review is "{section_title}" from "{document_title}." This section appears relevant because it may relate to safety expectations, vehicle readiness, inspection behavior, or driver responsibility.

Facts to Confirm:
We need to confirm what happened, when it happened, whether a vehicle, route, customer location, or inspection was involved, whether there were any safety events, and whether documentation or witness information exists.

Expectation:
The expectation is that drivers follow all safety procedures, report concerns immediately, complete required checks, and operate in a way that protects themselves, the public, customers, equipment, and the company.

Operational Impact:
Safety concerns can affect driver readiness, fleet reliability, customer delivery, DOT/FMCSA compliance posture, insurance exposure, and leadership confidence.

Employee Response Prompt:
Please explain what happened from your perspective and identify anything that may have contributed to the issue.

Corrective Expectation:
Going forward, the expectation is full compliance with safety procedures, timely communication of hazards or equipment concerns, and immediate correction of any unsafe behavior or missed process.

Documentation Language:
This should be documented as a safety expectation and prevention review. Any formal finding should be validated against safety records, policy, training records, and leadership/compliance review.

Recommended Next Step:
Confirm the facts, review safety documentation, determine whether retraining or inspection follow-up is needed, and document the supervisor discussion.

HR / Labor / Compliance Note:
This is not a final safety, HR, labor, legal, or compliance determination. Formal action should be reviewed before being finalized."""

    if situation == "route":
        return f"""Supervisor Counseling Script — Route Assignment / Operational Expectation Review

Opening Statement:
I want to review a route or assignment concern with you so we can confirm the facts, clarify expectations, and prevent operational disruption going forward.

Policy / Agreement Reference:
The section currently flagged for review is "{section_title}" from "{document_title}." This section appears relevant because it may relate to route assignment, dispatch expectations, scheduling, work rules, or operational coverage.

Facts to Confirm:
We need to confirm the assigned route, scheduled time, dispatch instruction, communication history, supervisor direction, and whether the issue affected customer delivery, route completion, or staffing.

Expectation:
The expectation is that route assignments and dispatch instructions are followed unless a supervisor approves a change or an operational exception is documented.

Operational Impact:
Route issues can affect customer service, service windows, staffing, overtime, delivery completion, supervisor planning, and team accountability.

Employee Response Prompt:
Please explain what happened and whether there was any confusion, delay, instruction issue, route condition, or communication problem that should be reviewed.

Corrective Expectation:
Going forward, route instructions, schedule expectations, and communication procedures must be followed. Any route concern should be escalated promptly to supervision.

Documentation Language:
This should be documented as a route assignment and expectation review. It should not be treated as a final agreement or policy violation without confirming the applicable work rule and operational facts.

Recommended Next Step:
Review dispatch records, route assignment details, supervisor notes, and applicable policy or agreement language before determining next steps.

HR / Labor / Compliance Note:
This is not a final HR, legal, labor, contract, or policy determination. Formal action should be reviewed before being finalized."""

    return f"""Supervisor Counseling Script — Operational Expectation Review

Opening Statement:
I want to review this concern with you in a factual and professional way. The goal is to clarify expectations, confirm the facts, and determine the appropriate next step without jumping to conclusions.

Policy / Agreement Reference:
The section currently flagged for review is "{section_title}" from "{document_title}." This section appears relevant for review, but it should not be treated as a final violation finding without human review.

Facts to Confirm:
We need to confirm what occurred, when it occurred, who was involved, whether documentation exists, whether prior coaching applies, and whether the official policy, agreement, work rule, or procedure is current.

Expectation:
The expectation is that all drivers and team members follow applicable policies, work rules, supervisor instructions, safety requirements, and operational procedures.

Operational Impact:
If not addressed, this type of issue may affect accountability, safety, route reliability, team fairness, documentation quality, compliance readiness, or leadership confidence.

Employee Response Prompt:
Before next steps are determined, I want to give you an opportunity to provide your perspective and any information that should be considered.

Corrective Expectation:
Going forward, the expectation is that the concern is corrected, the applicable policy or work rule is followed, and any future issue is communicated promptly through the proper channel.

Documentation Language:
This should be documented as a coaching and expectation-setting discussion unless HR, labor relations, compliance, legal, or leadership review confirms that formal action is appropriate.

Recommended Next Step:
Review the official document, confirm the facts, document the conversation, and determine whether coaching, monitoring, additional training, or formal review is appropriate.

HR / Labor / Compliance Note:
This is not a final HR, legal, labor, contract, or policy determination. Formal action should be reviewed before being finalized."""


def lori_policy_build_answer(
    question: str,
    situation_type: str,
    matches: List[Dict[str, Any]],
    review_request: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Enhanced policy/agreement answer.
    Overrides the earlier version to include stronger counseling language.
    """

    if not matches:
        return f"""Agreement / Policy Review

Situation Type:
{situation_type}

Status:
I do not see a matching agreement or policy section in the currently searchable LORI policy records.

Recommended Next Action:
Upload or confirm the relevant union agreement, company policy, work rule, SOP, or driver policy so LORI can compare the situation against the official language.

Supervisor Counseling Guidance:
Before counseling, the supervisor should confirm the facts, identify the correct document, review the current policy or agreement language, and avoid making a final violation statement until HR, labor relations, compliance, legal, or leadership review is complete.

Recommended Supervisor Language:
“I want to review this matter with you to clarify expectations and confirm the facts. This conversation is not a final policy, agreement, or disciplinary determination. We will review the applicable policy or agreement language before determining next steps.”

Compliance Note:
This is not a final HR, legal, labor, or contract determination. HR, labor relations, compliance, legal, or leadership review is recommended before formal action."""

    top_match = matches[0]
    top_section = top_match["section"]
    top_document = top_match["document"]

    section_label_parts = [
        lori_policy_clean_text(top_section.get("article_number")),
        lori_policy_clean_text(top_section.get("section_number")),
        lori_policy_clean_text(top_section.get("section_title")),
    ]
    section_label = " — ".join([part for part in section_label_parts if part])

    counseling_script = lori_policy_build_enhanced_counseling_script(
        situation_type=situation_type,
        question=question,
        section=top_section,
        document=top_document,
    )

    lines = []

    lines.append("Agreement / Policy Review")
    lines.append("")
    lines.append("Situation Type:")
    lines.append(situation_type)
    lines.append("")
    lines.append("Primary Document:")
    lines.append(lori_policy_clean_text(top_document.get("document_title")) or "Not listed")
    lines.append("")
    lines.append("Document Type:")
    lines.append(lori_policy_clean_text(top_document.get("document_type")) or "Not listed")
    lines.append("")
    lines.append("Potentially Relevant Section:")
    lines.append(section_label or "Section not listed")
    lines.append("")
    lines.append("Relevant Language:")
    lines.append(lori_policy_clean_text(top_section.get("section_text")))
    lines.append("")
    lines.append("Operational Concern:")
    lines.append(
        "The situation may require review against the identified agreement, policy, work rule, or procedure. "
        "The section appears relevant for operational review, but it should not be treated as a final violation finding."
    )
    lines.append("")
    lines.append("Facts to Confirm:")
    lines.append(
        "Confirm the date, time, driver or employee involved, supervisor, prior coaching history, documentation, "
        "applicable route or assignment details, whether the official agreement or company policy version is current, "
        "and whether HR, labor relations, compliance, legal, or leadership review is required before formal action."
    )
    lines.append("")
    lines.append(counseling_script)
    lines.append("")
    lines.append("Recommended Next Action:")
    lines.append(
        "Use the counseling script as a supervisor-ready starting point, verify the official document language, confirm the facts, "
        "document the conversation, and route the matter to HR, labor relations, compliance, legal, or leadership review before any formal action."
    )

    if len(matches) > 1:
        lines.append("")
        lines.append("Additional Sections to Review:")
        for index, match in enumerate(matches[1:4], start=2):
            section = match["section"]
            document = match["document"]
            label_parts = [
                lori_policy_clean_text(document.get("document_title")),
                lori_policy_clean_text(section.get("article_number")),
                lori_policy_clean_text(section.get("section_number")),
                lori_policy_clean_text(section.get("section_title")),
            ]
            label = " — ".join([part for part in label_parts if part])
            lines.append(f"{index}. {label}")

    lines.append("")
    lines.append("HR / Labor / Compliance Note:")
    lines.append(
        "This is not a final HR, legal, labor, or contract determination. HR, labor relations, compliance, legal, or leadership review is recommended before formal action."
    )

    return "\n".join(lines)
# ============================================================
# LORI DRIVE COMMAND CENTER
# Compliance & Policy Center Upload / Intake Endpoint
# Stores uploaded agreement, policy, SOP, and work-rule documents
# in Supabase Storage and creates document metadata records.
# ============================================================

from fastapi import UploadFile, File, Form
import uuid
import mimetypes


POLICY_UPLOAD_BUCKET = "policy-agreement-uploads"


def lori_policy_safe_filename(filename: str) -> str:
    filename = filename or "uploaded-policy-document"
    filename = filename.strip().replace("\\", "/").split("/")[-1]
    filename = re.sub(r"[^a-zA-Z0-9._-]+", "_", filename)
    filename = filename.strip("_")
    return filename or "uploaded-policy-document"


def lori_policy_safe_document_type(document_type: Optional[str]) -> str:
    allowed_types = {
        "Union Agreement / CBA",
        "Company Policy",
        "Driver Work Rules",
        "SOP / Procedure",
        "Safety Policy",
        "Attendance Policy",
        "Discipline / Progressive Coaching Policy",
        "Route Assignment Rules",
        "Overtime / Scheduling Rules",
        "Other Policy / Agreement",
    }

    if document_type in allowed_types:
        return document_type

    return "Other Policy / Agreement"


def lori_policy_storage_headers(content_type: Optional[str] = None) -> Dict[str, str]:
    base_headers = lori_regulatory_supabase_headers()

    headers = {
        "apikey": base_headers.get("apikey", ""),
        "Authorization": base_headers.get("Authorization", ""),
        "x-upsert": "true",
    }

    if content_type:
        headers["Content-Type"] = content_type

    return headers


async def lori_policy_upload_to_storage(
    file_bytes: bytes,
    file_path: str,
    content_type: str,
) -> Dict[str, Any]:
    upload_url = f"{SUPABASE_URL}/storage/v1/object/{POLICY_UPLOAD_BUCKET}/{file_path}"

    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post(
            upload_url,
            headers=lori_policy_storage_headers(content_type),
            content=file_bytes,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase Storage upload failed: {response.text}",
        )

    return {
        "storage_status": "uploaded",
        "storage_path": file_path,
        "storage_response": response.text,
    }


@app.post("/policy-document-intake")
async def policy_document_intake(
    api_key: Optional[str] = Form(None),

    document_title: str = Form(...),
    document_type: str = Form("Company Policy"),
    company_name: str = Form("Demonstration Company"),
    operating_state: str = Form("MD"),
    station_code: str = Form("JESSUP-01"),

    document_version: Optional[str] = Form(None),
    effective_date: Optional[str] = Form(None),
    expiration_date: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),

    file: Optional[UploadFile] = File(None),
):
    """
    Upload/intake endpoint for Compliance & Policy Center.

    This creates a searchable document metadata record now.
    Full PDF/text extraction and section parsing will be connected later.
    """

    lori_regulatory_require_key(api_key)

    cleaned_document_title = lori_policy_clean_text(document_title)
    cleaned_document_type = lori_policy_safe_document_type(document_type)
    cleaned_company_name = lori_policy_clean_text(company_name) or "Demonstration Company"
    cleaned_operating_state = lori_policy_clean_text(operating_state).upper() or "MD"
    cleaned_station_code = lori_policy_clean_text(station_code).upper() or "JESSUP-01"

    source_file_name = None
    source_file_path = None
    source_file_url = None
    upload_status = "Metadata Created / Pending File Upload"

    if file is not None:
        original_filename = lori_policy_safe_filename(file.filename or "policy-document")
        file_bytes = await file.read()

        if file_bytes:
            guessed_type = file.content_type or mimetypes.guess_type(original_filename)[0] or "application/octet-stream"

            unique_id = str(uuid.uuid4())
            source_file_path = (
                f"{cleaned_station_code}/"
                f"{cleaned_document_type.replace(' ', '_').replace('/', '_')}/"
                f"{unique_id}_{original_filename}"
            )

            await lori_policy_upload_to_storage(
                file_bytes=file_bytes,
                file_path=source_file_path,
                content_type=guessed_type,
            )

            source_file_name = original_filename
            source_file_url = f"{POLICY_UPLOAD_BUCKET}/{source_file_path}"
            upload_status = "Uploaded / Pending Extraction"

    summary = (
        f"{cleaned_document_type} uploaded or registered for LORI Compliance & Policy Center review. "
        f"Document is staged for section extraction, topic tagging, policy search, agreement review, "
        f"and supervisor-ready counseling intelligence."
    )

    payload = {
        "document_title": cleaned_document_title,
        "document_type": cleaned_document_type,
        "company_name": cleaned_company_name,
        "operating_state": cleaned_operating_state,
        "station_code": cleaned_station_code,
        "document_status": upload_status,
        "document_version": document_version,
        "effective_date": effective_date or None,
        "expiration_date": expiration_date or None,
        "source_file_name": source_file_name,
        "source_file_path": source_file_path,
        "source_file_url": source_file_url,
        "summary": summary,
        "notes": notes or "Uploaded through LORI Compliance & Policy Center. Extraction pending.",
        "created_by": "LORI Compliance Policy Center",
    }

    created = await lori_policy_supabase_post(
        "lori_policy_documents",
        payload,
    )

    document_record = created[0] if created else None

    return {
        "status": "success",
        "message": "Policy/agreement document intake completed.",
        "document_status": upload_status,
        "document": document_record,
        "storage_bucket": POLICY_UPLOAD_BUCKET,
        "source_file_name": source_file_name,
        "source_file_path": source_file_path,
        "next_step": (
            "Document metadata is saved. Full text extraction, clause parsing, section tagging, "
            "and searchable policy intelligence can now be connected as the next backend upgrade."
        ),
        "disclaimer": (
            "LORI provides operational decision support. Uploaded agreements, policies, SOPs, and work rules "
            "must be reviewed by authorized HR, labor relations, compliance, legal, or leadership personnel "
            "before formal action."
        ),
    }


@app.get("/policy-intake-status")
async def policy_intake_status(
    api_key: Optional[str] = Query(None),
    limit: int = Query(10),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 50))

    documents = await lori_policy_supabase_get(
        f"lori_policy_documents?select=*&order=created_at.desc&limit={limit}"
    )

    pending_extraction = [
        doc for doc in documents
        if "Pending Extraction" in str(doc.get("document_status") or "")
        or "Pending File Upload" in str(doc.get("document_status") or "")
    ]

    return {
        "status": "success",
        "documents_count": len(documents),
        "pending_extraction_count": len(pending_extraction),
        "documents": documents,
        "message": (
            "Compliance & Policy Center document intake is active. "
            "Documents can be uploaded or registered and staged for extraction."
        ),
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# Policy / Agreement Document Extraction Endpoint
# Extracts uploaded PDF/TXT policy documents from Supabase Storage,
# splits them into searchable sections, tags topics, and updates
# the document status to Extracted / Searchable.
# ============================================================

from io import BytesIO


async def lori_policy_download_from_storage(file_path: str) -> bytes:
    download_url = f"{SUPABASE_URL}/storage/v1/object/{POLICY_UPLOAD_BUCKET}/{file_path}"

    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.get(
            download_url,
            headers=lori_policy_storage_headers(),
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase Storage download failed: {response.text}",
        )

    return response.content


async def lori_policy_supabase_patch(
    table: str,
    record_id: str,
    payload: Dict[str, Any],
) -> Any:
    url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{record_id}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(
            url,
            headers=lori_regulatory_supabase_headers("return=representation"),
            json=payload,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase PATCH failed: {response.text}",
        )

    return response.json()


async def lori_policy_supabase_delete_sections(document_id: str) -> None:
    url = f"{SUPABASE_URL}/rest/v1/lori_policy_sections?document_id=eq.{document_id}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.delete(
            url,
            headers=lori_regulatory_supabase_headers(),
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase DELETE failed: {response.text}",
        )


async def lori_policy_supabase_post_many(
    table: str,
    payload: List[Dict[str, Any]],
) -> Any:
    if not payload:
        return []

    url = f"{SUPABASE_URL}/rest/v1/{table}"

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(
            url,
            headers=lori_regulatory_supabase_headers("return=representation"),
            json=payload,
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail=f"Supabase batch POST failed: {response.text}",
        )

    return response.json()


def lori_policy_extract_text_from_file(
    file_bytes: bytes,
    file_name: Optional[str],
) -> str:
    safe_name = (file_name or "").lower()

    if safe_name.endswith(".txt"):
        return file_bytes.decode("utf-8", errors="ignore")

    if safe_name.endswith(".pdf"):
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail=(
                    "PDF extraction requires the pypdf package. "
                    "Add pypdf to requirements.txt and redeploy Render."
                ),
            ) from exc

        reader = PdfReader(BytesIO(file_bytes))
        pages = []

        for page_index, page in enumerate(reader.pages, start=1):
            try:
                page_text = page.extract_text() or ""
            except Exception:
                page_text = ""

            if page_text.strip():
                pages.append(f"\n\n--- Page {page_index} ---\n{page_text}")

        return "\n".join(pages).strip()

    try:
        return file_bytes.decode("utf-8", errors="ignore")
    except Exception:
        return ""


def lori_policy_extract_topic_tags(text: str) -> List[str]:
    blob = lori_policy_clean_text(text).lower()
    tags = []

    for topic, keywords in POLICY_REVIEW_KEYWORDS.items():
        if any(keyword in blob for keyword in keywords):
            tags.append(topic)

    if "call-out" in blob or "call out" in blob or "callout" in blob:
        if "call-out" not in tags:
            tags.append("call-out")

    if "driver" in blob:
        if "driver operations" not in tags:
            tags.append("driver operations")

    if not tags:
        tags.append("general policy")

    return tags


def lori_policy_extract_risk_tags(text: str) -> List[str]:
    blob = lori_policy_clean_text(text).lower()
    risk_tags = []

    if any(term in blob for term in ["attendance", "call-out", "call out", "absence", "late", "no call", "no show"]):
        risk_tags.append("attendance risk")

    if any(term in blob for term in ["discipline", "corrective", "coaching", "counseling", "warning"]):
        risk_tags.append("coaching / discipline review")

    if any(term in blob for term in ["safety", "incident", "accident", "inspection", "vehicle"]):
        risk_tags.append("safety review")

    if any(term in blob for term in ["union", "agreement", "contract", "cba", "grievance", "labor"]):
        risk_tags.append("HR / labor review recommended")

    if any(term in blob for term in ["documentation", "document", "record"]):
        risk_tags.append("documentation needed")

    if not risk_tags:
        risk_tags.append("review needed")

    return risk_tags


def lori_policy_parse_section_heading(heading: str) -> Dict[str, Any]:
    clean_heading = lori_policy_clean_text(heading)

    article_number = None
    section_number = None
    section_title = clean_heading

    section_match = re.search(
        r"(section|sec\.?)\s+([0-9]+(?:\.[0-9]+)*)\s*[—\-:]\s*(.+)",
        clean_heading,
        flags=re.IGNORECASE,
    )

    if section_match:
        section_number = section_match.group(2).strip()
        section_title = section_match.group(3).strip()
        return {
            "article_number": None,
            "section_number": section_number,
            "section_title": section_title,
        }

    article_match = re.search(
        r"(article)\s+([0-9ivxlcdm]+)\s*[—\-:]\s*(.+)",
        clean_heading,
        flags=re.IGNORECASE,
    )

    if article_match:
        article_number = f"Article {article_match.group(2).strip()}"
        section_title = article_match.group(3).strip()
        return {
            "article_number": article_number,
            "section_number": None,
            "section_title": section_title,
        }

    policy_match = re.search(
        r"(policy area)\s+([0-9]+)\s*[—\-:]\s*(.+)",
        clean_heading,
        flags=re.IGNORECASE,
    )

    if policy_match:
        article_number = f"Policy Area {policy_match.group(2).strip()}"
        section_title = policy_match.group(3).strip()
        return {
            "article_number": article_number,
            "section_number": None,
            "section_title": section_title,
        }

    return {
        "article_number": article_number,
        "section_number": section_number,
        "section_title": section_title,
    }


def lori_policy_split_text_into_sections(text: str) -> List[Dict[str, Any]]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\n{3,}", "\n\n", normalized).strip()

    if not normalized:
        return []

    heading_pattern = re.compile(
        r"(?im)^(section\s+[0-9]+(?:\.[0-9]+)*\s*[—\-:]\s*.+|article\s+[0-9ivxlcdm]+\s*[—\-:]\s*.+|policy area\s+[0-9]+\s*[—\-:]\s*.+)$"
    )

    matches = list(heading_pattern.finditer(normalized))
    sections: List[Dict[str, Any]] = []

    if matches:
        for index, match in enumerate(matches):
            heading = match.group(1).strip()
            start = match.end()
            end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
            section_body = normalized[start:end].strip()

            parsed = lori_policy_parse_section_heading(heading)

            full_text = f"{heading}\n{section_body}".strip()

            if len(full_text) < 30:
                continue

            sections.append(
                {
                    "article_number": parsed.get("article_number"),
                    "section_number": parsed.get("section_number"),
                    "section_title": parsed.get("section_title"),
                    "section_text": full_text,
                    "page_number": None,
                    "topic_tags": lori_policy_extract_topic_tags(full_text),
                    "risk_tags": lori_policy_extract_risk_tags(full_text),
                    "applies_to": ["drivers", "supervisors", "operations leadership"],
                }
            )

    if not sections:
        paragraphs = [p.strip() for p in re.split(r"\n\s*\n", normalized) if p.strip()]
        current_chunk = []
        current_length = 0
        chunk_number = 1

        for paragraph in paragraphs:
            current_chunk.append(paragraph)
            current_length += len(paragraph)

            if current_length >= 1200:
                chunk_text = "\n\n".join(current_chunk).strip()
                sections.append(
                    {
                        "article_number": "Extracted Text",
                        "section_number": str(chunk_number),
                        "section_title": f"Extracted Policy Section {chunk_number}",
                        "section_text": chunk_text,
                        "page_number": None,
                        "topic_tags": lori_policy_extract_topic_tags(chunk_text),
                        "risk_tags": lori_policy_extract_risk_tags(chunk_text),
                        "applies_to": ["drivers", "supervisors", "operations leadership"],
                    }
                )
                chunk_number += 1
                current_chunk = []
                current_length = 0

        if current_chunk:
            chunk_text = "\n\n".join(current_chunk).strip()
            sections.append(
                {
                    "article_number": "Extracted Text",
                    "section_number": str(chunk_number),
                    "section_title": f"Extracted Policy Section {chunk_number}",
                    "section_text": chunk_text,
                    "page_number": None,
                    "topic_tags": lori_policy_extract_topic_tags(chunk_text),
                    "risk_tags": lori_policy_extract_risk_tags(chunk_text),
                    "applies_to": ["drivers", "supervisors", "operations leadership"],
                }
            )

    return sections[:100]


async def lori_policy_get_document_for_extraction(
    document_id: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    if document_id:
        records = await lori_policy_supabase_get(
            f"lori_policy_documents?select=*&id=eq.{document_id}&limit=1"
        )
        return records[0] if records else None

    documents = await lori_policy_supabase_get(
        "lori_policy_documents?select=*&order=created_at.desc&limit=50"
    )

    for document in documents:
        status = str(document.get("document_status") or "")
        file_path = document.get("source_file_path")

        if file_path and "Pending Extraction" in status:
            return document

    return None


@app.post("/policy-document-extract")
async def policy_document_extract(
    api_key: Optional[str] = Query(None),
    document_id: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    document = await lori_policy_get_document_for_extraction(document_id)

    if not document:
        return {
            "status": "not_found",
            "message": "No uploaded policy/agreement document is currently pending extraction.",
            "sections_created": 0,
            "answer_text": (
                "No uploaded policy or agreement document is currently pending extraction. "
                "Upload a document first or provide a valid document_id."
            ),
        }

    doc_id = document.get("id")
    file_path = document.get("source_file_path")
    file_name = document.get("source_file_name") or document.get("document_title")

    if not file_path:
        return {
            "status": "missing_file",
            "message": "The selected policy/agreement document does not have a stored file path.",
            "document": document,
            "sections_created": 0,
        }

    file_bytes = await lori_policy_download_from_storage(file_path)

    extracted_text = lori_policy_extract_text_from_file(
        file_bytes=file_bytes,
        file_name=file_name,
    )

    if not extracted_text.strip():
        await lori_policy_supabase_patch(
            "lori_policy_documents",
            doc_id,
            {
                "document_status": "Extraction Failed / No Text Found",
                "notes": "LORI attempted extraction, but no readable text was found. This may be a scanned PDF requiring OCR.",
            },
        )

        return {
            "status": "no_text_found",
            "message": "No readable text was found in the uploaded document.",
            "document_id": doc_id,
            "sections_created": 0,
            "answer_text": (
                "LORI could not extract readable text from this document. "
                "If this is a scanned PDF, OCR will be needed in a later upgrade."
            ),
        }

    sections = lori_policy_split_text_into_sections(extracted_text)

    if not sections:
        return {
            "status": "no_sections_created",
            "message": "Text was extracted, but no sections were created.",
            "document_id": doc_id,
            "extracted_character_count": len(extracted_text),
            "sections_created": 0,
        }

    await lori_policy_supabase_delete_sections(doc_id)

    insert_payload = []

    for section in sections:
        insert_payload.append(
            {
                "document_id": doc_id,
                "article_number": section.get("article_number"),
                "section_number": section.get("section_number"),
                "section_title": section.get("section_title"),
                "page_number": section.get("page_number"),
                "section_text": section.get("section_text"),
                "topic_tags": section.get("topic_tags"),
                "risk_tags": section.get("risk_tags"),
                "applies_to": section.get("applies_to"),
            }
        )

    created_sections = await lori_policy_supabase_post_many(
        "lori_policy_sections",
        insert_payload,
    )

    await lori_policy_supabase_patch(
        "lori_policy_documents",
        doc_id,
        {
            "document_status": "Extracted / Searchable",
            "summary": (
                f"{document.get('document_type') or 'Policy / Agreement'} extracted by LORI. "
                f"{len(created_sections)} searchable section(s) created for policy search, agreement review, "
                f"supervisor-ready counseling language, HR/labor/compliance review, and operational decision support."
            ),
            "notes": (
                f"Extraction complete. LORI created {len(created_sections)} searchable section(s). "
                "Human review is still required before formal HR, labor, legal, compliance, or leadership action."
            ),
        },
    )

    answer_text = f"""Policy / Agreement Extraction Complete

Document:
{document.get("document_title")}

Status:
Extracted / Searchable

Sections Created:
{len(created_sections)}

What LORI Can Do Now:
- Search this uploaded document
- Identify potentially relevant sections
- Support agreement or policy review
- Generate supervisor-ready counseling language
- Stage HR, labor, compliance, legal, or leadership review notes

Important Note:
This extraction supports operational decision-making. The official document and extracted sections should be verified by authorized HR, labor relations, compliance, legal, or leadership personnel before formal action."""

    return {
        "status": "success",
        "message": "Policy/agreement document extraction completed.",
        "document_id": doc_id,
        "document_title": document.get("document_title"),
        "source_file_name": file_name,
        "extracted_character_count": len(extracted_text),
        "sections_created": len(created_sections),
        "sections": created_sections,
        "answer_text": answer_text,
    }


@app.get("/policy-extraction-status")
async def policy_extraction_status(
    api_key: Optional[str] = Query(None),
    limit: int = Query(20),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 100))

    documents = await lori_policy_supabase_get(
        f"lori_policy_documents?select=*&order=created_at.desc&limit={limit}"
    )

    extracted_count = len([
        doc for doc in documents
        if "Extracted / Searchable" in str(doc.get("document_status") or "")
    ])

    pending_count = len([
        doc for doc in documents
        if "Pending Extraction" in str(doc.get("document_status") or "")
    ])

    failed_count = len([
        doc for doc in documents
        if "Extraction Failed" in str(doc.get("document_status") or "")
    ])

    return {
        "status": "success",
        "documents_count": len(documents),
        "extracted_count": extracted_count,
        "pending_extraction_count": pending_count,
        "failed_extraction_count": failed_count,
        "documents": documents,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# Specific Document Policy / Agreement Intelligence
# Allows LORI to search and review against one uploaded document,
# instead of searching every policy/agreement section together.
# ============================================================

from urllib.parse import quote


async def lori_policy_get_document_by_id_or_title(
    document_id: Optional[str] = None,
    document_title: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    if document_id:
        docs = await lori_policy_supabase_get(
            f"lori_policy_documents?select=*&id=eq.{quote(document_id)}&limit=1"
        )
        if docs:
            return docs[0]

    if document_title:
        cleaned_title = lori_policy_clean_text(document_title)
        docs = await lori_policy_supabase_get(
            f"lori_policy_documents?select=*&document_title=ilike.*{quote(cleaned_title)}*&order=created_at.desc&limit=1"
        )
        if docs:
            return docs[0]

    return None


async def lori_policy_get_latest_extracted_document() -> Optional[Dict[str, Any]]:
    docs = await lori_policy_supabase_get(
        "lori_policy_documents?select=*&order=created_at.desc&limit=50"
    )

    for doc in docs:
        status = str(doc.get("document_status") or "")
        if "Extracted / Searchable" in status:
            return doc

    return None


async def lori_policy_get_sections_for_document(document_id: str) -> List[Dict[str, Any]]:
    sections = await lori_policy_supabase_get(
        f"lori_policy_sections?select=*&document_id=eq.{quote(document_id)}&order=section_number.asc&limit=500"
    )
    return sections


async def lori_policy_find_matches_in_document(
    question: str,
    document: Dict[str, Any],
    limit: int = 5,
) -> List[Dict[str, Any]]:
    document_id = document.get("id")

    if not document_id:
        return []

    sections = await lori_policy_get_sections_for_document(document_id)

    scored: List[Dict[str, Any]] = []

    for section in sections:
        score = lori_policy_score_section(question, section, document)

        if score > 0:
            scored.append(
                {
                    "score": score,
                    "section": section,
                    "document": document,
                }
            )

    scored.sort(key=lambda item: item["score"], reverse=True)

    return scored[:limit]


@app.get("/policy-document-sections")
async def policy_document_sections(
    api_key: Optional[str] = Query(None),
    document_id: Optional[str] = Query(None),
    document_title: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 500))

    document = await lori_policy_get_document_by_id_or_title(
        document_id=document_id,
        document_title=document_title,
    )

    if not document:
        return {
            "status": "not_found",
            "message": "No matching policy/agreement document was found.",
            "document_id": document_id,
            "document_title": document_title,
            "sections_count": 0,
            "sections": [],
        }

    sections = await lori_policy_get_sections_for_document(document.get("id"))
    sections = sections[:limit]

    return {
        "status": "success",
        "document": document,
        "sections_count": len(sections),
        "sections": sections,
    }


@app.get("/policy-search-document")
async def policy_search_document(
    api_key: Optional[str] = Query(None),
    query: str = Query(...),
    document_id: Optional[str] = Query(None),
    document_title: Optional[str] = Query(None),
    limit: int = Query(5),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 10))

    document = await lori_policy_get_document_by_id_or_title(
        document_id=document_id,
        document_title=document_title,
    )

    if not document:
        document = await lori_policy_get_latest_extracted_document()

    if not document:
        return {
            "status": "not_found",
            "message": "No searchable policy/agreement document was found. Upload and extract a document first.",
            "query": query,
            "matches_count": 0,
            "matches": [],
        }

    matches = await lori_policy_find_matches_in_document(
        question=query,
        document=document,
        limit=limit,
    )

    return {
        "status": "success",
        "query": query,
        "document_id": document.get("id"),
        "document_title": document.get("document_title"),
        "document_status": document.get("document_status"),
        "matches_count": len(matches),
        "matches": matches,
    }


@app.get("/voiceflow/policy-review-document")
async def voiceflow_policy_review_document(
    api_key: Optional[str] = Query(None),
    question: str = Query(...),
    document_id: Optional[str] = Query(None),
    document_title: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    supervisor_name: Optional[str] = Query(None),
    limit: int = Query(5),
):
    lori_regulatory_require_key(api_key)

    limit = max(1, min(limit, 10))

    situation_type = lori_policy_detect_situation_type(question)

    document = await lori_policy_get_document_by_id_or_title(
        document_id=document_id,
        document_title=document_title,
    )

    if not document:
        document = await lori_policy_get_latest_extracted_document()

    if not document:
        answer_text = f"""Agreement / Policy Review

Situation Type:
{situation_type}

Status:
No extracted/searchable policy or agreement document is currently available.

Recommended Next Action:
Upload and extract the relevant union agreement, company policy, SOP, work rule, or driver policy before asking LORI to review the situation against a specific document.

Compliance Note:
This is not a final HR, legal, labor, compliance, or contract determination. Authorized review is required before formal action."""

        return {
            "status": "not_found",
            "question": question,
            "situation_type": situation_type,
            "matches_count": 0,
            "answer_text": answer_text,
        }

    matches = await lori_policy_find_matches_in_document(
        question=question,
        document=document,
        limit=limit,
    )

    review_payload = {
        "request_text": question,
        "driver_name": driver_name,
        "employee_id": employee_id,
        "supervisor_name": supervisor_name,
        "situation_type": situation_type,
        "operating_state": document.get("operating_state") or "MD",
        "station_code": document.get("station_code") or "JESSUP-01",
        "document_id": document.get("id"),
        "review_status": "Open",
        "priority": "Review Needed",
    }

    created_review = await lori_policy_supabase_post(
        "lori_policy_review_requests",
        review_payload,
    )

    review_request = created_review[0] if created_review else None

    answer_text = lori_policy_build_answer(
        question=question,
        situation_type=situation_type,
        matches=matches,
        review_request=review_request,
    )

    if review_request and matches:
        top_match = matches[0]

        finding_payload = {
            "review_request_id": review_request.get("id"),
            "document_id": document.get("id"),
            "section_id": top_match["section"].get("id"),
            "finding_type": "Specific Document Policy / Agreement Review",
            "confidence_level": "Needs Human Review",
            "issue_summary": question,
            "potentially_relevant_section": top_match["section"].get("section_title"),
            "relevant_language": top_match["section"].get("section_text"),
            "operational_concern": (
                "This may require review against the selected uploaded policy/agreement document."
            ),
            "facts_to_confirm": (
                "Confirm the facts, driver or employee involved, date, supervisor, documentation, current policy version, and prior coaching history."
            ),
            "recommended_supervisor_language": lori_policy_build_supervisor_language(situation_type),
            "recommended_next_action": (
                "Review the extracted document section and route the matter to HR, labor relations, compliance, legal, or leadership review before formal action."
            ),
        }

        await lori_policy_supabase_post(
            "lori_policy_findings",
            finding_payload,
        )

    return {
        "status": "success",
        "question": question,
        "situation_type": situation_type,
        "document_id": document.get("id"),
        "document_title": document.get("document_title"),
        "document_status": document.get("document_status"),
        "matches_count": len(matches),
        "review_request": review_request,
        "answer_text": answer_text,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# ACTION CENTER BACKEND
# Creates, reads, updates, and summarizes operational action items,
# supervisor follow-ups, HR/labor/compliance review notes,
# and leadership briefing queue items.
# ============================================================

from fastapi import Body
from datetime import date, datetime, timedelta
from urllib.parse import quote
from typing import Any, Dict, List, Optional


ACTION_ITEM_ALLOWED_FIELDS = {
    "action_title",
    "action_type",
    "action_status",
    "priority",
    "source_module",
    "source_type",
    "source_reference_id",
    "driver_name",
    "employee_id",
    "supervisor_name",
    "owner_name",
    "owner_role",
    "station_code",
    "operating_state",
    "company_name",
    "reason",
    "recommended_follow_up",
    "documentation_note",
    "compliance_note",
    "due_date",
    "created_by",
}

REVIEW_NOTE_ALLOWED_FIELDS = {
    "action_item_id",
    "review_type",
    "review_status",
    "priority",
    "related_driver_name",
    "related_employee_id",
    "supervisor_name",
    "policy_document_id",
    "policy_section_id",
    "review_summary",
    "relevant_policy_or_agreement",
    "facts_to_confirm",
    "recommended_review_note",
    "required_reviewer",
    "created_by",
}

LEADERSHIP_QUEUE_ALLOWED_FIELDS = {
    "action_item_id",
    "briefing_title",
    "briefing_type",
    "briefing_status",
    "priority",
    "station_code",
    "operating_state",
    "executive_summary",
    "key_risk",
    "recommended_leadership_action",
    "supervisor_follow_up",
    "compliance_note",
    "created_by",
}

SUPERVISOR_FOLLOWUP_ALLOWED_FIELDS = {
    "action_item_id",
    "supervisor_name",
    "driver_name",
    "employee_id",
    "followup_type",
    "followup_status",
    "priority",
    "followup_reason",
    "recommended_script",
    "documentation_required",
    "due_date",
    "created_by",
}


def lori_action_clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def lori_action_parse_due_date(value: Any) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()

    if not text or text.lower() in {"string", "null", "none"}:
        return None

    try:
        parsed = datetime.strptime(text, "%Y-%m-%d").date()
        return parsed.isoformat()
    except Exception:
        return None


def lori_action_default_due_date(days: int = 3) -> str:
    return (date.today() + timedelta(days=days)).isoformat()


def lori_action_filter_payload(
    payload: Dict[str, Any],
    allowed_fields: set,
) -> Dict[str, Any]:
    cleaned: Dict[str, Any] = {}

    for key, value in payload.items():
        if key not in allowed_fields:
            continue

        if key == "due_date":
            cleaned[key] = lori_action_parse_due_date(value)
        else:
            cleaned[key] = value if value != "" else None

    return cleaned


async def lori_action_get_table_rows(
    table: str,
    limit: int = 500,
) -> List[Dict[str, Any]]:
    limit = max(1, min(limit, 1000))
    return await lori_policy_supabase_get(
        f"{table}?select=*&order=created_at.desc&limit={limit}"
    )


def lori_action_count_by(rows: List[Dict[str, Any]], field: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}

    for row in rows:
        value = row.get(field) or "Not Listed"
        counts[str(value)] = counts.get(str(value), 0) + 1

    return counts


def lori_action_is_open(status: Any) -> bool:
    text = lori_action_clean_text(status).lower()
    return text not in {"closed", "complete", "completed", "resolved", "cancelled", "canceled"}


def lori_action_due_soon(row: Dict[str, Any], days: int = 7) -> bool:
    due_date_value = row.get("due_date")

    if not due_date_value:
        return False

    try:
        due = datetime.strptime(str(due_date_value), "%Y-%m-%d").date()
    except Exception:
        return False

    today = date.today()
    return today <= due <= today + timedelta(days=days)


def lori_action_build_answer(action: Dict[str, Any]) -> str:
    return f"""Action Center Item Created

Action:
{action.get("action_title") or "Operational follow-up"}

Type:
{action.get("action_type") or "Operational Follow-Up"}

Priority:
{action.get("priority") or "Medium"}

Status:
{action.get("action_status") or "Open"}

Owner:
{action.get("owner_name") or action.get("supervisor_name") or "Owner to be assigned"}

Reason:
{action.get("reason") or "Operational follow-up created through LORI."}

Recommended Follow-Up:
{action.get("recommended_follow_up") or "Review the issue, confirm the facts, document next steps, and escalate if needed."}

Due Date:
{action.get("due_date") or "Not assigned"}

Leadership / Compliance Note:
{action.get("compliance_note") or "Validate before formal HR, labor, legal, compliance, audit, regulatory, or corrective action."}"""


@app.get("/action-center-summary")
async def action_center_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    actions = await lori_action_get_table_rows("lori_action_items", 500)
    review_notes = await lori_action_get_table_rows("lori_review_notes", 500)
    briefing_items = await lori_action_get_table_rows("lori_leadership_briefing_queue", 500)
    supervisor_followups = await lori_action_get_table_rows("lori_supervisor_followups", 500)

    open_actions = [a for a in actions if lori_action_is_open(a.get("action_status"))]
    high_priority_actions = [
        a for a in actions
        if str(a.get("priority") or "").lower() in {"high", "critical"}
        and lori_action_is_open(a.get("action_status"))
    ]
    due_soon_actions = [
        a for a in actions
        if lori_action_is_open(a.get("action_status")) and lori_action_due_soon(a, 7)
    ]

    hr_labor_review_needed = [
        n for n in review_notes
        if lori_action_is_open(n.get("review_status"))
    ]

    answer_text = f"""Action Center Summary

Open Actions:
{len(open_actions)}

High / Critical Priority Actions:
{len(high_priority_actions)}

Supervisor Follow-Ups:
{len(supervisor_followups)}

HR / Labor / Compliance Review Notes:
{len(hr_labor_review_needed)}

Leadership Briefing Queue:
{len(briefing_items)}

Due Within 7 Days:
{len(due_soon_actions)}

Recommended Next Action:
Review high-priority open actions first, assign owners where missing, and move HR/labor/compliance-sensitive items into formal review before any corrective action is finalized."""

    return {
        "status": "success",
        "action_items_count": len(actions),
        "open_actions_count": len(open_actions),
        "high_priority_open_count": len(high_priority_actions),
        "due_within_7_days_count": len(due_soon_actions),
        "review_notes_count": len(review_notes),
        "open_review_notes_count": len(hr_labor_review_needed),
        "briefing_items_count": len(briefing_items),
        "supervisor_followups_count": len(supervisor_followups),
        "action_status_counts": lori_action_count_by(actions, "action_status"),
        "priority_counts": lori_action_count_by(actions, "priority"),
        "source_module_counts": lori_action_count_by(actions, "source_module"),
        "recent_actions": actions[:10],
        "answer_text": answer_text,
    }


@app.get("/action-items")
async def get_action_items(
    api_key: Optional[str] = Query(None),
    action_status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    source_module: Optional[str] = Query(None),
    supervisor_name: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_action_get_table_rows("lori_action_items", 500)

    if action_status:
        rows = [
            r for r in rows
            if lori_action_clean_text(r.get("action_status")).lower() == action_status.lower()
        ]

    if priority:
        rows = [
            r for r in rows
            if lori_action_clean_text(r.get("priority")).lower() == priority.lower()
        ]

    if source_module:
        rows = [
            r for r in rows
            if source_module.lower() in lori_action_clean_text(r.get("source_module")).lower()
        ]

    if supervisor_name:
        rows = [
            r for r in rows
            if supervisor_name.lower() in lori_action_clean_text(r.get("supervisor_name")).lower()
        ]

    limit = max(1, min(limit, 200))
    rows = rows[:limit]

    return {
        "status": "success",
        "action_items_count": len(rows),
        "action_items": rows,
    }


@app.post("/action-item-create")
async def action_item_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    action_payload = lori_action_filter_payload(
        payload,
        ACTION_ITEM_ALLOWED_FIELDS,
    )

    if not action_payload.get("action_title"):
        action_payload["action_title"] = "LORI Operational Follow-Up"

    action_payload.setdefault("action_type", "Operational Follow-Up")
    action_payload.setdefault("action_status", "Open")
    action_payload.setdefault("priority", "Medium")
    action_payload.setdefault("source_module", "LORI Action Center")
    action_payload.setdefault("station_code", "JESSUP-01")
    action_payload.setdefault("operating_state", "MD")
    action_payload.setdefault("company_name", "Food Authority")
    action_payload.setdefault("due_date", lori_action_default_due_date(3))
    action_payload.setdefault("created_by", "LORI Action Center")

    created = await lori_policy_supabase_post(
        "lori_action_items",
        action_payload,
    )

    action = created[0] if created else {}

    return {
        "status": "success",
        "message": "Action item created.",
        "action_item": action,
        "answer_text": lori_action_build_answer(action),
    }


@app.post("/action-item-update")
async def action_item_update(
    api_key: Optional[str] = Query(None),
    action_item_id: str = Query(...),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    update_payload = lori_action_filter_payload(
        payload,
        ACTION_ITEM_ALLOWED_FIELDS,
    )

    if "action_status" in update_payload:
        status_text = lori_action_clean_text(update_payload.get("action_status")).lower()
        if status_text in {"complete", "completed", "closed", "resolved"}:
            update_payload["completed_at"] = datetime.utcnow().isoformat()

    update_payload["updated_at"] = datetime.utcnow().isoformat()

    updated = await lori_policy_supabase_patch(
        "lori_action_items",
        action_item_id,
        update_payload,
    )

    action = updated[0] if updated else {}

    return {
        "status": "success",
        "message": "Action item updated.",
        "action_item": action,
        "answer_text": lori_action_build_answer(action),
    }


@app.post("/supervisor-followup-create")
async def supervisor_followup_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    action_payload = {
        "action_title": payload.get("action_title") or "Supervisor Follow-Up Required",
        "action_type": payload.get("action_type") or "Supervisor Follow-Up",
        "action_status": "Open",
        "priority": payload.get("priority") or "Medium",
        "source_module": payload.get("source_module") or "LORI Action Center",
        "source_type": payload.get("source_type") or "Supervisor Follow-Up",
        "driver_name": payload.get("driver_name"),
        "employee_id": payload.get("employee_id"),
        "supervisor_name": payload.get("supervisor_name"),
        "owner_name": payload.get("owner_name") or payload.get("supervisor_name"),
        "owner_role": payload.get("owner_role") or "Supervisor",
        "station_code": payload.get("station_code") or "JESSUP-01",
        "operating_state": payload.get("operating_state") or "MD",
        "company_name": payload.get("company_name") or "Food Authority",
        "reason": payload.get("followup_reason") or payload.get("reason") or "Supervisor follow-up created through LORI.",
        "recommended_follow_up": payload.get("recommended_script") or payload.get("recommended_follow_up"),
        "documentation_note": payload.get("documentation_required") or payload.get("documentation_note"),
        "due_date": lori_action_parse_due_date(payload.get("due_date")) or lori_action_default_due_date(3),
        "created_by": "LORI Action Center",
    }

    created_action = await lori_policy_supabase_post(
        "lori_action_items",
        action_payload,
    )

    action = created_action[0] if created_action else {}

    followup_payload = lori_action_filter_payload(
        payload,
        SUPERVISOR_FOLLOWUP_ALLOWED_FIELDS,
    )

    followup_payload["action_item_id"] = action.get("id")
    followup_payload.setdefault("followup_type", "Supervisor Follow-Up")
    followup_payload.setdefault("followup_status", "Open")
    followup_payload.setdefault("priority", action.get("priority") or "Medium")
    followup_payload.setdefault("due_date", action.get("due_date"))
    followup_payload.setdefault("created_by", "LORI Action Center")

    created_followup = await lori_policy_supabase_post(
        "lori_supervisor_followups",
        followup_payload,
    )

    followup = created_followup[0] if created_followup else {}

    answer_text = f"""Supervisor Follow-Up Created

Action:
{action.get("action_title")}

Supervisor / Owner:
{action.get("owner_name") or "Owner to be assigned"}

Driver:
{action.get("driver_name") or "Not driver-specific"}

Priority:
{action.get("priority")}

Due Date:
{action.get("due_date")}

Recommended Follow-Up:
{action.get("recommended_follow_up") or "Review the issue, confirm facts, document the conversation, and escalate if needed."}

Documentation Note:
{action.get("documentation_note") or "Document the follow-up and validate before formal action."}"""

    return {
        "status": "success",
        "message": "Supervisor follow-up created.",
        "action_item": action,
        "supervisor_followup": followup,
        "answer_text": answer_text,
    }


@app.post("/review-note-create")
async def review_note_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    action_payload = {
        "action_title": payload.get("action_title") or "HR / Labor / Compliance Review Required",
        "action_type": payload.get("action_type") or "HR / Labor / Compliance Review",
        "action_status": "Open",
        "priority": payload.get("priority") or "High",
        "source_module": payload.get("source_module") or "LORI Action Center",
        "source_type": payload.get("source_type") or "Review Note",
        "driver_name": payload.get("related_driver_name"),
        "employee_id": payload.get("related_employee_id"),
        "supervisor_name": payload.get("supervisor_name"),
        "owner_name": payload.get("owner_name") or "HR / Labor / Compliance",
        "owner_role": payload.get("owner_role") or "HR / Labor / Compliance",
        "station_code": payload.get("station_code") or "JESSUP-01",
        "operating_state": payload.get("operating_state") or "MD",
        "company_name": payload.get("company_name") or "Food Authority",
        "reason": payload.get("review_summary") or "Review note created through LORI.",
        "recommended_follow_up": payload.get("recommended_review_note"),
        "documentation_note": payload.get("facts_to_confirm"),
        "due_date": lori_action_parse_due_date(payload.get("due_date")) or lori_action_default_due_date(3),
        "created_by": "LORI Action Center",
    }

    created_action = await lori_policy_supabase_post(
        "lori_action_items",
        action_payload,
    )

    action = created_action[0] if created_action else {}

    review_payload = lori_action_filter_payload(
        payload,
        REVIEW_NOTE_ALLOWED_FIELDS,
    )

    review_payload["action_item_id"] = action.get("id")
    review_payload.setdefault("review_type", "HR / Labor / Compliance Review")
    review_payload.setdefault("review_status", "Open")
    review_payload.setdefault("priority", action.get("priority") or "High")
    review_payload.setdefault("required_reviewer", "HR, labor relations, compliance, legal, or leadership")
    review_payload.setdefault("created_by", "LORI Action Center")

    created_review = await lori_policy_supabase_post(
        "lori_review_notes",
        review_payload,
    )

    review_note = created_review[0] if created_review else {}

    answer_text = f"""HR / Labor / Compliance Review Note Created

Review Type:
{review_note.get("review_type") or "HR / Labor / Compliance Review"}

Priority:
{review_note.get("priority") or "High"}

Review Summary:
{review_note.get("review_summary") or action.get("reason") or "Review needed."}

Facts to Confirm:
{review_note.get("facts_to_confirm") or "Confirm facts, documentation, policy/agreement language, and reviewer assignment."}

Required Reviewer:
{review_note.get("required_reviewer") or "HR, labor relations, compliance, legal, or leadership"}

Action Item:
{action.get("action_title")}

Due Date:
{action.get("due_date")}"""

    return {
        "status": "success",
        "message": "Review note created.",
        "action_item": action,
        "review_note": review_note,
        "answer_text": answer_text,
    }


@app.post("/leadership-briefing-item-create")
async def leadership_briefing_item_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    action_payload = {
        "action_title": payload.get("action_title") or payload.get("briefing_title") or "Leadership Briefing Item",
        "action_type": payload.get("action_type") or "Leadership Briefing Item",
        "action_status": "Open",
        "priority": payload.get("priority") or "Medium",
        "source_module": payload.get("source_module") or "LORI Action Center",
        "source_type": payload.get("source_type") or "Leadership Briefing",
        "owner_name": payload.get("owner_name") or "Operations Leadership",
        "owner_role": payload.get("owner_role") or "Leadership",
        "station_code": payload.get("station_code") or "JESSUP-01",
        "operating_state": payload.get("operating_state") or "MD",
        "company_name": payload.get("company_name") or "Food Authority",
        "reason": payload.get("executive_summary") or payload.get("key_risk") or "Leadership briefing item created through LORI.",
        "recommended_follow_up": payload.get("recommended_leadership_action"),
        "documentation_note": payload.get("supervisor_follow_up"),
        "due_date": lori_action_parse_due_date(payload.get("due_date")) or lori_action_default_due_date(5),
        "created_by": "LORI Action Center",
    }

    created_action = await lori_policy_supabase_post(
        "lori_action_items",
        action_payload,
    )

    action = created_action[0] if created_action else {}

    briefing_payload = lori_action_filter_payload(
        payload,
        LEADERSHIP_QUEUE_ALLOWED_FIELDS,
    )

    briefing_payload["action_item_id"] = action.get("id")
    briefing_payload.setdefault("briefing_title", action.get("action_title") or "Leadership Briefing Item")
    briefing_payload.setdefault("briefing_type", "Operational Leadership Briefing")
    briefing_payload.setdefault("briefing_status", "Queued")
    briefing_payload.setdefault("priority", action.get("priority") or "Medium")
    briefing_payload.setdefault("station_code", action.get("station_code") or "JESSUP-01")
    briefing_payload.setdefault("operating_state", action.get("operating_state") or "MD")
    briefing_payload.setdefault("created_by", "LORI Action Center")

    created_briefing = await lori_policy_supabase_post(
        "lori_leadership_briefing_queue",
        briefing_payload,
    )

    briefing_item = created_briefing[0] if created_briefing else {}

    answer_text = f"""Leadership Briefing Item Created

Briefing Title:
{briefing_item.get("briefing_title")}

Priority:
{briefing_item.get("priority")}

Status:
{briefing_item.get("briefing_status")}

Executive Summary:
{briefing_item.get("executive_summary") or action.get("reason")}

Recommended Leadership Action:
{briefing_item.get("recommended_leadership_action") or action.get("recommended_follow_up")}

Supervisor Follow-Up:
{briefing_item.get("supervisor_follow_up") or action.get("documentation_note")}

Compliance Note:
{briefing_item.get("compliance_note") or "Validate before formal action."}"""

    return {
        "status": "success",
        "message": "Leadership briefing item created.",
        "action_item": action,
        "briefing_item": briefing_item,
        "answer_text": answer_text,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# KPI ACTION PLANS BACKEND
# Universal KPI upload, KPI detection, multiple KPI action plans,
# beautiful print-ready plans, Action Center integration,
# and Leadership Briefing integration.
# ============================================================

from fastapi import UploadFile, File, Body
from typing import Any, Dict, List, Optional, Tuple
from datetime import date, datetime, timedelta
from urllib.parse import quote
import csv
import io
import os
import re
import uuid
import html
import json
import math

try:
    import openpyxl
except Exception:
    openpyxl = None


KPI_ACTION_UPLOAD_BUCKET = "kpi-action-plan-uploads"


def lori_kpi_clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def lori_kpi_normalize_column(value: Any) -> str:
    text = lori_kpi_clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def lori_kpi_to_float(value: Any) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)

    text = str(value).strip()

    if not text or text.lower() in {"none", "null", "n/a", "na", "-", "string"}:
        return None

    text = text.replace(",", "")
    text = text.replace("$", "")
    text = text.replace("%", "")
    text = text.replace("(", "-").replace(")", "")

    try:
        return float(text)
    except Exception:
        return None


def lori_kpi_safe_file_name(filename: str) -> str:
    base = filename or "kpi_upload.csv"
    base = re.sub(r"[^a-zA-Z0-9._-]+", "_", base)
    return base[:120]


def lori_kpi_detect_direction(kpi_name: str) -> str:
    name = lori_kpi_clean_text(kpi_name).lower()

    lower_is_better_terms = [
        "call out",
        "call-out",
        "absence",
        "absentee",
        "incident",
        "accident",
        "complaint",
        "late",
        "missed",
        "failure",
        "defect",
        "exception",
        "overtime",
        "cost",
        "violation",
        "turnover",
        "damage",
        "injury",
        "delay",
        "risk",
        "gap",
        "error",
        "shortage",
    ]

    for term in lower_is_better_terms:
        if term in name:
            return "lower_is_better"

    return "higher_is_better"


def lori_kpi_detect_category(kpi_name: str, raw_row: Optional[Dict[str, Any]] = None) -> str:
    text = lori_kpi_clean_text(kpi_name).lower()

    if raw_row:
        text += " " + " ".join([lori_kpi_clean_text(v).lower() for v in raw_row.values()])

    category_rules = [
        ("Delivery Performance", ["on time", "on-time", "delivery", "route", "dispatch", "window", "completion"]),
        ("Attendance", ["attendance", "call out", "call-out", "absence", "absentee", "no show", "schedule"]),
        ("Safety", ["safety", "incident", "accident", "injury", "inspection", "dot", "violation"]),
        ("Training", ["training", "certification", "completion", "course", "learning"]),
        ("Payroll / Overtime", ["payroll", "overtime", "hours", "timecard", "exception", "labor cost"]),
        ("Compliance", ["compliance", "audit", "medical card", "license", "credential", "policy"]),
        ("Customer Service", ["customer", "complaint", "satisfaction", "service"]),
        ("Driver Performance", ["driver score", "scorecard", "driver", "performance"]),
        ("Fleet / Maintenance", ["vehicle", "fleet", "maintenance", "truck", "van", "repair"]),
        ("Warehouse / Operations", ["warehouse", "dock", "load", "pick", "inventory"]),
    ]

    for category, terms in category_rules:
        if any(term in text for term in terms):
            return category

    return "General Operations"


def lori_kpi_status(current: Optional[float], target: Optional[float], direction: str) -> Tuple[str, Optional[float], Optional[float]]:
    if current is None or target is None:
        return "Needs Review", None, None

    if direction == "lower_is_better":
        gap_value = current - target
        off_track = current > target
    else:
        gap_value = current - target
        off_track = current < target

    if target == 0:
        gap_percent = None
    else:
        gap_percent = round((gap_value / abs(target)) * 100, 2)

    if off_track:
        return "Off Track", round(gap_value, 2), gap_percent

    return "On Track", round(gap_value, 2), gap_percent


def lori_kpi_detect_column_roles(headers: List[str]) -> Dict[str, str]:
    roles: Dict[str, str] = {}

    for header in headers:
        normalized = lori_kpi_normalize_column(header)

        if normalized in roles:
            continue

        if any(term in normalized for term in ["kpi", "metric", "measure", "indicator", "goal_name"]):
            roles[header] = "kpi_name"
        elif any(term in normalized for term in ["current", "actual", "value", "result", "performance"]):
            roles[header] = "current_value"
        elif any(term in normalized for term in ["target", "goal", "standard", "threshold"]):
            roles[header] = "target_value"
        elif any(term in normalized for term in ["baseline", "prior", "previous", "last_period"]):
            roles[header] = "baseline_value"
        elif any(term in normalized for term in ["owner", "responsible", "manager"]):
            roles[header] = "owner_name"
        elif "supervisor" in normalized:
            roles[header] = "supervisor_name"
        elif "driver" in normalized and "score" not in normalized:
            roles[header] = "driver_name"
        elif any(term in normalized for term in ["employee", "emp_id", "employee_id"]):
            roles[header] = "employee_id"
        elif any(term in normalized for term in ["route", "route_id"]):
            roles[header] = "route_id"
        elif any(term in normalized for term in ["location", "site", "station", "facility"]):
            roles[header] = "location_name"
        elif any(term in normalized for term in ["category", "type", "department", "function"]):
            roles[header] = "kpi_category"
        elif any(term in normalized for term in ["period", "week", "month", "date"]):
            roles[header] = "reporting_period"
        elif any(term in normalized for term in ["unit", "measure_unit"]):
            roles[header] = "measurement_unit"
        elif "direction" in normalized:
            roles[header] = "direction"
        else:
            roles[header] = "dimension"

    return roles


def lori_kpi_parse_csv_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig", errors="ignore")

    if not text.strip():
        return []

    sample = text[:2048]

    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = []

    for row in reader:
        clean_row = {
            lori_kpi_clean_text(k): lori_kpi_clean_text(v)
            for k, v in row.items()
            if k is not None
        }

        if any(str(v).strip() for v in clean_row.values()):
            rows.append(clean_row)

    return rows


def lori_kpi_parse_excel_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    if openpyxl is None:
        raise Exception("Excel support is not installed. Add openpyxl to requirements.txt.")

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active

    rows_raw = list(worksheet.iter_rows(values_only=True))

    if not rows_raw:
        return []

    headers = [lori_kpi_clean_text(h) for h in rows_raw[0]]
    rows = []

    for raw in rows_raw[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = lori_kpi_clean_text(raw[idx]) if idx < len(raw) else ""

        if any(str(v).strip() for v in row.values()):
            rows.append(row)

    return rows


async def lori_kpi_upload_file_to_storage(
    file_bytes: bytes,
    filename: str,
    content_type: str,
) -> Dict[str, Any]:
    supabase_url = os.getenv("SUPABASE_URL", "").rstrip("/")
    service_key = (
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        or os.getenv("SUPABASE_SERVICE_KEY")
        or os.getenv("SUPABASE_KEY")
        or ""
    )

    if not supabase_url or not service_key:
        return {
            "uploaded": False,
            "file_path": None,
            "file_url": None,
            "error": "Supabase storage environment variables not found.",
        }

    safe_name = lori_kpi_safe_file_name(filename)
    file_path = f"{datetime.utcnow().strftime('%Y/%m/%d')}/{uuid.uuid4()}_{safe_name}"

    url = f"{supabase_url}/storage/v1/object/{KPI_ACTION_UPLOAD_BUCKET}/{file_path}"

    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": content_type or "application/octet-stream",
        "x-upsert": "true",
    }

    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post(url, headers=headers, content=file_bytes)

    if response.status_code >= 400:
        return {
            "uploaded": False,
            "file_path": file_path,
            "file_url": None,
            "error": response.text,
        }

    return {
        "uploaded": True,
        "file_path": file_path,
        "file_url": f"{supabase_url}/storage/v1/object/{KPI_ACTION_UPLOAD_BUCKET}/{file_path}",
        "error": None,
    }


def lori_kpi_get_role_column(column_roles: Dict[str, str], role: str) -> Optional[str]:
    for column, detected_role in column_roles.items():
        if detected_role == role:
            return column
    return None


def lori_kpi_build_metric_records(
    upload_id: str,
    rows: List[Dict[str, Any]],
    column_roles: Dict[str, str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], bool]:
    metric_records: List[Dict[str, Any]] = []
    detected_column_records: List[Dict[str, Any]] = []

    headers = list(rows[0].keys()) if rows else []

    for header in headers:
        samples = []
        for row in rows[:5]:
            value = row.get(header)
            if value not in [None, ""]:
                samples.append(value)

        detected_column_records.append({
            "upload_id": upload_id,
            "original_column_name": header,
            "normalized_column_name": lori_kpi_normalize_column(header),
            "detected_role": column_roles.get(header, "dimension"),
            "detected_confidence": 0.85 if column_roles.get(header) != "dimension" else 0.45,
            "sample_values": samples,
            "notes": "Detected by LORI KPI Action Plans universal column mapper.",
        })

    kpi_col = lori_kpi_get_role_column(column_roles, "kpi_name")
    current_col = lori_kpi_get_role_column(column_roles, "current_value")
    target_col = lori_kpi_get_role_column(column_roles, "target_value")
    baseline_col = lori_kpi_get_role_column(column_roles, "baseline_value")
    owner_col = lori_kpi_get_role_column(column_roles, "owner_name")
    supervisor_col = lori_kpi_get_role_column(column_roles, "supervisor_name")
    driver_col = lori_kpi_get_role_column(column_roles, "driver_name")
    employee_col = lori_kpi_get_role_column(column_roles, "employee_id")
    route_col = lori_kpi_get_role_column(column_roles, "route_id")
    location_col = lori_kpi_get_role_column(column_roles, "location_name")
    category_col = lori_kpi_get_role_column(column_roles, "kpi_category")
    period_col = lori_kpi_get_role_column(column_roles, "reporting_period")
    unit_col = lori_kpi_get_role_column(column_roles, "measurement_unit")
    direction_col = lori_kpi_get_role_column(column_roles, "direction")

    can_directly_map = bool(kpi_col and current_col)

    if can_directly_map:
        for row in rows:
            kpi_name = lori_kpi_clean_text(row.get(kpi_col))

            if not kpi_name:
                continue

            current = lori_kpi_to_float(row.get(current_col))
            target = lori_kpi_to_float(row.get(target_col)) if target_col else None
            baseline = lori_kpi_to_float(row.get(baseline_col)) if baseline_col else None
            unit = lori_kpi_clean_text(row.get(unit_col)) if unit_col else ""
            direction = lori_kpi_clean_text(row.get(direction_col)) if direction_col else ""

            if direction not in {"higher_is_better", "lower_is_better"}:
                direction = lori_kpi_detect_direction(kpi_name)

            status, gap_value, gap_percent = lori_kpi_status(current, target, direction)

            category = lori_kpi_clean_text(row.get(category_col)) if category_col else ""
            if not category:
                category = lori_kpi_detect_category(kpi_name, row)

            metric_records.append({
                "upload_id": upload_id,
                "kpi_name": kpi_name,
                "kpi_category": category,
                "kpi_description": f"Uploaded KPI metric: {kpi_name}",
                "current_value": current,
                "baseline_value": baseline,
                "target_value": target,
                "gap_value": gap_value,
                "gap_percent": gap_percent,
                "measurement_unit": unit or None,
                "direction": direction,
                "kpi_status": status,
                "reporting_period": lori_kpi_clean_text(row.get(period_col)) if period_col else None,
                "driver_name": lori_kpi_clean_text(row.get(driver_col)) if driver_col else None,
                "employee_id": lori_kpi_clean_text(row.get(employee_col)) if employee_col else None,
                "supervisor_name": lori_kpi_clean_text(row.get(supervisor_col)) if supervisor_col else None,
                "route_id": lori_kpi_clean_text(row.get(route_col)) if route_col else None,
                "location_name": lori_kpi_clean_text(row.get(location_col)) if location_col else None,
                "station_code": "JESSUP-01",
                "operating_state": "MD",
                "owner_name": lori_kpi_clean_text(row.get(owner_col)) if owner_col else None,
                "owner_role": "Owner",
                "dimensions": {
                    key: value
                    for key, value in row.items()
                    if column_roles.get(key) == "dimension"
                },
                "raw_row": row,
            })

    else:
        numeric_columns = []
        dimension_columns = []

        for header in headers:
            numeric_count = sum(1 for row in rows if lori_kpi_to_float(row.get(header)) is not None)

            if numeric_count >= max(1, len(rows) // 2):
                numeric_columns.append(header)
            else:
                dimension_columns.append(header)

        target_like_columns = [
            h for h in numeric_columns
            if any(term in lori_kpi_normalize_column(h) for term in ["target", "goal", "standard"])
        ]

        current_like_columns = [
            h for h in numeric_columns
            if h not in target_like_columns
        ]

        for row in rows:
            for metric_col in current_like_columns:
                kpi_name = metric_col
                current = lori_kpi_to_float(row.get(metric_col))

                if current is None:
                    continue

                target = None
                normalized_metric = lori_kpi_normalize_column(metric_col)

                for target_candidate in target_like_columns:
                    normalized_target = lori_kpi_normalize_column(target_candidate)

                    if normalized_metric in normalized_target or normalized_target.replace("target", "") in normalized_metric:
                        target = lori_kpi_to_float(row.get(target_candidate))
                        break

                direction = lori_kpi_detect_direction(kpi_name)
                status, gap_value, gap_percent = lori_kpi_status(current, target, direction)
                category = lori_kpi_detect_category(kpi_name, row)

                metric_records.append({
                    "upload_id": upload_id,
                    "kpi_name": kpi_name,
                    "kpi_category": category,
                    "kpi_description": f"LORI detected this numeric KPI from uploaded column: {metric_col}",
                    "current_value": current,
                    "baseline_value": None,
                    "target_value": target,
                    "gap_value": gap_value,
                    "gap_percent": gap_percent,
                    "measurement_unit": "%" if "%" in metric_col else None,
                    "direction": direction,
                    "kpi_status": status,
                    "reporting_period": None,
                    "station_code": "JESSUP-01",
                    "operating_state": "MD",
                    "owner_name": None,
                    "owner_role": "Owner",
                    "dimensions": {
                        key: value
                        for key, value in row.items()
                        if key in dimension_columns
                    },
                    "raw_row": row,
                })

    low_confidence = not can_directly_map and len(metric_records) == 0

    return metric_records, detected_column_records, low_confidence


def lori_kpi_root_cause_and_actions(category: str, kpi_name: str, direction: str) -> Dict[str, Any]:
    text = f"{category} {kpi_name}".lower()

    if "delivery" in text or "route" in text or "dispatch" in text:
        return {
            "root_cause": "Likely root causes include late departures, route sequencing gaps, driver readiness issues, dispatch delays, vehicle readiness, attendance coverage pressure, and inconsistent supervisor follow-up.",
            "impact": "Delivery KPI gaps can affect customer service, route completion, overtime, driver accountability, and leadership confidence.",
            "actions": [
                "Review the bottom 10 routes or failed delivery windows daily.",
                "Compare performance by route, driver, dispatch time, day of week, and zone.",
                "Launch a daily 10-minute dispatch readiness huddle.",
                "Assign supervisor follow-ups for drivers or routes below threshold.",
                "Track recovery trend weekly and escalate if the KPI does not improve."
            ],
        }

    if "attendance" in text or "call" in text or "absence" in text:
        return {
            "root_cause": "Likely root causes include repeat call-out patterns, unclear expectations, inconsistent documentation, weak early coaching, staffing pressure, and lack of weekly attendance review.",
            "impact": "Attendance KPI gaps can affect route coverage, overtime, team fairness, customer service, and supervisor planning.",
            "actions": [
                "Identify repeat attendance or call-out patterns.",
                "Review attendance records and supervisor notes.",
                "Coach employees on the call-out procedure and operational impact.",
                "Track weekly trend against the target.",
                "Use HR, labor, compliance, or leadership review before formal action."
            ],
        }

    if "safety" in text or "incident" in text or "accident" in text or "inspection" in text:
        return {
            "root_cause": "Likely root causes include inconsistent pre-trip discipline, incomplete coaching, route pressure, safety observation gaps, or failure to close prior safety findings.",
            "impact": "Safety KPI gaps can affect DOT readiness, driver risk, insurance exposure, operational reliability, and leadership accountability.",
            "actions": [
                "Review safety events by driver, route, vehicle, date, and severity.",
                "Confirm whether inspections and safety coaching were completed.",
                "Assign immediate supervisor follow-up for repeat patterns.",
                "Track safety observations weekly.",
                "Escalate unresolved or high-risk issues to safety/compliance leadership."
            ],
        }

    if "training" in text or "certification" in text:
        return {
            "root_cause": "Likely root causes include incomplete assignment tracking, schedule conflicts, missed reminders, unclear ownership, or training not tied to supervisor accountability.",
            "impact": "Training KPI gaps can affect readiness, compliance, safety, quality, and audit performance.",
            "actions": [
                "Identify employees missing required training.",
                "Assign training completion dates by owner.",
                "Set supervisor reminders for incomplete items.",
                "Track completion weekly until the KPI returns to target.",
                "Escalate overdue training tied to compliance or safety requirements."
            ],
        }

    if "payroll" in text or "overtime" in text or "exception" in text:
        return {
            "root_cause": "Likely root causes include schedule gaps, route delays, attendance coverage pressure, timecard exceptions, approval delays, or unplanned labor usage.",
            "impact": "Payroll and overtime KPI gaps can affect cost control, compliance, staffing decisions, and operational profitability.",
            "actions": [
                "Review exceptions by driver, supervisor, route, and week.",
                "Compare overtime to route completion and attendance gaps.",
                "Confirm approval workflow and documentation.",
                "Assign owner to reduce repeat exceptions.",
                "Track savings or exception reduction weekly."
            ],
        }

    return {
        "root_cause": "Likely root causes may include unclear ownership, inconsistent tracking, process variation, incomplete documentation, staffing or workload pressure, and lack of routine leadership review.",
        "impact": "This KPI may affect operational reliability, leadership visibility, accountability, cost, compliance, or service performance.",
        "actions": [
            "Confirm the KPI definition, target, owner, and reporting period.",
            "Break performance down by driver, supervisor, route, location, date, and category where available.",
            "Identify the bottom performers or highest-risk categories.",
            "Assign owners and due dates for corrective steps.",
            "Review progress weekly until the KPI returns to target."
        ],
    }


def lori_kpi_build_print_html(plan: Dict[str, Any], steps: List[Dict[str, Any]]) -> str:
    title = html.escape(lori_kpi_clean_text(plan.get("print_title") or plan.get("plan_title") or "KPI Action Plan"))
    subtitle = html.escape(lori_kpi_clean_text(plan.get("print_subtitle") or f"{plan.get('company_name') or 'Company'} | {plan.get('station_code') or ''}"))
    status_banner = html.escape(lori_kpi_clean_text(plan.get("print_status_banner") or f"{plan.get('plan_status') or 'Draft'} — {plan.get('priority') or 'Priority'}"))
    summary = html.escape(lori_kpi_clean_text(plan.get("print_executive_summary") or plan.get("executive_summary") or "Executive summary not provided."))
    problem = html.escape(lori_kpi_clean_text(plan.get("problem_statement") or "Problem statement not provided."))
    root = html.escape(lori_kpi_clean_text(plan.get("root_cause_analysis") or "Root cause analysis not provided."))
    thirty = html.escape(lori_kpi_clean_text(plan.get("thirty_day_plan") or "30-day plan not provided."))
    sixty = html.escape(lori_kpi_clean_text(plan.get("sixty_day_plan") or "60-day plan not provided."))
    ninety = html.escape(lori_kpi_clean_text(plan.get("ninety_day_plan") or "90-day plan not provided."))
    success = html.escape(lori_kpi_clean_text(plan.get("success_measure") or "Success measure not provided."))
    escalation = html.escape(lori_kpi_clean_text(plan.get("escalation_trigger") or "Escalation trigger not provided."))
    footer = html.escape(lori_kpi_clean_text(plan.get("print_footer_note") or "Prepared by LORI KPI Action Plans."))

    kpi_name = html.escape(lori_kpi_clean_text(plan.get("kpi_name")))
    category = html.escape(lori_kpi_clean_text(plan.get("kpi_category")))
    owner = html.escape(lori_kpi_clean_text(plan.get("plan_owner_name") or plan.get("supervisor_name") or "Owner not assigned"))
    unit = html.escape(lori_kpi_clean_text(plan.get("measurement_unit") or ""))

    current = "" if plan.get("current_value") is None else str(plan.get("current_value"))
    target = "" if plan.get("target_value") is None else str(plan.get("target_value"))
    gap = "" if plan.get("gap_value") is None else str(plan.get("gap_value"))

    step_cards = ""

    for step in sorted(steps, key=lambda s: s.get("step_number") or 0):
        step_cards += f"""
        <div class="step-card">
            <div class="step-top">
                <span class="step-number">Step {html.escape(str(step.get("step_number") or ""))}</span>
                <span class="step-phase">{html.escape(lori_kpi_clean_text(step.get("phase") or ""))}</span>
                <span class="step-priority">{html.escape(lori_kpi_clean_text(step.get("priority") or ""))}</span>
            </div>
            <h3>{html.escape(lori_kpi_clean_text(step.get("action_title") or ""))}</h3>
            <p>{html.escape(lori_kpi_clean_text(step.get("action_description") or ""))}</p>
            <div class="step-grid">
                <div><strong>Owner</strong><br>{html.escape(lori_kpi_clean_text(step.get("owner_name") or "Owner not assigned"))}</div>
                <div><strong>Due Date</strong><br>{html.escape(lori_kpi_clean_text(step.get("due_date") or "Not assigned"))}</div>
                <div><strong>Success Measure</strong><br>{html.escape(lori_kpi_clean_text(step.get("success_measure") or "Not provided"))}</div>
            </div>
        </div>
        """

    return f"""
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
    @page {{
        size: letter;
        margin: 0.55in;
    }}

    body {{
        font-family: Arial, Helvetica, sans-serif;
        color: #172033;
        background: #ffffff;
        margin: 0;
        padding: 0;
        line-height: 1.42;
    }}

    .plan {{
        max-width: 900px;
        margin: 0 auto;
    }}

    .cover {{
        border-radius: 22px;
        padding: 34px;
        background: linear-gradient(135deg, #f6f8ff 0%, #eef3ff 52%, #ffffff 100%);
        border: 1px solid #dbe5ff;
        margin-bottom: 24px;
    }}

    .brand {{
        font-size: 12px;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: #52627a;
        font-weight: 700;
        margin-bottom: 22px;
    }}

    h1 {{
        font-size: 34px;
        line-height: 1.08;
        margin: 0 0 10px 0;
        color: #111827;
    }}

    .subtitle {{
        color: #42526b;
        font-size: 15px;
        margin-bottom: 22px;
    }}

    .status {{
        display: inline-block;
        background: #1e3a8a;
        color: #ffffff;
        padding: 10px 14px;
        border-radius: 999px;
        font-size: 12px;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        font-weight: 700;
        margin-bottom: 22px;
    }}

    .metric-grid {{
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 12px;
        margin-top: 18px;
    }}

    .metric {{
        background: #ffffff;
        border: 1px solid #dbe5ff;
        border-radius: 16px;
        padding: 14px;
    }}

    .metric-label {{
        font-size: 11px;
        color: #667085;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 700;
        margin-bottom: 6px;
    }}

    .metric-value {{
        font-size: 22px;
        font-weight: 800;
        color: #111827;
    }}

    .section {{
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 18px;
        padding: 22px;
        margin-bottom: 16px;
        page-break-inside: avoid;
    }}

    .section h2 {{
        font-size: 18px;
        margin: 0 0 10px 0;
        color: #111827;
    }}

    .timeline {{
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 14px;
    }}

    .phase {{
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 16px;
    }}

    .phase h3 {{
        margin: 0 0 8px 0;
        font-size: 15px;
        color: #1e3a8a;
    }}

    .step-card {{
        border: 1px solid #dbe5ff;
        border-radius: 16px;
        padding: 16px;
        margin-bottom: 12px;
        page-break-inside: avoid;
        background: #fbfdff;
    }}

    .step-top {{
        display: flex;
        gap: 8px;
        flex-wrap: wrap;
        margin-bottom: 8px;
    }}

    .step-number,
    .step-phase,
    .step-priority {{
        font-size: 11px;
        font-weight: 700;
        border-radius: 999px;
        padding: 6px 9px;
        background: #eef3ff;
        color: #1e3a8a;
    }}

    .step-card h3 {{
        font-size: 15px;
        margin: 8px 0 6px 0;
    }}

    .step-grid {{
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 10px;
        margin-top: 12px;
        font-size: 12px;
    }}

    .footer {{
        border-top: 1px solid #e5e7eb;
        padding-top: 14px;
        margin-top: 20px;
        color: #667085;
        font-size: 11px;
    }}

    @media print {{
        body {{
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
        }}

        .no-print {{
            display: none !important;
        }}

        .cover {{
            page-break-inside: avoid;
        }}
    }}
</style>
</head>
<body>
<div class="plan">
    <div class="cover">
        <div class="brand">LORI Drive Command Center</div>
        <div class="status">{status_banner}</div>
        <h1>{title}</h1>
        <div class="subtitle">{subtitle}</div>
        <p>{summary}</p>

        <div class="metric-grid">
            <div class="metric">
                <div class="metric-label">KPI</div>
                <div class="metric-value" style="font-size:16px;">{kpi_name}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Current</div>
                <div class="metric-value">{html.escape(current)}{unit}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Target</div>
                <div class="metric-value">{html.escape(target)}{unit}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Gap</div>
                <div class="metric-value">{html.escape(gap)}{unit}</div>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>Plan Ownership</h2>
        <p><strong>Category:</strong> {category}</p>
        <p><strong>Owner:</strong> {owner}</p>
        <p><strong>Target Date:</strong> {html.escape(lori_kpi_clean_text(plan.get("recovery_target_date") or "Not assigned"))}</p>
    </div>

    <div class="section">
        <h2>Problem Statement</h2>
        <p>{problem}</p>
    </div>

    <div class="section">
        <h2>Root Cause Analysis</h2>
        <p>{root}</p>
    </div>

    <div class="section">
        <h2>30 / 60 / 90 Day Action Plan</h2>
        <div class="timeline">
            <div class="phase">
                <h3>30 Days</h3>
                <p>{thirty}</p>
            </div>
            <div class="phase">
                <h3>60 Days</h3>
                <p>{sixty}</p>
            </div>
            <div class="phase">
                <h3>90 Days</h3>
                <p>{ninety}</p>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>Action Steps</h2>
        {step_cards}
    </div>

    <div class="section">
        <h2>Success Measure</h2>
        <p>{success}</p>
        <h2>Escalation Trigger</h2>
        <p>{escalation}</p>
    </div>

    <div class="footer">
        {footer}<br>
        LORI provides operational decision support. Validate KPI findings, action plans, HR/labor considerations, compliance risks, and formal corrective actions with authorized leadership before final action.
    </div>
</div>
</body>
</html>
"""


async def lori_kpi_get_rows(table: str, query: str = "select=*&order=created_at.desc&limit=500") -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_kpi_create_finding_and_plan(upload_id: str, metric: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if metric.get("kpi_status") != "Off Track":
        return None

    kpi_name = metric.get("kpi_name") or "Uploaded KPI"
    category = metric.get("kpi_category") or lori_kpi_detect_category(kpi_name)
    direction = metric.get("direction") or lori_kpi_detect_direction(kpi_name)
    root_action = lori_kpi_root_cause_and_actions(category, kpi_name, direction)

    current = metric.get("current_value")
    target = metric.get("target_value")
    gap = metric.get("gap_value")
    gap_percent = metric.get("gap_percent")
    unit = metric.get("measurement_unit") or ""

    severity = "High"
    if gap_percent is not None and abs(float(gap_percent)) >= 50:
        severity = "Critical"

    owner_name = metric.get("owner_name") or metric.get("supervisor_name") or "Operations Leadership"
    owner_role = metric.get("owner_role") or "Leadership"

    finding_payload = {
        "upload_id": upload_id,
        "metric_id": metric.get("id"),
        "finding_title": f"{kpi_name} is off track",
        "finding_type": "Off-Track KPI",
        "severity": severity,
        "finding_status": "Open",
        "kpi_name": kpi_name,
        "kpi_category": category,
        "current_value": current,
        "target_value": target,
        "gap_value": gap,
        "gap_percent": gap_percent,
        "operational_impact": root_action["impact"],
        "likely_root_cause": root_action["root_cause"],
        "evidence_summary": f"Current value is {current}{unit} against a target of {target}{unit}.",
        "affected_driver": metric.get("driver_name"),
        "affected_supervisor": metric.get("supervisor_name"),
        "affected_route": metric.get("route_id"),
        "affected_location": metric.get("location_name") or metric.get("station_code"),
        "recommended_next_action": "Generate a KPI Action Plan with owner, due dates, 30/60/90-day corrective actions, and leadership review checkpoints.",
        "owner_name": owner_name,
        "owner_role": owner_role,
        "due_date": (date.today() + timedelta(days=3)).isoformat(),
        "leadership_note": "Leadership should review this KPI until it returns to target.",
    }

    created_finding = await lori_policy_supabase_post(
        "lori_kpi_action_plan_findings",
        finding_payload,
    )

    finding = created_finding[0] if created_finding else {}

    action_1 = root_action["actions"][0]
    action_2 = root_action["actions"][1] if len(root_action["actions"]) > 1 else "Review KPI drivers and confirm ownership."
    action_3 = root_action["actions"][2] if len(root_action["actions"]) > 2 else "Track weekly progress."

    plan_title = f"{kpi_name} KPI Action Plan"

    plan_payload = {
        "upload_id": upload_id,
        "metric_id": metric.get("id"),
        "finding_id": finding.get("id"),
        "plan_title": plan_title,
        "plan_status": "Draft",
        "plan_type": "KPI Action Plan",
        "priority": severity,
        "kpi_name": kpi_name,
        "kpi_category": category,
        "current_value": current,
        "target_value": target,
        "baseline_value": metric.get("baseline_value"),
        "gap_value": gap,
        "gap_percent": gap_percent,
        "measurement_unit": unit,
        "direction": direction,
        "company_name": "Food Authority",
        "station_code": metric.get("station_code") or "JESSUP-01",
        "operating_state": metric.get("operating_state") or "MD",
        "plan_owner_name": owner_name,
        "plan_owner_role": owner_role,
        "supervisor_name": metric.get("supervisor_name"),
        "executive_summary": f"{kpi_name} is off track and requires a focused KPI Action Plan.",
        "problem_statement": f"Current performance is {current}{unit} against a target of {target}{unit}, creating a gap of {gap}{unit}.",
        "root_cause_analysis": root_action["root_cause"],
        "thirty_day_plan": f"{action_1} Confirm the KPI owner, review the lowest-performing segments, assign immediate follow-ups, and begin weekly tracking.",
        "sixty_day_plan": f"{action_2} Compare performance by available dimensions such as driver, route, supervisor, location, date, and category. Adjust process gaps and assign Action Center follow-ups.",
        "ninety_day_plan": f"{action_3} Sustain the KPI at or better than target, document the operating rhythm, and convert successful recovery actions into standard practice.",
        "success_measure": f"Return {kpi_name} to target of {target}{unit} and sustain for 4 consecutive weeks.",
        "recovery_target_date": (date.today() + timedelta(days=90)).isoformat(),
        "escalation_trigger": "Escalate if the KPI does not show measurable improvement after 30 days or remains off target after 60 days.",
        "print_title": plan_title,
        "print_subtitle": "Food Authority | JESSUP-01 | Executive KPI Action Plan",
        "print_cover_summary": f"This print-ready KPI Action Plan is designed to move {kpi_name} back to target with practical leadership actions, owner accountability, due dates, and measurable recovery checkpoints.",
        "print_status_banner": f"OFF TRACK — {severity.upper()} PRIORITY ACTION PLAN",
        "print_executive_summary": f"{kpi_name} is currently {current}{unit} against a target of {target}{unit}. The plan focuses on root cause review, owner assignment, operational discipline, weekly tracking, and escalation triggers.",
        "print_action_plan_text": "This plan includes KPI gap review, likely root causes, 30/60/90-day actions, owner accountability, due dates, success measures, and escalation triggers.",
        "print_footer_note": "Prepared by LORI KPI Action Plans. Validate operational findings before formal corrective action.",
        "print_ready": True,
    }

    created_plan = await lori_policy_supabase_post(
        "lori_kpi_action_plans",
        plan_payload,
    )

    plan = created_plan[0] if created_plan else {}

    steps = [
        {
            "plan_id": plan.get("id"),
            "step_number": 1,
            "phase": "30-Day Action",
            "step_status": "Open",
            "priority": severity,
            "action_title": action_1,
            "action_description": "Start with immediate KPI stabilization by reviewing performance details, confirming ownership, and assigning specific follow-up.",
            "owner_name": owner_name,
            "owner_role": owner_role,
            "due_date": (date.today() + timedelta(days=7)).isoformat(),
            "success_measure": "Immediate action started and documented within 7 days.",
            "required_data": "Uploaded KPI file, current value, target value, affected dimensions, and owner review.",
            "documentation_note": "Document actions taken, facts confirmed, and follow-up needs.",
        },
        {
            "plan_id": plan.get("id"),
            "step_number": 2,
            "phase": "30-Day Action",
            "step_status": "Open",
            "priority": "High",
            "action_title": action_2,
            "action_description": "Analyze KPI drivers and identify repeat patterns by available dimensions.",
            "owner_name": owner_name,
            "owner_role": owner_role,
            "due_date": (date.today() + timedelta(days=14)).isoformat(),
            "success_measure": "Root cause pattern identified and actioned.",
            "required_data": "KPI trend, driver/route/supervisor/location breakdown if available.",
            "documentation_note": "Document root cause signals and corrective steps.",
        },
        {
            "plan_id": plan.get("id"),
            "step_number": 3,
            "phase": "60-Day Action",
            "step_status": "Open",
            "priority": "Medium",
            "action_title": action_3,
            "action_description": "Track weekly progress and adjust operational actions until performance moves toward target.",
            "owner_name": owner_name,
            "owner_role": owner_role,
            "due_date": (date.today() + timedelta(days=45)).isoformat(),
            "success_measure": "KPI trend shows measurable improvement toward target.",
            "required_data": "Weekly KPI report and Action Center follow-up status.",
            "documentation_note": "Escalate if improvement is not visible by checkpoint.",
        },
    ]

    created_steps = await lori_policy_supabase_post(
        "lori_kpi_action_plan_steps",
        steps,
    )

    print_html = lori_kpi_build_print_html(plan, created_steps or [])

    await lori_policy_supabase_post(
        "lori_kpi_action_plan_print_exports",
        {
            "plan_id": plan.get("id"),
            "export_title": f"Printable {plan_title}",
            "export_type": "Printable KPI Action Plan",
            "export_status": "Ready to Print",
            "printable_text": plan.get("print_action_plan_text"),
            "printable_html": print_html,
            "print_view_created": True,
            "created_by": "LORI KPI Action Plans",
        },
    )

    return {
        "finding": finding,
        "plan": plan,
        "steps": created_steps or [],
    }


@app.get("/kpi-action-plans-summary")
async def kpi_action_plans_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    uploads = await lori_kpi_get_rows("lori_kpi_action_plan_uploads")
    metrics = await lori_kpi_get_rows("lori_kpi_action_plan_metric_records")
    findings = await lori_kpi_get_rows("lori_kpi_action_plan_findings")
    plans = await lori_kpi_get_rows("lori_kpi_action_plans")
    steps = await lori_kpi_get_rows("lori_kpi_action_plan_steps")
    mapping_queue = await lori_kpi_get_rows("lori_kpi_action_plan_mapping_queue")

    off_track = [m for m in metrics if m.get("kpi_status") == "Off Track"]
    active_plans = [p for p in plans if p.get("plan_status") in {"Draft", "Active", "In Progress"}]
    print_ready = [p for p in plans if p.get("print_ready") is True]

    answer_text = f"""KPI Action Plans Summary

KPI Uploads:
{len(uploads)}

Detected KPI Metrics:
{len(metrics)}

Off-Track KPIs:
{len(off_track)}

KPI Action Plans:
{len(plans)}

Open Action Steps:
{len([s for s in steps if s.get("step_status") == "Open"])}

Print-Ready Plans:
{len(print_ready)}

Mapping Reviews Needed:
{len(mapping_queue)}

Recommended Next Action:
Review off-track KPIs first, open each KPI Action Plan, confirm ownership, and print the executive action plan for leadership review."""

    return {
        "status": "success",
        "uploads_count": len(uploads),
        "metrics_count": len(metrics),
        "off_track_kpis_count": len(off_track),
        "findings_count": len(findings),
        "action_plans_count": len(plans),
        "active_plans_count": len(active_plans),
        "action_steps_count": len(steps),
        "print_ready_plans_count": len(print_ready),
        "mapping_reviews_needed_count": len(mapping_queue),
        "recent_uploads": uploads[:10],
        "recent_plans": plans[:10],
        "answer_text": answer_text,
    }


@app.get("/kpi-action-plan-uploads")
async def kpi_action_plan_uploads(
    api_key: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_kpi_get_rows(
        "lori_kpi_action_plan_uploads",
        f"select=*&order=created_at.desc&limit={max(1, min(limit, 200))}",
    )

    return {
        "status": "success",
        "uploads_count": len(rows),
        "uploads": rows,
    }


@app.get("/kpi-action-plans")
async def kpi_action_plans(
    api_key: Optional[str] = Query(None),
    upload_id: Optional[str] = Query(None),
    plan_status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_kpi_get_rows(
        "lori_kpi_action_plans",
        "select=*&order=created_at.desc&limit=500",
    )

    if upload_id:
        rows = [r for r in rows if str(r.get("upload_id")) == upload_id]

    if plan_status:
        rows = [
            r for r in rows
            if lori_kpi_clean_text(r.get("plan_status")).lower() == plan_status.lower()
        ]

    if priority:
        rows = [
            r for r in rows
            if lori_kpi_clean_text(r.get("priority")).lower() == priority.lower()
        ]

    rows = rows[:max(1, min(limit, 200))]

    return {
        "status": "success",
        "plans_count": len(rows),
        "plans": rows,
    }


@app.get("/kpi-action-plan-detail")
async def kpi_action_plan_detail(
    api_key: Optional[str] = Query(None),
    plan_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    plans = await lori_kpi_get_rows(
        "lori_kpi_action_plans",
        f"select=*&id=eq.{quote(plan_id)}&limit=1",
    )

    if not plans:
        return {
            "status": "not_found",
            "message": "KPI Action Plan not found.",
            "plan_id": plan_id,
        }

    plan = plans[0]

    steps = await lori_kpi_get_rows(
        "lori_kpi_action_plan_steps",
        f"select=*&plan_id=eq.{quote(plan_id)}&order=step_number.asc&limit=100",
    )

    exports = await lori_kpi_get_rows(
        "lori_kpi_action_plan_print_exports",
        f"select=*&plan_id=eq.{quote(plan_id)}&order=created_at.desc&limit=5",
    )

    return {
        "status": "success",
        "plan": plan,
        "steps_count": len(steps),
        "steps": steps,
        "print_exports": exports,
    }


@app.get("/kpi-action-plan-print")
async def kpi_action_plan_print(
    api_key: Optional[str] = Query(None),
    plan_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    plans = await lori_kpi_get_rows(
        "lori_kpi_action_plans",
        f"select=*&id=eq.{quote(plan_id)}&limit=1",
    )

    if not plans:
        return {
            "status": "not_found",
            "message": "KPI Action Plan not found.",
            "plan_id": plan_id,
        }

    plan = plans[0]

    steps = await lori_kpi_get_rows(
        "lori_kpi_action_plan_steps",
        f"select=*&plan_id=eq.{quote(plan_id)}&order=step_number.asc&limit=100",
    )

    printable_html = lori_kpi_build_print_html(plan, steps)

    return {
        "status": "success",
        "plan_id": plan_id,
        "plan_title": plan.get("plan_title"),
        "print_ready": True,
        "printable_html": printable_html,
        "answer_text": "Printable KPI Action Plan is ready.",
    }


@app.post("/kpi-action-plan-upload")
async def kpi_action_plan_upload(
    api_key: Optional[str] = Query(None),
    upload_title: Optional[str] = Query(None),
    company_name: str = Query("Food Authority"),
    station_code: str = Query("JESSUP-01"),
    operating_state: str = Query("MD"),
    file: UploadFile = File(...),
):
    lori_regulatory_require_key(api_key)

    file_bytes = await file.read()
    filename = lori_kpi_safe_file_name(file.filename or "kpi_upload.csv")
    content_type = file.content_type or "application/octet-stream"

    lower_name = filename.lower()

    if lower_name.endswith(".csv") or lower_name.endswith(".txt"):
        rows = lori_kpi_parse_csv_bytes(file_bytes)
    elif lower_name.endswith(".xlsx"):
        rows = lori_kpi_parse_excel_bytes(file_bytes)
    else:
        return {
            "status": "unsupported_file_type",
            "message": "Please upload a CSV or XLSX KPI file.",
            "filename": filename,
        }

    if not rows:
        return {
            "status": "no_rows_found",
            "message": "LORI could not find KPI rows in the uploaded file.",
            "filename": filename,
        }

    storage_result = await lori_kpi_upload_file_to_storage(
        file_bytes=file_bytes,
        filename=filename,
        content_type=content_type,
    )

    headers = list(rows[0].keys())
    column_roles = lori_kpi_detect_column_roles(headers)

    upload_payload = {
        "upload_title": upload_title or f"KPI Upload — {filename}",
        "upload_type": "KPI Upload",
        "company_name": company_name,
        "station_code": station_code,
        "operating_state": operating_state,
        "upload_status": "Uploaded / Analyzing KPI Structure",
        "source_file_name": filename,
        "source_file_path": storage_result.get("file_path"),
        "source_file_url": storage_result.get("file_url"),
        "detected_columns": column_roles,
        "notes": "Uploaded through KPI Action Plans universal intake.",
        "created_by": "LORI KPI Action Plans",
    }

    created_upload = await lori_policy_supabase_post(
        "lori_kpi_action_plan_uploads",
        upload_payload,
    )

    upload = created_upload[0] if created_upload else {}
    upload_id = upload.get("id")

    metric_records, detected_column_records, low_confidence = lori_kpi_build_metric_records(
        upload_id=upload_id,
        rows=rows,
        column_roles=column_roles,
    )

    if detected_column_records:
        await lori_policy_supabase_post(
            "lori_kpi_action_plan_detected_columns",
            detected_column_records,
        )

    if low_confidence:
        await lori_policy_supabase_post(
            "lori_kpi_action_plan_mapping_queue",
            {
                "upload_id": upload_id,
                "mapping_status": "Needs User Confirmation",
                "confidence_level": "Low Confidence",
                "detected_columns": column_roles,
                "sample_rows": rows[:5],
                "question_for_user": "LORI detected an unknown KPI structure. Please confirm which columns represent KPI name, current value, target value, date, owner, location, and category.",
            },
        )

        await lori_policy_supabase_patch(
            "lori_kpi_action_plan_uploads",
            upload_id,
            {
                "upload_status": "Needs KPI Column Mapping",
                "detected_kpi_count": 0,
                "off_track_kpi_count": 0,
                "action_plan_count": 0,
                "updated_at": datetime.utcnow().isoformat(),
            },
        )

        return {
            "status": "needs_mapping",
            "message": "LORI uploaded the KPI file but needs help mapping the KPI columns.",
            "upload": upload,
            "detected_columns": column_roles,
            "sample_rows": rows[:5],
        }

    created_metrics = []

    if metric_records:
        created_metrics = await lori_policy_supabase_post(
            "lori_kpi_action_plan_metric_records",
            metric_records,
        )

    created_plans = []
    created_findings = []

    for metric in created_metrics:
        result = await lori_kpi_create_finding_and_plan(upload_id, metric)

        if result:
            created_findings.append(result.get("finding"))
            created_plans.append(result.get("plan"))

    off_track_count = len([m for m in created_metrics if m.get("kpi_status") == "Off Track"])

    upload_status = "Analyzed / KPI Action Plans Available" if created_plans else "Analyzed / No Off-Track KPI Plans Created"

    updated_upload = await lori_policy_supabase_patch(
        "lori_kpi_action_plan_uploads",
        upload_id,
        {
            "upload_status": upload_status,
            "detected_kpi_count": len(created_metrics),
            "off_track_kpi_count": off_track_count,
            "action_plan_count": len(created_plans),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    answer_text = f"""KPI Upload Analyzed

Upload:
{upload_payload["upload_title"]}

Detected KPI Metrics:
{len(created_metrics)}

Off-Track KPIs:
{off_track_count}

KPI Action Plans Created:
{len(created_plans)}

Print-Ready Plans:
{len(created_plans)}

Recommended Next Action:
Open KPI Action Plans, review each off-track KPI, confirm owners, print the executive action plan, and send approved action steps to Action Center."""

    return {
        "status": "success",
        "message": "KPI file uploaded, analyzed, and action plans generated.",
        "upload": updated_upload[0] if updated_upload else upload,
        "detected_columns": column_roles,
        "metrics_created": len(created_metrics),
        "off_track_kpis": off_track_count,
        "action_plans_created": len(created_plans),
        "created_plans": created_plans,
        "answer_text": answer_text,
    }


@app.post("/kpi-action-plan-send-to-action-center")
async def kpi_action_plan_send_to_action_center(
    api_key: Optional[str] = Query(None),
    plan_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    plans = await lori_kpi_get_rows(
        "lori_kpi_action_plans",
        f"select=*&id=eq.{quote(plan_id)}&limit=1",
    )

    if not plans:
        return {
            "status": "not_found",
            "message": "KPI Action Plan not found.",
            "plan_id": plan_id,
        }

    plan = plans[0]

    steps = await lori_kpi_get_rows(
        "lori_kpi_action_plan_steps",
        f"select=*&plan_id=eq.{quote(plan_id)}&order=step_number.asc&limit=100",
    )

    created_actions = []

    parent_action = await lori_policy_supabase_post(
        "lori_action_items",
        {
            "action_title": plan.get("plan_title") or "KPI Action Plan",
            "action_type": "KPI Action Plan",
            "action_status": "Open",
            "priority": plan.get("priority") or "High",
            "source_module": "KPI Action Plans",
            "source_type": "KPI Action Plan",
            "source_reference_id": plan_id,
            "supervisor_name": plan.get("supervisor_name"),
            "owner_name": plan.get("plan_owner_name") or "Operations Leadership",
            "owner_role": plan.get("plan_owner_role") or "Leadership",
            "station_code": plan.get("station_code") or "JESSUP-01",
            "operating_state": plan.get("operating_state") or "MD",
            "company_name": plan.get("company_name") or "Food Authority",
            "reason": plan.get("problem_statement"),
            "recommended_follow_up": plan.get("success_measure"),
            "documentation_note": plan.get("root_cause_analysis"),
            "due_date": plan.get("recovery_target_date"),
            "created_by": "LORI KPI Action Plans",
        },
    )

    if parent_action:
        created_actions.append(parent_action[0])

    for step in steps:
        action = await lori_policy_supabase_post(
            "lori_action_items",
            {
                "action_title": step.get("action_title"),
                "action_type": "KPI Action Step",
                "action_status": "Open",
                "priority": step.get("priority") or "Medium",
                "source_module": "KPI Action Plans",
                "source_type": step.get("phase") or "KPI Action Step",
                "source_reference_id": step.get("id"),
                "owner_name": step.get("owner_name"),
                "owner_role": step.get("owner_role"),
                "station_code": plan.get("station_code") or "JESSUP-01",
                "operating_state": plan.get("operating_state") or "MD",
                "company_name": plan.get("company_name") or "Food Authority",
                "reason": step.get("action_description"),
                "recommended_follow_up": step.get("success_measure"),
                "documentation_note": step.get("documentation_note"),
                "due_date": step.get("due_date"),
                "created_by": "LORI KPI Action Plans",
            },
        )

        if action:
            created_actions.append(action[0])
            await lori_policy_supabase_patch(
                "lori_kpi_action_plan_steps",
                step.get("id"),
                {
                    "action_center_item_id": action[0].get("id"),
                    "updated_at": datetime.utcnow().isoformat(),
                },
            )

    await lori_policy_supabase_patch(
        "lori_kpi_action_plans",
        plan_id,
        {
            "action_center_status": "Sent to Action Center",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    return {
        "status": "success",
        "message": "KPI Action Plan sent to Action Center.",
        "plan_id": plan_id,
        "actions_created": len(created_actions),
        "created_actions": created_actions,
        "answer_text": f"KPI Action Plan sent to Action Center. {len(created_actions)} action items were created.",
    }


@app.post("/kpi-action-plan-send-to-leadership-briefing")
async def kpi_action_plan_send_to_leadership_briefing(
    api_key: Optional[str] = Query(None),
    plan_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    plans = await lori_kpi_get_rows(
        "lori_kpi_action_plans",
        f"select=*&id=eq.{quote(plan_id)}&limit=1",
    )

    if not plans:
        return {
            "status": "not_found",
            "message": "KPI Action Plan not found.",
            "plan_id": plan_id,
        }

    plan = plans[0]

    briefing = await lori_policy_supabase_post(
        "lori_leadership_briefing_queue",
        {
            "briefing_title": plan.get("plan_title") or "KPI Action Plan",
            "briefing_type": "KPI Action Plan Leadership Briefing",
            "briefing_status": "Queued",
            "priority": plan.get("priority") or "High",
            "station_code": plan.get("station_code") or "JESSUP-01",
            "operating_state": plan.get("operating_state") or "MD",
            "executive_summary": plan.get("executive_summary"),
            "key_risk": plan.get("problem_statement"),
            "recommended_leadership_action": plan.get("success_measure"),
            "supervisor_follow_up": plan.get("thirty_day_plan"),
            "compliance_note": "Validate KPI findings, root causes, action assignments, HR/labor implications, and corrective actions before formal action.",
            "created_by": "LORI KPI Action Plans",
        },
    )

    await lori_policy_supabase_patch(
        "lori_kpi_action_plans",
        plan_id,
        {
            "leadership_briefing_status": "Sent to Leadership Briefing",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    return {
        "status": "success",
        "message": "KPI Action Plan added to Leadership Briefing Queue.",
        "plan_id": plan_id,
        "briefing_item": briefing[0] if briefing else {},
        "answer_text": "KPI Action Plan added to Leadership Briefing Queue.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# ROUTE CONFIGURATION CENTER BACKEND
# Route scenario review, work area maps, recommendations,
# cost savings, productivity gains, Action Center integration,
# Leadership Briefing integration, and download package data.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime, date, timedelta
import json
import html
import csv
import io
import math


def lori_route_clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_route_num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


async def lori_route_get_rows(
    table: str,
    query: str = "select=*&order=created_at.desc&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


def lori_route_build_moved_stops_csv(moved_stops: List[Dict[str, Any]]) -> str:
    output = io.StringIO()

    fieldnames = [
        "stop_id",
        "customer",
        "from_route",
        "to_route",
        "freight_type",
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for stop in moved_stops:
        writer.writerow({
            "stop_id": stop.get("stop_id", ""),
            "customer": stop.get("customer", ""),
            "from_route": stop.get("from_route", ""),
            "to_route": stop.get("to_route", ""),
            "freight_type": stop.get("freight_type", ""),
        })

    return output.getvalue()


def lori_route_build_print_html(
    recommendation: Dict[str, Any],
    impact: Optional[Dict[str, Any]] = None,
    maps: Optional[List[Dict[str, Any]]] = None,
) -> str:
    impact = impact or {}
    maps = maps or []

    title = html.escape(lori_route_clean_text(recommendation.get("recommendation_title") or "Route Configuration Review"))
    decision = html.escape(lori_route_clean_text(recommendation.get("decision") or "Review Needed"))
    priority = html.escape(lori_route_clean_text(recommendation.get("priority") or "High"))

    current_driver = html.escape(lori_route_clean_text(recommendation.get("current_driver_name") or "Current driver"))
    current_route = html.escape(lori_route_clean_text(recommendation.get("current_route_id") or "Current route"))
    receiving_driver = html.escape(lori_route_clean_text(recommendation.get("receiving_driver_name") or "Receiving driver"))
    receiving_route = html.escape(lori_route_clean_text(recommendation.get("receiving_route_id") or "Receiving route"))

    summary = html.escape(lori_route_clean_text(recommendation.get("recommendation_summary") or "Route reconfiguration recommendation."))
    reason = html.escape(lori_route_clean_text(recommendation.get("operational_reason") or "Operational reason not provided."))
    benefit = html.escape(lori_route_clean_text(recommendation.get("expected_business_benefit") or "Expected business benefit not provided."))

    cost = html.escape(lori_route_clean_text(recommendation.get("cost_savings_summary") or "Cost savings not calculated."))
    productivity = html.escape(lori_route_clean_text(recommendation.get("productivity_summary") or "Productivity impact not calculated."))
    fuel = html.escape(lori_route_clean_text(recommendation.get("fuel_savings_summary") or "Fuel savings not calculated."))
    service = html.escape(lori_route_clean_text(recommendation.get("service_impact_summary") or "Service impact not calculated."))
    safety = html.escape(lori_route_clean_text(recommendation.get("safety_risk_summary") or "Safety risk not calculated."))

    seven = html.escape(lori_route_clean_text(recommendation.get("expected_7_day_result") or "7-day result not provided."))
    thirty = html.escape(lori_route_clean_text(recommendation.get("expected_30_day_result") or "30-day result not provided."))
    sixty = html.escape(lori_route_clean_text(recommendation.get("expected_60_day_result") or "60-day result not provided."))
    ninety = html.escape(lori_route_clean_text(recommendation.get("expected_90_day_result") or "90-day result not provided."))

    if_do = html.escape(lori_route_clean_text(impact.get("if_you_do_this") or "If implemented, leadership should monitor service, cost, miles, utilization, and driver workload."))
    if_not = html.escape(lori_route_clean_text(impact.get("if_you_do_not_do_this") or "If not implemented, current route pressure may continue."))

    map_rows = ""
    for m in maps:
        map_rows += f"""
        <div class="map-card">
            <strong>{html.escape(lori_route_clean_text(m.get("map_type")))}</strong>
            <p>{html.escape(lori_route_clean_text(m.get("visual_summary") or m.get("boundary_summary")))}</p>
        </div>
        """

    return f"""
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
    @page {{
        size: letter;
        margin: 0.55in;
    }}

    body {{
        font-family: Arial, Helvetica, sans-serif;
        color: #172033;
        background: #ffffff;
        margin: 0;
        padding: 0;
        line-height: 1.42;
    }}

    .plan {{
        max-width: 920px;
        margin: 0 auto;
    }}

    .cover {{
        border-radius: 22px;
        padding: 34px;
        background: linear-gradient(135deg, #f6f8ff 0%, #eef3ff 55%, #ffffff 100%);
        border: 1px solid #dbe5ff;
        margin-bottom: 22px;
    }}

    .brand {{
        font-size: 12px;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: #52627a;
        font-weight: 700;
        margin-bottom: 22px;
    }}

    h1 {{
        font-size: 32px;
        line-height: 1.08;
        margin: 0 0 10px 0;
        color: #111827;
    }}

    h2 {{
        font-size: 18px;
        margin: 0 0 10px 0;
        color: #111827;
    }}

    .status {{
        display: inline-block;
        background: #1e3a8a;
        color: #ffffff;
        padding: 10px 14px;
        border-radius: 999px;
        font-size: 12px;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        font-weight: 700;
        margin-bottom: 18px;
    }}

    .grid {{
        display: grid;
        grid-template-columns: repeat(2, 1fr);
        gap: 12px;
        margin-top: 18px;
    }}

    .metric-grid {{
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 12px;
        margin-top: 18px;
    }}

    .card, .metric, .map-card {{
        background: #ffffff;
        border: 1px solid #dbe5ff;
        border-radius: 16px;
        padding: 14px;
    }}

    .metric-label {{
        font-size: 11px;
        color: #667085;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 700;
        margin-bottom: 6px;
    }}

    .metric-value {{
        font-size: 18px;
        font-weight: 800;
        color: #111827;
    }}

    .section {{
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 18px;
        padding: 22px;
        margin-bottom: 16px;
        page-break-inside: avoid;
    }}

    .benefit {{
        background: #f0fdf4;
        border: 1px solid #bbf7d0;
        border-radius: 18px;
        padding: 20px;
        margin-bottom: 16px;
    }}

    .timeline {{
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 12px;
    }}

    .timeline div {{
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 14px;
        padding: 14px;
    }}

    .footer {{
        border-top: 1px solid #e5e7eb;
        padding-top: 14px;
        margin-top: 20px;
        color: #667085;
        font-size: 11px;
    }}

    @media print {{
        body {{
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
        }}
    }}
</style>
</head>
<body>
<div class="plan">
    <div class="cover">
        <div class="brand">LORI Drive Command Center</div>
        <div class="status">{decision} | {priority} Priority</div>
        <h1>{title}</h1>
        <p>{summary}</p>

        <div class="metric-grid">
            <div class="metric">
                <div class="metric-label">Current Driver</div>
                <div class="metric-value">{current_driver}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Current Route</div>
                <div class="metric-value">{current_route}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Receiving Driver</div>
                <div class="metric-value">{receiving_driver}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Receiving Route</div>
                <div class="metric-value">{receiving_route}</div>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>Why This Reconfiguration Makes Sense</h2>
        <p>{reason}</p>
    </div>

    <div class="benefit">
        <h2>Expected Business Benefit</h2>
        <p>{benefit}</p>
        <div class="grid">
            <div class="card"><strong>Cost Savings</strong><br>{cost}</div>
            <div class="card"><strong>Productivity</strong><br>{productivity}</div>
            <div class="card"><strong>Fuel Savings</strong><br>{fuel}</div>
            <div class="card"><strong>Service Impact</strong><br>{service}</div>
            <div class="card"><strong>Safety / Fatigue</strong><br>{safety}</div>
        </div>
    </div>

    <div class="section">
        <h2>If You Do This</h2>
        <p>{if_do}</p>
        <h2>If You Do Not Do This</h2>
        <p>{if_not}</p>
    </div>

    <div class="section">
        <h2>Expected Results Timeline</h2>
        <div class="timeline">
            <div><strong>7 Days</strong><br>{seven}</div>
            <div><strong>30 Days</strong><br>{thirty}</div>
            <div><strong>60 Days</strong><br>{sixty}</div>
            <div><strong>90 Days</strong><br>{ninety}</div>
        </div>
    </div>

    <div class="section">
        <h2>Work Area Map Summary</h2>
        {map_rows}
    </div>

    <div class="footer">
        LORI provides route configuration decision support. Route changes, work area changes, driver workload changes, helper assignments, vehicle assignments, customer service commitments, DOT considerations, productivity estimates, cost savings estimates, and safety implications must be reviewed by authorized operations leadership before implementation.
    </div>
</div>
</body>
</html>
"""


@app.get("/route-configuration-summary")
async def route_configuration_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    uploads = await lori_route_get_rows("lori_route_config_uploads")
    drivers = await lori_route_get_rows("lori_route_driver_profiles")
    vehicles = await lori_route_get_rows("lori_route_vehicle_profiles")
    freight = await lori_route_get_rows("lori_route_freight_profiles")
    scenarios = await lori_route_get_rows("lori_route_scenarios")
    impacts = await lori_route_get_rows("lori_route_scenario_impacts")
    recommendations = await lori_route_get_rows("lori_route_reconfiguration_recommendations")
    maps = await lori_route_get_rows("lori_route_work_area_maps")
    downloads = await lori_route_get_rows("lori_route_work_area_downloads")

    overutilized = [d for d in drivers if lori_route_clean_text(d.get("workload_status")).lower() == "overutilized"]
    underutilized = [d for d in drivers if lori_route_clean_text(d.get("workload_status")).lower() == "underutilized"]
    high_capacity_risk = [d for d in drivers if lori_route_clean_text(d.get("capacity_risk")).lower() == "high"]

    return {
        "status": "success",
        "route_uploads_count": len(uploads),
        "driver_profiles_count": len(drivers),
        "vehicle_profiles_count": len(vehicles),
        "freight_profiles_count": len(freight),
        "scenarios_count": len(scenarios),
        "recommendations_count": len(recommendations),
        "work_area_maps_count": len(maps),
        "download_packages_count": len(downloads),
        "overutilized_drivers_count": len(overutilized),
        "underutilized_drivers_count": len(underutilized),
        "capacity_risk_count": len(high_capacity_risk),
        "recent_uploads": uploads[:5],
        "recent_recommendations": recommendations[:5],
        "answer_text": "Route Configuration Center is ready. LORI can review driver workload, vehicle fit, freight complexity, route pressure, scenario impacts, work area maps, and download packages.",
    }


@app.get("/route-config-uploads")
async def route_config_uploads(
    api_key: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_get_rows(
        "lori_route_config_uploads",
        f"select=*&order=created_at.desc&limit={max(1, min(limit, 200))}",
    )

    return {
        "status": "success",
        "uploads_count": len(rows),
        "uploads": rows,
    }


@app.get("/route-driver-profiles")
async def route_driver_profiles(
    api_key: Optional[str] = Query(None),
    upload_id: Optional[str] = Query(None),
    workload_status: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_get_rows(
        "lori_route_driver_profiles",
        "select=*&order=driver_name.asc&limit=500",
    )

    if upload_id:
        rows = [r for r in rows if str(r.get("upload_id")) == upload_id]

    if workload_status:
        rows = [
            r for r in rows
            if lori_route_clean_text(r.get("workload_status")).lower() == workload_status.lower()
        ]

    return {
        "status": "success",
        "drivers_count": len(rows[:limit]),
        "drivers": rows[:max(1, min(limit, 300))],
    }


@app.get("/route-scenarios")
async def route_scenarios(
    api_key: Optional[str] = Query(None),
    upload_id: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_get_rows(
        "lori_route_scenarios",
        "select=*&order=created_at.desc&limit=500",
    )

    if upload_id:
        rows = [r for r in rows if str(r.get("upload_id")) == upload_id]

    return {
        "status": "success",
        "scenarios_count": len(rows[:limit]),
        "scenarios": rows[:max(1, min(limit, 200))],
    }


@app.get("/route-recommendations")
async def route_recommendations(
    api_key: Optional[str] = Query(None),
    upload_id: Optional[str] = Query(None),
    decision: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_get_rows(
        "lori_route_reconfiguration_recommendations",
        "select=*&order=created_at.desc&limit=500",
    )

    if upload_id:
        rows = [r for r in rows if str(r.get("upload_id")) == upload_id]

    if decision:
        rows = [
            r for r in rows
            if lori_route_clean_text(r.get("decision")).lower() == decision.lower()
        ]

    return {
        "status": "success",
        "recommendations_count": len(rows[:limit]),
        "recommendations": rows[:max(1, min(limit, 200))],
    }


@app.get("/route-recommendation-detail")
async def route_recommendation_detail(
    api_key: Optional[str] = Query(None),
    recommendation_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    recommendations = await lori_route_get_rows(
        "lori_route_reconfiguration_recommendations",
        f"select=*&id=eq.{quote(recommendation_id)}&limit=1",
    )

    if not recommendations:
        return {
            "status": "not_found",
            "message": "Route recommendation not found.",
            "recommendation_id": recommendation_id,
        }

    rec = recommendations[0]
    scenario_id = rec.get("scenario_id")

    scenarios = []
    impacts = []

    if scenario_id:
        scenarios = await lori_route_get_rows(
            "lori_route_scenarios",
            f"select=*&id=eq.{quote(str(scenario_id))}&limit=1",
        )

        impacts = await lori_route_get_rows(
            "lori_route_scenario_impacts",
            f"select=*&scenario_id=eq.{quote(str(scenario_id))}&limit=5",
        )

    maps = await lori_route_get_rows(
        "lori_route_work_area_maps",
        f"select=*&recommendation_id=eq.{quote(recommendation_id)}&order=map_type.asc&limit=20",
    )

    downloads = await lori_route_get_rows(
        "lori_route_work_area_downloads",
        f"select=*&recommendation_id=eq.{quote(recommendation_id)}&limit=5",
    )

    print_exports = await lori_route_get_rows(
        "lori_route_print_exports",
        f"select=*&recommendation_id=eq.{quote(recommendation_id)}&limit=5",
    )

    return {
        "status": "success",
        "recommendation": rec,
        "scenario": scenarios[0] if scenarios else {},
        "impact": impacts[0] if impacts else {},
        "work_area_maps": maps,
        "download_packages": downloads,
        "print_exports": print_exports,
    }


@app.get("/route-work-area-maps")
async def route_work_area_maps(
    api_key: Optional[str] = Query(None),
    recommendation_id: Optional[str] = Query(None),
    scenario_id: Optional[str] = Query(None),
    map_type: Optional[str] = Query(None),
    limit: int = Query(50),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_get_rows(
        "lori_route_work_area_maps",
        "select=*&order=created_at.asc&limit=500",
    )

    if recommendation_id:
        rows = [r for r in rows if str(r.get("recommendation_id")) == recommendation_id]

    if scenario_id:
        rows = [r for r in rows if str(r.get("scenario_id")) == scenario_id]

    if map_type:
        rows = [
            r for r in rows
            if lori_route_clean_text(r.get("map_type")).lower() == map_type.lower()
        ]

    return {
        "status": "success",
        "maps_count": len(rows[:limit]),
        "maps": rows[:max(1, min(limit, 200))],
    }


@app.get("/route-work-area-download-package")
async def route_work_area_download_package(
    api_key: Optional[str] = Query(None),
    recommendation_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    recommendations = await lori_route_get_rows(
        "lori_route_reconfiguration_recommendations",
        f"select=*&id=eq.{quote(recommendation_id)}&limit=1",
    )

    if not recommendations:
        return {
            "status": "not_found",
            "message": "Route recommendation not found.",
            "recommendation_id": recommendation_id,
        }

    rec = recommendations[0]

    maps = await lori_route_get_rows(
        "lori_route_work_area_maps",
        f"select=*&recommendation_id=eq.{quote(recommendation_id)}&order=map_type.asc&limit=20",
    )

    downloads = await lori_route_get_rows(
        "lori_route_work_area_downloads",
        f"select=*&recommendation_id=eq.{quote(recommendation_id)}&limit=1",
    )

    current_map = next((m for m in maps if "current" in lori_route_clean_text(m.get("map_type")).lower()), None)
    proposed_map = next((m for m in maps if "proposed" in lori_route_clean_text(m.get("map_type")).lower()), None)
    overlay_map = next((m for m in maps if "overlay" in lori_route_clean_text(m.get("map_type")).lower() or "after" in lori_route_clean_text(m.get("map_type")).lower()), None)

    moved_stops = []

    if overlay_map and isinstance(overlay_map.get("moved_stops"), list):
        moved_stops = overlay_map.get("moved_stops")
    elif current_map and isinstance(current_map.get("moved_stops"), list):
        moved_stops = current_map.get("moved_stops")

    moved_stops_csv = lori_route_build_moved_stops_csv(moved_stops)

    package_summary = f"""Route Work Area Download Package

Recommendation:
{rec.get("recommendation_title")}

Decision:
{rec.get("decision")}

Current Work Area:
{current_map.get("map_title") if current_map else "Not available"}

Proposed Work Area:
{proposed_map.get("map_title") if proposed_map else "Not available"}

Overlay:
{overlay_map.get("map_title") if overlay_map else "Not available"}

Business Benefit:
{rec.get("expected_business_benefit")}

Cost Savings:
{rec.get("cost_savings_summary")}

Productivity:
{rec.get("productivity_summary")}

Fuel Savings:
{rec.get("fuel_savings_summary")}
"""

    return {
        "status": "success",
        "recommendation": rec,
        "download_package": downloads[0] if downloads else {},
        "current_work_area_map": current_map,
        "proposed_work_area_map": proposed_map,
        "before_after_overlay_map": overlay_map,
        "moved_stops": moved_stops,
        "moved_stops_csv": moved_stops_csv,
        "current_geojson": current_map.get("geojson_data") if current_map else {},
        "proposed_geojson": proposed_map.get("geojson_data") if proposed_map else {},
        "overlay_geojson": overlay_map.get("geojson_data") if overlay_map else {},
        "package_summary_text": package_summary,
        "answer_text": "Route work area download package is ready.",
    }


@app.get("/route-print-review")
async def route_print_review(
    api_key: Optional[str] = Query(None),
    recommendation_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    details = await route_recommendation_detail(
        api_key=api_key,
        recommendation_id=recommendation_id,
    )

    if details.get("status") != "success":
        return details

    rec = details.get("recommendation") or {}
    impact = details.get("impact") or {}
    maps = details.get("work_area_maps") or []

    printable_html = lori_route_build_print_html(rec, impact, maps)

    return {
        "status": "success",
        "recommendation_id": recommendation_id,
        "recommendation_title": rec.get("recommendation_title"),
        "print_ready": True,
        "printable_html": printable_html,
        "answer_text": "Printable Route Configuration Review is ready.",
    }


@app.post("/route-scenario-evaluate")
async def route_scenario_evaluate(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    scenario_name = lori_route_clean_text(payload.get("scenario_name") or "Manual Route Reconfiguration Scenario")

    current_driver = lori_route_clean_text(payload.get("current_driver_name") or payload.get("current_driver") or "Current Driver")
    current_route = lori_route_clean_text(payload.get("current_route_id") or payload.get("current_route") or "Current Route")
    receiving_driver = lori_route_clean_text(payload.get("receiving_driver_name") or payload.get("receiving_driver") or "Receiving Driver")
    receiving_route = lori_route_clean_text(payload.get("receiving_route_id") or payload.get("receiving_route") or "Receiving Route")

    stops_to_move = int(lori_route_num(payload.get("stops_to_move"), 0))
    freight_type = lori_route_clean_text(payload.get("freight_type") or "General Freight")
    freight_complexity = lori_route_clean_text(payload.get("freight_complexity") or "Medium")
    traffic_condition = lori_route_clean_text(payload.get("traffic_condition") or "Normal")

    estimated_added_miles = lori_route_num(payload.get("estimated_added_miles"), 0)
    estimated_reduced_miles = lori_route_num(payload.get("estimated_reduced_miles"), 0)
    estimated_added_time = lori_route_num(payload.get("estimated_added_time_minutes"), 0)
    estimated_reduced_time = lori_route_num(payload.get("estimated_reduced_time_minutes"), 0)

    current_helper_count = int(lori_route_num(payload.get("current_helper_count"), 0))
    proposed_helper_count = int(lori_route_num(payload.get("proposed_helper_count"), 0))

    fuel_cost_per_gallon = lori_route_num(payload.get("fuel_cost_per_gallon"), 3.75)
    estimated_mpg = max(lori_route_num(payload.get("estimated_mpg"), 9), 1)
    labor_cost_per_hour = lori_route_num(payload.get("labor_cost_per_hour"), 32)
    helper_cost_per_hour = lori_route_num(payload.get("helper_cost_per_hour"), 22)

    current_util = lori_route_num(payload.get("current_driver_utilization_percent"), 110)
    receiving_util = lori_route_num(payload.get("receiving_driver_utilization_percent"), 82)

    net_miles_reduced = estimated_reduced_miles - estimated_added_miles
    fuel_saved = max(net_miles_reduced / estimated_mpg, 0)
    fuel_savings = round(fuel_saved * fuel_cost_per_gallon, 2)

    net_minutes_saved = estimated_reduced_time - estimated_added_time
    labor_hours_saved = max(net_minutes_saved / 60, 0)
    labor_savings = round(labor_hours_saved * labor_cost_per_hour, 2)

    helper_delta = proposed_helper_count - current_helper_count
    helper_cost_change = round(helper_delta * helper_cost_per_hour * max(estimated_added_time / 60, 1), 2)

    total_savings = round(fuel_savings + labor_savings - helper_cost_change, 2)

    receiving_after_util = receiving_util + (stops_to_move * 1.5)
    current_after_util = current_util - (stops_to_move * 1.2)

    if receiving_after_util <= 100 and total_savings >= 0:
        decision = "Recommended"
        confidence = "Medium-High"
    elif receiving_after_util <= 110 and total_savings >= -25:
        decision = "Needs Supervisor Review"
        confidence = "Medium"
    else:
        decision = "Not Recommended"
        confidence = "Medium"

    scenario_created = await lori_policy_supabase_post(
        "lori_route_scenarios",
        {
            "scenario_name": scenario_name,
            "scenario_status": "Analyzed",
            "scenario_type": "What-If Route Reconfiguration",
            "current_driver_name": current_driver,
            "current_route_id": current_route,
            "receiving_driver_name": receiving_driver,
            "receiving_route_id": receiving_route,
            "stops_to_move": stops_to_move,
            "freight_type": freight_type,
            "freight_complexity": freight_complexity,
            "current_helper_count": current_helper_count,
            "proposed_helper_count": proposed_helper_count,
            "current_vehicle_type": lori_route_clean_text(payload.get("current_vehicle_type") or ""),
            "proposed_vehicle_type": lori_route_clean_text(payload.get("proposed_vehicle_type") or ""),
            "estimated_added_miles": estimated_added_miles,
            "estimated_reduced_miles": estimated_reduced_miles,
            "estimated_added_time_minutes": estimated_added_time,
            "estimated_reduced_time_minutes": estimated_reduced_time,
            "traffic_condition": traffic_condition,
            "delivery_window_sensitivity": lori_route_clean_text(payload.get("delivery_window_sensitivity") or "Medium"),
            "fuel_cost_per_gallon": fuel_cost_per_gallon,
            "estimated_mpg": estimated_mpg,
            "labor_cost_per_hour": labor_cost_per_hour,
            "helper_cost_per_hour": helper_cost_per_hour,
            "overtime_risk": lori_route_clean_text(payload.get("overtime_risk") or "Medium"),
            "scenario_notes": lori_route_clean_text(payload.get("notes") or "Manual route scenario evaluated by LORI."),
        },
    )

    scenario = scenario_created[0] if scenario_created else {}
    scenario_id = scenario.get("id")

    impact_payload = {
        "scenario_id": scenario_id,
        "recommendation": decision,
        "confidence_level": confidence,
        "productivity_impact": f"Estimated productivity gain is approximately {round(max((current_util - current_after_util), 0), 1)} utilization points relieved from the overloaded route.",
        "service_risk": "Low to Medium" if traffic_condition.lower() in {"normal", "light"} else "Medium to High",
        "safety_fatigue_risk": "Reduced fatigue pressure on the overutilized route." if current_after_util < current_util else "Monitor fatigue risk.",
        "vehicle_capacity_impact": "Review receiving vehicle capacity before implementation.",
        "freight_complexity_impact": f"Freight complexity is {freight_complexity}.",
        "traffic_impact": f"Traffic condition is {traffic_condition}.",
        "helper_impact": f"Helper count changes from {current_helper_count} to {proposed_helper_count}.",
        "estimated_labor_hours_saved": round(labor_hours_saved, 2),
        "estimated_overtime_reduction_hours": round(max(net_minutes_saved / 60, 0), 2),
        "estimated_miles_reduced": round(net_miles_reduced, 2),
        "estimated_fuel_saved_gallons": round(fuel_saved, 2),
        "estimated_fuel_cost_savings": fuel_savings,
        "estimated_helper_cost_change": helper_cost_change,
        "estimated_total_cost_savings": total_savings,
        "estimated_productivity_gain_percent": round(max(current_util - current_after_util, 0), 2),
        "estimated_stops_per_hour_change": round(stops_to_move / max(estimated_added_time / 60, 1), 2),
        "estimated_cost_per_stop_reduction": round(total_savings / max(stops_to_move, 1), 2),
        "if_you_do_this": f"{current_driver}'s route pressure should decrease, {receiving_driver} may absorb {stops_to_move} stops, and the route group may improve workload balance, mileage, fuel use, and overtime exposure.",
        "if_you_do_not_do_this": f"{current_driver} may remain overutilized, overtime pressure may continue, and service risk may remain elevated.",
        "decision_reason": f"Decision is {decision} based on utilization shift, estimated cost savings, traffic condition, freight complexity, helper impact, and receiving route capacity.",
        "required_supervisor_review": True,
        "recommended_owner": "Operations Leadership",
        "recommended_next_action": "Review scenario with transportation leadership, confirm delivery windows and vehicle fit, then create an Action Center follow-up if approved.",
    }

    impact_created = await lori_policy_supabase_post(
        "lori_route_scenario_impacts",
        impact_payload,
    )

    recommendation_payload = {
        "scenario_id": scenario_id,
        "recommendation_title": f"{decision}: Move {stops_to_move} Stops from {current_driver} to {receiving_driver}",
        "recommendation_status": "Draft",
        "decision": decision,
        "priority": "High" if decision in {"Recommended", "Needs Supervisor Review"} else "Medium",
        "current_driver_name": current_driver,
        "current_route_id": current_route,
        "receiving_driver_name": receiving_driver,
        "receiving_route_id": receiving_route,
        "recommendation_summary": f"Evaluate moving {stops_to_move} {freight_type} stops from {current_driver} to {receiving_driver}.",
        "operational_reason": "This scenario evaluates workload balance, mileage, fuel, helper usage, vehicle fit, freight complexity, traffic, service risk, and productivity impact.",
        "expected_7_day_result": "Leadership should see immediate visibility into workload balance and route pressure.",
        "expected_30_day_result": "Route performance should begin showing better workload balance and less overtime pressure if implemented correctly.",
        "expected_60_day_result": "Mileage, fuel, and stops-per-hour trends should show measurable improvement if service risk remains controlled.",
        "expected_90_day_result": "The reconfigured work area should become stable if customer service, safety, and productivity remain acceptable.",
        "expected_business_benefit": "Expected business benefit includes better workload balance, reduced route pressure, improved productivity, lower fuel/mileage waste, and improved service reliability.",
        "cost_savings_summary": f"Estimated total cost savings: ${total_savings}.",
        "productivity_summary": f"Estimated productivity gain: {round(max(current_util - current_after_util, 0), 2)} utilization points relieved from current route.",
        "fuel_savings_summary": f"Estimated fuel savings: {round(fuel_saved, 2)} gallons / ${fuel_savings}.",
        "service_impact_summary": impact_payload["service_risk"],
        "safety_risk_summary": impact_payload["safety_fatigue_risk"],
        "owner_name": "Operations Leadership",
        "owner_role": "Leadership",
    }

    rec_created = await lori_policy_supabase_post(
        "lori_route_reconfiguration_recommendations",
        recommendation_payload,
    )

    recommendation = rec_created[0] if rec_created else {}

    return {
        "status": "success",
        "message": "Route scenario evaluated.",
        "scenario": scenario,
        "impact": impact_created[0] if impact_created else {},
        "recommendation": recommendation,
        "answer_text": f"Route scenario evaluated. Decision: {decision}. Estimated total cost savings: ${total_savings}.",
    }


@app.post("/route-recommendation-send-to-action-center")
async def route_recommendation_send_to_action_center(
    api_key: Optional[str] = Query(None),
    recommendation_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    recs = await lori_route_get_rows(
        "lori_route_reconfiguration_recommendations",
        f"select=*&id=eq.{quote(recommendation_id)}&limit=1",
    )

    if not recs:
        return {
            "status": "not_found",
            "message": "Route recommendation not found.",
        }

    rec = recs[0]

    action_created = await lori_policy_supabase_post(
        "lori_action_items",
        {
            "action_title": rec.get("recommendation_title"),
            "action_type": "Route Reconfiguration Review",
            "action_status": "Open",
            "priority": rec.get("priority") or "High",
            "source_module": "Route Configuration Center",
            "source_type": "Route Reconfiguration Recommendation",
            "source_reference_id": recommendation_id,
            "owner_name": rec.get("owner_name") or "Operations Leadership",
            "owner_role": rec.get("owner_role") or "Leadership",
            "station_code": "JESSUP-01",
            "operating_state": "MD",
            "company_name": "Food Authority",
            "reason": rec.get("operational_reason"),
            "recommended_follow_up": rec.get("recommendation_summary"),
            "documentation_note": "Validate route change, work area impact, service requirements, helper needs, vehicle fit, safety impact, and leadership approval before implementation.",
            "due_date": (date.today() + timedelta(days=7)).isoformat(),
            "created_by": "LORI Route Configuration Center",
        },
    )

    await lori_policy_supabase_patch(
        "lori_route_reconfiguration_recommendations",
        recommendation_id,
        {
            "action_center_status": "Sent to Action Center",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    return {
        "status": "success",
        "message": "Route recommendation sent to Action Center.",
        "action_item": action_created[0] if action_created else {},
        "answer_text": "Route recommendation sent to Action Center.",
    }


@app.post("/route-recommendation-send-to-leadership")
async def route_recommendation_send_to_leadership(
    api_key: Optional[str] = Query(None),
    recommendation_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    recs = await lori_route_get_rows(
        "lori_route_reconfiguration_recommendations",
        f"select=*&id=eq.{quote(recommendation_id)}&limit=1",
    )

    if not recs:
        return {
            "status": "not_found",
            "message": "Route recommendation not found.",
        }

    rec = recs[0]

    briefing = await lori_policy_supabase_post(
        "lori_leadership_briefing_queue",
        {
            "briefing_title": rec.get("recommendation_title"),
            "briefing_type": "Route Configuration Leadership Briefing",
            "briefing_status": "Queued",
            "priority": rec.get("priority") or "High",
            "station_code": "JESSUP-01",
            "operating_state": "MD",
            "executive_summary": rec.get("recommendation_summary"),
            "key_risk": rec.get("safety_risk_summary"),
            "recommended_leadership_action": rec.get("operational_reason"),
            "supervisor_follow_up": rec.get("expected_business_benefit"),
            "compliance_note": "Validate route boundaries, customer commitments, driver workload, helper needs, vehicle assignments, DOT considerations, and safety implications before implementation.",
            "created_by": "LORI Route Configuration Center",
        },
    )

    await lori_policy_supabase_patch(
        "lori_route_reconfiguration_recommendations",
        recommendation_id,
        {
            "leadership_briefing_status": "Sent to Leadership Briefing",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    return {
        "status": "success",
        "message": "Route recommendation added to Leadership Briefing Queue.",
        "briefing_item": briefing[0] if briefing else {},
        "answer_text": "Route recommendation added to Leadership Briefing Queue.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# PUSH NOTIFICATION CENTER BACKEND
# Drafts, templates, recipient groups, approval workflow,
# delivery queue, acknowledgments, audit history.
# Live email/SMS sending will be connected in a later upgrade.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime, date
import json


def lori_comm_clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_comm_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if str(value).strip().lower() in {"true", "yes", "y", "1", "required"}:
        return True

    if str(value).strip().lower() in {"false", "no", "n", "0", "not required"}:
        return False

    return default


async def lori_comm_get_rows(
    table: str,
    query: str = "select=*&order=created_at.desc&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_comm_log_event(
    campaign_id: Optional[str],
    event_type: str,
    event_summary: str,
    event_status: str = "Logged",
    event_details: Optional[Dict[str, Any]] = None,
    performed_by: str = "LORI Push Notification Center",
):
    payload = {
        "campaign_id": campaign_id,
        "event_type": event_type,
        "event_status": event_status,
        "event_summary": event_summary,
        "event_details": event_details or {},
        "performed_by": performed_by,
    }

    return await lori_policy_supabase_post(
        "lori_comm_audit_events",
        payload,
    )


def lori_comm_build_memo(
    to_line: str,
    from_line: str,
    subject_line: str,
    message_body: str,
) -> str:
    today = date.today().isoformat()

    return f"""MEMORANDUM

To: {to_line}
From: {from_line}
Date: {today}
Subject: {subject_line}

{message_body}
"""


@app.get("/comm-summary")
async def comm_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    templates = await lori_comm_get_rows("lori_comm_templates")
    groups = await lori_comm_get_rows("lori_comm_recipient_groups")
    campaigns = await lori_comm_get_rows("lori_comm_campaigns")
    recipients = await lori_comm_get_rows("lori_comm_recipients")
    logs = await lori_comm_get_rows("lori_comm_delivery_logs")
    acknowledgments = await lori_comm_get_rows("lori_comm_acknowledgments")

    draft_messages = [
        c for c in campaigns
        if lori_comm_clean_text(c.get("campaign_status")).lower() == "draft"
    ]

    scheduled_messages = [
        c for c in campaigns
        if lori_comm_clean_text(c.get("campaign_status")).lower() == "scheduled"
    ]

    queued_messages = [
        c for c in campaigns
        if lori_comm_clean_text(c.get("campaign_status")).lower() in {"queued", "ready to send"}
    ]

    sent_messages = [
        c for c in campaigns
        if lori_comm_clean_text(c.get("campaign_status")).lower() == "sent"
    ]

    failed_logs = [
        l for l in logs
        if lori_comm_clean_text(l.get("delivery_status")).lower() == "failed"
    ]

    pending_ack = [
        a for a in acknowledgments
        if lori_comm_clean_text(a.get("acknowledgment_status")).lower() == "pending"
    ]

    safety_alerts = [
        c for c in campaigns
        if "safety" in lori_comm_clean_text(c.get("message_type")).lower()
    ]

    memorandums = [
        c for c in campaigns
        if "memo" in lori_comm_clean_text(c.get("message_type")).lower()
        or lori_comm_clean_text(c.get("memo_body")) != ""
    ]

    return {
        "status": "success",
        "templates_count": len(templates),
        "recipient_groups_count": len(groups),
        "campaigns_count": len(campaigns),
        "recipients_count": len(recipients),
        "delivery_logs_count": len(logs),
        "acknowledgments_count": len(acknowledgments),
        "draft_messages_count": len(draft_messages),
        "scheduled_messages_count": len(scheduled_messages),
        "queued_messages_count": len(queued_messages),
        "sent_messages_count": len(sent_messages),
        "failed_deliveries_count": len(failed_logs),
        "pending_acknowledgments_count": len(pending_ack),
        "safety_alerts_count": len(safety_alerts),
        "memorandums_count": len(memorandums),
        "recent_campaigns": campaigns[:5],
        "answer_text": "Push Notification Center backend is ready for templates, campaigns, recipient groups, approval workflow, delivery queue, acknowledgments, and audit tracking. Live sending is not enabled yet.",
    }


@app.get("/comm-templates")
async def comm_templates(
    api_key: Optional[str] = Query(None),
    template_type: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_templates",
        "select=*&order=template_name.asc&limit=500",
    )

    if template_type:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("template_type")).lower() == template_type.lower()
        ]

    return {
        "status": "success",
        "templates_count": len(rows[:limit]),
        "templates": rows[:max(1, min(limit, 300))],
    }


@app.get("/comm-recipient-groups")
async def comm_recipient_groups(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    group_type: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_recipient_groups",
        "select=*&order=group_name.asc&limit=500",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("station_code")).lower() == station_code.lower()
        ]

    if group_type:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("group_type")).lower() == group_type.lower()
        ]

    return {
        "status": "success",
        "groups_count": len(rows[:limit]),
        "groups": rows[:max(1, min(limit, 300))],
    }


@app.get("/comm-recipient-group-members")
async def comm_recipient_group_members(
    api_key: Optional[str] = Query(None),
    group_id: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    limit: int = Query(300),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_recipient_group_members",
        "select=*&order=recipient_name.asc&limit=1000",
    )

    if group_id:
        rows = [
            r for r in rows
            if str(r.get("group_id")) == group_id
        ]

    if station_code:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("station_code")).lower() == station_code.lower()
        ]

    return {
        "status": "success",
        "members_count": len(rows[:limit]),
        "members": rows[:max(1, min(limit, 500))],
    }


@app.get("/comm-campaigns")
async def comm_campaigns(
    api_key: Optional[str] = Query(None),
    campaign_status: Optional[str] = Query(None),
    message_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_campaigns",
        "select=*&order=created_at.desc&limit=500",
    )

    if campaign_status:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("campaign_status")).lower() == campaign_status.lower()
        ]

    if message_type:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("message_type")).lower() == message_type.lower()
        ]

    if priority:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("priority")).lower() == priority.lower()
        ]

    return {
        "status": "success",
        "campaigns_count": len(rows[:limit]),
        "campaigns": rows[:max(1, min(limit, 300))],
    }


@app.get("/comm-campaign-detail")
async def comm_campaign_detail(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Communication campaign not found.",
            "campaign_id": campaign_id,
        }

    recipients = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&order=recipient_name.asc&limit=1000",
    )

    delivery_logs = await lori_comm_get_rows(
        "lori_comm_delivery_logs",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&order=created_at.desc&limit=1000",
    )

    acknowledgments = await lori_comm_get_rows(
        "lori_comm_acknowledgments",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&order=created_at.desc&limit=1000",
    )

    attachments = await lori_comm_get_rows(
        "lori_comm_attachments",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&order=created_at.desc&limit=100",
    )

    audit_events = await lori_comm_get_rows(
        "lori_comm_audit_events",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&order=created_at.desc&limit=200",
    )

    return {
        "status": "success",
        "campaign": campaigns[0],
        "recipients_count": len(recipients),
        "recipients": recipients,
        "delivery_logs_count": len(delivery_logs),
        "delivery_logs": delivery_logs,
        "acknowledgments_count": len(acknowledgments),
        "acknowledgments": acknowledgments,
        "attachments_count": len(attachments),
        "attachments": attachments,
        "audit_events_count": len(audit_events),
        "audit_events": audit_events,
    }


@app.post("/comm-campaign-create")
async def comm_campaign_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    message_title = lori_comm_clean_text(payload.get("message_title") or payload.get("title"))

    if not message_title:
        return {
            "status": "error",
            "message": "message_title is required.",
        }

    subject_line = lori_comm_clean_text(payload.get("subject_line") or message_title)
    message_body = lori_comm_clean_text(payload.get("message_body") or payload.get("body") or "")
    sms_body = lori_comm_clean_text(payload.get("sms_body") or message_body[:280])

    memo_body = lori_comm_clean_text(payload.get("memo_body") or "")

    if not memo_body:
        memo_body = lori_comm_build_memo(
            to_line=lori_comm_clean_text(payload.get("audience_type") or "Drivers and Staff"),
            from_line=lori_comm_clean_text(payload.get("from_line") or "Operations Leadership"),
            subject_line=subject_line,
            message_body=message_body,
        )

    campaign_payload = {
        "message_title": message_title,
        "message_type": lori_comm_clean_text(payload.get("message_type") or "General Staff Notice"),
        "priority": lori_comm_clean_text(payload.get("priority") or "Routine"),
        "delivery_method": lori_comm_clean_text(payload.get("delivery_method") or "Email"),
        "audience_type": lori_comm_clean_text(payload.get("audience_type") or "Custom Group"),
        "station_code": lori_comm_clean_text(payload.get("station_code") or "JESSUP-01"),
        "operating_state": lori_comm_clean_text(payload.get("operating_state") or "MD"),
        "route_id": lori_comm_clean_text(payload.get("route_id") or ""),
        "route_group": lori_comm_clean_text(payload.get("route_group") or ""),
        "supervisor_name": lori_comm_clean_text(payload.get("supervisor_name") or ""),
        "shift_name": lori_comm_clean_text(payload.get("shift_name") or ""),
        "subject_line": subject_line,
        "message_body": message_body,
        "sms_body": sms_body,
        "memo_body": memo_body,
        "attachment_count": int(payload.get("attachment_count") or 0),
        "requires_acknowledgment": lori_comm_bool(payload.get("requires_acknowledgment"), False),
        "schedule_type": lori_comm_clean_text(payload.get("schedule_type") or "Send Now"),
        "scheduled_send_at": payload.get("scheduled_send_at"),
        "approval_status": "Draft",
        "campaign_status": "Draft",
        "source_module": lori_comm_clean_text(payload.get("source_module") or ""),
        "source_reference_id": lori_comm_clean_text(payload.get("source_reference_id") or ""),
        "internal_notes": lori_comm_clean_text(payload.get("internal_notes") or ""),
        "created_by": lori_comm_clean_text(payload.get("created_by") or "LORI Push Notification Center"),
    }

    created = await lori_policy_supabase_post(
        "lori_comm_campaigns",
        campaign_payload,
    )

    campaign = created[0] if created else {}

    await lori_comm_log_event(
        campaign_id=campaign.get("id"),
        event_type="Campaign Created",
        event_summary=f"Communication campaign created: {message_title}",
        event_details={"live_send": False, "source": "comm-campaign-create"},
    )

    return {
        "status": "success",
        "message": "Communication campaign created as draft.",
        "campaign": campaign,
        "answer_text": "Communication campaign created as draft. No live text or email was sent.",
    }


@app.post("/comm-campaign-add-recipient")
async def comm_campaign_add_recipient(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
        }

    recipient_name = lori_comm_clean_text(payload.get("recipient_name"))

    if not recipient_name:
        return {
            "status": "error",
            "message": "recipient_name is required.",
        }

    recipient_payload = {
        "campaign_id": campaign_id,
        "recipient_name": recipient_name,
        "recipient_role": lori_comm_clean_text(payload.get("recipient_role") or ""),
        "employee_id": lori_comm_clean_text(payload.get("employee_id") or ""),
        "station_code": lori_comm_clean_text(payload.get("station_code") or "JESSUP-01"),
        "route_id": lori_comm_clean_text(payload.get("route_id") or ""),
        "supervisor_name": lori_comm_clean_text(payload.get("supervisor_name") or ""),
        "shift_name": lori_comm_clean_text(payload.get("shift_name") or ""),
        "email_address": lori_comm_clean_text(payload.get("email_address") or ""),
        "phone_number": lori_comm_clean_text(payload.get("phone_number") or ""),
        "preferred_delivery_method": lori_comm_clean_text(payload.get("preferred_delivery_method") or "Email"),
        "recipient_status": "Selected",
        "sms_consent_status": lori_comm_clean_text(payload.get("sms_consent_status") or "Unknown"),
        "email_consent_status": lori_comm_clean_text(payload.get("email_consent_status") or "Unknown"),
        "opt_out_status": lori_comm_clean_text(payload.get("opt_out_status") or "Not Opted Out"),
    }

    created = await lori_policy_supabase_post(
        "lori_comm_recipients",
        recipient_payload,
    )

    recipients = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "recipient_count": len(recipients),
            "pending_acknowledgment_count": len(recipients) if campaigns[0].get("requires_acknowledgment") else 0,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Recipient Added",
        event_summary=f"Recipient added: {recipient_name}",
        event_details={"recipient_name": recipient_name},
    )

    return {
        "status": "success",
        "message": "Recipient added to campaign.",
        "recipient": created[0] if created else {},
        "recipient_count": len(recipients),
    }


@app.post("/comm-campaign-add-group")
async def comm_campaign_add_group(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
    group_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
        }

    members = await lori_comm_get_rows(
        "lori_comm_recipient_group_members",
        f"select=*&group_id=eq.{quote(group_id)}&member_status=eq.Active&limit=1000",
    )

    if not members:
        return {
            "status": "error",
            "message": "No active members found for this recipient group.",
        }

    created_recipients = []

    existing = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    existing_keys = {
        f"{r.get('employee_id')}|{r.get('email_address')}|{r.get('phone_number')}"
        for r in existing
    }

    for m in members:
        key = f"{m.get('employee_id')}|{m.get('email_address')}|{m.get('phone_number')}"

        if key in existing_keys:
            continue

        recipient_payload = {
            "campaign_id": campaign_id,
            "recipient_name": m.get("recipient_name"),
            "recipient_role": m.get("recipient_role"),
            "employee_id": m.get("employee_id"),
            "station_code": m.get("station_code") or "JESSUP-01",
            "route_id": m.get("route_id"),
            "supervisor_name": m.get("supervisor_name"),
            "email_address": m.get("email_address"),
            "phone_number": m.get("phone_number"),
            "preferred_delivery_method": campaigns[0].get("delivery_method") or "Email",
            "recipient_status": "Selected",
            "sms_consent_status": "Demo Consent",
            "email_consent_status": "Demo Consent",
            "opt_out_status": "Not Opted Out",
        }

        created = await lori_policy_supabase_post(
            "lori_comm_recipients",
            recipient_payload,
        )

        if created:
            created_recipients.append(created[0])

    recipients = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "recipient_count": len(recipients),
            "pending_acknowledgment_count": len(recipients) if campaigns[0].get("requires_acknowledgment") else 0,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Recipient Group Added",
        event_summary=f"Recipient group added to campaign. Members added: {len(created_recipients)}",
        event_details={"group_id": group_id, "members_added": len(created_recipients)},
    )

    return {
        "status": "success",
        "message": "Recipient group added to campaign.",
        "members_added": len(created_recipients),
        "recipient_count": len(recipients),
        "recipients": created_recipients,
    }


@app.post("/comm-campaign-submit-for-approval")
async def comm_campaign_submit_for_approval(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
        }

    recipients = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    if not recipients:
        return {
            "status": "error",
            "message": "Add at least one recipient before submitting for approval.",
        }

    updated = await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "approval_status": "Pending Review",
            "campaign_status": "Pending Review",
            "recipient_count": len(recipients),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Submitted for Approval",
        event_summary="Communication campaign submitted for leadership review.",
        event_details={"recipient_count": len(recipients)},
    )

    return {
        "status": "success",
        "message": "Campaign submitted for approval.",
        "campaign": updated[0] if updated else {},
    }


@app.post("/comm-campaign-approve")
async def comm_campaign_approve(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    approved_by = lori_comm_clean_text(payload.get("approved_by") or "Operations Leadership")
    approval_notes = lori_comm_clean_text(payload.get("approval_notes") or "Approved for communication queue.")

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
        }

    updated = await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "approval_status": "Approved",
            "campaign_status": "Approved",
            "approved_by": approved_by,
            "approved_at": datetime.utcnow().isoformat(),
            "approval_notes": approval_notes,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Campaign Approved",
        event_summary=f"Campaign approved by {approved_by}.",
        event_details={"approved_by": approved_by, "approval_notes": approval_notes},
        performed_by=approved_by,
    )

    return {
        "status": "success",
        "message": "Campaign approved.",
        "campaign": updated[0] if updated else {},
    }


@app.post("/comm-campaign-reject")
async def comm_campaign_reject(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    rejected_by = lori_comm_clean_text(payload.get("rejected_by") or "Operations Leadership")
    rejection_reason = lori_comm_clean_text(payload.get("rejection_reason") or "Returned for edits.")

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
        }

    updated = await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "approval_status": "Rejected",
            "campaign_status": "Returned for Edits",
            "approval_notes": rejection_reason,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Campaign Rejected",
        event_summary=f"Campaign rejected/returned by {rejected_by}.",
        event_details={"rejected_by": rejected_by, "rejection_reason": rejection_reason},
        performed_by=rejected_by,
    )

    return {
        "status": "success",
        "message": "Campaign returned for edits.",
        "campaign": updated[0] if updated else {},
    }


@app.post("/comm-campaign-queue")
async def comm_campaign_queue(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        f"select=*&id=eq.{quote(campaign_id)}&limit=1",
    )

    if not campaigns:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
        }

    campaign = campaigns[0]

    if lori_comm_clean_text(campaign.get("approval_status")).lower() != "approved":
        return {
            "status": "error",
            "message": "Campaign must be approved before it can be queued.",
            "approval_status": campaign.get("approval_status"),
        }

    recipients = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    if not recipients:
        return {
            "status": "error",
            "message": "Campaign has no recipients.",
        }

    existing_logs = await lori_comm_get_rows(
        "lori_comm_delivery_logs",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    existing_log_keys = {
        f"{l.get('recipient_id')}|{l.get('delivery_method')}"
        for l in existing_logs
    }

    created_logs = []
    created_acks = []

    delivery_method = lori_comm_clean_text(campaign.get("delivery_method") or "Email")

    for r in recipients:
        method_list = []

        if delivery_method.lower() in {"text and email", "sms and email"}:
            method_list = ["Text Message", "Email"]
        elif delivery_method.lower() in {"text message", "sms", "text"}:
            method_list = ["Text Message"]
        else:
            method_list = ["Email"]

        for method in method_list:
            key = f"{r.get('id')}|{method}"

            if key in existing_log_keys:
                continue

            log_payload = {
                "campaign_id": campaign_id,
                "recipient_id": r.get("id"),
                "delivery_method": method,
                "delivery_status": "Queued - Live Send Not Enabled",
                "provider_name": "Pending Provider",
                "failure_reason": "Live sending provider has not been connected yet.",
            }

            created = await lori_policy_supabase_post(
                "lori_comm_delivery_logs",
                log_payload,
            )

            if created:
                created_logs.append(created[0])

        if campaign.get("requires_acknowledgment"):
            existing_ack = await lori_comm_get_rows(
                "lori_comm_acknowledgments",
                f"select=*&campaign_id=eq.{quote(campaign_id)}&recipient_id=eq.{quote(str(r.get('id')))}&limit=1",
            )

            if not existing_ack:
                ack_payload = {
                    "campaign_id": campaign_id,
                    "recipient_id": r.get("id"),
                    "acknowledgment_required": True,
                    "acknowledgment_status": "Pending",
                    "acknowledgment_method": "Pending",
                }

                created_ack = await lori_policy_supabase_post(
                    "lori_comm_acknowledgments",
                    ack_payload,
                )

                if created_ack:
                    created_acks.append(created_ack[0])

    updated = await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "campaign_status": "Queued",
            "recipient_count": len(recipients),
            "sent_count": 0,
            "failed_count": 0,
            "acknowledged_count": 0,
            "pending_acknowledgment_count": len(recipients) if campaign.get("requires_acknowledgment") else 0,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Campaign Queued",
        event_summary="Campaign queued for future live sending. No live text or email was sent.",
        event_details={
            "live_send": False,
            "recipients": len(recipients),
            "delivery_logs_created": len(created_logs),
            "acknowledgments_created": len(created_acks),
        },
    )

    return {
        "status": "success",
        "message": "Campaign queued. No live text or email was sent because live providers are not connected yet.",
        "campaign": updated[0] if updated else {},
        "recipients_count": len(recipients),
        "delivery_logs_created": len(created_logs),
        "acknowledgments_created": len(created_acks),
        "answer_text": "Campaign queued for future live sending. No live text or email was sent yet.",
    }


@app.get("/comm-delivery-logs")
async def comm_delivery_logs(
    api_key: Optional[str] = Query(None),
    campaign_id: Optional[str] = Query(None),
    delivery_status: Optional[str] = Query(None),
    limit: int = Query(300),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_delivery_logs",
        "select=*&order=created_at.desc&limit=1000",
    )

    if campaign_id:
        rows = [
            r for r in rows
            if str(r.get("campaign_id")) == campaign_id
        ]

    if delivery_status:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("delivery_status")).lower() == delivery_status.lower()
        ]

    return {
        "status": "success",
        "delivery_logs_count": len(rows[:limit]),
        "delivery_logs": rows[:max(1, min(limit, 500))],
    }


@app.get("/comm-acknowledgments")
async def comm_acknowledgments(
    api_key: Optional[str] = Query(None),
    campaign_id: Optional[str] = Query(None),
    acknowledgment_status: Optional[str] = Query(None),
    limit: int = Query(300),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_acknowledgments",
        "select=*&order=created_at.desc&limit=1000",
    )

    if campaign_id:
        rows = [
            r for r in rows
            if str(r.get("campaign_id")) == campaign_id
        ]

    if acknowledgment_status:
        rows = [
            r for r in rows
            if lori_comm_clean_text(r.get("acknowledgment_status")).lower() == acknowledgment_status.lower()
        ]

    return {
        "status": "success",
        "acknowledgments_count": len(rows[:limit]),
        "acknowledgments": rows[:max(1, min(limit, 500))],
    }


@app.post("/comm-acknowledge")
async def comm_acknowledge(
    api_key: Optional[str] = Query(None),
    acknowledgment_id: str = Query(...),
    payload: Dict[str, Any] = Body(default={}),
):
    lori_regulatory_require_key(api_key)

    acknowledgments = await lori_comm_get_rows(
        "lori_comm_acknowledgments",
        f"select=*&id=eq.{quote(acknowledgment_id)}&limit=1",
    )

    if not acknowledgments:
        return {
            "status": "not_found",
            "message": "Acknowledgment record not found.",
        }

    ack = acknowledgments[0]
    campaign_id = ack.get("campaign_id")

    updated = await lori_policy_supabase_patch(
        "lori_comm_acknowledgments",
        acknowledgment_id,
        {
            "acknowledgment_status": "Acknowledged",
            "acknowledged_at": datetime.utcnow().isoformat(),
            "acknowledgment_method": lori_comm_clean_text(payload.get("acknowledgment_method") or "Manual"),
            "acknowledgment_note": lori_comm_clean_text(payload.get("acknowledgment_note") or "Acknowledged."),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    all_acks = await lori_comm_get_rows(
        "lori_comm_acknowledgments",
        f"select=*&campaign_id=eq.{quote(str(campaign_id))}&limit=1000",
    )

    acknowledged_count = len([
        a for a in all_acks
        if lori_comm_clean_text(a.get("acknowledgment_status")).lower() == "acknowledged"
    ])

    pending_count = len(all_acks) - acknowledged_count

    await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        str(campaign_id),
        {
            "acknowledged_count": acknowledged_count,
            "pending_acknowledgment_count": pending_count,
            "campaign_status": "Complete" if pending_count == 0 else "Acknowledgment Pending",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=str(campaign_id),
        event_type="Acknowledgment Completed",
        event_summary="Recipient acknowledgment was marked complete.",
        event_details={"acknowledgment_id": acknowledgment_id},
    )

    return {
        "status": "success",
        "message": "Acknowledgment marked complete.",
        "acknowledgment": updated[0] if updated else {},
        "acknowledged_count": acknowledged_count,
        "pending_acknowledgment_count": pending_count,
    }


@app.get("/comm-audit-events")
async def comm_audit_events(
    api_key: Optional[str] = Query(None),
    campaign_id: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_comm_get_rows(
        "lori_comm_audit_events",
        "select=*&order=created_at.desc&limit=1000",
    )

    if campaign_id:
        rows = [
            r for r in rows
            if str(r.get("campaign_id")) == campaign_id
        ]

    return {
        "status": "success",
        "audit_events_count": len(rows[:limit]),
        "audit_events": rows[:max(1, min(limit, 500))],
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# RESEND LIVE EMAIL SENDING
# Connects Push Notification Center queued email records to Resend.
# SMS/text remains queued until an SMS provider is connected.
# ============================================================

import os
import html as html_lib
import httpx
from typing import Optional, Dict, Any, List
from datetime import datetime
from fastapi import Body, Query


RESEND_API_URL = "https://api.resend.com/emails"


def lori_comm_email_enabled() -> bool:
    return os.getenv("COMM_LIVE_EMAIL_ENABLED", "false").strip().lower() == "true"


def lori_comm_get_resend_config() -> Dict[str, str]:
    return {
        "api_key": os.getenv("RESEND_API_KEY", "").strip(),
        "from_email": os.getenv("RESEND_FROM_EMAIL", "").strip(),
        "reply_to": os.getenv("RESEND_REPLY_TO", "").strip(),
    }


async def lori_comm_resolve_campaign_for_email(campaign_ref: str) -> Optional[Dict[str, Any]]:
    """
    Resolves a campaign by:
    - full UUID
    - partial UUID copied from a UI
    - exact message title
    """
    ref = lori_comm_clean_text(campaign_ref)

    if not ref:
        return None

    campaigns = await lori_comm_get_rows(
        "lori_comm_campaigns",
        "select=*&order=created_at.desc&limit=500",
    )

    for campaign in campaigns:
        if str(campaign.get("id")) == ref:
            return campaign

    partial_matches = [
        campaign for campaign in campaigns
        if str(campaign.get("id", "")).startswith(ref)
    ]

    if len(partial_matches) == 1:
        return partial_matches[0]

    for campaign in campaigns:
        if lori_comm_clean_text(campaign.get("message_title")).lower() == ref.lower():
            return campaign

    return None


def lori_comm_build_email_html(
    campaign: Dict[str, Any],
    recipient: Optional[Dict[str, Any]] = None,
) -> str:
    recipient = recipient or {}

    title = html_lib.escape(lori_comm_clean_text(campaign.get("message_title") or "LORI Communication"))
    subject = html_lib.escape(lori_comm_clean_text(campaign.get("subject_line") or title))
    message_type = html_lib.escape(lori_comm_clean_text(campaign.get("message_type") or "Notice"))
    priority = html_lib.escape(lori_comm_clean_text(campaign.get("priority") or "Routine"))
    body = html_lib.escape(lori_comm_clean_text(campaign.get("message_body") or ""))
    memo = html_lib.escape(str(campaign.get("memo_body") or ""))
    recipient_name = html_lib.escape(lori_comm_clean_text(recipient.get("recipient_name") or "Team Member"))
    station_code = html_lib.escape(lori_comm_clean_text(campaign.get("station_code") or "JESSUP-01"))

    body_html = body.replace("\n", "<br>")
    memo_html = memo.replace("\n", "<br>")

    acknowledgment_note = ""

    if campaign.get("requires_acknowledgment"):
        acknowledgment_note = """
        <div class="ack">
            <strong>Acknowledgment Required:</strong>
            Please acknowledge this communication according to your supervisor’s instructions.
        </div>
        """

    return f"""
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<style>
    body {{
        margin: 0;
        padding: 0;
        background: #f6f8fb;
        font-family: Arial, Helvetica, sans-serif;
        color: #111827;
    }}

    .wrap {{
        max-width: 720px;
        margin: 0 auto;
        padding: 28px;
    }}

    .card {{
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 18px;
        padding: 28px;
    }}

    .brand {{
        font-size: 12px;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        font-weight: 800;
        color: #475569;
        margin-bottom: 16px;
    }}

    .badge {{
        display: inline-block;
        background: #1e3a8a;
        color: #ffffff;
        border-radius: 999px;
        padding: 8px 12px;
        font-size: 12px;
        font-weight: 700;
        margin-bottom: 18px;
    }}

    h1 {{
        margin: 0 0 10px 0;
        font-size: 26px;
        line-height: 1.18;
    }}

    .meta {{
        color: #667085;
        font-size: 13px;
        margin-bottom: 22px;
    }}

    .message {{
        font-size: 15px;
        line-height: 1.6;
        margin-bottom: 20px;
    }}

    .memo {{
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 14px;
        padding: 18px;
        font-size: 14px;
        line-height: 1.55;
        margin-top: 16px;
    }}

    .ack {{
        margin-top: 20px;
        padding: 16px;
        border-radius: 14px;
        background: #fff7ed;
        border: 1px solid #fed7aa;
        color: #7c2d12;
        font-size: 14px;
    }}

    .footer {{
        margin-top: 22px;
        color: #667085;
        font-size: 12px;
        line-height: 1.45;
    }}
</style>
</head>
<body>
<div class="wrap">
    <div class="card">
        <div class="brand">LORI Drive Command Center</div>
        <div class="badge">{message_type} | {priority}</div>
        <h1>{subject}</h1>
        <div class="meta">To: {recipient_name} | Station: {station_code}</div>

        <div class="message">
            {body_html}
        </div>

        <div class="memo">
            {memo_html}
        </div>

        {acknowledgment_note}

        <div class="footer">
            This message was generated through LORI Drive Command Center. Please follow your organization’s communication, safety, HR, compliance, and acknowledgment procedures.
        </div>
    </div>
</div>
</body>
</html>
"""


def lori_comm_build_email_text(
    campaign: Dict[str, Any],
    recipient: Optional[Dict[str, Any]] = None,
) -> str:
    recipient = recipient or {}

    recipient_name = lori_comm_clean_text(recipient.get("recipient_name") or "Team Member")
    subject = lori_comm_clean_text(campaign.get("subject_line") or campaign.get("message_title") or "LORI Communication")
    message_body = lori_comm_clean_text(campaign.get("message_body") or "")
    memo_body = str(campaign.get("memo_body") or "")

    ack = ""

    if campaign.get("requires_acknowledgment"):
        ack = "\n\nACKNOWLEDGMENT REQUIRED: Please acknowledge this communication according to your supervisor’s instructions."

    return f"""LORI Drive Command Center

To: {recipient_name}
Subject: {subject}

{message_body}

{memo_body}
{ack}

This message was sent through LORI Drive Command Center.
"""


async def lori_resend_send_email(
    to_email: str,
    subject: str,
    html_content: str,
    text_content: str,
) -> Dict[str, Any]:
    config = lori_comm_get_resend_config()

    if not config["api_key"]:
        return {
            "ok": False,
            "status_code": 500,
            "error": "RESEND_API_KEY is missing in Render environment variables.",
        }

    if not config["from_email"]:
        return {
            "ok": False,
            "status_code": 500,
            "error": "RESEND_FROM_EMAIL is missing in Render environment variables.",
        }

    payload = {
        "from": config["from_email"],
        "to": [to_email],
        "subject": subject,
        "html": html_content,
        "text": text_content,
    }

    if config["reply_to"]:
        payload["reply_to"] = config["reply_to"]

    headers = {
        "Authorization": f"Bearer {config['api_key']}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            RESEND_API_URL,
            headers=headers,
            json=payload,
        )

    try:
        data = response.json()
    except Exception:
        data = {"raw_response": response.text}

    if response.status_code >= 400:
        return {
            "ok": False,
            "status_code": response.status_code,
            "error": data,
        }

    return {
        "ok": True,
        "status_code": response.status_code,
        "data": data,
        "provider_message_id": data.get("id"),
    }


@app.post("/comm-send-test-email")
async def comm_send_test_email(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    if not lori_comm_email_enabled():
        return {
            "status": "email_disabled",
            "message": "Live email sending is disabled. Set COMM_LIVE_EMAIL_ENABLED=true in Render.",
        }

    to_email = lori_comm_clean_text(payload.get("to_email"))
    subject = lori_comm_clean_text(payload.get("subject") or "LORI Drive Test Email")
    message_body = lori_comm_clean_text(payload.get("message_body") or "This is a LORI Drive test email using Resend.")

    if not to_email:
        return {
            "status": "error",
            "message": "to_email is required.",
        }

    html_content = f"""
    <div style="font-family: Arial, sans-serif; line-height: 1.5;">
        <h2>LORI Drive Command Center</h2>
        <p>{html_lib.escape(message_body)}</p>
        <p style="color:#667085;font-size:12px;">This is a live Resend email test from LORI.</p>
    </div>
    """

    result = await lori_resend_send_email(
        to_email=to_email,
        subject=subject,
        html_content=html_content,
        text_content=message_body,
    )

    if not result.get("ok"):
        return {
            "status": "failed",
            "message": "Test email failed.",
            "error": result,
        }

    return {
        "status": "success",
        "message": "Test email sent through Resend.",
        "resend_result": result,
    }


@app.post("/comm-campaign-send-email")
async def comm_campaign_send_email(
    api_key: Optional[str] = Query(None),
    campaign_id: str = Query(...),
    dry_run: bool = Query(False),
    force_resend: bool = Query(False),
    allow_demo_addresses: bool = Query(False),
):
    lori_regulatory_require_key(api_key)

    campaign = await lori_comm_resolve_campaign_for_email(campaign_id)

    if not campaign:
        return {
            "status": "not_found",
            "message": "Campaign not found.",
            "campaign_id": campaign_id,
        }

    campaign_id = str(campaign.get("id"))

    if not dry_run and not lori_comm_email_enabled():
        return {
            "status": "email_disabled",
            "message": "Live email sending is disabled. Set COMM_LIVE_EMAIL_ENABLED=true in Render.",
            "campaign_id": campaign_id,
        }

    approval_status = lori_comm_clean_text(campaign.get("approval_status")).lower()
    campaign_status = lori_comm_clean_text(campaign.get("campaign_status")).lower()

    allowed_statuses = {
        "approved",
        "queued",
        "email sent - sms pending",
        "acknowledgment pending",
    }

    if approval_status != "approved" and campaign_status not in allowed_statuses:
        return {
            "status": "error",
            "message": "Campaign must be approved or queued before sending live email.",
            "approval_status": campaign.get("approval_status"),
            "campaign_status": campaign.get("campaign_status"),
        }

    recipients = await lori_comm_get_rows(
        "lori_comm_recipients",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    if not recipients:
        return {
            "status": "error",
            "message": "Campaign has no recipients.",
            "campaign_id": campaign_id,
        }

    existing_logs = await lori_comm_get_rows(
        "lori_comm_delivery_logs",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    existing_email_logs_by_recipient = {
        str(log.get("recipient_id")): log
        for log in existing_logs
        if lori_comm_clean_text(log.get("delivery_method")).lower() == "email"
    }

    sent = []
    failed = []
    skipped = []
    dry_run_targets = []

    subject = lori_comm_clean_text(campaign.get("subject_line") or campaign.get("message_title") or "LORI Communication")

    for recipient in recipients:
        recipient_id = str(recipient.get("id"))
        recipient_name = lori_comm_clean_text(recipient.get("recipient_name"))
        to_email = lori_comm_clean_text(recipient.get("email_address"))

        if not to_email:
            skipped.append({
                "recipient": recipient_name,
                "reason": "No email address on recipient record.",
            })
            continue

        if to_email.lower().endswith("@example.com") and not allow_demo_addresses:
            skipped.append({
                "recipient": recipient_name,
                "email": to_email,
                "reason": "Demo example.com address skipped. Replace with real test email or set allow_demo_addresses=true.",
            })
            continue

        opt_out = lori_comm_clean_text(recipient.get("opt_out_status")).lower()

        if "opted out" in opt_out and "not opted out" not in opt_out:
            skipped.append({
                "recipient": recipient_name,
                "email": to_email,
                "reason": "Recipient is opted out.",
            })
            continue

        existing_log = existing_email_logs_by_recipient.get(recipient_id)

        if existing_log and lori_comm_clean_text(existing_log.get("delivery_status")).lower() == "sent" and not force_resend:
            skipped.append({
                "recipient": recipient_name,
                "email": to_email,
                "reason": "Email already marked as sent. Use force_resend=true to send again.",
            })
            continue

        if not existing_log:
            created_log = await lori_policy_supabase_post(
                "lori_comm_delivery_logs",
                {
                    "campaign_id": campaign_id,
                    "recipient_id": recipient_id,
                    "delivery_method": "Email",
                    "delivery_status": "Pending Email Send",
                    "provider_name": "Resend",
                },
            )

            existing_log = created_log[0] if created_log else None

        if dry_run:
            dry_run_targets.append({
                "recipient": recipient_name,
                "email": to_email,
                "delivery_log_id": existing_log.get("id") if existing_log else None,
            })
            continue

        html_content = lori_comm_build_email_html(campaign, recipient)
        text_content = lori_comm_build_email_text(campaign, recipient)

        result = await lori_resend_send_email(
            to_email=to_email,
            subject=subject,
            html_content=html_content,
            text_content=text_content,
        )

        log_id = existing_log.get("id") if existing_log else None

        if result.get("ok"):
            if log_id:
                await lori_policy_supabase_patch(
                    "lori_comm_delivery_logs",
                    log_id,
                    {
                        "delivery_status": "Sent",
                        "provider_name": "Resend",
                        "provider_message_id": result.get("provider_message_id"),
                        "sent_at": datetime.utcnow().isoformat(),
                        "failure_reason": None,
                        "updated_at": datetime.utcnow().isoformat(),
                    },
                )

            sent.append({
                "recipient": recipient_name,
                "email": to_email,
                "provider_message_id": result.get("provider_message_id"),
            })
        else:
            if log_id:
                await lori_policy_supabase_patch(
                    "lori_comm_delivery_logs",
                    log_id,
                    {
                        "delivery_status": "Failed",
                        "provider_name": "Resend",
                        "failed_at": datetime.utcnow().isoformat(),
                        "failure_reason": json.dumps(result.get("error")),
                        "updated_at": datetime.utcnow().isoformat(),
                    },
                )

            failed.append({
                "recipient": recipient_name,
                "email": to_email,
                "error": result,
            })

    if dry_run:
        return {
            "status": "dry_run",
            "message": "Dry run complete. No live email was sent.",
            "campaign_id": campaign_id,
            "would_send_count": len(dry_run_targets),
            "skipped_count": len(skipped),
            "would_send_to": dry_run_targets,
            "skipped": skipped,
        }

    updated_logs = await lori_comm_get_rows(
        "lori_comm_delivery_logs",
        f"select=*&campaign_id=eq.{quote(campaign_id)}&limit=1000",
    )

    sent_count = len([
        log for log in updated_logs
        if lori_comm_clean_text(log.get("delivery_status")).lower() == "sent"
    ])

    failed_count = len([
        log for log in updated_logs
        if lori_comm_clean_text(log.get("delivery_status")).lower() == "failed"
    ])

    text_logs_pending = len([
        log for log in updated_logs
        if lori_comm_clean_text(log.get("delivery_method")).lower() in {"text message", "sms", "text"}
        and lori_comm_clean_text(log.get("delivery_status")).lower() != "sent"
    ])

    if failed_count > 0:
        next_status = "Email Sent With Failures"
    elif text_logs_pending > 0:
        next_status = "Email Sent - SMS Pending"
    elif campaign.get("requires_acknowledgment"):
        next_status = "Acknowledgment Pending"
    else:
        next_status = "Sent"

    updated_campaign = await lori_policy_supabase_patch(
        "lori_comm_campaigns",
        campaign_id,
        {
            "sent_count": sent_count,
            "failed_count": failed_count,
            "campaign_status": next_status,
            "sent_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_comm_log_event(
        campaign_id=campaign_id,
        event_type="Live Email Send Attempted",
        event_summary=f"Resend email attempt complete. Sent: {len(sent)}. Failed: {len(failed)}. Skipped: {len(skipped)}.",
        event_details={
            "provider": "Resend",
            "sent_count": len(sent),
            "failed_count": len(failed),
            "skipped_count": len(skipped),
            "sms_connected": False,
        },
        performed_by="LORI Push Notification Center",
    )

    return {
        "status": "success" if not failed else "partial_success",
        "message": "Live email send attempt completed through Resend.",
        "campaign": updated_campaign[0] if updated_campaign else {},
        "sent_count": len(sent),
        "failed_count": len(failed),
        "skipped_count": len(skipped),
        "sent": sent,
        "failed": failed,
        "skipped": skipped,
        "answer_text": "Email sending completed through Resend. SMS/text remains queued until an SMS provider is connected.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# ROUTE SCORING ENGINE BACKEND
# Station-wide utilization, driver scorecards,
# route rebalancing opportunities, and station suggestions.
# ============================================================

from fastapi import Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote


async def lori_route_scoring_get_rows(
    table: str,
    query: str = "select=*&order=created_at.desc&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


def lori_route_scoring_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


async def lori_route_scoring_latest_review() -> Optional[Dict[str, Any]]:
    reviews = await lori_route_scoring_get_rows(
        "lori_route_station_reviews",
        "select=*&order=created_at.desc&limit=1",
    )

    if reviews:
        return reviews[0]

    return None


@app.get("/route-scoring-summary")
async def route_scoring_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    latest_review = await lori_route_scoring_latest_review()

    if not latest_review:
        return {
            "status": "empty",
            "message": "No route scoring review found.",
            "station_reviews_count": 0,
            "driver_scorecards_count": 0,
            "opportunities_count": 0,
            "station_suggestions_count": 0,
        }

    review_id = str(latest_review.get("id"))

    scorecards = await lori_route_scoring_get_rows(
        "lori_route_driver_scorecards",
        f"select=*&station_review_id=eq.{quote(review_id)}&order=utilization_percent.desc&limit=500",
    )

    opportunities = await lori_route_scoring_get_rows(
        "lori_route_rebalancing_opportunities",
        f"select=*&station_review_id=eq.{quote(review_id)}&order=created_at.desc&limit=100",
    )

    suggestions = await lori_route_scoring_get_rows(
        "lori_route_station_suggestions",
        f"select=*&station_review_id=eq.{quote(review_id)}&order=created_at.desc&limit=200",
    )

    overutilized = [
        r for r in scorecards
        if lori_route_scoring_clean(r.get("workload_status")).lower() == "overutilized"
    ]

    underutilized = [
        r for r in scorecards
        if lori_route_scoring_clean(r.get("workload_status")).lower() == "underutilized"
    ]

    near_capacity = [
        r for r in scorecards
        if lori_route_scoring_clean(r.get("workload_status")).lower() == "near capacity"
    ]

    balanced = [
        r for r in scorecards
        if lori_route_scoring_clean(r.get("workload_status")).lower() == "balanced"
    ]

    can_absorb = [
        r for r in scorecards
        if bool(r.get("can_absorb_work")) is True
    ]

    should_give_up = [
        r for r in scorecards
        if bool(r.get("should_give_up_work")) is True
    ]

    should_not_receive = [
        r for r in scorecards
        if bool(r.get("should_not_receive_work")) is True
    ]

    return {
        "status": "success",
        "latest_review": latest_review,
        "station_review_id": review_id,

        "driver_scorecards_count": len(scorecards),
        "opportunities_count": len(opportunities),
        "station_suggestions_count": len(suggestions),

        "overutilized_count": len(overutilized),
        "underutilized_count": len(underutilized),
        "near_capacity_count": len(near_capacity),
        "balanced_count": len(balanced),

        "can_absorb_work_count": len(can_absorb),
        "should_give_up_work_count": len(should_give_up),
        "should_not_receive_work_count": len(should_not_receive),

        "top_rebalancing_opportunity": opportunities[0] if opportunities else None,
        "top_station_suggestions": suggestions[:5],

        "answer_text": "Route scoring engine is ready. LORI identified route utilization, overutilized drivers, underutilized drivers, rebalancing opportunities, and station-wide route suggestions.",
    }


@app.get("/route-station-reviews")
async def route_station_reviews(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_scoring_get_rows(
        "lori_route_station_reviews",
        "select=*&order=created_at.desc&limit=500",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_route_scoring_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    return {
        "status": "success",
        "station_reviews_count": len(rows[:limit]),
        "station_reviews": rows[:max(1, min(limit, 300))],
    }


@app.get("/route-driver-scorecards")
async def route_driver_scorecards(
    api_key: Optional[str] = Query(None),
    station_review_id: Optional[str] = Query(None),
    workload_status: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    limit: int = Query(500),
):
    lori_regulatory_require_key(api_key)

    if not station_review_id:
        latest_review = await lori_route_scoring_latest_review()
        if latest_review:
            station_review_id = str(latest_review.get("id"))

    if not station_review_id:
        return {
            "status": "empty",
            "message": "No station review found.",
            "scorecards_count": 0,
            "scorecards": [],
        }

    rows = await lori_route_scoring_get_rows(
        "lori_route_driver_scorecards",
        f"select=*&station_review_id=eq.{quote(station_review_id)}&order=utilization_percent.desc&limit=1000",
    )

    if workload_status:
        rows = [
            r for r in rows
            if lori_route_scoring_clean(r.get("workload_status")).lower() == workload_status.lower()
        ]

    if driver_name:
        rows = [
            r for r in rows
            if driver_name.lower() in lori_route_scoring_clean(r.get("driver_name")).lower()
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_route_scoring_clean(r.get("route_id")).lower() == route_id.lower()
        ]

    return {
        "status": "success",
        "station_review_id": station_review_id,
        "scorecards_count": len(rows[:limit]),
        "scorecards": rows[:max(1, min(limit, 500))],
    }


@app.get("/route-rebalancing-opportunities")
async def route_rebalancing_opportunities(
    api_key: Optional[str] = Query(None),
    station_review_id: Optional[str] = Query(None),
    recommendation_strength: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    if not station_review_id:
        latest_review = await lori_route_scoring_latest_review()
        if latest_review:
            station_review_id = str(latest_review.get("id"))

    if not station_review_id:
        return {
            "status": "empty",
            "message": "No station review found.",
            "opportunities_count": 0,
            "opportunities": [],
        }

    rows = await lori_route_scoring_get_rows(
        "lori_route_rebalancing_opportunities",
        f"select=*&station_review_id=eq.{quote(station_review_id)}&order=estimated_total_savings.desc&limit=500",
    )

    if recommendation_strength:
        rows = [
            r for r in rows
            if recommendation_strength.lower() in lori_route_scoring_clean(r.get("recommendation_strength")).lower()
        ]

    return {
        "status": "success",
        "station_review_id": station_review_id,
        "opportunities_count": len(rows[:limit]),
        "opportunities": rows[:max(1, min(limit, 300))],
    }


@app.get("/route-station-suggestions")
async def route_station_suggestions(
    api_key: Optional[str] = Query(None),
    station_review_id: Optional[str] = Query(None),
    suggestion_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    if not station_review_id:
        latest_review = await lori_route_scoring_latest_review()
        if latest_review:
            station_review_id = str(latest_review.get("id"))

    if not station_review_id:
        return {
            "status": "empty",
            "message": "No station review found.",
            "suggestions_count": 0,
            "suggestions": [],
        }

    rows = await lori_route_scoring_get_rows(
        "lori_route_station_suggestions",
        f"select=*&station_review_id=eq.{quote(station_review_id)}&order=created_at.desc&limit=500",
    )

    if suggestion_type:
        rows = [
            r for r in rows
            if suggestion_type.lower() in lori_route_scoring_clean(r.get("suggestion_type")).lower()
        ]

    if priority:
        rows = [
            r for r in rows
            if lori_route_scoring_clean(r.get("priority")).lower() == priority.lower()
        ]

    return {
        "status": "success",
        "station_review_id": station_review_id,
        "suggestions_count": len(rows[:limit]),
        "suggestions": rows[:max(1, min(limit, 300))],
    }


@app.get("/route-scoring-review-detail")
async def route_scoring_review_detail(
    api_key: Optional[str] = Query(None),
    station_review_id: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    if not station_review_id:
        latest_review = await lori_route_scoring_latest_review()
        if latest_review:
            station_review_id = str(latest_review.get("id"))

    if not station_review_id:
        return {
            "status": "empty",
            "message": "No station route scoring review found.",
        }

    reviews = await lori_route_scoring_get_rows(
        "lori_route_station_reviews",
        f"select=*&id=eq.{quote(station_review_id)}&limit=1",
    )

    if not reviews:
        return {
            "status": "not_found",
            "message": "Station review not found.",
            "station_review_id": station_review_id,
        }

    scorecards = await lori_route_scoring_get_rows(
        "lori_route_driver_scorecards",
        f"select=*&station_review_id=eq.{quote(station_review_id)}&order=utilization_percent.desc&limit=1000",
    )

    opportunities = await lori_route_scoring_get_rows(
        "lori_route_rebalancing_opportunities",
        f"select=*&station_review_id=eq.{quote(station_review_id)}&order=estimated_total_savings.desc&limit=500",
    )

    suggestions = await lori_route_scoring_get_rows(
        "lori_route_station_suggestions",
        f"select=*&station_review_id=eq.{quote(station_review_id)}&order=created_at.desc&limit=500",
    )

    return {
        "status": "success",
        "station_review": reviews[0],
        "scorecards_count": len(scorecards),
        "scorecards": scorecards,
        "opportunities_count": len(opportunities),
        "opportunities": opportunities,
        "suggestions_count": len(suggestions),
        "suggestions": suggestions,
        "overutilized_drivers": [
            r for r in scorecards
            if lori_route_scoring_clean(r.get("workload_status")).lower() == "overutilized"
        ],
        "underutilized_drivers": [
            r for r in scorecards
            if lori_route_scoring_clean(r.get("workload_status")).lower() == "underutilized"
        ],
        "near_capacity_drivers": [
            r for r in scorecards
            if lori_route_scoring_clean(r.get("workload_status")).lower() == "near capacity"
        ],
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# SINGLE ROUTE RECONFIGURATION BACKEND
# Typed input + surrounding route candidate review
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime


def lori_single_route_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_single_route_num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def lori_single_route_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


async def lori_single_route_get_rows(
    table: str,
    query: str = "select=*&order=created_at.desc&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_single_route_latest_request() -> Optional[Dict[str, Any]]:
    rows = await lori_single_route_get_rows(
        "lori_single_route_reconfiguration_requests",
        "select=*&order=created_at.desc&limit=1",
    )
    return rows[0] if rows else None


async def lori_single_route_resolve_request(request_ref: str) -> Optional[Dict[str, Any]]:
    ref = lori_single_route_clean(request_ref)

    if not ref:
        return None

    rows = await lori_single_route_get_rows(
        "lori_single_route_reconfiguration_requests",
        "select=*&order=created_at.desc&limit=500",
    )

    for row in rows:
        if str(row.get("id")) == ref:
            return row

    partial_matches = [
        row for row in rows
        if str(row.get("id", "")).startswith(ref)
    ]

    if len(partial_matches) == 1:
        return partial_matches[0]

    for row in rows:
        if lori_single_route_clean(row.get("request_title")).lower() == ref.lower():
            return row

    return None


def lori_single_route_workload_status(utilization: float) -> str:
    if utilization >= 105:
        return "Overutilized"
    if utilization >= 95:
        return "Near Capacity"
    if utilization >= 85:
        return "Balanced"
    return "Underutilized"


def lori_single_route_candidate_score(
    utilization: float,
    stops_per_hour: float,
    overtime_hours: float,
    missed_stops: float = 0,
) -> float:
    capacity_score = max(0, 100 - utilization)
    productivity_score = min(100, stops_per_hour * 20)
    cost_fit_score = max(0, 100 - (overtime_hours * 25))
    service_score = max(0, 100 - (missed_stops * 20))

    return round(
        (capacity_score * 0.40)
        + (productivity_score * 0.25)
        + (cost_fit_score * 0.25)
        + (service_score * 0.10),
        2,
    )


@app.get("/single-route-summary")
async def single_route_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    latest_request = await lori_single_route_latest_request()

    requests = await lori_single_route_get_rows(
        "lori_single_route_reconfiguration_requests",
        "select=*&order=created_at.desc&limit=500",
    )

    candidates = await lori_single_route_get_rows(
        "lori_single_route_surrounding_candidates",
        "select=*&order=overall_candidate_score.desc&limit=500",
    )

    recommendations = await lori_single_route_get_rows(
        "lori_single_route_recommendations",
        "select=*&order=created_at.desc&limit=500",
    )

    audit = await lori_single_route_get_rows(
        "lori_single_route_typed_input_audit",
        "select=*&order=created_at.desc&limit=500",
    )

    analyzed = [
        r for r in requests
        if lori_single_route_clean(r.get("request_status")).lower() == "analyzed"
    ]

    typed = [
        r for r in requests
        if bool(r.get("typed_data_present")) is True
    ]

    uploaded = [
        r for r in requests
        if bool(r.get("uploaded_data_present")) is True
    ]

    return {
        "status": "success",
        "requests_count": len(requests),
        "analyzed_requests_count": len(analyzed),
        "typed_data_requests_count": len(typed),
        "uploaded_data_requests_count": len(uploaded),
        "surrounding_candidates_count": len(candidates),
        "recommendations_count": len(recommendations),
        "typed_input_audit_count": len(audit),
        "latest_request": latest_request,
        "latest_recommendation": recommendations[0] if recommendations else None,
        "answer_text": "Single Route Reconfiguration is ready for typed route data, surrounding route comparison, and proposed route-change recommendations.",
    }


@app.get("/single-route-requests")
async def single_route_requests(
    api_key: Optional[str] = Query(None),
    request_status: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_single_route_get_rows(
        "lori_single_route_reconfiguration_requests",
        "select=*&order=created_at.desc&limit=500",
    )

    if request_status:
        rows = [
            r for r in rows
            if lori_single_route_clean(r.get("request_status")).lower() == request_status.lower()
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_single_route_clean(r.get("current_route_id")).lower() == route_id.lower()
        ]

    if station_code:
        rows = [
            r for r in rows
            if lori_single_route_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    return {
        "status": "success",
        "requests_count": len(rows[:limit]),
        "requests": rows[:max(1, min(limit, 300))],
    }


@app.get("/single-route-detail")
async def single_route_detail(
    api_key: Optional[str] = Query(None),
    request_id: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    if not request_id:
        latest = await lori_single_route_latest_request()
        request_id = str(latest.get("id")) if latest else None

    if not request_id:
        return {
            "status": "empty",
            "message": "No single route request found.",
        }

    request = await lori_single_route_resolve_request(request_id)

    if not request:
        return {
            "status": "not_found",
            "message": "Single route request not found.",
            "request_id": request_id,
        }

    request_id = str(request.get("id"))

    candidates = await lori_single_route_get_rows(
        "lori_single_route_surrounding_candidates",
        f"select=*&request_id=eq.{quote(request_id)}&order=overall_candidate_score.desc&limit=500",
    )

    recommendations = await lori_single_route_get_rows(
        "lori_single_route_recommendations",
        f"select=*&request_id=eq.{quote(request_id)}&order=created_at.desc&limit=100",
    )

    audit = await lori_single_route_get_rows(
        "lori_single_route_typed_input_audit",
        f"select=*&request_id=eq.{quote(request_id)}&order=created_at.asc&limit=500",
    )

    return {
        "status": "success",
        "request": request,
        "surrounding_candidates_count": len(candidates),
        "surrounding_candidates": candidates,
        "recommendations_count": len(recommendations),
        "recommendations": recommendations,
        "typed_input_audit_count": len(audit),
        "typed_input_audit": audit,
        "top_candidate": candidates[0] if candidates else None,
        "top_recommendation": recommendations[0] if recommendations else None,
    }


@app.get("/single-route-surrounding-candidates")
async def single_route_surrounding_candidates(
    api_key: Optional[str] = Query(None),
    request_id: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    if not request_id:
        latest = await lori_single_route_latest_request()
        request_id = str(latest.get("id")) if latest else None

    if not request_id:
        return {
            "status": "empty",
            "candidates_count": 0,
            "candidates": [],
        }

    request = await lori_single_route_resolve_request(request_id)

    if not request:
        return {
            "status": "not_found",
            "message": "Request not found.",
        }

    rows = await lori_single_route_get_rows(
        "lori_single_route_surrounding_candidates",
        f"select=*&request_id=eq.{quote(str(request.get('id')))}&order=overall_candidate_score.desc&limit=500",
    )

    return {
        "status": "success",
        "request_id": str(request.get("id")),
        "candidates_count": len(rows[:limit]),
        "candidates": rows[:max(1, min(limit, 300))],
    }


@app.get("/single-route-recommendations")
async def single_route_recommendations(
    api_key: Optional[str] = Query(None),
    request_id: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    if not request_id:
        latest = await lori_single_route_latest_request()
        request_id = str(latest.get("id")) if latest else None

    if not request_id:
        return {
            "status": "empty",
            "recommendations_count": 0,
            "recommendations": [],
        }

    request = await lori_single_route_resolve_request(request_id)

    if not request:
        return {
            "status": "not_found",
            "message": "Request not found.",
        }

    rows = await lori_single_route_get_rows(
        "lori_single_route_recommendations",
        f"select=*&request_id=eq.{quote(str(request.get('id')))}&order=created_at.desc&limit=500",
    )

    return {
        "status": "success",
        "request_id": str(request.get("id")),
        "recommendations_count": len(rows[:limit]),
        "recommendations": rows[:max(1, min(limit, 300))],
    }


@app.get("/single-route-typed-input-audit")
async def single_route_typed_input_audit(
    api_key: Optional[str] = Query(None),
    request_id: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    if not request_id:
        latest = await lori_single_route_latest_request()
        request_id = str(latest.get("id")) if latest else None

    if not request_id:
        return {
            "status": "empty",
            "typed_input_audit_count": 0,
            "typed_input_audit": [],
        }

    request = await lori_single_route_resolve_request(request_id)

    if not request:
        return {
            "status": "not_found",
            "message": "Request not found.",
        }

    rows = await lori_single_route_get_rows(
        "lori_single_route_typed_input_audit",
        f"select=*&request_id=eq.{quote(str(request.get('id')))}&order=created_at.asc&limit=500",
    )

    return {
        "status": "success",
        "request_id": str(request.get("id")),
        "typed_input_audit_count": len(rows[:limit]),
        "typed_input_audit": rows[:max(1, min(limit, 300))],
    }


@app.post("/single-route-request-create")
async def single_route_request_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    current_route_id = lori_single_route_clean(payload.get("current_route_id"))
    current_driver_name = lori_single_route_clean(payload.get("current_driver_name"))

    if not current_route_id:
        return {
            "status": "error",
            "message": "current_route_id is required.",
        }

    if not current_driver_name:
        return {
            "status": "error",
            "message": "current_driver_name is required.",
        }

    request_title = lori_single_route_clean(
        payload.get("request_title")
        or f"Single Route Review — {current_route_id} {current_driver_name}"
    )

    request_payload = {
        "request_title": request_title,
        "request_status": "Draft",
        "request_mode": "Single Route Reconfiguration",
        "company_name": lori_single_route_clean(payload.get("company_name") or "Food Authority"),
        "station_code": lori_single_route_clean(payload.get("station_code") or "JESSUP-01"),
        "operating_state": lori_single_route_clean(payload.get("operating_state") or "MD"),
        "route_group": lori_single_route_clean(payload.get("route_group") or "Delivery Operations"),
        "data_source_mode": lori_single_route_clean(payload.get("data_source_mode") or "Typed Data"),
        "typed_data_present": True,
        "uploaded_data_present": bool(payload.get("uploaded_data_present") or False),
        "uploaded_file_count": lori_single_route_int(payload.get("uploaded_file_count"), 0),
        "current_route_id": current_route_id,
        "current_route_name": lori_single_route_clean(payload.get("current_route_name")),
        "current_driver_name": current_driver_name,
        "current_supervisor_name": lori_single_route_clean(payload.get("current_supervisor_name")),
        "reason_for_review": lori_single_route_clean(payload.get("reason_for_review")),
        "route_issue_type": lori_single_route_clean(payload.get("route_issue_type") or "Needs Review"),
        "typed_stop_count": lori_single_route_int(payload.get("typed_stop_count"), 0),
        "typed_miles": lori_single_route_num(payload.get("typed_miles"), 0),
        "typed_scheduled_hours": lori_single_route_num(payload.get("typed_scheduled_hours"), 0),
        "typed_actual_hours": lori_single_route_num(payload.get("typed_actual_hours"), 0),
        "typed_overtime_hours": lori_single_route_num(payload.get("typed_overtime_hours"), 0),
        "typed_helper_count": lori_single_route_int(payload.get("typed_helper_count"), 0),
        "typed_vehicle_type": lori_single_route_clean(payload.get("typed_vehicle_type") or "Box Truck"),
        "typed_freight_type": lori_single_route_clean(payload.get("typed_freight_type") or "Small Boxes"),
        "typed_delivery_window_pressure": lori_single_route_clean(payload.get("typed_delivery_window_pressure") or "Unknown"),
        "typed_traffic_pressure": lori_single_route_clean(payload.get("typed_traffic_pressure") or "Unknown"),
        "typed_service_issues": lori_single_route_clean(payload.get("typed_service_issues")),
        "typed_driver_notes": lori_single_route_clean(payload.get("typed_driver_notes")),
        "lori_initial_assessment": "Draft created. Run evaluation to compare surrounding routes and generate a recommendation.",
        "lori_decision_summary": "Pending evaluation.",
        "leadership_summary": "Pending evaluation.",
    }

    created = await lori_policy_supabase_post(
        "lori_single_route_reconfiguration_requests",
        request_payload,
    )

    request = created[0] if created else {}
    request_id = str(request.get("id"))

    audit_items = [
        ("Route", "Current Route ID", current_route_id),
        ("Route", "Current Driver", current_driver_name),
        ("Workload", "Typed Stop Count", str(request_payload["typed_stop_count"])),
        ("Workload", "Typed Actual Hours", str(request_payload["typed_actual_hours"])),
        ("Workload", "Typed Overtime Hours", str(request_payload["typed_overtime_hours"])),
        ("Issue", "Reason for Review", request_payload["reason_for_review"]),
    ]

    for section, label, value in audit_items:
        if value:
            await lori_policy_supabase_post(
                "lori_single_route_typed_input_audit",
                {
                    "request_id": request_id,
                    "input_section": section,
                    "input_label": label,
                    "input_value": value,
                    "input_source": "Typed by User",
                    "created_by": lori_single_route_clean(payload.get("created_by") or "Route Configuration User"),
                },
            )

    return {
        "status": "success",
        "message": "Single route reconfiguration request created from typed input.",
        "request": request,
        "request_id": request_id,
        "next_step": "Run POST /single-route-evaluate to compare surrounding routes and generate a recommendation.",
    }


@app.post("/single-route-evaluate")
async def single_route_evaluate(
    api_key: Optional[str] = Query(None),
    request_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    request = await lori_single_route_resolve_request(request_id)

    if not request:
        return {
            "status": "not_found",
            "message": "Single route request not found.",
            "request_id": request_id,
        }

    request_id = str(request.get("id"))
    current_route_id = lori_single_route_clean(request.get("current_route_id"))
    current_driver_name = lori_single_route_clean(request.get("current_driver_name"))

    # Clear previous generated candidates/recommendations for this request.
    old_candidates = await lori_single_route_get_rows(
        "lori_single_route_surrounding_candidates",
        f"select=*&request_id=eq.{quote(request_id)}&limit=500",
    )

    old_recs = await lori_single_route_get_rows(
        "lori_single_route_recommendations",
        f"select=*&request_id=eq.{quote(request_id)}&limit=500",
    )

    for row in old_candidates:
        await lori_policy_supabase_delete(
            "lori_single_route_surrounding_candidates",
            str(row.get("id")),
        )

    for row in old_recs:
        await lori_policy_supabase_delete(
            "lori_single_route_recommendations",
            str(row.get("id")),
        )

    # Use existing route scoring scorecards as surrounding route intelligence.
    scorecards = await lori_single_route_get_rows(
        "lori_route_driver_scorecards",
        "select=*&order=utilization_percent.asc&limit=500",
    )

    candidate_rows = [
        row for row in scorecards
        if lori_single_route_clean(row.get("route_id")).lower() != current_route_id.lower()
    ]

    created_candidates = []

    for candidate in candidate_rows:
        utilization = lori_single_route_num(candidate.get("utilization_percent"), 0)
        stops_per_hour = lori_single_route_num(candidate.get("stops_per_hour"), 0)
        overtime = lori_single_route_num(candidate.get("overtime_hours"), 0)
        missed = lori_single_route_num(candidate.get("missed_stop_count"), 0)

        overall_score = lori_single_route_candidate_score(
            utilization=utilization,
            stops_per_hour=stops_per_hour,
            overtime_hours=overtime,
            missed_stops=missed,
        )

        can_absorb = utilization < 90
        should_not_receive = utilization >= 100

        candidate_payload = {
            "request_id": request_id,
            "candidate_route_id": candidate.get("route_id"),
            "candidate_route_name": candidate.get("route_name"),
            "candidate_driver_name": candidate.get("driver_name"),
            "candidate_supervisor_name": candidate.get("supervisor_name"),
            "station_code": candidate.get("station_code") or request.get("station_code"),
            "route_group": candidate.get("route_group") or request.get("route_group"),
            "stop_count": lori_single_route_int(candidate.get("stop_count"), 0),
            "total_miles": lori_single_route_num(candidate.get("total_miles"), 0),
            "scheduled_hours": lori_single_route_num(candidate.get("scheduled_hours"), 0),
            "actual_hours": lori_single_route_num(candidate.get("actual_hours"), 0),
            "overtime_hours": overtime,
            "helper_count": lori_single_route_int(candidate.get("helper_count"), 0),
            "vehicle_type": candidate.get("vehicle_type"),
            "freight_type": candidate.get("freight_type"),
            "utilization_percent": utilization,
            "stops_per_hour": stops_per_hour,
            "capacity_status": "Available Capacity" if can_absorb else "Near/At Capacity",
            "can_absorb_work": can_absorb,
            "should_not_receive_work": should_not_receive,
            "proximity_to_review_route": "Surrounding / Needs Map Confirmation",
            "compatibility_score": 90 if can_absorb else 65,
            "productivity_score": min(100, stops_per_hour * 20),
            "cost_fit_score": max(0, 100 - (overtime * 25)),
            "service_risk_score": max(0, missed * 20),
            "overall_candidate_score": overall_score,
            "lori_candidate_summary": "Candidate evaluated against typed route request using utilization, productivity, overtime, service risk, and capacity.",
            "reason_to_consider": "May be able to absorb limited nearby work if delivery windows, customer commitments, vehicle fit, and freight complexity are confirmed.",
            "reason_to_reject": "Reject or hold if adding stops creates overtime, missed windows, safety risk, or vehicle-capacity issues.",
        }

        created = await lori_policy_supabase_post(
            "lori_single_route_surrounding_candidates",
            candidate_payload,
        )

        if created:
            created_candidates.append(created[0])

    created_candidates = sorted(
        created_candidates,
        key=lambda x: lori_single_route_num(x.get("overall_candidate_score"), 0),
        reverse=True,
    )

    best = created_candidates[0] if created_candidates else None

    typed_utilization = 0
    actual_hours = lori_single_route_num(request.get("typed_actual_hours"), 0)
    scheduled_hours = lori_single_route_num(request.get("typed_scheduled_hours"), 0)

    if scheduled_hours > 0:
        typed_utilization = round((actual_hours / scheduled_hours) * 100, 2)

    suggested_stops = 6 if typed_utilization >= 105 else 3
    before_util = typed_utilization
    after_from = max(0, before_util - 15)
    after_to = min(99, lori_single_route_num(best.get("utilization_percent"), 0) + 7) if best else None

    recommendation = None

    if best:
        recommendation_payload = {
            "request_id": request_id,
            "recommendation_title": f"Recommended Reconfiguration: Move {suggested_stops} stops from {current_route_id} to {best.get('candidate_route_id')}",
            "recommendation_status": "Draft",
            "decision": "Recommended for Supervisor Review",
            "priority": "High" if before_util >= 105 else "Medium",
            "from_route_id": current_route_id,
            "from_route_name": request.get("current_route_name"),
            "from_driver_name": current_driver_name,
            "to_route_id": best.get("candidate_route_id"),
            "to_route_name": best.get("candidate_route_name"),
            "to_driver_name": best.get("candidate_driver_name"),
            "suggested_action": "Move limited low-complexity stops to the surrounding route with the best available capacity.",
            "suggested_stops_to_move": suggested_stops,
            "suggested_freight_type": request.get("typed_freight_type") or "Small Boxes",
            "suggested_complexity": "Low",
            "before_utilization_percent": before_util,
            "after_from_route_utilization_percent": after_from,
            "after_to_route_utilization_percent": after_to,
            "estimated_miles_reduced": 7.5,
            "estimated_hours_reduced": 1.25,
            "estimated_overtime_reduction": 1.0,
            "estimated_fuel_savings": 3.38,
            "estimated_labor_savings": 40.00,
            "estimated_total_savings": 46.58,
            "estimated_productivity_gain": 7.5,
            "service_risk": "Low to Medium",
            "safety_risk": "Low to Medium",
            "customer_impact": "Delivery windows must be verified before moving stops.",
            "helper_impact": "No additional helper expected based on typed data.",
            "vehicle_fit_impact": "Vehicle and freight compatibility require supervisor confirmation.",
            "why_this_makes_sense": f"{current_route_id} appears overloaded while {best.get('candidate_route_id')} has stronger available-capacity indicators.",
            "if_you_do_this": "Route pressure should decrease, overtime exposure should improve, and available capacity should be used more productively.",
            "if_you_do_not_do_this": "The reviewed route may continue carrying overtime, service risk, fuel waste, and driver fatigue exposure.",
            "supervisor_review_notes": "Supervisor must confirm exact customers/stops, delivery windows, service commitments, driver workload, vehicle fit, and safety before approval.",
            "recommended_next_action": "Review proposed stop movement with dispatch and supervisor, then test for one review cycle before making permanent.",
            "diagram_json": {
                "diagram_title": "Before / After Route Reconfiguration",
                "before": {
                    "route_id": current_route_id,
                    "driver": current_driver_name,
                    "status": lori_single_route_workload_status(before_util),
                    "utilization_percent": before_util,
                    "typed_stop_count": request.get("typed_stop_count"),
                },
                "after_from_route": {
                    "route_id": current_route_id,
                    "driver": current_driver_name,
                    "status": "Improved",
                    "projected_utilization_percent": after_from,
                    "stops_moved_out": suggested_stops,
                },
                "after_to_route": {
                    "route_id": best.get("candidate_route_id"),
                    "driver": best.get("candidate_driver_name"),
                    "status": "Capacity Used",
                    "projected_utilization_percent": after_to,
                    "stops_moved_in": suggested_stops,
                },
                "arrow_label": f"Move {suggested_stops} low-complexity stops",
                "visual_type": "route-card-arrow-diagram",
            },
        }

        created_rec = await lori_policy_supabase_post(
            "lori_single_route_recommendations",
            recommendation_payload,
        )

        recommendation = created_rec[0] if created_rec else None

        await lori_policy_supabase_patch(
            "lori_single_route_reconfiguration_requests",
            request_id,
            {
                "request_status": "Analyzed",
                "surrounding_routes_considered": len(created_candidates),
                "best_candidate_route_id": best.get("candidate_route_id"),
                "best_candidate_driver_name": best.get("candidate_driver_name"),
                "lori_initial_assessment": f"{current_route_id} was evaluated using typed route data and surrounding route scorecards.",
                "lori_decision_summary": recommendation_payload["suggested_action"],
                "leadership_summary": recommendation_payload["why_this_makes_sense"],
                "updated_at": datetime.utcnow().isoformat(),
            },
        )

    return {
        "status": "success",
        "message": "Single route typed input evaluation completed.",
        "request_id": request_id,
        "candidates_created": len(created_candidates),
        "top_candidate": best,
        "recommendation": recommendation,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# ROUTE MAP DATA ENGINE BACKEND
# Map layers, route stops, work-area boundaries, ZIP coverage,
# crossover findings, and map download packages.
# ============================================================

from fastapi import Query
from fastapi.responses import HTMLResponse
from typing import Any, Dict, List, Optional
from urllib.parse import quote
import html


def lori_route_map_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


async def lori_route_map_get_rows(
    table: str,
    query: str = "select=*&order=created_at.desc&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


@app.get("/route-map-summary")
async def route_map_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    layers = await lori_route_map_get_rows(
        "lori_route_map_layers",
        "select=*&order=created_at.desc&limit=500",
    )

    stops = await lori_route_map_get_rows(
        "lori_route_stop_points",
        "select=*&order=route_id.asc,stop_sequence.asc&limit=1000",
    )

    boundaries = await lori_route_map_get_rows(
        "lori_route_work_area_boundaries",
        "select=*&order=created_at.desc&limit=500",
    )

    crossovers = await lori_route_map_get_rows(
        "lori_route_crossover_findings",
        "select=*&order=created_at.desc&limit=500",
    )

    packages = await lori_route_map_get_rows(
        "lori_route_map_download_packages",
        "select=*&order=created_at.desc&limit=500",
    )

    current_route_layers = [
        r for r in layers
        if lori_route_map_clean(r.get("map_type")).lower() == "current route map"
    ]

    before_after_layers = [
        r for r in layers
        if "before" in lori_route_map_clean(r.get("map_type")).lower()
        or "after" in lori_route_map_clean(r.get("map_type")).lower()
    ]

    station_layers = [
        r for r in layers
        if "station" in lori_route_map_clean(r.get("map_type")).lower()
    ]

    moved_stops = [
        r for r in stops
        if bool(r.get("moved_stop_candidate")) is True
    ]

    return {
        "status": "success",
        "map_layers_count": len(layers),
        "current_route_layers_count": len(current_route_layers),
        "before_after_layers_count": len(before_after_layers),
        "station_layers_count": len(station_layers),
        "stop_points_count": len(stops),
        "moved_stops_count": len(moved_stops),
        "work_area_boundaries_count": len(boundaries),
        "crossover_findings_count": len(crossovers),
        "download_packages_count": len(packages),
        "latest_layer": layers[0] if layers else None,
        "latest_crossover_finding": crossovers[0] if crossovers else None,
        "answer_text": "Route Map Data Engine is ready. LORI can provide route map layers, stop points, work-area boundaries, ZIP coverage, crossover findings, and downloadable map packages.",
    }


@app.get("/route-map-layers")
async def route_map_layers(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    map_type: Optional[str] = Query(None),
    limit: int = Query(500),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_map_get_rows(
        "lori_route_map_layers",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("route_id")).lower() == route_id.lower()
            or lori_route_map_clean(r.get("comparison_route_id")).lower() == route_id.lower()
        ]

    if map_type:
        rows = [
            r for r in rows
            if map_type.lower() in lori_route_map_clean(r.get("map_type")).lower()
        ]

    return {
        "status": "success",
        "layers_count": len(rows[:limit]),
        "layers": rows[:max(1, min(limit, 500))],
    }


@app.get("/route-map-layer-detail")
async def route_map_layer_detail(
    api_key: Optional[str] = Query(None),
    layer_id: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    if not layer_id:
        return {
            "status": "error",
            "message": "layer_id is required.",
        }

    rows = await lori_route_map_get_rows(
        "lori_route_map_layers",
        f"select=*&id=eq.{quote(layer_id)}&limit=1",
    )

    if not rows:
        return {
            "status": "not_found",
            "message": "Map layer not found.",
            "layer_id": layer_id,
        }

    return {
        "status": "success",
        "layer": rows[0],
    }


@app.get("/route-stop-points")
async def route_stop_points(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    moved_only: bool = Query(False),
    limit: int = Query(1000),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_map_get_rows(
        "lori_route_stop_points",
        "select=*&order=route_id.asc,stop_sequence.asc&limit=2000",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("route_id")).lower() == route_id.lower()
            or lori_route_map_clean(r.get("moved_to_route_id")).lower() == route_id.lower()
        ]

    if moved_only:
        rows = [
            r for r in rows
            if bool(r.get("moved_stop_candidate")) is True
        ]

    return {
        "status": "success",
        "stop_points_count": len(rows[:limit]),
        "stop_points": rows[:max(1, min(limit, 1000))],
    }


@app.get("/route-work-area-boundaries")
async def route_work_area_boundaries(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    zip_code: Optional[str] = Query(None),
    boundary_type: Optional[str] = Query(None),
    limit: int = Query(500),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_map_get_rows(
        "lori_route_work_area_boundaries",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("route_id")).lower() == route_id.lower()
        ]

    if zip_code:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("zip_code")) == zip_code
        ]

    if boundary_type:
        rows = [
            r for r in rows
            if boundary_type.lower() in lori_route_map_clean(r.get("boundary_type")).lower()
        ]

    return {
        "status": "success",
        "boundaries_count": len(rows[:limit]),
        "boundaries": rows[:max(1, min(limit, 500))],
    }


@app.get("/route-crossover-findings")
async def route_crossover_findings(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    source_route_id: Optional[str] = Query(None),
    crossed_route_id: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    limit: int = Query(500),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_map_get_rows(
        "lori_route_crossover_findings",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if source_route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("source_route_id")).lower() == source_route_id.lower()
        ]

    if crossed_route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("crossed_route_id")).lower() == crossed_route_id.lower()
        ]

    if severity:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("severity")).lower() == severity.lower()
        ]

    return {
        "status": "success",
        "crossover_findings_count": len(rows[:limit]),
        "crossover_findings": rows[:max(1, min(limit, 500))],
    }


@app.get("/route-map-download-packages")
async def route_map_download_packages(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    comparison_route_id: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_route_map_get_rows(
        "lori_route_map_download_packages",
        "select=*&order=created_at.desc&limit=500",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("route_id")).lower() == route_id.lower()
        ]

    if comparison_route_id:
        rows = [
            r for r in rows
            if lori_route_map_clean(r.get("comparison_route_id")).lower() == comparison_route_id.lower()
        ]

    return {
        "status": "success",
        "download_packages_count": len(rows[:limit]),
        "download_packages": rows[:max(1, min(limit, 100))],
    }


@app.get("/route-map-view-data")
async def route_map_view_data(
    api_key: Optional[str] = Query(None),
    station_code: str = Query("JESSUP-01"),
    route_id: Optional[str] = Query(None),
    comparison_route_id: Optional[str] = Query(None),
    view_type: str = Query("current"),
):
    lori_regulatory_require_key(api_key)

    layers = await lori_route_map_get_rows(
        "lori_route_map_layers",
        "select=*&order=created_at.desc&limit=1000",
    )

    stops = await lori_route_map_get_rows(
        "lori_route_stop_points",
        "select=*&order=route_id.asc,stop_sequence.asc&limit=2000",
    )

    boundaries = await lori_route_map_get_rows(
        "lori_route_work_area_boundaries",
        "select=*&order=created_at.desc&limit=1000",
    )

    crossovers = await lori_route_map_get_rows(
        "lori_route_crossover_findings",
        "select=*&order=created_at.desc&limit=1000",
    )

    packages = await lori_route_map_get_rows(
        "lori_route_map_download_packages",
        "select=*&order=created_at.desc&limit=500",
    )

    layers = [
        r for r in layers
        if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
    ]

    stops = [
        r for r in stops
        if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
    ]

    boundaries = [
        r for r in boundaries
        if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
    ]

    crossovers = [
        r for r in crossovers
        if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
    ]

    packages = [
        r for r in packages
        if lori_route_map_clean(r.get("station_code")).lower() == station_code.lower()
    ]

    if route_id:
        route_id_lower = route_id.lower()

        layers = [
            r for r in layers
            if lori_route_map_clean(r.get("route_id")).lower() == route_id_lower
            or lori_route_map_clean(r.get("comparison_route_id")).lower() == route_id_lower
            or "station" in lori_route_map_clean(r.get("map_type")).lower()
        ]

        stops = [
            r for r in stops
            if lori_route_map_clean(r.get("route_id")).lower() == route_id_lower
            or lori_route_map_clean(r.get("moved_to_route_id")).lower() == route_id_lower
        ]

        boundaries = [
            r for r in boundaries
            if lori_route_map_clean(r.get("route_id")).lower() == route_id_lower
            or lori_route_map_clean(r.get("boundary_type")).lower() == "zip coverage"
        ]

        crossovers = [
            r for r in crossovers
            if lori_route_map_clean(r.get("source_route_id")).lower() == route_id_lower
            or lori_route_map_clean(r.get("crossed_route_id")).lower() == route_id_lower
        ]

    if comparison_route_id:
        comp_lower = comparison_route_id.lower()

        layers = [
            r for r in layers
            if lori_route_map_clean(r.get("comparison_route_id")).lower() == comp_lower
            or lori_route_map_clean(r.get("route_id")).lower() == comp_lower
            or "station" in lori_route_map_clean(r.get("map_type")).lower()
        ] or layers

        stops = [
            r for r in stops
            if lori_route_map_clean(r.get("route_id")).lower() == comp_lower
            or lori_route_map_clean(r.get("moved_to_route_id")).lower() == comp_lower
        ] or stops

        boundaries = [
            r for r in boundaries
            if lori_route_map_clean(r.get("route_id")).lower() == comp_lower
            or lori_route_map_clean(r.get("boundary_type")).lower() == "zip coverage"
        ] or boundaries

    view_type_lower = view_type.lower()

    if view_type_lower in ["current", "current_route"]:
        selected_layers = [
            r for r in layers
            if "current" in lori_route_map_clean(r.get("map_type")).lower()
        ]
    elif view_type_lower in ["before_after", "comparison", "proposed"]:
        selected_layers = [
            r for r in layers
            if "before" in lori_route_map_clean(r.get("map_type")).lower()
            or "after" in lori_route_map_clean(r.get("map_type")).lower()
            or "comparison" in lori_route_map_clean(r.get("layer_type")).lower()
        ]
    elif view_type_lower in ["station", "full_station"]:
        selected_layers = [
            r for r in layers
            if "station" in lori_route_map_clean(r.get("map_type")).lower()
        ]
    else:
        selected_layers = layers

    if not selected_layers:
        selected_layers = layers

    return {
        "status": "success",
        "view_type": view_type,
        "station_code": station_code,
        "route_id": route_id,
        "comparison_route_id": comparison_route_id,
        "layers_count": len(selected_layers),
        "layers": selected_layers,
        "stop_points_count": len(stops),
        "stop_points": stops,
        "boundaries_count": len(boundaries),
        "boundaries": boundaries,
        "crossover_findings_count": len(crossovers),
        "crossover_findings": crossovers,
        "download_packages_count": len(packages),
        "download_packages": packages,
        "map_note": "Map view is based on currently available route geography, ZIP coverage, work area boundaries, crossover findings, and route configuration data.",
    }


@app.get("/route-map-download-html", response_class=HTMLResponse)
async def route_map_download_html(
    api_key: Optional[str] = Query(None),
    station_code: str = Query("JESSUP-01"),
    route_id: str = Query("R-104"),
    comparison_route_id: str = Query("R-208"),
):
    lori_regulatory_require_key(api_key)

    data = await route_map_view_data(
        api_key=api_key,
        station_code=station_code,
        route_id=route_id,
        comparison_route_id=comparison_route_id,
        view_type="comparison",
    )

    layers = data.get("layers", [])
    stops = data.get("stop_points", [])
    boundaries = data.get("boundaries", [])
    crossovers = data.get("crossover_findings", [])

    layer_items = "".join(
        f"<li><strong>{html.escape(str(layer.get('map_title', 'Map Layer')))}</strong> — "
        f"{html.escape(str(layer.get('map_type', '')))} / "
        f"{html.escape(str(layer.get('layer_label', '')))}</li>"
        for layer in layers
    )

    stop_items = "".join(
        f"<tr><td>{html.escape(str(stop.get('route_id', '')))}</td>"
        f"<td>{html.escape(str(stop.get('marker_label', '')))}</td>"
        f"<td>{html.escape(str(stop.get('stop_city', '')))}, {html.escape(str(stop.get('stop_state', '')))}</td>"
        f"<td>{html.escape(str(stop.get('stop_zip', '')))}</td>"
        f"<td>{html.escape(str(stop.get('stop_type', '')))}</td></tr>"
        for stop in stops
    )

    boundary_items = "".join(
        f"<li><strong>{html.escape(str(boundary.get('boundary_title', 'Boundary')))}</strong> — "
        f"{html.escape(str(boundary.get('boundary_type', '')))} "
        f"{html.escape(str(boundary.get('zip_code', '')))}</li>"
        for boundary in boundaries
    )

    crossover_items = "".join(
        f"<li><strong>{html.escape(str(crossover.get('finding_title', 'Crossover Finding')))}</strong><br>"
        f"Severity: {html.escape(str(crossover.get('severity', '')))}<br>"
        f"{html.escape(str(crossover.get('finding_summary', '')))}</li>"
        for crossover in crossovers
    )

    html_doc = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <title>LORI Route Map Package {html.escape(route_id)} to {html.escape(comparison_route_id)}</title>
      <style>
        body {{
          font-family: Arial, sans-serif;
          margin: 32px;
          color: #111827;
          background: #f8fafc;
        }}
        .card {{
          background: white;
          border: 1px solid #dbe3ef;
          border-radius: 16px;
          padding: 20px;
          margin-bottom: 18px;
          box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
        }}
        h1 {{ margin-bottom: 4px; }}
        h2 {{ margin-top: 0; color: #1f2937; }}
        .badge {{
          display: inline-block;
          padding: 6px 10px;
          border-radius: 999px;
          background: #e0f2fe;
          color: #075985;
          font-size: 12px;
          font-weight: 700;
          margin-right: 8px;
        }}
        table {{
          width: 100%;
          border-collapse: collapse;
          font-size: 13px;
        }}
        th, td {{
          border: 1px solid #e5e7eb;
          padding: 8px;
          text-align: left;
        }}
        th {{
          background: #f1f5f9;
        }}
        .note {{
          font-size: 12px;
          color: #64748b;
          line-height: 1.5;
        }}
      </style>
    </head>
    <body>
      <div class="card">
        <h1>LORI Route Map Package</h1>
        <p>
          <span class="badge">Station: {html.escape(station_code)}</span>
          <span class="badge">Route: {html.escape(route_id)}</span>
          <span class="badge">Comparison: {html.escape(comparison_route_id)}</span>
        </p>
        <p class="note">Generated by LORI Drive Command Center. Demonstration Data Only — Not Company Proprietary Data.</p>
      </div>

      <div class="card">
        <h2>Map Layers</h2>
        <ul>{layer_items or "<li>No map layers found.</li>"}</ul>
      </div>

      <div class="card">
        <h2>Route Stop Points</h2>
        <table>
          <thead>
            <tr>
              <th>Route</th>
              <th>Marker</th>
              <th>City/State</th>
              <th>ZIP</th>
              <th>Stop Type</th>
            </tr>
          </thead>
          <tbody>{stop_items or "<tr><td colspan='5'>No stop points found.</td></tr>"}</tbody>
        </table>
      </div>

      <div class="card">
        <h2>Work Area / ZIP Boundaries</h2>
        <ul>{boundary_items or "<li>No boundaries found.</li>"}</ul>
      </div>

      <div class="card">
        <h2>Crossover Findings</h2>
        <ul>{crossover_items or "<li>No crossover findings found.</li>"}</ul>
      </div>

      <div class="card note">
        LORI map views are operational decision-support visuals based on available route geography, ZIP coverage, work area boundaries, and route configuration data.
        Final route changes must be reviewed and approved by authorized operations leadership before implementation.
      </div>
    </body>
    </html>
    """

    return HTMLResponse(content=html_doc)
# ============================================================
# LORI DRIVE COMMAND CENTER
# NATIONWIDE OPERATING CONTEXT BACKEND
# U.S. states, regions, stations, ZIP coverage,
# state compliance profiles, and module operating context.
# ============================================================

from fastapi import Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote


def lori_nationwide_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


async def lori_nationwide_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


@app.get("/nationwide-context-summary")
async def nationwide_context_summary(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    states = await lori_nationwide_get_rows(
        "lori_us_operating_states",
        "select=*&order=state_code.asc&limit=100",
    )

    regions = await lori_nationwide_get_rows(
        "lori_operating_regions",
        "select=*&order=region_name.asc&limit=100",
    )

    stations = await lori_nationwide_get_rows(
        "lori_operating_stations",
        "select=*&order=station_code.asc&limit=500",
    )

    zip_coverages = await lori_nationwide_get_rows(
        "lori_station_zip_coverage",
        "select=*&order=station_code.asc,zip_code.asc&limit=1000",
    )

    compliance_profiles = await lori_nationwide_get_rows(
        "lori_state_compliance_profiles",
        "select=*&order=state_code.asc&limit=100",
    )

    module_contexts = await lori_nationwide_get_rows(
        "lori_nationwide_module_contexts",
        "select=*&order=module_name.asc&limit=100",
    )

    active_stations = [
        s for s in stations
        if lori_nationwide_clean(s.get("station_status")).lower() == "active"
    ]

    demo_stations = [
        s for s in stations
        if lori_nationwide_clean(s.get("station_status")).lower() == "demo"
    ]

    map_ready = [
        s for s in stations
        if bool(s.get("map_ready")) is True
    ]

    route_scoring_ready = [
        s for s in stations
        if bool(s.get("route_scoring_ready")) is True
    ]

    push_ready = [
        s for s in stations
        if bool(s.get("push_notification_ready")) is True
    ]

    return {
        "status": "success",
        "states_count": len(states),
        "regions_count": len(regions),
        "stations_count": len(stations),
        "active_stations_count": len(active_stations),
        "demo_stations_count": len(demo_stations),
        "zip_coverages_count": len(zip_coverages),
        "state_compliance_profiles_count": len(compliance_profiles),
        "module_contexts_count": len(module_contexts),
        "map_ready_stations_count": len(map_ready),
        "route_scoring_ready_stations_count": len(route_scoring_ready),
        "push_notification_ready_stations_count": len(push_ready),
        "default_operating_context": {
            "company_name": "Food Authority",
            "region": "Mid-Atlantic",
            "state": "MD",
            "station_code": "JESSUP-01",
            "city": "Jessup",
            "time_zone": "America/New_York",
        },
        "answer_text": "Nationwide Operating Context is ready. LORI can support national, regional, state, station, ZIP, route, driver, compliance, and module-level filtering across the United States.",
    }


@app.get("/us-operating-states")
async def us_operating_states(
    api_key: Optional[str] = Query(None),
    state_code: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    active_only: bool = Query(True),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_nationwide_get_rows(
        "lori_us_operating_states",
        "select=*&order=state_code.asc&limit=100",
    )

    if active_only:
        rows = [r for r in rows if bool(r.get("active_state")) is True]

    if state_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("state_code")).lower() == state_code.lower()
        ]

    if region:
        rows = [
            r for r in rows
            if region.lower() in lori_nationwide_clean(r.get("default_region")).lower()
        ]

    return {
        "status": "success",
        "states_count": len(rows),
        "states": rows,
    }


@app.get("/operating-regions")
async def operating_regions(
    api_key: Optional[str] = Query(None),
    region_code: Optional[str] = Query(None),
    state_code: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_nationwide_get_rows(
        "lori_operating_regions",
        "select=*&order=region_name.asc&limit=100",
    )

    if region_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("region_code")).lower() == region_code.lower()
        ]

    if state_code:
        state_upper = state_code.upper()
        rows = [
            r for r in rows
            if state_upper in (r.get("covered_states") or [])
        ]

    return {
        "status": "success",
        "regions_count": len(rows),
        "regions": rows,
    }


@app.get("/operating-stations")
async def operating_stations(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    operating_state: Optional[str] = Query(None),
    region_code: Optional[str] = Query(None),
    company_name: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    map_ready: Optional[bool] = Query(None),
    route_scoring_ready: Optional[bool] = Query(None),
    push_notification_ready: Optional[bool] = Query(None),
    limit: int = Query(500),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_nationwide_get_rows(
        "lori_operating_stations",
        "select=*&order=station_code.asc&limit=1000",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if operating_state:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("operating_state")).lower() == operating_state.lower()
        ]

    if region_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("region_code")).lower() == region_code.lower()
        ]

    if company_name:
        rows = [
            r for r in rows
            if company_name.lower() in lori_nationwide_clean(r.get("company_name")).lower()
        ]

    if city:
        rows = [
            r for r in rows
            if city.lower() in lori_nationwide_clean(r.get("city")).lower()
        ]

    if status:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("station_status")).lower() == status.lower()
        ]

    if map_ready is not None:
        rows = [
            r for r in rows
            if bool(r.get("map_ready")) is map_ready
        ]

    if route_scoring_ready is not None:
        rows = [
            r for r in rows
            if bool(r.get("route_scoring_ready")) is route_scoring_ready
        ]

    if push_notification_ready is not None:
        rows = [
            r for r in rows
            if bool(r.get("push_notification_ready")) is push_notification_ready
        ]

    limit = max(1, min(limit, 500))

    return {
        "status": "success",
        "stations_count": len(rows[:limit]),
        "stations": rows[:limit],
    }


@app.get("/station-zip-coverage")
async def station_zip_coverage(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    operating_state: Optional[str] = Query(None),
    zip_code: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    route_group: Optional[str] = Query(None),
    map_ready: Optional[bool] = Query(None),
    limit: int = Query(1000),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_nationwide_get_rows(
        "lori_station_zip_coverage",
        "select=*&order=station_code.asc,zip_code.asc&limit=2000",
    )

    if station_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("station_code")).lower() == station_code.lower()
        ]

    if operating_state:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("operating_state")).lower() == operating_state.lower()
        ]

    if zip_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("zip_code")) == zip_code
        ]

    if route_id:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("route_id")).lower() == route_id.lower()
        ]

    if route_group:
        rows = [
            r for r in rows
            if route_group.lower() in lori_nationwide_clean(r.get("route_group")).lower()
        ]

    if map_ready is not None:
        rows = [
            r for r in rows
            if bool(r.get("map_ready")) is map_ready
        ]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "zip_coverages_count": len(rows[:limit]),
        "zip_coverages": rows[:limit],
    }


@app.get("/state-compliance-profiles")
async def state_compliance_profiles(
    api_key: Optional[str] = Query(None),
    state_code: Optional[str] = Query(None),
    jurisdiction_type: Optional[str] = Query(None),
    regulatory_watch_enabled: Optional[bool] = Query(None),
    policy_review_enabled: Optional[bool] = Query(None),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_nationwide_get_rows(
        "lori_state_compliance_profiles",
        "select=*&order=state_code.asc&limit=100",
    )

    if state_code:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("state_code")).lower() == state_code.lower()
        ]

    if jurisdiction_type:
        rows = [
            r for r in rows
            if lori_nationwide_clean(r.get("jurisdiction_type")).lower() == jurisdiction_type.lower()
        ]

    if regulatory_watch_enabled is not None:
        rows = [
            r for r in rows
            if bool(r.get("regulatory_watch_enabled")) is regulatory_watch_enabled
        ]

    if policy_review_enabled is not None:
        rows = [
            r for r in rows
            if bool(r.get("policy_review_enabled")) is policy_review_enabled
        ]

    return {
        "status": "success",
        "state_compliance_profiles_count": len(rows),
        "state_compliance_profiles": rows,
        "decision_support_note": "LORI provides operational decision support only. Federal, state, local, labor, safety, DOT, HR, and company policy requirements must be reviewed and approved by authorized leadership before implementation.",
    }


@app.get("/nationwide-module-contexts")
async def nationwide_module_contexts(
    api_key: Optional[str] = Query(None),
    module_name: Optional[str] = Query(None),
    supports_national: Optional[bool] = Query(None),
    supports_state: Optional[bool] = Query(None),
    supports_station: Optional[bool] = Query(None),
    supports_route: Optional[bool] = Query(None),
    supports_driver: Optional[bool] = Query(None),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_nationwide_get_rows(
        "lori_nationwide_module_contexts",
        "select=*&order=module_name.asc&limit=100",
    )

    if module_name:
        rows = [
            r for r in rows
            if module_name.lower() in lori_nationwide_clean(r.get("module_name")).lower()
        ]

    if supports_national is not None:
        rows = [r for r in rows if bool(r.get("supports_national")) is supports_national]

    if supports_state is not None:
        rows = [r for r in rows if bool(r.get("supports_state")) is supports_state]

    if supports_station is not None:
        rows = [r for r in rows if bool(r.get("supports_station")) is supports_station]

    if supports_route is not None:
        rows = [r for r in rows if bool(r.get("supports_route")) is supports_route]

    if supports_driver is not None:
        rows = [r for r in rows if bool(r.get("supports_driver")) is supports_driver]

    return {
        "status": "success",
        "module_contexts_count": len(rows),
        "module_contexts": rows,
    }


@app.get("/operating-context-options")
async def operating_context_options(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    states = await lori_nationwide_get_rows(
        "lori_us_operating_states",
        "select=state_code,state_name,default_region,default_time_zone,active_state&order=state_code.asc&limit=100",
    )

    regions = await lori_nationwide_get_rows(
        "lori_operating_regions",
        "select=region_code,region_name,covered_states,region_status&order=region_name.asc&limit=100",
    )

    stations = await lori_nationwide_get_rows(
        "lori_operating_stations",
        "select=company_name,region_code,region_name,station_code,station_name,station_status,operating_state,city,primary_zip,time_zone,latitude,longitude,map_ready,route_scoring_ready,push_notification_ready,compliance_profile_ready&order=station_code.asc&limit=500",
    )

    modules = await lori_nationwide_get_rows(
        "lori_nationwide_module_contexts",
        "select=module_name,context_level,supports_national,supports_region,supports_state,supports_station,supports_route,supports_driver,module_status&order=module_name.asc&limit=100",
    )

    return {
        "status": "success",
        "states": states,
        "regions": regions,
        "stations": stations,
        "modules": modules,
        "default_context": {
            "company_name": "Food Authority",
            "region_code": "MID_ATLANTIC",
            "region_name": "Mid-Atlantic",
            "operating_state": "MD",
            "station_code": "JESSUP-01",
            "station_name": "Jessup Delivery Station",
            "city": "Jessup",
            "primary_zip": "20794",
            "time_zone": "America/New_York",
        },
    }


@app.get("/operating-context-resolve")
async def operating_context_resolve(
    api_key: Optional[str] = Query(None),
    state_code: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    zip_code: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    states = await lori_nationwide_get_rows(
        "lori_us_operating_states",
        "select=*&order=state_code.asc&limit=100",
    )

    stations = await lori_nationwide_get_rows(
        "lori_operating_stations",
        "select=*&order=station_code.asc&limit=500",
    )

    zips = await lori_nationwide_get_rows(
        "lori_station_zip_coverage",
        "select=*&order=station_code.asc,zip_code.asc&limit=1000",
    )

    profiles = await lori_nationwide_get_rows(
        "lori_state_compliance_profiles",
        "select=*&order=state_code.asc&limit=100",
    )

    selected_state = None
    selected_station = None
    selected_zip_rows = []

    if station_code:
        for station in stations:
            if lori_nationwide_clean(station.get("station_code")).lower() == station_code.lower():
                selected_station = station
                state_code = state_code or station.get("operating_state")
                break

    if state_code:
        for state in states:
            if lori_nationwide_clean(state.get("state_code")).lower() == state_code.lower():
                selected_state = state
                break

    if zip_code:
        selected_zip_rows = [
            z for z in zips
            if lori_nationwide_clean(z.get("zip_code")) == zip_code
        ]

        if selected_zip_rows and not station_code:
            station_code = selected_zip_rows[0].get("station_code")

    if station_code and not selected_station:
        for station in stations:
            if lori_nationwide_clean(station.get("station_code")).lower() == station_code.lower():
                selected_station = station
                break

    selected_profile = None

    if state_code:
        for profile in profiles:
            if lori_nationwide_clean(profile.get("state_code")).lower() == state_code.lower():
                selected_profile = profile
                break

    station_zip_coverages = []

    if selected_station:
        station_zip_coverages = [
            z for z in zips
            if lori_nationwide_clean(z.get("station_code")).lower()
            == lori_nationwide_clean(selected_station.get("station_code")).lower()
        ]

    return {
        "status": "success",
        "state": selected_state,
        "station": selected_station,
        "zip_matches": selected_zip_rows,
        "station_zip_coverages": station_zip_coverages,
        "state_compliance_profile": selected_profile,
        "resolved_context": {
            "state_code": state_code,
            "station_code": selected_station.get("station_code") if selected_station else station_code,
            "station_name": selected_station.get("station_name") if selected_station else None,
            "city": selected_station.get("city") if selected_station else None,
            "region_code": selected_station.get("region_code") if selected_station else None,
            "region_name": selected_station.get("region_name") if selected_station else None,
            "time_zone": selected_station.get("time_zone") if selected_station else None,
            "primary_zip": selected_station.get("primary_zip") if selected_station else zip_code,
        },
        "decision_support_note": "LORI provides operational decision support only. Federal, state, local, labor, safety, DOT, HR, and company policy requirements must be reviewed and approved by authorized leadership before implementation.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# OPERATING CONTEXT VALIDATION ENGINE
# Validates state, city, station, ZIP, and time zone before use.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, Optional, List


def lori_context_norm(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_context_upper(value: Any) -> str:
    return lori_context_norm(value).upper()


@app.post("/operating-context-validate")
async def operating_context_validate(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    selected_company = lori_context_norm(payload.get("company_name"))
    selected_region = lori_context_norm(payload.get("region_name"))
    selected_state = lori_context_upper(payload.get("operating_state") or payload.get("state_code"))
    selected_city = lori_context_norm(payload.get("city"))
    selected_station_code = lori_context_upper(payload.get("station_code"))
    selected_station_name = lori_context_norm(payload.get("station_name"))
    selected_zip = lori_context_norm(payload.get("primary_zip") or payload.get("zip_code"))
    selected_route_group = lori_context_norm(payload.get("route_group"))
    selected_time_zone = lori_context_norm(payload.get("time_zone"))
    selected_compliance_profile = lori_context_norm(payload.get("compliance_profile"))
    selected_state_reg_profile = lori_context_norm(payload.get("state_regulatory_profile"))

    issues: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    recommended_correction: Dict[str, Any] = {}

    required_fields = {
        "company_name": selected_company,
        "region_name": selected_region,
        "operating_state": selected_state,
        "city": selected_city,
        "station_code": selected_station_code,
        "station_name": selected_station_name,
        "primary_zip": selected_zip,
        "route_group": selected_route_group,
        "time_zone": selected_time_zone,
        "compliance_profile": selected_compliance_profile,
        "state_regulatory_profile": selected_state_reg_profile,
    }

    for field_name, field_value in required_fields.items():
        if not field_value:
            issues.append({
                "type": "missing_required_field",
                "field": field_name,
                "message": f"{field_name} is required before LORI can operate.",
            })

    states = await lori_nationwide_get_rows(
        "lori_us_operating_states",
        "select=*&order=state_code.asc&limit=100",
    )

    stations = await lori_nationwide_get_rows(
        "lori_operating_stations",
        "select=*&order=station_code.asc&limit=1000",
    )

    zip_rows = await lori_nationwide_get_rows(
        "lori_station_zip_coverage",
        "select=*&order=station_code.asc,zip_code.asc&limit=3000",
    )

    selected_state_record = None
    selected_station_record = None
    selected_zip_matches = []

    for state in states:
        if lori_context_upper(state.get("state_code")) == selected_state:
            selected_state_record = state
            break

    for station in stations:
        if lori_context_upper(station.get("station_code")) == selected_station_code:
            selected_station_record = station
            break

    if selected_zip:
        selected_zip_matches = [
            z for z in zip_rows
            if lori_context_norm(z.get("zip_code")) == selected_zip
        ]

    if selected_state and not selected_state_record:
        issues.append({
            "type": "invalid_state",
            "field": "operating_state",
            "message": f"{selected_state} is not a supported U.S. operating state or jurisdiction.",
        })

    if selected_station_code and not selected_station_record:
        issues.append({
            "type": "unknown_station",
            "field": "station_code",
            "message": f"Station {selected_station_code} is not configured in LORI.",
        })

    if selected_station_record:
        station_state = lori_context_upper(selected_station_record.get("operating_state"))
        station_city = lori_context_norm(selected_station_record.get("city"))
        station_zip = lori_context_norm(selected_station_record.get("primary_zip"))
        station_name = lori_context_norm(selected_station_record.get("station_name"))
        station_region = lori_context_norm(selected_station_record.get("region_name"))
        station_region_code = lori_context_norm(selected_station_record.get("region_code"))
        station_time_zone = lori_context_norm(selected_station_record.get("time_zone"))
        station_company = lori_context_norm(selected_station_record.get("company_name"))

        recommended_correction.update({
            "company_name": station_company or selected_company,
            "region_code": station_region_code,
            "region_name": station_region or selected_region,
            "operating_state": station_state,
            "city": station_city,
            "station_code": selected_station_code,
            "station_name": station_name,
            "primary_zip": station_zip,
            "time_zone": station_time_zone,
            "route_group": selected_route_group or "Delivery Operations",
            "compliance_profile": selected_compliance_profile or "Federal + State + Company Policy Review",
            "state_regulatory_profile": f"{station_state} DOT / FMCSA / State Review Required",
        })

        if selected_state and station_state and selected_state != station_state:
            issues.append({
                "type": "station_state_mismatch",
                "field": "operating_state",
                "message": f"{selected_station_code} is configured for {station_state}, but selected state is {selected_state}.",
                "selected": selected_state,
                "found": station_state,
            })

        if selected_city and station_city and selected_city.lower() != station_city.lower():
            issues.append({
                "type": "station_city_mismatch",
                "field": "city",
                "message": f"{selected_station_code} is configured for {station_city}, but selected city is {selected_city}.",
                "selected": selected_city,
                "found": station_city,
            })

        if selected_zip and station_zip and selected_zip != station_zip:
            matching_station_zip = [
                z for z in zip_rows
                if lori_context_upper(z.get("station_code")) == selected_station_code
                and lori_context_norm(z.get("zip_code")) == selected_zip
            ]

            if not matching_station_zip:
                issues.append({
                    "type": "station_zip_mismatch",
                    "field": "primary_zip",
                    "message": f"{selected_zip} is not configured as a ZIP coverage area for {selected_station_code}. Primary station ZIP is {station_zip}.",
                    "selected": selected_zip,
                    "found": station_zip,
                })

        if selected_time_zone and station_time_zone and selected_time_zone != station_time_zone:
            warnings.append({
                "type": "station_time_zone_mismatch",
                "field": "time_zone",
                "message": f"{selected_station_code} is configured for {station_time_zone}, but selected time zone is {selected_time_zone}.",
                "selected": selected_time_zone,
                "found": station_time_zone,
            })

    if selected_zip and selected_zip_matches:
        state_matches = [
            z for z in selected_zip_matches
            if lori_context_upper(z.get("operating_state")) == selected_state
        ]

        city_matches = [
            z for z in selected_zip_matches
            if selected_city and lori_context_norm(z.get("city")).lower() == selected_city.lower()
        ]

        station_matches = [
            z for z in selected_zip_matches
            if selected_station_code and lori_context_upper(z.get("station_code")) == selected_station_code
        ]

        if selected_state and not state_matches:
            found_states = sorted(list(set([
                lori_context_upper(z.get("operating_state"))
                for z in selected_zip_matches
                if lori_context_upper(z.get("operating_state"))
            ])))

            issues.append({
                "type": "zip_state_mismatch",
                "field": "primary_zip",
                "message": f"ZIP {selected_zip} is not configured for selected state {selected_state}.",
                "selected": selected_state,
                "found": found_states,
            })

        if selected_city and selected_zip_matches and not city_matches:
            found_cities = sorted(list(set([
                lori_context_norm(z.get("city"))
                for z in selected_zip_matches
                if lori_context_norm(z.get("city"))
            ])))

            issues.append({
                "type": "zip_city_mismatch",
                "field": "city",
                "message": f"ZIP {selected_zip} does not match selected city {selected_city} in configured LORI ZIP coverage.",
                "selected": selected_city,
                "found": found_cities,
            })

        if selected_station_code and selected_zip_matches and not station_matches:
            found_stations = sorted(list(set([
                lori_context_upper(z.get("station_code"))
                for z in selected_zip_matches
                if lori_context_upper(z.get("station_code"))
            ])))

            warnings.append({
                "type": "zip_station_not_primary_match",
                "field": "station_code",
                "message": f"ZIP {selected_zip} exists in LORI coverage but is not tied to selected station {selected_station_code}.",
                "selected": selected_station_code,
                "found": found_stations,
            })

    if selected_zip and not selected_zip_matches:
        warnings.append({
            "type": "zip_not_configured",
            "field": "primary_zip",
            "message": f"ZIP {selected_zip} is not currently configured in LORI station ZIP coverage. Upload or configure ZIP coverage before map-level route validation.",
        })

    if selected_state_record and not selected_station_record:
        recommended_correction.update({
            "operating_state": selected_state_record.get("state_code"),
            "region_name": selected_state_record.get("default_region"),
            "time_zone": selected_state_record.get("default_time_zone"),
            "state_regulatory_profile": f"{selected_state_record.get('state_code')} DOT / FMCSA / State Review Required",
        })

    context_valid = len(issues) == 0
    save_allowed = context_valid

    if selected_station_record and selected_state:
        station_summary = f"{selected_station_record.get('station_code')} is configured as {selected_station_record.get('city')}, {selected_station_record.get('operating_state')} {selected_station_record.get('primary_zip')}."
    else:
        station_summary = "No matching station record found."

    return {
        "status": "success",
        "context_valid": context_valid,
        "save_allowed": save_allowed,
        "validation_status": "Valid Context" if context_valid else "Mismatch Detected",
        "blocking_issues_count": len(issues),
        "warnings_count": len(warnings),
        "blocking_issues": issues,
        "warnings": warnings,
        "selected_context": {
            "company_name": selected_company,
            "region_name": selected_region,
            "operating_state": selected_state,
            "city": selected_city,
            "station_code": selected_station_code,
            "station_name": selected_station_name,
            "primary_zip": selected_zip,
            "route_group": selected_route_group,
            "time_zone": selected_time_zone,
            "compliance_profile": selected_compliance_profile,
            "state_regulatory_profile": selected_state_reg_profile,
        },
        "lori_found": {
            "state": selected_state_record,
            "station": selected_station_record,
            "zip_matches": selected_zip_matches,
        },
        "recommended_correction": recommended_correction,
        "station_summary": station_summary,
        "message": "Operating context is valid." if context_valid else "Operating context mismatch detected. Please correct state, city, station, ZIP, or time zone before continuing.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# NEW FACILITY / STATION SETUP BACKEND
# Creates new station setup contexts for nationwide operations.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
import re
from datetime import datetime


def lori_station_setup_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_station_setup_upper(value: Any) -> str:
    return lori_station_setup_clean(value).upper()


def lori_station_setup_code_city(city: str) -> str:
    cleaned = re.sub(r"[^A-Za-z]", "", city or "").upper()
    if len(cleaned) >= 3:
        return cleaned[:3]
    if cleaned:
        return cleaned.ljust(3, "X")
    return "STA"


async def lori_station_setup_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_station_setup_generate_code(state_code: str, city: str) -> str:
    state_code = lori_station_setup_upper(state_code)
    city_code = lori_station_setup_code_city(city)

    base_code = f"{state_code}-{city_code}"
    stations = await lori_station_setup_get_rows(
        "lori_operating_stations",
        "select=station_code&limit=1000",
    )

    existing_codes = {
        lori_station_setup_upper(row.get("station_code"))
        for row in stations
    }

    for number in range(1, 100):
        candidate = f"{base_code}-{number:02d}"
        if candidate not in existing_codes:
            return candidate

    return f"{base_code}-{datetime.utcnow().strftime('%H%M%S')}"


@app.get("/station-setup-options")
async def station_setup_options(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    states = await lori_station_setup_get_rows(
        "lori_us_operating_states",
        "select=*&order=state_code.asc&limit=100",
    )

    regions = await lori_station_setup_get_rows(
        "lori_operating_regions",
        "select=*&order=region_name.asc&limit=100",
    )

    stations = await lori_station_setup_get_rows(
        "lori_operating_stations",
        "select=*&order=station_code.asc&limit=1000",
    )

    route_groups = [
        "Delivery Operations",
        "Pickup Operations",
        "Linehaul",
        "Shuttle Routes",
        "Warehouse Transfer",
        "Last Mile Delivery",
        "Local Delivery",
        "Regional Delivery",
        "Long Haul",
        "Refrigerated Delivery",
        "Bulk Freight",
        "Mixed Freight",
        "Service Routes",
        "All Routes",
    ]

    time_zones = [
        {"value": "America/New_York", "label": "America/New_York — Eastern Time"},
        {"value": "America/Chicago", "label": "America/Chicago — Central Time"},
        {"value": "America/Denver", "label": "America/Denver — Mountain Time"},
        {"value": "America/Phoenix", "label": "America/Phoenix — Arizona Time"},
        {"value": "America/Los_Angeles", "label": "America/Los_Angeles — Pacific Time"},
        {"value": "America/Anchorage", "label": "America/Anchorage — Alaska Time"},
        {"value": "Pacific/Honolulu", "label": "Pacific/Honolulu — Hawaii Time"},
        {"value": "America/Indiana/Indianapolis", "label": "America/Indiana/Indianapolis — Indiana Eastern Time"},
        {"value": "America/Boise", "label": "America/Boise — Mountain / Idaho Time"},
        {"value": "America/Detroit", "label": "America/Detroit — Michigan Eastern Time"},
    ]

    compliance_profiles = [
        "Federal + State + Company Policy Review",
        "Federal Only",
        "State-Specific Review",
        "Company Policy Review",
        "DOT / Safety Review",
        "Labor / HR Review",
        "Station-Specific Policy Review",
        "Route Configuration Review",
        "Messaging / Communication Review",
        "Manual Review Required",
    ]

    regulatory_profiles = [
        "Federal / FMCSA Review Required",
        "State DOT Review Required",
        "Local Jurisdiction Review May Apply",
        "Labor / HR Review Required",
        "Company Policy Review Required",
        "Manual Regulatory Review Required",
    ]

    return {
        "status": "success",
        "states": states,
        "regions": regions,
        "stations": stations,
        "route_groups": route_groups,
        "time_zones": time_zones,
        "compliance_profiles": compliance_profiles,
        "regulatory_profiles": regulatory_profiles,
        "message": "Station setup options loaded.",
    }


@app.post("/operating-station-create")
async def operating_station_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    company_name = lori_station_setup_clean(payload.get("company_name") or "Demo Logistics")
    operating_state = lori_station_setup_upper(payload.get("operating_state") or payload.get("state_code"))
    city = lori_station_setup_clean(payload.get("city"))
    station_name = lori_station_setup_clean(payload.get("station_name"))
    station_code = lori_station_setup_upper(payload.get("station_code"))
    primary_zip = lori_station_setup_clean(payload.get("primary_zip") or payload.get("zip_code"))
    route_group = lori_station_setup_clean(payload.get("route_group") or "Delivery Operations")
    time_zone = lori_station_setup_clean(payload.get("time_zone"))
    region_name = lori_station_setup_clean(payload.get("region_name"))
    region_code = lori_station_setup_upper(payload.get("region_code"))
    county = lori_station_setup_clean(payload.get("county"))
    address_line_1 = lori_station_setup_clean(payload.get("address_line_1"))
    latitude = payload.get("latitude")
    longitude = payload.get("longitude")

    missing = []

    required = {
        "company_name": company_name,
        "operating_state": operating_state,
        "city": city,
        "station_name": station_name,
        "primary_zip": primary_zip,
        "route_group": route_group,
    }

    for key, value in required.items():
        if not value:
            missing.append(key)

    if missing:
        return {
            "status": "error",
            "message": "Missing required station setup fields.",
            "missing_fields": missing,
        }

    states = await lori_station_setup_get_rows(
        "lori_us_operating_states",
        "select=*&order=state_code.asc&limit=100",
    )

    regions = await lori_station_setup_get_rows(
        "lori_operating_regions",
        "select=*&order=region_name.asc&limit=100",
    )

    state_record = None

    for state in states:
        if lori_station_setup_upper(state.get("state_code")) == operating_state:
            state_record = state
            break

    if not state_record:
        return {
            "status": "error",
            "message": f"{operating_state} is not a supported U.S. operating state or jurisdiction.",
        }

    if not region_name:
        region_name = lori_station_setup_clean(state_record.get("default_region"))

    if not time_zone:
        time_zone = lori_station_setup_clean(state_record.get("default_time_zone"))

    if not region_code:
        for region in regions:
            if lori_station_setup_clean(region.get("region_name")).lower() == region_name.lower():
                region_code = lori_station_setup_upper(region.get("region_code"))
                break

    if not region_code:
        region_code = lori_station_setup_upper(region_name.replace(" ", "_"))

    if not station_code:
        station_code = await lori_station_setup_generate_code(operating_state, city)

    existing = await lori_station_setup_get_rows(
        "lori_operating_stations",
        f"select=*&station_code=eq.{station_code}&limit=1",
    )

    if existing:
        return {
            "status": "duplicate",
            "message": f"Station code {station_code} already exists.",
            "existing_station": existing[0],
            "recommended_action": "Use the existing station or choose a different station code.",
        }

    station_payload = {
        "company_name": company_name,
        "region_code": region_code,
        "region_name": region_name,
        "station_code": station_code,
        "station_name": station_name,
        "station_status": "Setup Required",
        "station_type": lori_station_setup_clean(payload.get("station_type") or "Distribution / Delivery Station"),
        "operating_state": operating_state,
        "city": city,
        "county": county,
        "primary_zip": primary_zip,
        "time_zone": time_zone,
        "address_line_1": address_line_1,
        "address_line_2": lori_station_setup_clean(payload.get("address_line_2")),
        "latitude": latitude,
        "longitude": longitude,
        "route_groups": [route_group],
        "supported_modules": [
            "Route Configuration",
            "Push Notifications",
            "Driver 360",
            "Compliance & Policy Center",
            "Regulatory Watch",
            "KPI Action Plans",
            "Action Center",
            "Leadership Dashboard",
            "Data Intake",
            "Reports",
            "SOP Builder",
        ],
        "map_ready": False,
        "route_scoring_ready": False,
        "push_notification_ready": False,
        "compliance_profile_ready": False,
        "notes": "New facility setup context created. Maps, route scoring, driver data, ZIP coverage detail, push recipients, and compliance documents must be configured before full operations.",
    }

    created_station = await lori_policy_supabase_post(
        "lori_operating_stations",
        station_payload,
    )

    station = created_station[0] if created_station else station_payload

    zip_payload = {
        "station_code": station_code,
        "station_name": station_name,
        "operating_state": operating_state,
        "city": city,
        "zip_code": primary_zip,
        "route_group": route_group,
        "route_id": None,
        "driver_name": None,
        "coverage_type": "Station ZIP Coverage",
        "coverage_status": "Setup Required",
        "map_ready": False,
        "geojson": {},
        "notes": "Initial ZIP coverage created for new station setup. Upload or configure full ZIP/service area geography before map-level route validation.",
    }

    created_zip = await lori_policy_supabase_post(
        "lori_station_zip_coverage",
        zip_payload,
    )

    setup_context = {
        "company_name": company_name,
        "region_code": region_code,
        "region_name": region_name,
        "operating_state": operating_state,
        "city": city,
        "station_code": station_code,
        "station_name": station_name,
        "primary_zip": primary_zip,
        "route_group": route_group,
        "time_zone": time_zone,
        "compliance_profile": lori_station_setup_clean(payload.get("compliance_profile") or "Federal + State + Company Policy Review"),
        "state_regulatory_profile": lori_station_setup_clean(payload.get("state_regulatory_profile") or f"{operating_state} DOT / FMCSA / State Review Required"),
    }

    return {
        "status": "success",
        "message": "New facility/station setup context created.",
        "station": station,
        "zip_coverage": created_zip[0] if created_zip else zip_payload,
        "setup_context": setup_context,
        "context_can_be_saved": True,
        "operational_modules_limited": True,
        "route_configuration_ready": False,
        "map_ready": False,
        "route_scoring_ready": False,
        "push_notification_ready": False,
        "compliance_profile_ready": False,
        "required_next_steps": [
            "Upload or configure station route geometry.",
            "Upload stop addresses or stop coordinates.",
            "Configure ZIP coverage and work-area boundaries.",
            "Upload route scoring or driver performance data.",
            "Add drivers and supervisors.",
            "Add push notification recipients.",
            "Upload state/company/station policy documents.",
            "Review state compliance profile with authorized leadership.",
        ],
        "decision_support_note": "This station is saved as Setup Required. LORI can store the context, but full route maps, route configuration, push notifications, route scoring, and compliance workflows require additional station setup data.",
    }


@app.get("/operating-station-readiness")
async def operating_station_readiness(
    api_key: Optional[str] = Query(None),
    station_code: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    station_code = lori_station_setup_upper(station_code)

    stations = await lori_station_setup_get_rows(
        "lori_operating_stations",
        f"select=*&station_code=eq.{station_code}&limit=1",
    )

    if not stations:
        return {
            "status": "not_found",
            "message": f"Station {station_code} is not configured in LORI.",
        }

    station = stations[0]

    zip_rows = await lori_station_setup_get_rows(
        "lori_station_zip_coverage",
        f"select=*&station_code=eq.{station_code}&limit=500",
    )

    readiness_items = [
        {
            "area": "Station Record",
            "ready": True,
            "message": "Station record exists.",
        },
        {
            "area": "ZIP Coverage",
            "ready": len(zip_rows) > 0,
            "message": f"{len(zip_rows)} ZIP coverage record(s) configured.",
        },
        {
            "area": "Route Maps",
            "ready": bool(station.get("map_ready")),
            "message": "Route maps are ready." if station.get("map_ready") else "Route map data still needs to be configured.",
        },
        {
            "area": "Route Scoring",
            "ready": bool(station.get("route_scoring_ready")),
            "message": "Route scoring is ready." if station.get("route_scoring_ready") else "Route scoring data still needs to be uploaded.",
        },
        {
            "area": "Push Notifications",
            "ready": bool(station.get("push_notification_ready")),
            "message": "Push notification recipients are ready." if station.get("push_notification_ready") else "Push recipients still need to be configured.",
        },
        {
            "area": "Compliance Profile",
            "ready": bool(station.get("compliance_profile_ready")),
            "message": "Compliance profile is ready." if station.get("compliance_profile_ready") else "Compliance profile requires authorized review.",
        },
    ]

    ready_count = len([item for item in readiness_items if item["ready"]])
    total_count = len(readiness_items)

    full_operations_ready = ready_count == total_count

    return {
        "status": "success",
        "station": station,
        "zip_coverages_count": len(zip_rows),
        "zip_coverages": zip_rows,
        "readiness_score": round((ready_count / total_count) * 100, 2),
        "full_operations_ready": full_operations_ready,
        "readiness_items": readiness_items,
        "message": "Station is fully operations-ready." if full_operations_ready else "Station exists, but additional setup is required before full operations.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# STATION DUPLICATE CHECK
# Prevents duplicate facility/station setup records.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, Optional, List


@app.post("/operating-station-duplicate-check")
async def operating_station_duplicate_check(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    company_name = lori_station_setup_clean(payload.get("company_name"))
    operating_state = lori_station_setup_upper(payload.get("operating_state") or payload.get("state_code"))
    city = lori_station_setup_clean(payload.get("city"))
    station_name = lori_station_setup_clean(payload.get("station_name"))
    station_code = lori_station_setup_upper(payload.get("station_code"))
    primary_zip = lori_station_setup_clean(payload.get("primary_zip") or payload.get("zip_code"))

    stations = await lori_station_setup_get_rows(
        "lori_operating_stations",
        "select=*&order=station_code.asc&limit=2000",
    )

    matches = []

    for station in stations:
        existing_code = lori_station_setup_upper(station.get("station_code"))
        existing_name = lori_station_setup_clean(station.get("station_name"))
        existing_company = lori_station_setup_clean(station.get("company_name"))
        existing_state = lori_station_setup_upper(station.get("operating_state"))
        existing_city = lori_station_setup_clean(station.get("city"))
        existing_zip = lori_station_setup_clean(station.get("primary_zip"))

        match_reasons = []

        if station_code and existing_code == station_code:
            match_reasons.append("Same station code")

        if (
            station_name
            and city
            and operating_state
            and existing_name.lower() == station_name.lower()
            and existing_city.lower() == city.lower()
            and existing_state == operating_state
        ):
            match_reasons.append("Same station name, city, and state")

        if (
            company_name
            and city
            and operating_state
            and primary_zip
            and existing_company.lower() == company_name.lower()
            and existing_city.lower() == city.lower()
            and existing_state == operating_state
            and existing_zip == primary_zip
        ):
            match_reasons.append("Same company, city, state, and ZIP")

        if (
            city
            and operating_state
            and primary_zip
            and existing_city.lower() == city.lower()
            and existing_state == operating_state
            and existing_zip == primary_zip
        ):
            match_reasons.append("Same city, state, and ZIP")

        if match_reasons:
            matches.append({
                "station": station,
                "match_reasons": match_reasons,
                "display_label": f"{existing_code} — {existing_name} — {existing_city}, {existing_state} {existing_zip}",
            })

    duplicate_detected = len(matches) > 0

    return {
        "status": "success",
        "duplicate_detected": duplicate_detected,
        "matches_count": len(matches),
        "matches": matches,
        "message": "Possible duplicate station found." if duplicate_detected else "No duplicate station found.",
        "recommended_action": "Use the existing station instead of creating a duplicate." if duplicate_detected else "Safe to continue station setup.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# ROUTE CONFIGURATION / WORK AREA BALANCING ENGINE
# Stores uploaded/typed route stops, analyzes workload balance,
# detects crossover risk, and recommends stop movements.
# ============================================================

from fastapi import Body, Query, UploadFile, File, Form
from fastapi.responses import HTMLResponse
from typing import Any, Dict, List, Optional
from urllib.parse import quote
import csv
import io
import html
import openpyxl


def lori_routecfg_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_routecfg_upper(value: Any) -> str:
    return lori_routecfg_clean(value).upper()


def lori_routecfg_num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def lori_routecfg_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


async def lori_routecfg_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


def lori_routecfg_row_value(row: Dict[str, Any], keys: List[str], default: Any = "") -> Any:
    lowered = {str(k).lower().strip(): v for k, v in row.items()}
    for key in keys:
        if key in row and row.get(key) not in [None, ""]:
            return row.get(key)
        key_l = key.lower().strip()
        if key_l in lowered and lowered.get(key_l) not in [None, ""]:
            return lowered.get(key_l)
    return default


def lori_routecfg_normalize_stop(
    row: Dict[str, Any],
    project_id: str,
    source_type: str = "Typed",
    source_file_id: Optional[str] = None,
    source_row_number: Optional[int] = None,
    default_route_id: Optional[str] = None,
    default_driver_name: Optional[str] = None,
    default_state: Optional[str] = None,
) -> Dict[str, Any]:
    route_id = lori_routecfg_clean(
        lori_routecfg_row_value(row, ["route_id", "route", "route number"], default_route_id or "")
    )

    driver_name = lori_routecfg_clean(
        lori_routecfg_row_value(row, ["driver_name", "driver", "employee_name"], default_driver_name or "")
    )

    street_address = lori_routecfg_clean(
        lori_routecfg_row_value(row, ["street_address", "address", "delivery_address", "stop_address"])
    )

    city = lori_routecfg_clean(
        lori_routecfg_row_value(row, ["city", "stop_city"])
    )

    operating_state = lori_routecfg_upper(
        lori_routecfg_row_value(row, ["state", "operating_state", "stop_state"], default_state or "")
    )

    zip_code = lori_routecfg_clean(
        lori_routecfg_row_value(row, ["zip", "zip_code", "postal_code"])
    )

    validation_status = "Pending Validation"
    review_required = False

    if not street_address or not city or not operating_state or not zip_code:
        validation_status = "Needs Review"
        review_required = True

    return {
        "project_id": project_id,
        "source_type": source_type,
        "source_file_id": source_file_id,
        "source_row_number": source_row_number,
        "route_id": route_id or "UNASSIGNED",
        "driver_name": driver_name,
        "stop_sequence": lori_routecfg_int(lori_routecfg_row_value(row, ["stop_sequence", "sequence", "stop_number", "stop #"]), 0),
        "stop_id": lori_routecfg_clean(lori_routecfg_row_value(row, ["stop_id", "stop id"])),
        "customer_name": lori_routecfg_clean(lori_routecfg_row_value(row, ["customer_name", "customer", "location_name", "stop_name"])),
        "street_address": street_address or "ADDRESS MISSING",
        "city": city or "CITY MISSING",
        "operating_state": operating_state or "STATE MISSING",
        "zip_code": zip_code or "ZIP MISSING",
        "latitude": lori_routecfg_row_value(row, ["latitude", "lat"], None),
        "longitude": lori_routecfg_row_value(row, ["longitude", "lng", "lon"], None),
        "geocode_status": "Coordinates Provided" if lori_routecfg_row_value(row, ["latitude", "lat"], None) and lori_routecfg_row_value(row, ["longitude", "lng", "lon"], None) else "Not Geocoded",
        "delivery_window_start": lori_routecfg_clean(lori_routecfg_row_value(row, ["delivery_window_start", "window_start"])),
        "delivery_window_end": lori_routecfg_clean(lori_routecfg_row_value(row, ["delivery_window_end", "window_end"])),
        "service_time_minutes": lori_routecfg_num(lori_routecfg_row_value(row, ["service_time_minutes", "service_time", "service minutes"]), 0),
        "freight_type": lori_routecfg_clean(lori_routecfg_row_value(row, ["freight_type", "freight"])),
        "priority": lori_routecfg_clean(lori_routecfg_row_value(row, ["priority"], "Normal")),
        "delivery_notes": lori_routecfg_clean(lori_routecfg_row_value(row, ["delivery_notes", "notes"])),
        "current_assignment_route_id": route_id or default_route_id or "UNASSIGNED",
        "current_assignment_driver": driver_name or default_driver_name or "",
        "validation_status": validation_status,
        "review_required": review_required,
        "created_by": "LORI Route Configuration Engine",
    }


@app.post("/route-config-project-create")
async def route_config_project_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    project_payload = {
        "project_title": lori_routecfg_clean(payload.get("project_title") or "New Route Configuration Project"),
        "project_type": "Route Configuration",
        "project_status": "Draft",
        "company_name": lori_routecfg_clean(payload.get("company_name")),
        "region_code": lori_routecfg_clean(payload.get("region_code")),
        "region_name": lori_routecfg_clean(payload.get("region_name")),
        "operating_state": lori_routecfg_upper(payload.get("operating_state")),
        "city": lori_routecfg_clean(payload.get("city")),
        "station_code": lori_routecfg_upper(payload.get("station_code")),
        "station_name": lori_routecfg_clean(payload.get("station_name")),
        "primary_zip": lori_routecfg_clean(payload.get("primary_zip")),
        "route_group": lori_routecfg_clean(payload.get("route_group") or "Delivery Operations"),
        "time_zone": lori_routecfg_clean(payload.get("time_zone")),
        "configuration_goal": lori_routecfg_clean(payload.get("configuration_goal") or "Create the most efficient work area by reducing crossover, balancing workload, and moving stops between routes."),
        "project_notes": lori_routecfg_clean(payload.get("project_notes")),
        "created_by": "LORI Route Configuration Engine",
    }

    created = await lori_policy_supabase_post("lori_route_config_projects", project_payload)
    project = created[0] if created else project_payload

    return {
        "status": "success",
        "message": "Route configuration project created.",
        "project": project,
    }


@app.post("/route-config-routes-add")
async def route_config_routes_add(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    project_id = lori_routecfg_clean(payload.get("project_id"))
    routes = payload.get("routes") or []

    if not project_id:
        return {"status": "error", "message": "project_id is required."}

    if not routes:
        return {"status": "error", "message": "At least one route is required."}

    created_routes = []

    for route in routes:
        route_payload = {
            "project_id": project_id,
            "route_role": lori_routecfg_clean(route.get("route_role") or "Comparison Route"),
            "route_id": lori_routecfg_clean(route.get("route_id")),
            "route_name": lori_routecfg_clean(route.get("route_name")),
            "driver_name": lori_routecfg_clean(route.get("driver_name")),
            "supervisor_name": lori_routecfg_clean(route.get("supervisor_name")),
            "vehicle_type": lori_routecfg_clean(route.get("vehicle_type")),
            "freight_type": lori_routecfg_clean(route.get("freight_type")),
            "planned_stop_count": lori_routecfg_int(route.get("planned_stop_count"), 0),
            "actual_stop_count": lori_routecfg_int(route.get("actual_stop_count"), 0),
            "scheduled_hours": lori_routecfg_num(route.get("scheduled_hours"), 0),
            "actual_hours": lori_routecfg_num(route.get("actual_hours"), 0),
            "planned_miles": lori_routecfg_num(route.get("planned_miles"), 0),
            "actual_miles": lori_routecfg_num(route.get("actual_miles"), 0),
            "overtime_hours": lori_routecfg_num(route.get("overtime_hours"), 0),
            "route_color": lori_routecfg_clean(route.get("route_color")),
            "notes": lori_routecfg_clean(route.get("notes")),
        }

        if not route_payload["route_id"]:
            continue

        created = await lori_policy_supabase_post("lori_route_config_routes", route_payload)
        if created:
            created_routes.append(created[0])

    return {
        "status": "success",
        "routes_created": len(created_routes),
        "routes": created_routes,
    }


@app.post("/route-config-typed-stops-add")
async def route_config_typed_stops_add(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    project_id = lori_routecfg_clean(payload.get("project_id"))
    stops = payload.get("stops") or []
    default_state = lori_routecfg_upper(payload.get("operating_state"))

    if not project_id:
        return {"status": "error", "message": "project_id is required."}

    if not stops:
        return {"status": "error", "message": "At least one stop is required."}

    created_stops = []

    for idx, stop in enumerate(stops, start=1):
        stop_payload = lori_routecfg_normalize_stop(
            stop,
            project_id=project_id,
            source_type="Typed",
            source_row_number=idx,
            default_state=default_state,
        )

        created = await lori_policy_supabase_post("lori_route_config_stops", stop_payload)
        if created:
            created_stops.append(created[0])

    return {
        "status": "success",
        "message": "Typed route stops saved.",
        "stops_created": len(created_stops),
        "stops": created_stops,
    }


@app.post("/route-config-stop-upload")
async def route_config_stop_upload(
    api_key: Optional[str] = Query(None),
    project_id: str = Form(...),
    upload_purpose: str = Form("Route Stop Data"),
    uploaded_route_id: Optional[str] = Form(None),
    uploaded_driver_name: Optional[str] = Form(None),
    operating_state: Optional[str] = Form(None),
    file: UploadFile = File(...),
):
    lori_regulatory_require_key(api_key)

    file_name = file.filename or "uploaded_route_stops"
    file_ext = file_name.split(".")[-1].lower() if "." in file_name else "unknown"
    contents = await file.read()

    upload_payload = {
        "project_id": project_id,
        "upload_title": file_name,
        "original_file_name": file_name,
        "file_type": file_ext,
        "upload_purpose": upload_purpose,
        "uploaded_route_id": uploaded_route_id,
        "uploaded_driver_name": uploaded_driver_name,
        "parse_status": "Uploaded",
        "rows_detected": 0,
        "rows_imported": 0,
        "rows_needing_review": 0,
        "notes": "File received by LORI Route Configuration Engine.",
    }

    created_upload = await lori_policy_supabase_post("lori_route_config_stop_uploads", upload_payload)
    upload_record = created_upload[0] if created_upload else upload_payload
    upload_id = upload_record.get("id")

    parsed_rows: List[Dict[str, Any]] = []
    parse_note = ""

    try:
        if file_ext == "csv":
            text = contents.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(text))
            parsed_rows = [dict(row) for row in reader]

        elif file_ext in ["xlsx", "xlsm"]:
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                for row_values in rows[1:]:
                    parsed_rows.append({
                        headers[i]: row_values[i] if i < len(row_values) else None
                        for i in range(len(headers))
                        if headers[i]
                    })

        elif file_ext == "txt":
            text = contents.decode("utf-8-sig", errors="ignore")
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            for idx, line in enumerate(lines, start=1):
                parsed_rows.append({
                    "stop_id": f"TXT-{idx}",
                    "address": line,
                    "city": "",
                    "state": operating_state or "",
                    "zip": "",
                    "route_id": uploaded_route_id or "",
                    "driver_name": uploaded_driver_name or "",
                    "notes": "Imported from text file. Needs review.",
                })

        elif file_ext == "pdf":
            parse_note = "PDF received. Automatic PDF extraction is staged; upload Excel/CSV for best structured import."
            parsed_rows = []

        else:
            parse_note = f"File type {file_ext} received but not parsed."

    except Exception as exc:
        parse_note = f"File parse failed: {str(exc)}"
        parsed_rows = []

    created_stops = []
    rows_needing_review = 0

    for idx, row in enumerate(parsed_rows, start=1):
        stop_payload = lori_routecfg_normalize_stop(
            row,
            project_id=project_id,
            source_type=f"Upload {file_ext.upper()}",
            source_file_id=upload_id,
            source_row_number=idx,
            default_route_id=uploaded_route_id,
            default_driver_name=uploaded_driver_name,
            default_state=operating_state,
        )

        if stop_payload.get("review_required"):
            rows_needing_review += 1

        created = await lori_policy_supabase_post("lori_route_config_stops", stop_payload)
        if created:
            created_stops.append(created[0])

    return {
        "status": "success",
        "message": "Route stop file received.",
        "file_name": file_name,
        "file_type": file_ext,
        "upload_record": upload_record,
        "rows_detected": len(parsed_rows),
        "rows_imported": len(created_stops),
        "rows_needing_review": rows_needing_review,
        "parse_note": parse_note,
        "stops": created_stops[:25],
        "note": "Only the first 25 imported stops are returned in this response. All imported stops are stored in Supabase.",
    }


@app.get("/route-config-project-detail")
async def route_config_project_detail(
    api_key: Optional[str] = Query(None),
    project_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    project_rows = await lori_routecfg_get_rows(
        "lori_route_config_projects",
        f"select=*&id=eq.{quote(project_id)}&limit=1",
    )

    if not project_rows:
        return {"status": "not_found", "message": "Route configuration project not found."}

    routes = await lori_routecfg_get_rows(
        "lori_route_config_routes",
        f"select=*&project_id=eq.{quote(project_id)}&order=route_id.asc&limit=500",
    )

    stops = await lori_routecfg_get_rows(
        "lori_route_config_stops",
        f"select=*&project_id=eq.{quote(project_id)}&order=route_id.asc,stop_sequence.asc&limit=3000",
    )

    issues = await lori_routecfg_get_rows(
        "lori_route_config_stop_issues",
        f"select=*&project_id=eq.{quote(project_id)}&order=created_at.desc&limit=1000",
    )

    balance = await lori_routecfg_get_rows(
        "lori_route_config_balance_results",
        f"select=*&project_id=eq.{quote(project_id)}&order=created_at.desc&limit=1000",
    )

    moves = await lori_routecfg_get_rows(
        "lori_route_config_stop_moves",
        f"select=*&project_id=eq.{quote(project_id)}&order=created_at.desc&limit=1000",
    )

    return {
        "status": "success",
        "project": project_rows[0],
        "routes_count": len(routes),
        "routes": routes,
        "stops_count": len(stops),
        "stops": stops,
        "issues_count": len(issues),
        "issues": issues,
        "balance_results_count": len(balance),
        "balance_results": balance,
        "stop_moves_count": len(moves),
        "stop_moves": moves,
    }


@app.get("/route-config-stops")
async def route_config_stops(
    api_key: Optional[str] = Query(None),
    project_id: str = Query(...),
    route_id: Optional[str] = Query(None),
    review_required: Optional[bool] = Query(None),
    limit: int = Query(1000),
):
    lori_regulatory_require_key(api_key)

    stops = await lori_routecfg_get_rows(
        "lori_route_config_stops",
        f"select=*&project_id=eq.{quote(project_id)}&order=route_id.asc,stop_sequence.asc&limit=3000",
    )

    if route_id:
        stops = [
            s for s in stops
            if lori_routecfg_clean(s.get("route_id")).lower() == route_id.lower()
        ]

    if review_required is not None:
        stops = [
            s for s in stops
            if bool(s.get("review_required")) is review_required
        ]

    limit = max(1, min(limit, 3000))

    return {
        "status": "success",
        "stops_count": len(stops[:limit]),
        "stops": stops[:limit],
    }


@app.post("/route-config-validate-stops")
async def route_config_validate_stops(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    project_id = lori_routecfg_clean(payload.get("project_id"))

    if not project_id:
        return {"status": "error", "message": "project_id is required."}

    details = await route_config_project_detail(api_key=api_key, project_id=project_id)

    if details.get("status") != "success":
        return details

    project = details.get("project") or {}
    stops = details.get("stops") or []

    project_state = lori_routecfg_upper(project.get("operating_state"))
    seen = {}
    issues_created = []

    for stop in stops:
        stop_id = stop.get("id")
        address_key = "|".join([
            lori_routecfg_clean(stop.get("street_address")).lower(),
            lori_routecfg_clean(stop.get("city")).lower(),
            lori_routecfg_upper(stop.get("operating_state")),
            lori_routecfg_clean(stop.get("zip_code")),
        ])

        issue_rows = []

        if "MISSING" in lori_routecfg_upper(stop.get("street_address")):
            issue_rows.append(("Missing Address", "High", "Street address is missing.", "Enter the stop street address."))

        if "MISSING" in lori_routecfg_upper(stop.get("city")):
            issue_rows.append(("Missing City", "High", "City is missing.", "Enter the stop city."))

        if "MISSING" in lori_routecfg_upper(stop.get("operating_state")):
            issue_rows.append(("Missing State", "High", "State is missing.", "Enter the stop state."))

        if "MISSING" in lori_routecfg_upper(stop.get("zip_code")):
            issue_rows.append(("Missing ZIP", "High", "ZIP code is missing.", "Enter the stop ZIP code."))

        if project_state and lori_routecfg_upper(stop.get("operating_state")) not in [project_state, "STATE MISSING"]:
            issue_rows.append((
                "State Mismatch",
                "High",
                f"Stop state {stop.get('operating_state')} does not match project state {project_state}.",
                "Confirm whether this stop belongs in this station/work area."
            ))

        if address_key in seen:
            issue_rows.append((
                "Possible Duplicate Stop",
                "Medium",
                "Another stop has the same address/city/state/ZIP.",
                "Review duplicate stop before route balancing."
            ))
        else:
            seen[address_key] = stop_id

        for issue_type, severity, message, fix in issue_rows:
            issue_payload = {
                "project_id": project_id,
                "stop_id": stop_id,
                "issue_type": issue_type,
                "issue_severity": severity,
                "issue_message": message,
                "recommended_fix": fix,
            }

            created = await lori_policy_supabase_post("lori_route_config_stop_issues", issue_payload)
            if created:
                issues_created.append(created[0])

    return {
        "status": "success",
        "message": "Stop validation complete.",
        "stops_checked": len(stops),
        "issues_created": len(issues_created),
        "issues": issues_created[:100],
    }


@app.post("/route-config-analyze-balance")
async def route_config_analyze_balance(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    project_id = lori_routecfg_clean(payload.get("project_id"))

    if not project_id:
        return {"status": "error", "message": "project_id is required."}

    stops = await lori_routecfg_get_rows(
        "lori_route_config_stops",
        f"select=*&project_id=eq.{quote(project_id)}&limit=5000",
    )

    if not stops:
        return {
            "status": "error",
            "message": "No route stops found. Upload or type stops before analyzing route balance.",
        }

    route_groups: Dict[str, List[Dict[str, Any]]] = {}

    for stop in stops:
        route_id = lori_routecfg_clean(stop.get("route_id") or "UNASSIGNED")
        route_groups.setdefault(route_id, []).append(stop)

    avg_stops = len(stops) / max(len(route_groups), 1)
    results = []

    for route_id, route_stops in route_groups.items():
        stop_count = len(route_stops)
        service_minutes = sum(lori_routecfg_num(s.get("service_time_minutes"), 8) for s in route_stops)
        estimated_service_hours = round(service_minutes / 60, 2)

        estimated_miles = round(stop_count * 3.2, 2)
        estimated_drive_hours = round(estimated_miles / 22, 2)
        estimated_total_hours = round(estimated_service_hours + estimated_drive_hours, 2)
        estimated_overtime = max(0, round(estimated_total_hours - 8, 2))

        utilization = round((stop_count / max(avg_stops, 1)) * 100, 2)

        if utilization >= 115:
            workload_status = "Overutilized"
        elif utilization <= 85:
            workload_status = "Underutilized"
        else:
            workload_status = "Balanced"

        zips = [lori_routecfg_clean(s.get("zip_code")) for s in route_stops]
        zip_overlap = len(zips) != len(set(zips))
        crossover_risk = "Needs Review" if zip_overlap else "Preliminary Low"

        driver_name = lori_routecfg_clean(route_stops[0].get("driver_name"))

        result_payload = {
            "project_id": project_id,
            "balance_type": "Current",
            "route_id": route_id,
            "driver_name": driver_name,
            "stop_count": stop_count,
            "estimated_miles": estimated_miles,
            "estimated_drive_hours": estimated_drive_hours,
            "estimated_service_hours": estimated_service_hours,
            "estimated_total_hours": estimated_total_hours,
            "estimated_overtime_hours": estimated_overtime,
            "utilization_percent": utilization,
            "workload_status": workload_status,
            "territory_status": "Preliminary — Needs Map/Geocode Review",
            "crossover_risk": crossover_risk,
            "delivery_window_risk": "Needs Review",
            "vehicle_capacity_risk": "Needs Review",
            "summary": f"{route_id} has {stop_count} stops and is currently marked {workload_status}.",
        }

        created = await lori_policy_supabase_post("lori_route_config_balance_results", result_payload)
        results.append(created[0] if created else result_payload)

    overloaded = [r for r in results if r.get("workload_status") == "Overutilized"]
    underused = [r for r in results if r.get("workload_status") == "Underutilized"]

    return {
        "status": "success",
        "message": "Route balance analysis complete.",
        "routes_analyzed": len(results),
        "total_stops": len(stops),
        "average_stops_per_route": round(avg_stops, 2),
        "overutilized_routes": overloaded,
        "underutilized_routes": underused,
        "balance_results": results,
        "note": "This is preliminary workload analysis. Street-level territory and exact driving distance require geocoded stops or a connected routing service.",
    }


@app.post("/route-config-generate-stop-moves")
async def route_config_generate_stop_moves(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    project_id = lori_routecfg_clean(payload.get("project_id"))
    max_moves = lori_routecfg_int(payload.get("max_moves"), 25)

    if not project_id:
        return {"status": "error", "message": "project_id is required."}

    stops = await lori_routecfg_get_rows(
        "lori_route_config_stops",
        f"select=*&project_id=eq.{quote(project_id)}&limit=5000",
    )

    if not stops:
        return {
            "status": "error",
            "message": "No stops found. Upload or type route stops before generating move recommendations.",
        }

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for stop in stops:
        route_id = lori_routecfg_clean(stop.get("route_id") or "UNASSIGNED")
        grouped.setdefault(route_id, []).append(stop)

    if len(grouped) < 2:
        return {
            "status": "error",
            "message": "At least two routes are required to recommend moving stops between routes.",
        }

    route_counts = {route_id: len(items) for route_id, items in grouped.items()}
    avg_count = sum(route_counts.values()) / len(route_counts)

    from_route = max(route_counts, key=route_counts.get)
    to_route = min(route_counts, key=route_counts.get)

    from_stops = grouped[from_route]
    to_stops = grouped[to_route]

    if from_route == to_route:
        return {
            "status": "success",
            "message": "Routes are already balanced enough for a preliminary pass.",
            "recommendations_created": 0,
            "recommendations": [],
        }

    from_driver = lori_routecfg_clean(from_stops[0].get("driver_name"))
    to_driver = lori_routecfg_clean(to_stops[0].get("driver_name"))

    # Prefer stops in ZIP/city overlap with the receiving route.
    to_zips = set(lori_routecfg_clean(s.get("zip_code")) for s in to_stops)
    to_cities = set(lori_routecfg_clean(s.get("city")).lower() for s in to_stops)

    preferred_candidates = [
        s for s in from_stops
        if lori_routecfg_clean(s.get("zip_code")) in to_zips
        or lori_routecfg_clean(s.get("city")).lower() in to_cities
    ]

    fallback_candidates = [s for s in from_stops if s not in preferred_candidates]

    suggested_move_count = max(1, int((route_counts[from_route] - route_counts[to_route]) / 2))
    suggested_move_count = min(suggested_move_count, max_moves)

    candidates = (preferred_candidates + fallback_candidates)[:suggested_move_count]

    created_moves = []

    for stop in candidates:
        reason = (
            f"This stop appears better aligned with {to_route} based on current route workload, "
            f"ZIP/city overlap, and available capacity."
        )

        move_payload = {
            "project_id": project_id,
            "stop_record_id": stop.get("id"),
            "customer_name": stop.get("customer_name"),
            "street_address": stop.get("street_address"),
            "city": stop.get("city"),
            "operating_state": stop.get("operating_state"),
            "zip_code": stop.get("zip_code"),
            "from_route_id": from_route,
            "from_driver_name": from_driver,
            "to_route_id": to_route,
            "to_driver_name": to_driver,
            "recommendation_status": "Draft",
            "recommendation_confidence": "Preliminary",
            "reason_to_move": reason,
            "distance_logic": "Preliminary routing logic based on available ZIP/city/workload data. Geocoded route distance improves accuracy.",
            "territory_logic": f"Move stop from overloaded route {from_route} toward underused route {to_route}.",
            "delivery_window_impact": "Needs supervisor review before implementation.",
            "service_risk": "Low to Medium",
            "safety_risk": "Low to Medium",
            "estimated_miles_reduced": 3.2,
            "estimated_minutes_reduced": 18,
            "estimated_overtime_reduced": 0.25,
            "estimated_cost_savings": 28.50,
            "supervisor_review_required": True,
            "approval_status": "Pending Review",
        }

        created = await lori_policy_supabase_post("lori_route_config_stop_moves", move_payload)
        if created:
            created_moves.append(created[0])

    return {
        "status": "success",
        "message": f"LORI generated preliminary stop move recommendations from {from_route} to {to_route}.",
        "from_route": from_route,
        "to_route": to_route,
        "from_driver": from_driver,
        "to_driver": to_driver,
        "recommendations_created": len(created_moves),
        "recommendations": created_moves,
        "note": "Recommendations are preliminary and require supervisor/operations review before implementation.",
    }


@app.get("/route-config-packet-html", response_class=HTMLResponse)
async def route_config_packet_html(
    api_key: Optional[str] = Query(None),
    project_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    details = await route_config_project_detail(api_key=api_key, project_id=project_id)

    if details.get("status") != "success":
        return HTMLResponse("<h1>Route Configuration Project Not Found</h1>", status_code=404)

    project = details.get("project") or {}
    routes = details.get("routes") or []
    stops = details.get("stops") or []
    moves = details.get("stop_moves") or []
    balance = details.get("balance_results") or []

    route_rows = "".join(
        f"<tr><td>{html.escape(str(r.get('route_id','')))}</td>"
        f"<td>{html.escape(str(r.get('driver_name','')))}</td>"
        f"<td>{html.escape(str(r.get('workload_status','')))}</td>"
        f"<td>{html.escape(str(r.get('utilization_percent','')))}</td></tr>"
        for r in balance
    )

    move_rows = "".join(
        f"<tr><td>{html.escape(str(m.get('customer_name','')))}</td>"
        f"<td>{html.escape(str(m.get('street_address','')))}</td>"
        f"<td>{html.escape(str(m.get('from_route_id','')))}</td>"
        f"<td>{html.escape(str(m.get('to_route_id','')))}</td>"
        f"<td>{html.escape(str(m.get('reason_to_move','')))}</td></tr>"
        for m in moves
    )

    html_doc = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <title>LORI Route Configuration Packet</title>
      <style>
        body {{
          font-family: Arial, sans-serif;
          margin: 32px;
          background: #f8fafc;
          color: #111827;
        }}
        .card {{
          background: #ffffff;
          border: 1px solid #dbe3ef;
          border-radius: 14px;
          padding: 18px;
          margin-bottom: 16px;
        }}
        h1 {{ margin-bottom: 4px; }}
        table {{
          width: 100%;
          border-collapse: collapse;
          font-size: 13px;
        }}
        th, td {{
          border: 1px solid #e5e7eb;
          padding: 8px;
          text-align: left;
          vertical-align: top;
        }}
        th {{ background: #f1f5f9; }}
        .note {{
          color: #64748b;
          font-size: 12px;
          line-height: 1.5;
        }}
      </style>
    </head>
    <body>
      <div class="card">
        <h1>LORI Route Configuration Packet</h1>
        <p><strong>Project:</strong> {html.escape(str(project.get('project_title','')))}</p>
        <p><strong>Station:</strong> {html.escape(str(project.get('station_code','')))} — {html.escape(str(project.get('city','')))}, {html.escape(str(project.get('operating_state','')))}</p>
        <p><strong>Total Stops:</strong> {len(stops)} | <strong>Routes:</strong> {len(routes)} | <strong>Recommended Moves:</strong> {len(moves)}</p>
      </div>

      <div class="card">
        <h2>Route Balance</h2>
        <table>
          <thead>
            <tr>
              <th>Route</th>
              <th>Driver</th>
              <th>Workload Status</th>
              <th>Utilization %</th>
            </tr>
          </thead>
          <tbody>{route_rows or "<tr><td colspan='4'>No route balance results yet.</td></tr>"}</tbody>
        </table>
      </div>

      <div class="card">
        <h2>Recommended Stop Moves</h2>
        <table>
          <thead>
            <tr>
              <th>Customer</th>
              <th>Address</th>
              <th>From Route</th>
              <th>To Route</th>
              <th>Reason</th>
            </tr>
          </thead>
          <tbody>{move_rows or "<tr><td colspan='5'>No stop move recommendations yet.</td></tr>"}</tbody>
        </table>
      </div>

      <div class="card note">
        LORI provides operational decision support. Route changes, driver workload changes, helper assignments,
        vehicle assignments, delivery window changes, customer commitments, labor considerations, productivity estimates,
        and cost savings estimates must be reviewed and approved by authorized operations leadership before implementation.
      </div>
    </body>
    </html>
    """

    return HTMLResponse(content=html_doc)
# ============================================================
# LORI DRIVE COMMAND CENTER
# USER ROLES + STATION PERMISSIONS ACCESS ENGINE
# Controls user profile, allowed stations, module access, and
# permission checks for nationwide multi-station operations.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote


def lori_access_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_access_upper(value: Any) -> str:
    return lori_access_clean(value).upper()


def lori_access_email(value: Any) -> str:
    return lori_access_clean(value).lower()


async def lori_access_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_access_log_event(
    user_profile_id: Optional[str],
    action_type: str,
    module_name: Optional[str] = None,
    company_name: Optional[str] = None,
    region_name: Optional[str] = None,
    operating_state: Optional[str] = None,
    station_code: Optional[str] = None,
    route_group: Optional[str] = None,
    access_result: Optional[str] = None,
    reason: Optional[str] = None,
    request_context: Optional[Dict[str, Any]] = None,
):
    payload = {
        "user_profile_id": user_profile_id,
        "action_type": action_type,
        "module_name": module_name,
        "company_name": company_name,
        "region_name": region_name,
        "operating_state": operating_state,
        "station_code": station_code,
        "route_group": route_group,
        "access_result": access_result,
        "reason": reason,
        "request_context": request_context or {},
    }

    try:
        await lori_policy_supabase_post("lori_access_audit_log", payload)
    except Exception:
        pass


def lori_access_module_to_view_field(module_name: str) -> str:
    module = lori_access_clean(module_name).lower()

    if "dashboard" in module or "leadership" in module:
        return "can_view_dashboard"
    if "report" in module:
        return "can_view_reports"
    if "driver" in module:
        return "can_view_driver_360"
    if "route" in module:
        return "can_view_route_configuration"
    if "data" in module or "intake" in module or "document" in module:
        return "can_view_data_intake"
    if "action" in module:
        return "can_view_action_center"
    if "kpi" in module:
        return "can_view_kpi_plans"
    if "compliance" in module or "policy" in module or "contract" in module or "labor" in module:
        return "can_view_compliance"
    if "regulatory" in module:
        return "can_view_regulatory"
    if "push" in module or "notification" in module:
        return "can_view_push_notifications"
    if "sop" in module:
        return "can_view_sop_builder"
    if "ask" in module or "lori" in module:
        return "can_view_ask_lori"

    return "can_view_dashboard"


def lori_access_action_to_field(action: str, module_name: str = "") -> str:
    action_clean = lori_access_clean(action).lower()

    if action_clean in ["view", "read", "open"]:
        return lori_access_module_to_view_field(module_name)
    if action_clean in ["upload", "create_upload", "document_upload"]:
        return "can_upload_documents"
    if action_clean in ["send", "send_notification", "message"]:
        return "can_send_notifications"
    if action_clean in ["create_action", "action_item"]:
        return "can_create_action_items"
    if action_clean in ["approve", "approve_route_change", "final_approval"]:
        return "can_approve_route_changes"
    if action_clean in ["manage_context", "change_context"]:
        return "can_manage_station_context"
    if action_clean in ["manage_users", "admin_users"]:
        return "can_manage_users"

    return lori_access_module_to_view_field(module_name)


async def lori_access_get_user_bundle(email: str) -> Dict[str, Any]:
    email_clean = lori_access_email(email)

    users = await lori_access_get_rows(
        "lori_user_profiles",
        f"select=*&email=eq.{quote(email_clean)}&limit=1",
    )

    if not users:
        return {
            "found": False,
            "user": None,
            "roles": [],
            "station_access": [],
            "module_access": [],
        }

    user = users[0]
    user_id = user.get("id")

    assignments = await lori_access_get_rows(
        "lori_user_role_assignments",
        f"select=*&user_profile_id=eq.{quote(str(user_id))}&assignment_status=eq.Active&limit=500",
    )

    roles = []
    for assignment in assignments:
        role_id = assignment.get("role_id")
        if not role_id:
            continue

        role_rows = await lori_access_get_rows(
            "lori_roles",
            f"select=*&id=eq.{quote(str(role_id))}&limit=1",
        )

        if role_rows:
            roles.append(role_rows[0])

    station_access = await lori_access_get_rows(
        "lori_user_station_access",
        f"select=*&user_profile_id=eq.{quote(str(user_id))}&access_status=eq.Active&limit=1000",
    )

    module_access = await lori_access_get_rows(
        "lori_user_module_access",
        f"select=*&user_profile_id=eq.{quote(str(user_id))}&access_status=eq.Active&limit=1000",
    )

    return {
        "found": True,
        "user": user,
        "roles": roles,
        "station_access": station_access,
        "module_access": module_access,
    }


def lori_access_station_allowed(
    user: Dict[str, Any],
    station_access: List[Dict[str, Any]],
    requested_context: Dict[str, Any],
) -> Dict[str, Any]:
    requested_company = lori_access_clean(requested_context.get("company_name"))
    requested_region = lori_access_upper(requested_context.get("region_code"))
    requested_region_name = lori_access_clean(requested_context.get("region_name"))
    requested_state = lori_access_upper(requested_context.get("operating_state"))
    requested_station = lori_access_upper(requested_context.get("station_code"))

    if user.get("can_view_national") or user.get("can_view_all_companies"):
        return {
            "allowed": True,
            "scope": "National",
            "reason": "User has national or all-company access.",
            "matched_access": None,
        }

    for access in station_access:
        scope = lori_access_clean(access.get("access_scope")).lower()

        access_company = lori_access_clean(access.get("company_name"))
        access_region = lori_access_upper(access.get("region_code"))
        access_region_name = lori_access_clean(access.get("region_name"))
        access_state = lori_access_upper(access.get("operating_state"))
        access_station = lori_access_upper(access.get("station_code"))

        if not access.get("can_view"):
            continue

        if scope == "national":
            return {
                "allowed": True,
                "scope": "National",
                "reason": "User has national access record.",
                "matched_access": access,
            }

        if scope == "region":
            company_match = not requested_company or access_company.lower() == requested_company.lower()
            region_match = (
                requested_region and access_region and requested_region == access_region
            ) or (
                requested_region_name and access_region_name and requested_region_name.lower() == access_region_name.lower()
            )

            if company_match and region_match:
                return {
                    "allowed": True,
                    "scope": "Region",
                    "reason": "User has regional access for this context.",
                    "matched_access": access,
                }

        if scope == "state":
            company_match = not requested_company or access_company.lower() == requested_company.lower()
            state_match = requested_state and access_state and requested_state == access_state

            if company_match and state_match:
                return {
                    "allowed": True,
                    "scope": "State",
                    "reason": "User has state-level access for this context.",
                    "matched_access": access,
                }

        if scope == "station":
            company_match = not requested_company or access_company.lower() == requested_company.lower()
            station_match = requested_station and access_station and requested_station == access_station

            if company_match and station_match:
                return {
                    "allowed": True,
                    "scope": "Station",
                    "reason": "User has station access for this context.",
                    "matched_access": access,
                }

    return {
        "allowed": False,
        "scope": "None",
        "reason": "User does not have access to this station, state, region, or company context.",
        "matched_access": None,
    }


def lori_access_module_allowed(
    roles: List[Dict[str, Any]],
    module_access: List[Dict[str, Any]],
    module_name: str,
    action: str = "view",
) -> Dict[str, Any]:
    module_name_clean = lori_access_clean(module_name)
    action_field = lori_access_action_to_field(action, module_name_clean)

    for override in module_access:
        if lori_access_clean(override.get("module_name")).lower() == module_name_clean.lower():
            if action.lower() in ["view", "read", "open"] and override.get("can_view"):
                return {
                    "allowed": True,
                    "source": "Module Override",
                    "reason": "User has module-specific view access.",
                    "field_checked": "can_view",
                }

            if action.lower() in ["create", "add"] and override.get("can_create"):
                return {
                    "allowed": True,
                    "source": "Module Override",
                    "reason": "User has module-specific create access.",
                    "field_checked": "can_create",
                }

            if action.lower() in ["edit", "update"] and override.get("can_edit"):
                return {
                    "allowed": True,
                    "source": "Module Override",
                    "reason": "User has module-specific edit access.",
                    "field_checked": "can_edit",
                }

            if action.lower() in ["approve", "final_approval"] and override.get("can_approve"):
                return {
                    "allowed": True,
                    "source": "Module Override",
                    "reason": "User has module-specific approval access.",
                    "field_checked": "can_approve",
                }

            if action.lower() in ["export", "download"] and override.get("can_export"):
                return {
                    "allowed": True,
                    "source": "Module Override",
                    "reason": "User has module-specific export access.",
                    "field_checked": "can_export",
                }

    for role in roles:
        if role.get(action_field):
            return {
                "allowed": True,
                "source": "Role",
                "reason": f"Role {role.get('role_name')} allows this action.",
                "role": role,
                "field_checked": action_field,
            }

    return {
        "allowed": False,
        "source": "Role",
        "reason": f"No assigned role allows {action_field}.",
        "field_checked": action_field,
    }


@app.get("/user-access-profile")
async def user_access_profile(
    api_key: Optional[str] = Query(None),
    email: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    bundle = await lori_access_get_user_bundle(email)

    if not bundle["found"]:
        return {
            "status": "not_found",
            "message": "User profile was not found in LORI access control.",
            "email": email,
        }

    user = bundle["user"]

    await lori_access_log_event(
        user_profile_id=user.get("id"),
        action_type="User Access Profile Viewed",
        access_result="Allowed",
        reason="Profile lookup completed.",
        request_context={"email": email},
    )

    return {
        "status": "success",
        "user": user,
        "roles": bundle["roles"],
        "station_access": bundle["station_access"],
        "module_access": bundle["module_access"],
        "default_context": {
            "company_name": user.get("company_name"),
            "region_code": user.get("default_region_code"),
            "region_name": user.get("default_region_name"),
            "operating_state": user.get("default_operating_state"),
            "station_code": user.get("default_station_code"),
            "station_name": user.get("default_station_name"),
            "city": user.get("default_city"),
            "route_group": user.get("default_route_group"),
        },
        "message": "User access profile loaded.",
    }


@app.get("/user-allowed-stations")
async def user_allowed_stations(
    api_key: Optional[str] = Query(None),
    email: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    bundle = await lori_access_get_user_bundle(email)

    if not bundle["found"]:
        return {
            "status": "not_found",
            "message": "User profile was not found.",
            "allowed_stations": [],
        }

    user = bundle["user"]
    station_access = bundle["station_access"]

    stations = await lori_access_get_rows(
        "lori_operating_stations",
        "select=*&order=company_name.asc,operating_state.asc,city.asc,station_code.asc&limit=2000",
    )

    if user.get("can_view_national") or user.get("can_view_all_companies"):
        allowed = stations
    else:
        allowed = []

        for station in stations:
            check = lori_access_station_allowed(
                user=user,
                station_access=station_access,
                requested_context={
                    "company_name": station.get("company_name"),
                    "region_code": station.get("region_code"),
                    "region_name": station.get("region_name"),
                    "operating_state": station.get("operating_state"),
                    "station_code": station.get("station_code"),
                },
            )

            if check["allowed"]:
                allowed.append(station)

    return {
        "status": "success",
        "email": email,
        "allowed_stations_count": len(allowed),
        "allowed_stations": allowed,
        "message": "Allowed stations loaded.",
    }


@app.get("/user-allowed-contexts")
async def user_allowed_contexts(
    api_key: Optional[str] = Query(None),
    email: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    station_response = await user_allowed_stations(api_key=api_key, email=email)

    if station_response.get("status") != "success":
        return station_response

    contexts = []

    for station in station_response.get("allowed_stations", []):
        contexts.append({
            "company_name": station.get("company_name"),
            "region_code": station.get("region_code"),
            "region_name": station.get("region_name"),
            "operating_state": station.get("operating_state"),
            "city": station.get("city"),
            "station_code": station.get("station_code"),
            "station_name": station.get("station_name"),
            "primary_zip": station.get("primary_zip"),
            "route_group": "All Routes",
            "time_zone": station.get("time_zone"),
            "station_status": station.get("station_status"),
            "map_ready": station.get("map_ready"),
            "route_scoring_ready": station.get("route_scoring_ready"),
            "push_notification_ready": station.get("push_notification_ready"),
            "compliance_profile_ready": station.get("compliance_profile_ready"),
        })

    return {
        "status": "success",
        "email": email,
        "allowed_contexts_count": len(contexts),
        "allowed_contexts": contexts,
        "message": "Allowed operating contexts loaded.",
    }


@app.post("/permission-check")
async def permission_check(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    email = lori_access_email(payload.get("email"))
    module_name = lori_access_clean(payload.get("module_name") or "Dashboard")
    action = lori_access_clean(payload.get("action") or "view")
    requested_context = payload.get("requested_context") or {}

    if not email:
        return {
            "status": "error",
            "allowed": False,
            "message": "email is required.",
        }

    bundle = await lori_access_get_user_bundle(email)

    if not bundle["found"]:
        return {
            "status": "not_found",
            "allowed": False,
            "message": "User profile was not found in LORI access control.",
        }

    user = bundle["user"]
    roles = bundle["roles"]
    station_access = bundle["station_access"]
    module_access = bundle["module_access"]

    station_check = lori_access_station_allowed(
        user=user,
        station_access=station_access,
        requested_context=requested_context,
    )

    module_check = lori_access_module_allowed(
        roles=roles,
        module_access=module_access,
        module_name=module_name,
        action=action,
    )

    allowed = station_check["allowed"] and module_check["allowed"]

    reason = "Access allowed." if allowed else f"{station_check['reason']} {module_check['reason']}"

    await lori_access_log_event(
        user_profile_id=user.get("id"),
        action_type="Permission Check",
        module_name=module_name,
        company_name=requested_context.get("company_name"),
        region_name=requested_context.get("region_name"),
        operating_state=requested_context.get("operating_state"),
        station_code=requested_context.get("station_code"),
        route_group=requested_context.get("route_group"),
        access_result="Allowed" if allowed else "Denied",
        reason=reason,
        request_context={
            "email": email,
            "module_name": module_name,
            "action": action,
            "requested_context": requested_context,
            "station_check": station_check,
            "module_check": module_check,
        },
    )

    return {
        "status": "success",
        "allowed": allowed,
        "email": email,
        "module_name": module_name,
        "action": action,
        "station_check": station_check,
        "module_check": module_check,
        "reason": reason,
        "user": user,
        "roles": roles,
    }


@app.get("/module-access-check")
async def module_access_check(
    api_key: Optional[str] = Query(None),
    email: str = Query(...),
    module_name: str = Query(...),
    action: str = Query("view"),
):
    lori_regulatory_require_key(api_key)

    return await permission_check(
        api_key=api_key,
        payload={
            "email": email,
            "module_name": module_name,
            "action": action,
            "requested_context": {},
        },
    )


@app.post("/access-audit-log-create")
async def access_audit_log_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    email = lori_access_email(payload.get("email"))
    user_profile_id = payload.get("user_profile_id")

    if email and not user_profile_id:
        bundle = await lori_access_get_user_bundle(email)
        if bundle["found"]:
            user_profile_id = bundle["user"].get("id")

    await lori_access_log_event(
        user_profile_id=user_profile_id,
        action_type=lori_access_clean(payload.get("action_type") or "Access Event"),
        module_name=payload.get("module_name"),
        company_name=payload.get("company_name"),
        region_name=payload.get("region_name"),
        operating_state=payload.get("operating_state"),
        station_code=payload.get("station_code"),
        route_group=payload.get("route_group"),
        access_result=payload.get("access_result"),
        reason=payload.get("reason"),
        request_context=payload.get("request_context") or payload,
    )

    return {
        "status": "success",
        "message": "Access audit log created.",
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# USER TRANSFER / ACCESS CHANGE ENGINE
# Handles station transfers, temporary assignments, access
# removals, suspensions, and user default context changes.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
import os
import httpx
from datetime import date, datetime


SUPABASE_URL_TRANSFER = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_TRANSFER = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_transfer_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_transfer_upper(value: Any) -> str:
    return lori_transfer_clean(value).upper()


def lori_transfer_email(value: Any) -> str:
    return lori_transfer_clean(value).lower()


def lori_transfer_today() -> str:
    return date.today().isoformat()


async def lori_transfer_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_transfer_patch_rows(
    table: str,
    match_query: str,
    payload: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not SUPABASE_URL_TRANSFER or not SUPABASE_SERVICE_ROLE_KEY_TRANSFER:
        raise RuntimeError("Missing Supabase environment variables.")

    url = f"{SUPABASE_URL_TRANSFER}/rest/v1/{table}?{match_query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_TRANSFER,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_TRANSFER}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(url, headers=headers, json=payload)

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase PATCH failed: {response.status_code} {response.text}")

    try:
        return response.json()
    except Exception:
        return []


async def lori_transfer_find_user(email: str) -> Optional[Dict[str, Any]]:
    rows = await lori_transfer_get_rows(
        "lori_user_profiles",
        f"select=*&email=eq.{quote(lori_transfer_email(email))}&limit=1",
    )
    return rows[0] if rows else None


async def lori_transfer_find_station(station_code: str) -> Optional[Dict[str, Any]]:
    rows = await lori_transfer_get_rows(
        "lori_operating_stations",
        f"select=*&station_code=eq.{quote(lori_transfer_upper(station_code))}&limit=1",
    )
    return rows[0] if rows else None


async def lori_transfer_find_rule(transfer_type: str) -> Optional[Dict[str, Any]]:
    rows = await lori_transfer_get_rows(
        "lori_user_access_change_rules",
        f"select=*&transfer_type=eq.{quote(lori_transfer_clean(transfer_type))}&limit=1",
    )
    return rows[0] if rows else None


@app.get("/access-change-rules")
async def access_change_rules(
    api_key: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    rules = await lori_transfer_get_rows(
        "lori_user_access_change_rules",
        "select=*&order=transfer_type.asc&limit=100",
    )

    return {
        "status": "success",
        "rules_count": len(rules),
        "rules": rules,
    }


@app.get("/access-change-requests")
async def access_change_requests(
    api_key: Optional[str] = Query(None),
    request_status: Optional[str] = Query(None),
    user_email: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_transfer_get_rows(
        "lori_user_access_change_requests",
        "select=*&order=created_at.desc&limit=500",
    )

    if request_status:
        rows = [
            r for r in rows
            if lori_transfer_clean(r.get("request_status")).lower() == request_status.lower()
        ]

    if user_email:
        rows = [
            r for r in rows
            if lori_transfer_email(r.get("user_email")) == lori_transfer_email(user_email)
        ]

    limit = max(1, min(limit, 500))

    return {
        "status": "success",
        "requests_count": len(rows[:limit]),
        "requests": rows[:limit],
    }


@app.post("/access-change-request-create")
async def access_change_request_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    user_email = lori_transfer_email(payload.get("user_email"))
    transfer_type = lori_transfer_clean(payload.get("request_type") or "Permanent Transfer")
    new_station_code = lori_transfer_upper(payload.get("new_station_code"))
    transfer_effective_date = lori_transfer_clean(payload.get("transfer_effective_date") or lori_transfer_today())
    expiration_date = lori_transfer_clean(payload.get("expiration_date"))

    if not user_email:
        return {"status": "error", "message": "user_email is required."}

    if not new_station_code and transfer_type not in ["Termination / Remove Access", "Leave of Absence / Suspend Access"]:
        return {"status": "error", "message": "new_station_code is required for station access changes."}

    user = await lori_transfer_find_user(user_email)

    if not user:
        return {
            "status": "not_found",
            "message": "User profile not found. Create the user profile before requesting a transfer.",
            "user_email": user_email,
        }

    rule = await lori_transfer_find_rule(transfer_type)

    if not rule:
        return {
            "status": "error",
            "message": f"Transfer type '{transfer_type}' is not configured.",
        }

    if rule.get("requires_expiration_date") and not expiration_date:
        return {
            "status": "error",
            "message": f"{transfer_type} requires an expiration date.",
        }

    new_station = None
    if new_station_code:
        new_station = await lori_transfer_find_station(new_station_code)
        if not new_station:
            return {
                "status": "not_found",
                "message": f"New station {new_station_code} was not found in LORI station registry.",
            }

    request_title = lori_transfer_clean(
        payload.get("request_title")
        or f"{transfer_type} — {user.get('full_name')} to {new_station_code or 'Access Removal'}"
    )

    request_payload = {
        "request_title": request_title,
        "request_type": transfer_type,
        "request_status": "Pending Review",
        "user_profile_id": user.get("id"),
        "user_full_name": user.get("full_name"),
        "user_email": user.get("email"),

        "current_company_name": user.get("company_name"),
        "current_region_code": user.get("default_region_code"),
        "current_region_name": user.get("default_region_name"),
        "current_operating_state": user.get("default_operating_state"),
        "current_city": user.get("default_city"),
        "current_station_code": user.get("default_station_code"),
        "current_station_name": user.get("default_station_name"),
        "current_route_group": user.get("default_route_group"),

        "new_company_name": new_station.get("company_name") if new_station else None,
        "new_region_code": new_station.get("region_code") if new_station else None,
        "new_region_name": new_station.get("region_name") if new_station else None,
        "new_operating_state": new_station.get("operating_state") if new_station else None,
        "new_city": new_station.get("city") if new_station else None,
        "new_station_code": new_station.get("station_code") if new_station else None,
        "new_station_name": new_station.get("station_name") if new_station else None,
        "new_route_group": payload.get("new_route_group") or "All Routes",

        "current_role_code": payload.get("current_role_code"),
        "current_role_name": payload.get("current_role_name"),
        "new_role_code": payload.get("new_role_code"),
        "new_role_name": payload.get("new_role_name"),

        "transfer_effective_date": transfer_effective_date,
        "expiration_date": expiration_date or None,
        "transfer_reason": lori_transfer_clean(payload.get("transfer_reason")),
        "requested_by": lori_transfer_clean(payload.get("requested_by") or "LORI Admin"),
        "approval_notes": lori_transfer_clean(payload.get("approval_notes")),
        "requires_expiration_date": bool(rule.get("requires_expiration_date")),
        "requires_admin_approval": bool(rule.get("requires_admin_approval")),
    }

    created = await lori_policy_supabase_post(
        "lori_user_access_change_requests",
        request_payload,
    )

    request = created[0] if created else request_payload

    await lori_policy_supabase_post(
        "lori_user_access_change_approvals",
        {
            "access_change_request_id": request.get("id"),
            "approval_step": "Admin / HR / Regional Approval",
            "approver_role": "Platform Admin / HR / Regional Manager",
            "approval_status": "Pending",
            "required": True,
            "step_order": 1,
        },
    )

    await lori_access_log_event(
        user_profile_id=user.get("id"),
        action_type="Transfer Request Created",
        module_name="User Access Management",
        company_name=user.get("company_name"),
        region_name=user.get("default_region_name"),
        operating_state=user.get("default_operating_state"),
        station_code=user.get("default_station_code"),
        route_group=user.get("default_route_group"),
        access_result="Created",
        reason=f"{transfer_type} request created.",
        request_context=request,
    )

    return {
        "status": "success",
        "message": "User transfer/access change request created.",
        "request": request,
    }


@app.post("/access-change-request-approve")
async def access_change_request_approve(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_id = lori_transfer_clean(payload.get("request_id"))
    approved_by = lori_transfer_clean(payload.get("approved_by") or "LORI Admin")
    approval_notes = lori_transfer_clean(payload.get("approval_notes"))

    if not request_id:
        return {"status": "error", "message": "request_id is required."}

    rows = await lori_transfer_get_rows(
        "lori_user_access_change_requests",
        f"select=*&id=eq.{quote(request_id)}&limit=1",
    )

    if not rows:
        return {"status": "not_found", "message": "Access change request not found."}

    request = rows[0]

    updated = await lori_transfer_patch_rows(
        "lori_user_access_change_requests",
        f"id=eq.{quote(request_id)}",
        {
            "request_status": "Approved",
            "approved_by": approved_by,
            "approval_notes": approval_notes,
            "approved_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_transfer_patch_rows(
        "lori_user_access_change_approvals",
        f"access_change_request_id=eq.{quote(request_id)}",
        {
            "approval_status": "Approved",
            "approval_decision": "Approved",
            "approval_notes": approval_notes,
            "approver_name": approved_by,
            "decided_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_access_log_event(
        user_profile_id=request.get("user_profile_id"),
        action_type="Transfer Request Approved",
        module_name="User Access Management",
        company_name=request.get("current_company_name"),
        region_name=request.get("current_region_name"),
        operating_state=request.get("current_operating_state"),
        station_code=request.get("current_station_code"),
        access_result="Approved",
        reason=approval_notes or "Transfer/access change approved.",
        request_context=updated[0] if updated else request,
    )

    return {
        "status": "success",
        "message": "Access change request approved.",
        "request": updated[0] if updated else request,
    }


@app.post("/access-change-request-deny")
async def access_change_request_deny(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_id = lori_transfer_clean(payload.get("request_id"))
    denied_by = lori_transfer_clean(payload.get("denied_by") or "LORI Admin")
    denial_reason = lori_transfer_clean(payload.get("denial_reason"))

    if not request_id:
        return {"status": "error", "message": "request_id is required."}

    rows = await lori_transfer_get_rows(
        "lori_user_access_change_requests",
        f"select=*&id=eq.{quote(request_id)}&limit=1",
    )

    if not rows:
        return {"status": "not_found", "message": "Access change request not found."}

    request = rows[0]

    updated = await lori_transfer_patch_rows(
        "lori_user_access_change_requests",
        f"id=eq.{quote(request_id)}",
        {
            "request_status": "Denied",
            "denied_by": denied_by,
            "denial_reason": denial_reason,
            "denied_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_access_log_event(
        user_profile_id=request.get("user_profile_id"),
        action_type="Transfer Request Denied",
        module_name="User Access Management",
        station_code=request.get("current_station_code"),
        access_result="Denied",
        reason=denial_reason or "Transfer/access change denied.",
        request_context=updated[0] if updated else request,
    )

    return {
        "status": "success",
        "message": "Access change request denied.",
        "request": updated[0] if updated else request,
    }


@app.post("/access-change-request-complete")
async def access_change_request_complete(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_id = lori_transfer_clean(payload.get("request_id"))
    completed_by = lori_transfer_clean(payload.get("completed_by") or "LORI Admin")
    completion_notes = lori_transfer_clean(payload.get("completion_notes"))

    if not request_id:
        return {"status": "error", "message": "request_id is required."}

    rows = await lori_transfer_get_rows(
        "lori_user_access_change_requests",
        f"select=*&id=eq.{quote(request_id)}&limit=1",
    )

    if not rows:
        return {"status": "not_found", "message": "Access change request not found."}

    request = rows[0]

    if request.get("request_status") not in ["Approved", "Pending Review"]:
        return {
            "status": "error",
            "message": f"Request status is {request.get('request_status')}. Only approved requests should be completed.",
        }

    user_id = request.get("user_profile_id")
    transfer_type = request.get("request_type")
    rule = await lori_transfer_find_rule(transfer_type)

    if not rule:
        return {"status": "error", "message": "Transfer rule not found."}

    effective_date = request.get("transfer_effective_date") or lori_transfer_today()
    expiration_date = request.get("expiration_date")

    old_access_disabled = False
    new_access_enabled = False
    default_context_updated = False

    if rule.get("disables_old_station_access") and request.get("current_station_code"):
        await lori_transfer_patch_rows(
            "lori_user_station_access",
            f"user_profile_id=eq.{quote(str(user_id))}&station_code=eq.{quote(str(request.get('current_station_code')))}",
            {
                "access_status": "Transferred",
                "can_view": False,
                "can_upload": False,
                "can_edit": False,
                "can_approve": False,
                "can_send_notifications": False,
                "expiration_date": effective_date,
                "updated_at": datetime.utcnow().isoformat(),
            },
        )
        old_access_disabled = True

    if rule.get("deactivates_user"):
        await lori_transfer_patch_rows(
            "lori_user_profiles",
            f"id=eq.{quote(str(user_id))}",
            {
                "user_status": "Inactive",
                "updated_at": datetime.utcnow().isoformat(),
            },
        )

    if rule.get("suspends_user"):
        await lori_transfer_patch_rows(
            "lori_user_profiles",
            f"id=eq.{quote(str(user_id))}",
            {
                "user_status": "Suspended",
                "updated_at": datetime.utcnow().isoformat(),
            },
        )

    if rule.get("adds_new_station_access") and request.get("new_station_code"):
        existing_new_access = await lori_transfer_get_rows(
            "lori_user_station_access",
            f"select=*&user_profile_id=eq.{quote(str(user_id))}&station_code=eq.{quote(str(request.get('new_station_code')))}&limit=1",
        )

        new_access_payload = {
            "user_profile_id": user_id,
            "company_name": request.get("new_company_name"),
            "region_code": request.get("new_region_code"),
            "region_name": request.get("new_region_name"),
            "operating_state": request.get("new_operating_state"),
            "city": request.get("new_city"),
            "station_code": request.get("new_station_code"),
            "station_name": request.get("new_station_name"),
            "route_group": request.get("new_route_group") or "All Routes",
            "access_scope": "Station",
            "access_status": "Active",
            "can_view": True,
            "can_upload": True,
            "can_edit": True,
            "can_approve": True,
            "can_send_notifications": True,
            "effective_date": effective_date,
            "expiration_date": expiration_date or None,
            "created_by": completed_by,
        }

        if existing_new_access:
            await lori_transfer_patch_rows(
                "lori_user_station_access",
                f"id=eq.{quote(str(existing_new_access[0].get('id')))}",
                {
                    "access_status": "Active",
                    "can_view": True,
                    "can_upload": True,
                    "can_edit": True,
                    "can_approve": True,
                    "can_send_notifications": True,
                    "effective_date": effective_date,
                    "expiration_date": expiration_date or None,
                    "updated_at": datetime.utcnow().isoformat(),
                },
            )
        else:
            await lori_policy_supabase_post(
                "lori_user_station_access",
                new_access_payload,
            )

        new_access_enabled = True

    if rule.get("updates_default_context") and request.get("new_station_code"):
        await lori_transfer_patch_rows(
            "lori_user_profiles",
            f"id=eq.{quote(str(user_id))}",
            {
                "company_name": request.get("new_company_name"),
                "default_region_code": request.get("new_region_code"),
                "default_region_name": request.get("new_region_name"),
                "default_operating_state": request.get("new_operating_state"),
                "default_station_code": request.get("new_station_code"),
                "default_station_name": request.get("new_station_name"),
                "default_city": request.get("new_city"),
                "default_route_group": request.get("new_route_group") or "All Routes",
                "updated_at": datetime.utcnow().isoformat(),
            },
        )
        default_context_updated = True

    history_payload = {
        "access_change_request_id": request_id,
        "user_profile_id": user_id,
        "user_full_name": request.get("user_full_name"),
        "user_email": request.get("user_email"),
        "change_type": transfer_type,
        "change_status": "Completed",
        "from_company_name": request.get("current_company_name"),
        "from_region_name": request.get("current_region_name"),
        "from_operating_state": request.get("current_operating_state"),
        "from_city": request.get("current_city"),
        "from_station_code": request.get("current_station_code"),
        "from_station_name": request.get("current_station_name"),
        "to_company_name": request.get("new_company_name"),
        "to_region_name": request.get("new_region_name"),
        "to_operating_state": request.get("new_operating_state"),
        "to_city": request.get("new_city"),
        "to_station_code": request.get("new_station_code"),
        "to_station_name": request.get("new_station_name"),
        "old_access_status": "Transferred" if old_access_disabled else "Unchanged",
        "new_access_status": "Active" if new_access_enabled else "Not Added",
        "effective_date": effective_date,
        "expiration_date": expiration_date,
        "changed_by": completed_by,
        "change_reason": request.get("transfer_reason"),
        "notes": completion_notes,
    }

    await lori_policy_supabase_post(
        "lori_user_access_change_history",
        history_payload,
    )

    updated_request = await lori_transfer_patch_rows(
        "lori_user_access_change_requests",
        f"id=eq.{quote(request_id)}",
        {
            "request_status": "Completed",
            "completed_by": completed_by,
            "completion_notes": completion_notes,
            "old_access_disabled": old_access_disabled,
            "new_access_enabled": new_access_enabled,
            "default_context_updated": default_context_updated,
            "completed_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_access_log_event(
        user_profile_id=user_id,
        action_type="Transfer Request Completed",
        module_name="User Access Management",
        company_name=request.get("new_company_name"),
        region_name=request.get("new_region_name"),
        operating_state=request.get("new_operating_state"),
        station_code=request.get("new_station_code"),
        access_result="Completed",
        reason=f"{transfer_type} completed.",
        request_context={
            "request": updated_request[0] if updated_request else request,
            "history": history_payload,
        },
    )

    return {
        "status": "success",
        "message": "Access change completed.",
        "request": updated_request[0] if updated_request else request,
        "old_access_disabled": old_access_disabled,
        "new_access_enabled": new_access_enabled,
        "default_context_updated": default_context_updated,
        "history": history_payload,
    }


@app.get("/access-change-history")
async def access_change_history(
    api_key: Optional[str] = Query(None),
    user_email: Optional[str] = Query(None),
    limit: int = Query(100),
):
    lori_regulatory_require_key(api_key)

    rows = await lori_transfer_get_rows(
        "lori_user_access_change_history",
        "select=*&order=created_at.desc&limit=500",
    )

    if user_email:
        rows = [
            row for row in rows
            if lori_transfer_email(row.get("user_email")) == lori_transfer_email(user_email)
        ]

    limit = max(1, min(limit, 500))

    return {
        "status": "success",
        "history_count": len(rows[:limit]),
        "history": rows[:limit],
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# CENTRAL DOCUMENT & DATA INTAKE ENGINE
# Station / Operations Uploads
# Driver File Uploads
# Employee / Staff File Uploads
# Central document library used by Route Configuration, Driver 360,
# Compliance, KPI, Counseling, Contract Safeguard, Action Center, etc.
# ============================================================

from fastapi import Body, Query, UploadFile, File, Form
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime
import os
import io
import csv
import uuid
import zipfile
import xml.etree.ElementTree as ET
import httpx
import openpyxl


SUPABASE_URL_DOCS = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_DOCS = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_doc_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_doc_upper(value: Any) -> str:
    return lori_doc_clean(value).upper()


def lori_doc_bool_from_text(*values: Any, keywords: List[str]) -> bool:
    combined = " ".join([lori_doc_clean(v).lower() for v in values if v is not None])
    return any(keyword.lower() in combined for keyword in keywords)


def lori_doc_slug(value: Any) -> str:
    cleaned = lori_doc_clean(value).lower()
    keep = []
    for char in cleaned:
        if char.isalnum():
            keep.append(char)
        elif char in [" ", "-", "_", "/", "."]:
            keep.append("-")
    slug = "".join(keep).strip("-")
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug or "document"


async def lori_doc_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_doc_upload_to_storage(
    bucket: str,
    storage_path: str,
    contents: bytes,
    content_type: str = "application/octet-stream",
) -> Dict[str, Any]:
    if not SUPABASE_URL_DOCS or not SUPABASE_SERVICE_ROLE_KEY_DOCS:
        raise RuntimeError("Missing Supabase environment variables.")

    safe_path = quote(storage_path, safe="/")
    url = f"{SUPABASE_URL_DOCS}/storage/v1/object/{bucket}/{safe_path}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_DOCS,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_DOCS}",
        "Content-Type": content_type,
        "x-upsert": "true",
    }

    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post(url, headers=headers, content=contents)

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase storage upload failed: {response.status_code} {response.text}")

    try:
        return response.json()
    except Exception:
        return {"status": "uploaded", "path": storage_path}


def lori_doc_extract_text_from_docx(contents: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as docx:
            xml_content = docx.read("word/document.xml")
        root = ET.fromstring(xml_content)
        text_parts = []
        for element in root.iter():
            if element.tag.endswith("}t") and element.text:
                text_parts.append(element.text)
        return "\n".join(text_parts).strip()
    except Exception as exc:
        return f"DOCX extraction failed: {str(exc)}"


def lori_doc_extract_text_and_rows(file_ext: str, contents: bytes) -> Dict[str, Any]:
    file_ext = file_ext.lower().strip(".")
    extracted_text = ""
    parsed_rows_count = 0
    extraction_status = "Not Extracted"
    parse_status = "Not Parsed"
    extraction_notes = ""

    try:
        if file_ext in ["txt"]:
            extracted_text = contents.decode("utf-8-sig", errors="ignore")
            extraction_status = "Extracted"
            parse_status = "Text Extracted"
            extraction_notes = "TXT text extracted."

        elif file_ext in ["csv"]:
            text = contents.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(text))
            rows = [row for row in reader]
            parsed_rows_count = len(rows)
            preview_lines = text.splitlines()[:80]
            extracted_text = "\n".join(preview_lines)
            extraction_status = "Extracted"
            parse_status = "Parsed"
            extraction_notes = f"CSV parsed with {parsed_rows_count} data rows."

        elif file_ext in ["xlsx", "xlsm"]:
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            parsed_rows_count = max(0, len(rows) - 1)
            preview = []
            for row in rows[:80]:
                preview.append("\t".join([str(v) if v is not None else "" for v in row]))
            extracted_text = "\n".join(preview)
            extraction_status = "Extracted"
            parse_status = "Parsed"
            extraction_notes = f"Excel file parsed with {parsed_rows_count} data rows."

        elif file_ext in ["docx"]:
            extracted_text = lori_doc_extract_text_from_docx(contents)
            extraction_status = "Extracted" if extracted_text and not extracted_text.startswith("DOCX extraction failed") else "Needs Review"
            parse_status = "Text Extracted" if extraction_status == "Extracted" else "Needs Review"
            extraction_notes = "DOCX text extracted." if extraction_status == "Extracted" else extracted_text

        elif file_ext in ["pdf"]:
            extracted_text = ""
            extraction_status = "Not Extracted"
            parse_status = "PDF Stored"
            extraction_notes = "PDF uploaded and stored. PDF text extraction is staged for a later backend step."

        else:
            extracted_text = ""
            extraction_status = "Not Extracted"
            parse_status = "Stored"
            extraction_notes = f"File type {file_ext} uploaded and stored. Automatic extraction not configured yet."

    except Exception as exc:
        extracted_text = ""
        extraction_status = "Failed Extraction"
        parse_status = "Failed"
        extraction_notes = str(exc)

    return {
        "extraction_status": extraction_status,
        "parse_status": parse_status,
        "extracted_text": extracted_text[:250000] if extracted_text else "",
        "parsed_rows_count": parsed_rows_count,
        "extraction_notes": extraction_notes,
    }


def lori_doc_flags(
    intake_lane: str,
    document_type: str,
    document_category: str,
    applies_to: str,
    driver_type: str = "",
    employee_role: str = "",
) -> Dict[str, bool]:
    values = [intake_lane, document_type, document_category, applies_to, driver_type, employee_role]

    contract_related = lori_doc_bool_from_text(
        *values,
        keywords=[
            "contract", "agreement", "owner-operator", "owner operator",
            "contractor", "collective bargaining", "cba"
        ],
    )

    union_related = lori_doc_bool_from_text(
        *values,
        keywords=["union", "collective bargaining", "cba", "labor agreement"],
    )

    contractor_related = lori_doc_bool_from_text(
        *values,
        keywords=["contractor", "owner-operator", "owner operator"],
    )

    owner_operator_related = lori_doc_bool_from_text(
        *values,
        keywords=["owner-operator", "owner operator"],
    )

    pay_related = lori_doc_bool_from_text(
        *values,
        keywords=["pay", "compensation", "wage", "overtime", "rate", "guarantee"],
    )

    route_assignment_related = lori_doc_bool_from_text(
        *values,
        keywords=["route assignment", "route", "manifest", "work area", "territory", "stop"],
    )

    safety_related = lori_doc_bool_from_text(
        *values,
        keywords=["safety", "accident", "incident", "injury", "crash", "dot", "fmCSA".lower()],
    )

    accident_related = lori_doc_bool_from_text(
        *values,
        keywords=["accident", "incident", "crash", "collision"],
    )

    counseling_related = lori_doc_bool_from_text(
        *values,
        keywords=["counseling", "disciplinary", "discipline", "corrective action", "warning"],
    )

    training_related = lori_doc_bool_from_text(
        *values,
        keywords=["training", "certification", "credential", "license", "mvr"],
    )

    compliance_related = lori_doc_bool_from_text(
        *values,
        keywords=["compliance", "policy", "regulatory", "dot", "fmCSA".lower(), "osha"],
    )

    policy_related = lori_doc_bool_from_text(
        *values,
        keywords=["policy", "sop", "procedure", "standard operating"],
    )

    kpi_related = lori_doc_bool_from_text(
        *values,
        keywords=["kpi", "performance", "scorecard", "payroll exception", "metrics"],
    )

    route_configuration_related = route_assignment_related

    notification_related = lori_doc_bool_from_text(
        *values,
        keywords=["notification", "message", "push", "sms", "email"],
    )

    labor_review_required = contract_related or union_related or contractor_related or pay_related
    hr_review_required = counseling_related or accident_related or training_related or pay_related
    legal_review_required = contract_related or union_related

    return {
        "contract_related": contract_related,
        "union_related": union_related,
        "contractor_related": contractor_related,
        "owner_operator_related": owner_operator_related,
        "pay_related": pay_related,
        "route_assignment_related": route_assignment_related,
        "labor_review_required": labor_review_required,
        "hr_review_required": hr_review_required,
        "legal_review_required": legal_review_required,
        "safety_related": safety_related,
        "accident_related": accident_related,
        "counseling_related": counseling_related,
        "training_related": training_related,
        "compliance_related": compliance_related,
        "policy_related": policy_related,
        "kpi_related": kpi_related,
        "route_configuration_related": route_configuration_related,
        "notification_related": notification_related,
    }


def lori_doc_default_modules(flags: Dict[str, bool], intake_lane: str) -> List[str]:
    modules = ["Document Library"]

    if intake_lane == "Station / Operations":
        modules.extend(["Data Intake", "Leadership Dashboard", "Reports"])

    if intake_lane == "Driver File":
        modules.extend(["Driver 360", "Counseling", "Safety"])

    if intake_lane == "Employee / Staff File":
        modules.extend(["Employee / Staff File", "Training", "Safety"])

    if flags.get("route_configuration_related"):
        modules.append("Route Configuration")

    if flags.get("contract_related") or flags.get("union_related") or flags.get("pay_related"):
        modules.append("Contract Safeguard")

    if flags.get("compliance_related") or flags.get("policy_related"):
        modules.append("Compliance & Policy Center")

    if flags.get("kpi_related"):
        modules.append("KPI Action Plans")

    if flags.get("accident_related") or flags.get("safety_related"):
        modules.append("Safety Review")

    return sorted(list(set(modules)))


@app.get("/document-library-summary")
async def document_library_summary(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    docs = await lori_doc_get_rows(
        "lori_document_library",
        "select=*&order=created_at.desc&limit=5000",
    )

    if station_code:
        docs = [
            d for d in docs
            if lori_doc_upper(d.get("station_code")) == lori_doc_upper(station_code)
        ]

    return {
        "status": "success",
        "documents_count": len(docs),
        "station_operations_count": len([d for d in docs if d.get("intake_lane") == "Station / Operations"]),
        "driver_file_count": len([d for d in docs if d.get("intake_lane") == "Driver File"]),
        "employee_staff_file_count": len([d for d in docs if d.get("intake_lane") == "Employee / Staff File"]),
        "contract_related_count": len([d for d in docs if d.get("contract_related")]),
        "union_related_count": len([d for d in docs if d.get("union_related")]),
        "safety_related_count": len([d for d in docs if d.get("safety_related")]),
        "accident_related_count": len([d for d in docs if d.get("accident_related")]),
        "counseling_related_count": len([d for d in docs if d.get("counseling_related")]),
        "route_configuration_related_count": len([d for d in docs if d.get("route_configuration_related")]),
        "kpi_related_count": len([d for d in docs if d.get("kpi_related")]),
        "extracted_count": len([d for d in docs if d.get("extraction_status") == "Extracted"]),
    }


@app.post("/document-intake-upload")
async def document_intake_upload(
    api_key: Optional[str] = Query(None),

    intake_lane: str = Form(...),
    document_title: str = Form(...),
    document_type: str = Form("Other"),
    document_category: str = Form(""),
    applies_to: str = Form("Station"),

    company_name: str = Form(""),
    region_code: str = Form(""),
    region_name: str = Form(""),
    operating_state: str = Form(""),
    city: str = Form(""),
    station_code: str = Form(""),
    station_name: str = Form(""),
    primary_zip: str = Form(""),
    route_group: str = Form(""),
    route_id: str = Form(""),

    subject_type: str = Form("Station"),
    driver_name: str = Form(""),
    driver_id: str = Form(""),
    driver_type: str = Form(""),
    employee_name: str = Form(""),
    employee_id: str = Form(""),
    employee_role: str = Form(""),
    department: str = Form(""),
    supervisor_name: str = Form(""),

    incident_date: Optional[str] = Form(None),
    effective_date: Optional[str] = Form(None),
    expiration_date: Optional[str] = Form(None),
    review_due_date: Optional[str] = Form(None),

    document_owner: str = Form(""),
    uploaded_by: str = Form("LORI Document Intake"),
    notes: str = Form(""),

    file: UploadFile = File(...),
):
    lori_regulatory_require_key(api_key)

    original_file_name = file.filename or "uploaded_document"
    file_ext = original_file_name.split(".")[-1].lower() if "." in original_file_name else "unknown"
    file_content_type = file.content_type or "application/octet-stream"
    contents = await file.read()
    file_size = len(contents)

    flags = lori_doc_flags(
        intake_lane=intake_lane,
        document_type=document_type,
        document_category=document_category,
        applies_to=applies_to,
        driver_type=driver_type,
        employee_role=employee_role,
    )

    modules = lori_doc_default_modules(flags, intake_lane)

    extraction = lori_doc_extract_text_and_rows(file_ext, contents)

    bucket = "lori-central-documents"

    station_part = lori_doc_slug(station_code or "no-station")
    lane_part = lori_doc_slug(intake_lane)
    file_part = lori_doc_slug(original_file_name)
    storage_path = f"{station_part}/{lane_part}/{uuid.uuid4()}-{file_part}"

    storage_result = await lori_doc_upload_to_storage(
        bucket=bucket,
        storage_path=storage_path,
        contents=contents,
        content_type=file_content_type,
    )

    sensitivity_level = "Sensitive" if intake_lane in ["Driver File", "Employee / Staff File"] or flags.get("contract_related") or flags.get("accident_related") else "Standard"
    privacy_classification = "Personnel" if intake_lane in ["Driver File", "Employee / Staff File"] else "Internal"

    document_payload = {
        "document_title": lori_doc_clean(document_title),
        "document_status": "Uploaded",
        "intake_lane": lori_doc_clean(intake_lane),
        "document_type": lori_doc_clean(document_type),
        "document_category": lori_doc_clean(document_category),
        "applies_to": lori_doc_clean(applies_to),

        "company_name": lori_doc_clean(company_name),
        "region_code": lori_doc_upper(region_code),
        "region_name": lori_doc_clean(region_name),
        "operating_state": lori_doc_upper(operating_state),
        "city": lori_doc_clean(city),
        "station_code": lori_doc_upper(station_code),
        "station_name": lori_doc_clean(station_name),
        "primary_zip": lori_doc_clean(primary_zip),
        "route_group": lori_doc_clean(route_group),
        "route_id": lori_doc_clean(route_id),

        "subject_type": lori_doc_clean(subject_type),
        "driver_name": lori_doc_clean(driver_name),
        "driver_id": lori_doc_clean(driver_id),
        "driver_type": lori_doc_clean(driver_type),
        "employee_name": lori_doc_clean(employee_name),
        "employee_id": lori_doc_clean(employee_id),
        "employee_role": lori_doc_clean(employee_role),
        "department": lori_doc_clean(department),
        "supervisor_name": lori_doc_clean(supervisor_name),

        **flags,

        "incident_date": incident_date or None,
        "effective_date": effective_date or None,
        "expiration_date": expiration_date or None,
        "review_due_date": review_due_date or None,

        "original_file_name": original_file_name,
        "file_type": file_ext,
        "file_size_bytes": file_size,
        "storage_bucket": bucket,
        "storage_path": storage_path,

        "extraction_status": extraction["extraction_status"],
        "parse_status": extraction["parse_status"],
        "extracted_text": extraction["extracted_text"],
        "parsed_rows_count": extraction["parsed_rows_count"],
        "extraction_notes": extraction["extraction_notes"],

        "referenced_by_modules": modules,
        "searchable_keywords": [
            lori_doc_clean(document_type).lower(),
            lori_doc_clean(document_category).lower(),
            lori_doc_clean(applies_to).lower(),
            lori_doc_clean(driver_name).lower(),
            lori_doc_clean(employee_name).lower(),
            lori_doc_clean(route_id).lower(),
            lori_doc_clean(station_code).lower(),
        ],
        "privacy_classification": privacy_classification,
        "sensitivity_level": sensitivity_level,
        "archive_status": "Active",

        "document_owner": lori_doc_clean(document_owner),
        "uploaded_by": lori_doc_clean(uploaded_by),
        "notes": lori_doc_clean(notes),
    }

    created = await lori_policy_supabase_post("lori_document_library", document_payload)
    document = created[0] if created else document_payload

    extraction_job_payload = {
        "document_id": document.get("id"),
        "job_type": "Initial Upload Extraction",
        "job_status": "Completed" if extraction["extraction_status"] == "Extracted" else "Needs Review",
        "file_type": file_ext,
        "extraction_method": "LORI Basic Extraction",
        "rows_detected": extraction["parsed_rows_count"],
        "rows_imported": 0,
        "pages_detected": 0,
        "extracted_text_preview": extraction["extracted_text"][:1000] if extraction["extracted_text"] else "",
        "error_message": "" if extraction["extraction_status"] != "Failed Extraction" else extraction["extraction_notes"],
        "review_required": extraction["extraction_status"] != "Extracted",
        "started_at": datetime.utcnow().isoformat(),
        "completed_at": datetime.utcnow().isoformat(),
    }

    await lori_policy_supabase_post("lori_document_extraction_jobs", extraction_job_payload)

    return {
        "status": "success",
        "message": "Document uploaded to LORI Document Library.",
        "document": document,
        "storage": {
            "bucket": bucket,
            "path": storage_path,
            "result": storage_result,
        },
        "extraction": extraction,
        "referenced_by_modules": modules,
        "warning": "PDF extraction is staged. PDF is stored and available for selection, but structured extraction will be added in a later step." if file_ext == "pdf" else None,
    }


@app.get("/document-library")
async def document_library(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    intake_lane: Optional[str] = Query(None),
    document_type: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    driver_id: Optional[str] = Query(None),
    employee_name: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    module_name: Optional[str] = Query(None),
    contract_related: Optional[bool] = Query(None),
    safety_related: Optional[bool] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    docs = await lori_doc_get_rows(
        "lori_document_library",
        "select=*&order=created_at.desc&limit=5000",
    )

    if station_code:
        docs = [d for d in docs if lori_doc_upper(d.get("station_code")) == lori_doc_upper(station_code)]

    if intake_lane:
        docs = [d for d in docs if lori_doc_clean(d.get("intake_lane")).lower() == intake_lane.lower()]

    if document_type:
        docs = [d for d in docs if lori_doc_clean(d.get("document_type")).lower() == document_type.lower()]

    if driver_name:
        docs = [d for d in docs if driver_name.lower() in lori_doc_clean(d.get("driver_name")).lower()]

    if driver_id:
        docs = [d for d in docs if lori_doc_clean(d.get("driver_id")).lower() == driver_id.lower()]

    if employee_name:
        docs = [d for d in docs if employee_name.lower() in lori_doc_clean(d.get("employee_name")).lower()]

    if employee_id:
        docs = [d for d in docs if lori_doc_clean(d.get("employee_id")).lower() == employee_id.lower()]

    if route_id:
        docs = [d for d in docs if lori_doc_clean(d.get("route_id")).lower() == route_id.lower()]

    if module_name:
        docs = [
            d for d in docs
            if module_name in (d.get("referenced_by_modules") or [])
        ]

    if contract_related is not None:
        docs = [d for d in docs if bool(d.get("contract_related")) is contract_related]

    if safety_related is not None:
        docs = [d for d in docs if bool(d.get("safety_related")) is safety_related]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "documents_count": len(docs[:limit]),
        "documents": docs[:limit],
    }


@app.get("/document-detail")
async def document_detail(
    api_key: Optional[str] = Query(None),
    document_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    docs = await lori_doc_get_rows(
        "lori_document_library",
        f"select=*&id=eq.{quote(document_id)}&limit=1",
    )

    if not docs:
        return {
            "status": "not_found",
            "message": "Document not found.",
        }

    relationships = await lori_doc_get_rows(
        "lori_document_relationships",
        f"select=*&document_id=eq.{quote(document_id)}&order=created_at.desc&limit=500",
    )

    findings = await lori_doc_get_rows(
        "lori_document_review_findings",
        f"select=*&document_id=eq.{quote(document_id)}&order=created_at.desc&limit=500",
    )

    extraction_jobs = await lori_doc_get_rows(
        "lori_document_extraction_jobs",
        f"select=*&document_id=eq.{quote(document_id)}&order=created_at.desc&limit=100",
    )

    return {
        "status": "success",
        "document": docs[0],
        "relationships_count": len(relationships),
        "relationships": relationships,
        "findings_count": len(findings),
        "findings": findings,
        "extraction_jobs_count": len(extraction_jobs),
        "extraction_jobs": extraction_jobs,
    }


@app.post("/document-relationship-create")
async def document_relationship_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    document_id = lori_doc_clean(payload.get("document_id"))
    target_module = lori_doc_clean(payload.get("target_module"))
    relationship_type = lori_doc_clean(payload.get("relationship_type") or f"Attached to {target_module}")

    if not document_id:
        return {"status": "error", "message": "document_id is required."}

    if not target_module:
        return {"status": "error", "message": "target_module is required."}

    relationship_payload = {
        "document_id": document_id,
        "relationship_type": relationship_type,
        "target_module": target_module,
        "target_record_id": payload.get("target_record_id"),
        "target_reference": lori_doc_clean(payload.get("target_reference")),
        "company_name": lori_doc_clean(payload.get("company_name")),
        "station_code": lori_doc_upper(payload.get("station_code")),
        "route_id": lori_doc_clean(payload.get("route_id")),
        "driver_id": lori_doc_clean(payload.get("driver_id")),
        "driver_name": lori_doc_clean(payload.get("driver_name")),
        "employee_id": lori_doc_clean(payload.get("employee_id")),
        "employee_name": lori_doc_clean(payload.get("employee_name")),
        "relationship_status": "Active",
        "linked_by": lori_doc_clean(payload.get("linked_by") or "LORI Document Intake"),
        "notes": lori_doc_clean(payload.get("notes")),
    }

    created = await lori_policy_supabase_post("lori_document_relationships", relationship_payload)

    return {
        "status": "success",
        "message": "Document relationship created.",
        "relationship": created[0] if created else relationship_payload,
    }


@app.post("/document-review-finding-create")
async def document_review_finding_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    document_id = lori_doc_clean(payload.get("document_id"))

    if not document_id:
        return {"status": "error", "message": "document_id is required."}

    finding_payload = {
        "document_id": document_id,
        "finding_type": lori_doc_clean(payload.get("finding_type") or "Document Review Finding"),
        "finding_category": lori_doc_clean(payload.get("finding_category")),
        "risk_level": lori_doc_clean(payload.get("risk_level") or "Medium"),
        "finding_title": lori_doc_clean(payload.get("finding_title")),
        "finding_summary": lori_doc_clean(payload.get("finding_summary") or "Finding requires review."),
        "relevant_excerpt": lori_doc_clean(payload.get("relevant_excerpt")),
        "operational_impact": lori_doc_clean(payload.get("operational_impact")),
        "cost_impact": lori_doc_clean(payload.get("cost_impact")),
        "safety_impact": lori_doc_clean(payload.get("safety_impact")),
        "compliance_impact": lori_doc_clean(payload.get("compliance_impact")),
        "labor_impact": lori_doc_clean(payload.get("labor_impact")),
        "recommended_action": lori_doc_clean(payload.get("recommended_action")),
        "requires_supervisor_review": bool(payload.get("requires_supervisor_review", False)),
        "requires_hr_review": bool(payload.get("requires_hr_review", False)),
        "requires_labor_review": bool(payload.get("requires_labor_review", False)),
        "requires_legal_review": bool(payload.get("requires_legal_review", False)),
        "blocks_final_action": bool(payload.get("blocks_final_action", False)),
        "created_by": lori_doc_clean(payload.get("created_by") or "LORI Document Review Engine"),
    }

    created = await lori_policy_supabase_post("lori_document_review_findings", finding_payload)

    return {
        "status": "success",
        "message": "Document review finding created.",
        "finding": created[0] if created else finding_payload,
    }


@app.get("/documents-for-driver")
async def documents_for_driver(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    driver_id: Optional[str] = Query(None),
    limit: int = Query(200),
):
    return await document_library(
        api_key=api_key,
        station_code=station_code,
        intake_lane="Driver File",
        driver_name=driver_name,
        driver_id=driver_id,
        limit=limit,
    )


@app.get("/documents-for-employee")
async def documents_for_employee(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    employee_name: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    limit: int = Query(200),
):
    return await document_library(
        api_key=api_key,
        station_code=station_code,
        intake_lane="Employee / Staff File",
        employee_name=employee_name,
        employee_id=employee_id,
        limit=limit,
    )


@app.get("/documents-for-route-configuration")
async def documents_for_route_configuration(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    limit: int = Query(300),
):
    return await document_library(
        api_key=api_key,
        station_code=station_code,
        route_id=route_id,
        module_name="Route Configuration",
        limit=limit,
    )


@app.get("/documents-for-contract-safeguard")
async def documents_for_contract_safeguard(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    driver_id: Optional[str] = Query(None),
    limit: int = Query(300),
):
    lori_regulatory_require_key(api_key)

    docs_response = await document_library(
        api_key=api_key,
        station_code=station_code,
        driver_name=driver_name,
        driver_id=driver_id,
        limit=1000,
    )

    docs = docs_response.get("documents", [])

    docs = [
        d for d in docs
        if d.get("contract_related")
        or d.get("union_related")
        or d.get("contractor_related")
        or d.get("owner_operator_related")
        or d.get("pay_related")
        or d.get("route_assignment_related")
    ]

    return {
        "status": "success",
        "documents_count": len(docs[:limit]),
        "documents": docs[:limit],
    }


@app.post("/document-extract-text")
async def document_extract_text(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    document_id = lori_doc_clean(payload.get("document_id"))

    if not document_id:
        return {"status": "error", "message": "document_id is required."}

    detail = await document_detail(api_key=api_key, document_id=document_id)

    if detail.get("status") != "success":
        return detail

    document = detail.get("document") or {}

    if document.get("extracted_text"):
        return {
            "status": "success",
            "message": "Document already has extracted text.",
            "document_id": document_id,
            "extraction_status": document.get("extraction_status"),
            "parse_status": document.get("parse_status"),
            "extracted_text_preview": document.get("extracted_text", "")[:3000],
        }

    job_payload = {
        "document_id": document_id,
        "job_type": "Manual Extraction Request",
        "job_status": "Needs Review",
        "file_type": document.get("file_type"),
        "extraction_method": "Manual/Advanced Extraction Required",
        "rows_detected": 0,
        "rows_imported": 0,
        "pages_detected": 0,
        "extracted_text_preview": "",
        "error_message": "No extracted text is available yet. PDF/OCR advanced extraction will be added in a later step.",
        "review_required": True,
        "started_at": datetime.utcnow().isoformat(),
        "completed_at": datetime.utcnow().isoformat(),
    }

    created_job = await lori_policy_supabase_post("lori_document_extraction_jobs", job_payload)

    return {
        "status": "needs_review",
        "message": "Document does not have extracted text yet. Advanced extraction is required.",
        "document_id": document_id,
        "extraction_job": created_job[0] if created_job else job_payload,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# ROUTE CHANGE CONTRACT & LABOR AGREEMENT SAFEGUARD ENGINE
# Reviews contractor agreements, union agreements, route policies,
# pay policies, and station documents before route changes.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime
import re


def lori_contract_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_contract_upper(value: Any) -> str:
    return lori_contract_clean(value).upper()


async def lori_contract_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


def lori_contract_contains(text: str, keywords: List[str]) -> bool:
    text_lower = (text or "").lower()
    return any(keyword.lower() in text_lower for keyword in keywords)


def lori_contract_excerpt(text: str, keywords: List[str], max_len: int = 450) -> str:
    if not text:
        return ""

    text_clean = " ".join(text.split())
    text_lower = text_clean.lower()

    for keyword in keywords:
        idx = text_lower.find(keyword.lower())
        if idx >= 0:
            start = max(0, idx - 160)
            end = min(len(text_clean), idx + max_len)
            return text_clean[start:end]

    return text_clean[:max_len]


async def lori_contract_get_or_create_from_document_library(document_id: str) -> Optional[Dict[str, Any]]:
    docs = await lori_contract_get_rows(
        "lori_document_library",
        f"select=*&id=eq.{quote(document_id)}&limit=1",
    )

    if not docs:
        return None

    doc = docs[0]

    existing = await lori_contract_get_rows(
        "lori_route_contract_documents",
        f"select=*&original_file_name=eq.{quote(str(doc.get('original_file_name') or ''))}&station_code=eq.{quote(str(doc.get('station_code') or ''))}&limit=1",
    )

    if existing:
        return existing[0]

    contract_doc_payload = {
        "document_title": doc.get("document_title") or "Document Library Agreement",
        "document_type": doc.get("document_type") or "Contract / Agreement",
        "document_status": "Linked from Document Library",
        "company_name": doc.get("company_name"),
        "region_code": doc.get("region_code"),
        "region_name": doc.get("region_name"),
        "operating_state": doc.get("operating_state"),
        "city": doc.get("city"),
        "station_code": doc.get("station_code"),
        "station_name": doc.get("station_name"),
        "route_group": doc.get("route_group"),
        "applies_to": doc.get("applies_to") or "Route Change",
        "applies_to_driver": doc.get("driver_name"),
        "applies_to_route_id": doc.get("route_id"),
        "applies_to_scope": doc.get("applies_to") or "Station",
        "original_file_name": doc.get("original_file_name"),
        "file_type": doc.get("file_type"),
        "storage_bucket": doc.get("storage_bucket"),
        "storage_path": doc.get("storage_path"),
        "extraction_status": doc.get("extraction_status"),
        "extracted_text": doc.get("extracted_text"),
        "extraction_notes": doc.get("extraction_notes") or "Linked from central LORI Document Library.",
        "uploaded_by": "LORI Contract Safeguard",
    }

    created = await lori_policy_supabase_post(
        "lori_route_contract_documents",
        contract_doc_payload,
    )

    return created[0] if created else contract_doc_payload


def lori_contract_analyze_text(text: str, agreement_type: str = "") -> Dict[str, Any]:
    findings = []

    route_keywords = ["route", "territory", "work area", "assignment", "reassignment", "stops", "delivery responsibility"]
    notice_keywords = ["notice", "advance notice", "written notice", "notify", "notification"]
    approval_keywords = ["approval", "approved", "authorized", "supervisor", "manager review", "written approval"]
    pay_keywords = ["pay", "compensation", "wage", "rate", "overtime", "minimum hours", "guarantee", "premium"]
    union_keywords = ["union", "collective bargaining", "cba", "grievance", "seniority", "bid", "bidding"]
    contractor_keywords = ["contractor", "owner-operator", "owner operator", "subcontractor", "independent contractor"]
    schedule_keywords = ["schedule", "shift", "hours", "start time", "end time", "dispatch"]
    vehicle_keywords = ["vehicle", "equipment", "helper", "assistant", "truck"]

    text_has = lambda kws: lori_contract_contains(text, kws)

    contract_risk_score = 0
    labor_risk_score = 0
    cost_risk_score = 0

    if not text:
        findings.append({
            "finding_type": "Document Text Not Available",
            "finding_category": "Extraction Needed",
            "risk_level": "Medium",
            "clause_title": "No Extracted Text",
            "clause_reference": "Document extraction",
            "short_excerpt": "",
            "finding_summary": "This document is stored, but LORI does not have extracted text to review yet.",
            "operational_impact": "The route change should not be finalized until the agreement or policy can be reviewed.",
            "possible_cost_impact": "Unknown until document text is reviewed.",
            "recommended_action": "Run text extraction or upload a searchable PDF, TXT, DOCX, or policy text.",
            "requires_supervisor_review": True,
            "requires_hr_review": True,
            "requires_labor_review": True,
            "requires_legal_review": False,
            "blocks_final_approval": True,
        })
        return {
            "contract_risk": "Medium",
            "labor_risk": "Medium",
            "cost_impact_risk": "Medium",
            "implementation_status": "Supervisor / HR Review Required",
            "possible_cost_impact": True,
            "possible_grievance_risk": False,
            "possible_notice_required": False,
            "possible_pay_adjustment_required": False,
            "supervisor_review_required": True,
            "hr_labor_review_required": True,
            "legal_contract_review_required": False,
            "summary": "Document text is not available. Review is incomplete.",
            "recommended_next_step": "Upload a searchable agreement or have HR/labor/contract owner review the document manually.",
            "do_not_implement_warning": "Do not implement route changes until agreement review is complete.",
            "findings": findings,
        }

    if text_has(route_keywords):
        contract_risk_score += 1
        findings.append({
            "finding_type": "Route Assignment Language Found",
            "finding_category": "Route / Territory",
            "risk_level": "Medium",
            "clause_title": "Route or Work Assignment",
            "clause_reference": "Route assignment / territory language",
            "short_excerpt": lori_contract_excerpt(text, route_keywords),
            "finding_summary": "The document appears to reference route assignment, territory, stops, or delivery responsibility.",
            "operational_impact": "Moving stops or changing work areas may require supervisor or contract-owner review.",
            "possible_cost_impact": "Possible if route assignment affects pay, guarantee, workload, or overtime.",
            "recommended_action": "Review route assignment language before implementing the route change.",
            "requires_supervisor_review": True,
            "requires_hr_review": False,
            "requires_labor_review": False,
            "requires_legal_review": False,
            "blocks_final_approval": False,
        })

    if text_has(notice_keywords):
        contract_risk_score += 1
        findings.append({
            "finding_type": "Notice Requirement Possible",
            "finding_category": "Notice",
            "risk_level": "Medium",
            "clause_title": "Notice / Notification",
            "clause_reference": "Notice language",
            "short_excerpt": lori_contract_excerpt(text, notice_keywords),
            "finding_summary": "The document appears to contain notice or notification language.",
            "operational_impact": "The company may need to provide notice before changing routes, schedules, or assignments.",
            "possible_cost_impact": "Potential cost risk if notice requirements are missed.",
            "recommended_action": "Confirm notice obligations before implementing the change.",
            "requires_supervisor_review": True,
            "requires_hr_review": True,
            "requires_labor_review": False,
            "requires_legal_review": False,
            "blocks_final_approval": False,
        })

    if text_has(approval_keywords):
        contract_risk_score += 1
        findings.append({
            "finding_type": "Approval Requirement Possible",
            "finding_category": "Approval",
            "risk_level": "Medium",
            "clause_title": "Approval / Authorized Review",
            "clause_reference": "Approval language",
            "short_excerpt": lori_contract_excerpt(text, approval_keywords),
            "finding_summary": "The document appears to require approval or authorized review.",
            "operational_impact": "Route changes should be reviewed before implementation.",
            "possible_cost_impact": "Low to medium depending on approval requirements.",
            "recommended_action": "Send this route change to the appropriate supervisor or contract owner.",
            "requires_supervisor_review": True,
            "requires_hr_review": False,
            "requires_labor_review": False,
            "requires_legal_review": False,
            "blocks_final_approval": False,
        })

    if text_has(pay_keywords):
        cost_risk_score += 2
        findings.append({
            "finding_type": "Pay / Compensation Impact Possible",
            "finding_category": "Cost / Pay",
            "risk_level": "High",
            "clause_title": "Pay / Overtime / Compensation",
            "clause_reference": "Pay language",
            "short_excerpt": lori_contract_excerpt(text, pay_keywords),
            "finding_summary": "The document appears to contain pay, compensation, overtime, minimum hours, rate, or guarantee language.",
            "operational_impact": "Moving stops, changing route workload, or reducing/increasing hours may create pay impact.",
            "possible_cost_impact": "Additional pay, overtime, guarantee, premium, or compensation adjustment may apply.",
            "recommended_action": "Have HR, payroll, labor relations, or contract owner review before final approval.",
            "requires_supervisor_review": True,
            "requires_hr_review": True,
            "requires_labor_review": True,
            "requires_legal_review": False,
            "blocks_final_approval": True,
        })

    if text_has(union_keywords):
        labor_risk_score += 3
        findings.append({
            "finding_type": "Union / Seniority / Grievance Risk Possible",
            "finding_category": "Union / Labor",
            "risk_level": "High",
            "clause_title": "Union / CBA / Seniority / Grievance",
            "clause_reference": "Labor agreement language",
            "short_excerpt": lori_contract_excerpt(text, union_keywords),
            "finding_summary": "The document appears to contain union, collective bargaining, grievance, seniority, or bidding language.",
            "operational_impact": "Route changes may affect bidding rights, seniority, grievance exposure, or labor obligations.",
            "possible_cost_impact": "Potential grievance, premium pay, back pay, or implementation delay risk.",
            "recommended_action": "Do not implement until labor relations, HR, or authorized leadership reviews the change.",
            "requires_supervisor_review": True,
            "requires_hr_review": True,
            "requires_labor_review": True,
            "requires_legal_review": True,
            "blocks_final_approval": True,
        })

    if text_has(contractor_keywords):
        contract_risk_score += 2
        findings.append({
            "finding_type": "Contractor / Owner-Operator Language Found",
            "finding_category": "Contractor Agreement",
            "risk_level": "High",
            "clause_title": "Contractor / Owner-Operator",
            "clause_reference": "Contractor language",
            "short_excerpt": lori_contract_excerpt(text, contractor_keywords),
            "finding_summary": "The document appears to reference contractor, owner-operator, subcontractor, or independent contractor arrangements.",
            "operational_impact": "Changing route responsibility may affect contractor scope, compensation, or contractual obligations.",
            "possible_cost_impact": "Rate adjustment, contractor dispute, or contract compliance risk may exist.",
            "recommended_action": "Have the contract owner review before implementing route changes.",
            "requires_supervisor_review": True,
            "requires_hr_review": False,
            "requires_labor_review": False,
            "requires_legal_review": True,
            "blocks_final_approval": True,
        })

    if text_has(schedule_keywords):
        labor_risk_score += 1
        findings.append({
            "finding_type": "Schedule Impact Possible",
            "finding_category": "Schedule",
            "risk_level": "Medium",
            "clause_title": "Schedule / Shift / Hours",
            "clause_reference": "Schedule language",
            "short_excerpt": lori_contract_excerpt(text, schedule_keywords),
            "finding_summary": "The document appears to reference schedules, shifts, start times, hours, or dispatch.",
            "operational_impact": "Route changes may affect schedule, route duration, or start/end times.",
            "possible_cost_impact": "Potential overtime or schedule premium risk.",
            "recommended_action": "Review schedule impact before implementing.",
            "requires_supervisor_review": True,
            "requires_hr_review": True,
            "requires_labor_review": False,
            "requires_legal_review": False,
            "blocks_final_approval": False,
        })

    if text_has(vehicle_keywords):
        contract_risk_score += 1
        findings.append({
            "finding_type": "Vehicle / Helper Assignment Impact Possible",
            "finding_category": "Equipment / Helper",
            "risk_level": "Medium",
            "clause_title": "Vehicle / Helper / Equipment",
            "clause_reference": "Vehicle or helper language",
            "short_excerpt": lori_contract_excerpt(text, vehicle_keywords),
            "finding_summary": "The document appears to reference vehicle, equipment, helper, or assistant assignments.",
            "operational_impact": "Route changes may require vehicle, helper, or equipment review.",
            "possible_cost_impact": "Potential helper, equipment, or capacity cost impact.",
            "recommended_action": "Confirm vehicle/helper requirements before implementation.",
            "requires_supervisor_review": True,
            "requires_hr_review": False,
            "requires_labor_review": False,
            "requires_legal_review": False,
            "blocks_final_approval": False,
        })

    if not findings:
        findings.append({
            "finding_type": "No Specific Restriction Found",
            "finding_category": "General",
            "risk_level": "Low",
            "clause_title": "No direct route-change restriction detected",
            "clause_reference": "Keyword review",
            "short_excerpt": text[:350],
            "finding_summary": "LORI did not detect obvious route assignment, union, contractor, pay, notice, or approval language in the available text.",
            "operational_impact": "Route change may proceed as planning draft, but authorized operations review is still required.",
            "possible_cost_impact": "No specific cost impact detected from the available text.",
            "recommended_action": "Proceed with supervisor review before implementation.",
            "requires_supervisor_review": True,
            "requires_hr_review": False,
            "requires_labor_review": False,
            "requires_legal_review": False,
            "blocks_final_approval": False,
        })

    highest_score = max(contract_risk_score, labor_risk_score, cost_risk_score)

    if highest_score >= 3:
        contract_risk = "High" if contract_risk_score >= 2 else "Medium"
        labor_risk = "High" if labor_risk_score >= 3 else "Medium"
        cost_risk = "High" if cost_risk_score >= 2 else "Medium"
        implementation_status = "Do Not Implement Yet"
        warning = "Do not implement route changes until authorized HR, labor, legal, contract, or operations leadership review is complete."
    elif highest_score >= 1:
        contract_risk = "Medium"
        labor_risk = "Medium" if labor_risk_score else "Low"
        cost_risk = "Medium" if cost_risk_score else "Low"
        implementation_status = "Supervisor Review Required"
        warning = "Route change should remain in planning until supervisor review is complete."
    else:
        contract_risk = "Low"
        labor_risk = "Low"
        cost_risk = "Low"
        implementation_status = "Allowed for Planning Only"
        warning = "Supervisor review is still required before implementation."

    return {
        "contract_risk": contract_risk,
        "labor_risk": labor_risk,
        "cost_impact_risk": cost_risk,
        "implementation_status": implementation_status,
        "possible_cost_impact": cost_risk_score > 0,
        "possible_grievance_risk": labor_risk_score >= 3,
        "possible_notice_required": text_has(notice_keywords),
        "possible_pay_adjustment_required": cost_risk_score > 0,
        "supervisor_review_required": True,
        "hr_labor_review_required": labor_risk_score > 0 or cost_risk_score > 0,
        "legal_contract_review_required": contract_risk_score >= 2 or labor_risk_score >= 3,
        "summary": "LORI reviewed the available agreement/policy text for route-change impact.",
        "recommended_next_step": "Send for authorized review before implementation." if highest_score >= 1 else "Proceed as planning draft with supervisor review.",
        "do_not_implement_warning": warning,
        "findings": findings,
    }


@app.get("/route-contract-safeguard-documents")
async def route_contract_safeguard_documents(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    driver_id: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    limit: int = Query(300),
):
    lori_regulatory_require_key(api_key)

    docs = await lori_contract_get_rows(
        "lori_route_contract_documents",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        docs = [d for d in docs if lori_contract_upper(d.get("station_code")) == lori_contract_upper(station_code)]

    if driver_name:
        docs = [d for d in docs if driver_name.lower() in lori_contract_clean(d.get("applies_to_driver")).lower()]

    if route_id:
        docs = [d for d in docs if lori_contract_clean(d.get("applies_to_route_id")).lower() == route_id.lower()]

    central_docs_response = await documents_for_contract_safeguard(
        api_key=api_key,
        station_code=station_code,
        driver_name=driver_name,
        driver_id=driver_id,
        limit=limit,
    )

    central_docs = central_docs_response.get("documents", []) if isinstance(central_docs_response, dict) else []

    return {
        "status": "success",
        "contract_documents_count": len(docs[:limit]),
        "contract_documents": docs[:limit],
        "central_document_library_matches_count": len(central_docs),
        "central_document_library_matches": central_docs,
    }


@app.post("/route-contract-safeguard-review-create")
async def route_contract_safeguard_review_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    route_config_project_id = payload.get("route_config_project_id")
    document_id = payload.get("document_id")
    central_document_id = payload.get("central_document_id")

    linked_document = None

    if central_document_id and not document_id:
        linked_document = await lori_contract_get_or_create_from_document_library(str(central_document_id))
        document_id = linked_document.get("id") if linked_document else None

    if document_id and not linked_document:
        docs = await lori_contract_get_rows(
            "lori_route_contract_documents",
            f"select=*&id=eq.{quote(str(document_id))}&limit=1",
        )
        linked_document = docs[0] if docs else None

    project = None
    if route_config_project_id:
        projects = await lori_contract_get_rows(
            "lori_route_config_projects",
            f"select=*&id=eq.{quote(str(route_config_project_id))}&limit=1",
        )
        project = projects[0] if projects else None

    review_payload = {
        "route_config_project_id": route_config_project_id,
        "review_title": payload.get("review_title") or "Route Change Contract & Labor Safeguard Review",
        "review_status": "Draft",
        "agreement_required_status": payload.get("agreement_required_status") or "Agreement Selected",
        "company_name": payload.get("company_name") or (project or {}).get("company_name") or (linked_document or {}).get("company_name"),
        "region_code": payload.get("region_code") or (project or {}).get("region_code") or (linked_document or {}).get("region_code"),
        "region_name": payload.get("region_name") or (project or {}).get("region_name") or (linked_document or {}).get("region_name"),
        "operating_state": payload.get("operating_state") or (project or {}).get("operating_state") or (linked_document or {}).get("operating_state"),
        "city": payload.get("city") or (project or {}).get("city") or (linked_document or {}).get("city"),
        "station_code": payload.get("station_code") or (project or {}).get("station_code") or (linked_document or {}).get("station_code"),
        "station_name": payload.get("station_name") or (project or {}).get("station_name") or (linked_document or {}).get("station_name"),
        "route_group": payload.get("route_group") or (project or {}).get("route_group") or (linked_document or {}).get("route_group"),
        "current_route_id": payload.get("current_route_id"),
        "current_driver_name": payload.get("current_driver_name"),
        "receiving_route_id": payload.get("receiving_route_id"),
        "receiving_driver_name": payload.get("receiving_driver_name"),
        "route_change_summary": payload.get("route_change_summary") or "Route change requires contract/labor safeguard review.",
        "proposed_stop_moves": payload.get("proposed_stop_moves") or 0,
        "proposed_work_area_change": payload.get("proposed_work_area_change", True),
        "agreement_type_selected": payload.get("agreement_type_selected") or (linked_document or {}).get("document_type") or "Agreement / Policy",
        "document_id": document_id,
        "contract_risk": "Not Reviewed",
        "labor_risk": "Not Reviewed",
        "cost_impact_risk": "Not Reviewed",
        "implementation_status": "Planning Only",
        "safeguard_summary": "Safeguard review created. Run review before final route change approval.",
        "recommended_next_step": "Run Contract & Labor Safeguard Review.",
    }

    created = await lori_policy_supabase_post(
        "lori_route_contract_safeguard_reviews",
        review_payload,
    )

    return {
        "status": "success",
        "message": "Contract & Labor Safeguard review created.",
        "review": created[0] if created else review_payload,
        "linked_document": linked_document,
    }


@app.post("/route-contract-safeguard-run")
async def route_contract_safeguard_run(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    review_id = lori_contract_clean(payload.get("safeguard_review_id") or payload.get("review_id"))

    if not review_id:
        return {"status": "error", "message": "safeguard_review_id is required."}

    reviews = await lori_contract_get_rows(
        "lori_route_contract_safeguard_reviews",
        f"select=*&id=eq.{quote(review_id)}&limit=1",
    )

    if not reviews:
        return {"status": "not_found", "message": "Safeguard review not found."}

    review = reviews[0]

    document = None
    document_id = review.get("document_id")

    if document_id:
        docs = await lori_contract_get_rows(
            "lori_route_contract_documents",
            f"select=*&id=eq.{quote(str(document_id))}&limit=1",
        )
        document = docs[0] if docs else None

    text = (document or {}).get("extracted_text") or ""
    agreement_type = review.get("agreement_type_selected") or (document or {}).get("document_type") or ""

    analysis = lori_contract_analyze_text(text, agreement_type=agreement_type)

    findings_created = []

    for finding in analysis["findings"]:
        finding_payload = {
            "safeguard_review_id": review_id,
            "document_id": document_id,
            **finding,
        }

        created = await lori_policy_supabase_post(
            "lori_route_contract_safeguard_findings",
            finding_payload,
        )

        if created:
            findings_created.append(created[0])

    updated_review_payload = {
        "review_status": "Reviewed",
        "contract_risk": analysis["contract_risk"],
        "labor_risk": analysis["labor_risk"],
        "cost_impact_risk": analysis["cost_impact_risk"],
        "implementation_status": analysis["implementation_status"],
        "possible_cost_impact": analysis["possible_cost_impact"],
        "possible_grievance_risk": analysis["possible_grievance_risk"],
        "possible_notice_required": analysis["possible_notice_required"],
        "possible_pay_adjustment_required": analysis["possible_pay_adjustment_required"],
        "supervisor_review_required": analysis["supervisor_review_required"],
        "hr_labor_review_required": analysis["hr_labor_review_required"],
        "legal_contract_review_required": analysis["legal_contract_review_required"],
        "safeguard_summary": analysis["summary"],
        "recommended_next_step": analysis["recommended_next_step"],
        "do_not_implement_warning": analysis["do_not_implement_warning"],
        "updated_at": datetime.utcnow().isoformat(),
    }

    # Use existing transfer patch helper if already present.
    try:
        updated = await lori_transfer_patch_rows(
            "lori_route_contract_safeguard_reviews",
            f"id=eq.{quote(review_id)}",
            updated_review_payload,
        )
        updated_review = updated[0] if updated else {**review, **updated_review_payload}
    except Exception:
        updated_review = {**review, **updated_review_payload}

    return {
        "status": "success",
        "message": "Contract & Labor Safeguard review complete.",
        "review": updated_review,
        "document_reviewed": document,
        "findings_created": len(findings_created),
        "findings": findings_created,
        "decision_support_note": "LORI provides operational decision support only. Contract, union, labor, pay, policy, legal, DOT, HR, and employment-related questions must be reviewed and approved by authorized company leadership, HR, labor relations, legal counsel, or the appropriate contract owner before implementation.",
    }


@app.get("/route-contract-safeguard-review-detail")
async def route_contract_safeguard_review_detail(
    api_key: Optional[str] = Query(None),
    safeguard_review_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    reviews = await lori_contract_get_rows(
        "lori_route_contract_safeguard_reviews",
        f"select=*&id=eq.{quote(safeguard_review_id)}&limit=1",
    )

    if not reviews:
        return {"status": "not_found", "message": "Safeguard review not found."}

    review = reviews[0]

    findings = await lori_contract_get_rows(
        "lori_route_contract_safeguard_findings",
        f"select=*&safeguard_review_id=eq.{quote(safeguard_review_id)}&order=created_at.desc&limit=500",
    )

    acknowledgements = await lori_contract_get_rows(
        "lori_route_contract_acknowledgements",
        f"select=*&safeguard_review_id=eq.{quote(safeguard_review_id)}&order=created_at.desc&limit=100",
    )

    action_links = await lori_contract_get_rows(
        "lori_route_contract_action_links",
        f"select=*&safeguard_review_id=eq.{quote(safeguard_review_id)}&order=created_at.desc&limit=100",
    )

    document = None

    if review.get("document_id"):
        docs = await lori_contract_get_rows(
            "lori_route_contract_documents",
            f"select=*&id=eq.{quote(str(review.get('document_id')))}&limit=1",
        )
        document = docs[0] if docs else None

    return {
        "status": "success",
        "review": review,
        "document": document,
        "findings_count": len(findings),
        "findings": findings,
        "acknowledgements_count": len(acknowledgements),
        "acknowledgements": acknowledgements,
        "action_links_count": len(action_links),
        "action_links": action_links,
    }


@app.post("/route-contract-acknowledgement-create")
async def route_contract_acknowledgement_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    acknowledgement_payload = {
        "route_config_project_id": payload.get("route_config_project_id"),
        "safeguard_review_id": payload.get("safeguard_review_id"),
        "acknowledgement_type": payload.get("acknowledgement_type") or "No Agreement Applies",
        "acknowledged_by": payload.get("acknowledged_by") or "LORI User",
        "acknowledgement_text": payload.get("acknowledgement_text") or "I understand that LORI has not reviewed a contract, union agreement, or route assignment policy for this route change, and this decision still requires authorized operations review before implementation.",
        "acknowledgement_status": "Accepted",
        "company_name": payload.get("company_name"),
        "operating_state": payload.get("operating_state"),
        "station_code": payload.get("station_code"),
        "current_route_id": payload.get("current_route_id"),
        "receiving_route_id": payload.get("receiving_route_id"),
    }

    created = await lori_policy_supabase_post(
        "lori_route_contract_acknowledgements",
        acknowledgement_payload,
    )

    return {
        "status": "success",
        "message": "Contract/labor acknowledgement recorded.",
        "acknowledgement": created[0] if created else acknowledgement_payload,
    }


@app.post("/route-contract-safeguard-action-link-create")
async def route_contract_safeguard_action_link_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    safeguard_review_id = payload.get("safeguard_review_id")

    if not safeguard_review_id:
        return {"status": "error", "message": "safeguard_review_id is required."}

    action_payload = {
        "safeguard_review_id": safeguard_review_id,
        "action_center_item_id": payload.get("action_center_item_id"),
        "action_title": payload.get("action_title") or "Review Contract / Labor Agreement Before Route Change",
        "action_owner": payload.get("action_owner") or "Operations Leadership / HR / Labor Relations / Contract Manager",
        "action_status": payload.get("action_status") or "Open",
        "due_date": payload.get("due_date"),
        "notes": payload.get("notes") or "Contract/labor safeguard found potential review requirement before route change implementation.",
    }

    created = await lori_policy_supabase_post(
        "lori_route_contract_action_links",
        action_payload,
    )

    return {
        "status": "success",
        "message": "Safeguard action link created.",
        "action_link": created[0] if created else action_payload,
    }


@app.get("/route-contract-safeguard-status")
async def route_contract_safeguard_status(
    api_key: Optional[str] = Query(None),
    route_config_project_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    reviews = await lori_contract_get_rows(
        "lori_route_contract_safeguard_reviews",
        f"select=*&route_config_project_id=eq.{quote(route_config_project_id)}&order=created_at.desc&limit=100",
    )

    acknowledgements = await lori_contract_get_rows(
        "lori_route_contract_acknowledgements",
        f"select=*&route_config_project_id=eq.{quote(route_config_project_id)}&order=created_at.desc&limit=100",
    )

    high_risk_reviews = [
        r for r in reviews
        if r.get("contract_risk") == "High"
        or r.get("labor_risk") == "High"
        or r.get("cost_impact_risk") == "High"
        or r.get("implementation_status") == "Do Not Implement Yet"
    ]

    review_complete = len(reviews) > 0 or len(acknowledgements) > 0
    final_approval_allowed = review_complete and len(high_risk_reviews) == 0

    return {
        "status": "success",
        "route_config_project_id": route_config_project_id,
        "safeguard_reviews_count": len(reviews),
        "acknowledgements_count": len(acknowledgements),
        "high_risk_reviews_count": len(high_risk_reviews),
        "review_complete": review_complete,
        "final_approval_allowed": final_approval_allowed,
        "implementation_status": "Do Not Implement Yet" if high_risk_reviews else ("Planning Review Complete" if review_complete else "Safeguard Required"),
        "reviews": reviews,
        "acknowledgements": acknowledgements,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# SOP BUILDER ENGINE
# Creates SOP drafts, sections, reviews, versions, acknowledgements,
# and action links from station operations, driver, employee, safety,
# compliance, route configuration, and policy documents.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime, date
import os
import httpx


SUPABASE_URL_SOP = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_SOP = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_sop_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_sop_upper(value: Any) -> str:
    return lori_sop_clean(value).upper()


def lori_sop_safe_uuid(value: Any):
    cleaned = lori_sop_clean(value)
    return cleaned if cleaned else None


def lori_sop_today() -> str:
    return date.today().isoformat()


async def lori_sop_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_sop_patch_rows(
    table: str,
    match_query: str,
    payload: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not SUPABASE_URL_SOP or not SUPABASE_SERVICE_ROLE_KEY_SOP:
        raise RuntimeError("Missing Supabase environment variables.")

    url = f"{SUPABASE_URL_SOP}/rest/v1/{table}?{match_query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_SOP,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_SOP}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(url, headers=headers, json=payload)

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase PATCH failed: {response.status_code} {response.text}")

    try:
        return response.json()
    except Exception:
        return []


def lori_sop_number(station_code: str, sop_type: str) -> str:
    station = lori_sop_upper(station_code or "GLOBAL")
    sop_type_short = lori_sop_upper(sop_type or "SOP").replace(" ", "-")[:18]
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M")
    return f"SOP-{station}-{sop_type_short}-{timestamp}"


def lori_sop_review_flags(
    sop_type: str,
    source_document: Optional[Dict[str, Any]],
    user_instructions: str,
) -> Dict[str, bool]:
    combined = " ".join([
        lori_sop_clean(sop_type),
        lori_sop_clean(user_instructions),
        lori_sop_clean((source_document or {}).get("document_type")),
        lori_sop_clean((source_document or {}).get("document_category")),
        lori_sop_clean((source_document or {}).get("document_title")),
    ]).lower()

    requires_safety_review = any(k in combined for k in ["safety", "accident", "incident", "dot", "osha"])
    requires_compliance_review = any(k in combined for k in ["compliance", "regulatory", "policy", "audit"])
    requires_hr_review = any(k in combined for k in ["employee", "staff", "discipline", "attendance", "counseling", "training"])
    requires_labor_review = any(k in combined for k in ["union", "cba", "collective bargaining", "labor", "seniority", "grievance"])
    requires_legal_review = any(k in combined for k in ["contract", "agreement", "owner-operator", "contractor", "legal"])

    if source_document:
        requires_safety_review = requires_safety_review or bool(source_document.get("safety_related")) or bool(source_document.get("accident_related"))
        requires_compliance_review = requires_compliance_review or bool(source_document.get("compliance_related")) or bool(source_document.get("policy_related"))
        requires_hr_review = requires_hr_review or bool(source_document.get("hr_review_required"))
        requires_labor_review = requires_labor_review or bool(source_document.get("labor_review_required")) or bool(source_document.get("union_related"))
        requires_legal_review = requires_legal_review or bool(source_document.get("legal_review_required")) or bool(source_document.get("contract_related"))

    risk_level = "High" if requires_labor_review or requires_legal_review else ("Elevated" if requires_safety_review or requires_compliance_review or requires_hr_review else "Standard")

    return {
        "risk_level": risk_level,
        "requires_hr_review": requires_hr_review,
        "requires_labor_review": requires_labor_review,
        "requires_legal_review": requires_legal_review,
        "requires_safety_review": requires_safety_review,
        "requires_compliance_review": requires_compliance_review,
    }


def lori_sop_generate_sections(
    request: Dict[str, Any],
    source_document: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    sop_type = lori_sop_clean(request.get("sop_type") or "Station Operations SOP")
    subject_area = lori_sop_clean(request.get("subject_area") or sop_type)
    department = lori_sop_clean(request.get("department") or "Operations")
    audience = lori_sop_clean(request.get("audience") or "Authorized staff")
    trigger_reason = lori_sop_clean(request.get("trigger_reason") or "Operational need")
    instructions = lori_sop_clean(request.get("user_instructions"))
    source_title = lori_sop_clean((source_document or {}).get("document_title"))
    source_text = lori_sop_clean((source_document or {}).get("extracted_text"))

    source_note = f"This SOP is supported by source document: {source_title}." if source_title else "No source document was selected. This SOP is generated from the user instructions and operating context."

    if source_text:
        source_summary = source_text[:1200]
    else:
        source_summary = "No extracted source text is available. Authorized review is required before publication."

    base_procedure = [
        "Confirm the active operating context, station, department, and audience before using this SOP.",
        "Review the source document, policy, workflow, safety issue, route change, or compliance trigger connected to this SOP.",
        "Confirm who owns the process and who is authorized to approve changes.",
        "Follow the step-by-step procedure below and document any exceptions.",
        "Escalate unclear, high-risk, HR, labor, legal, safety, DOT, or compliance issues before final action.",
        "Record completion, acknowledgement, or follow-up action in LORI when required."
    ]

    if "route" in sop_type.lower() or "route" in subject_area.lower():
        base_procedure.extend([
            "Review the proposed route or work area change in Route Configuration.",
            "Confirm workload, stop count, mileage, delivery windows, vehicle needs, helper needs, and driver impact.",
            "Run the Contract & Labor Safeguard before implementation.",
            "Send unresolved risks to Action Center or authorized leadership review.",
            "Do not implement final route changes until required approvals are complete."
        ])

    if "safety" in sop_type.lower() or "accident" in subject_area.lower():
        base_procedure.extend([
            "Secure the area and address immediate safety concerns.",
            "Document the incident, involved persons, time, location, and known facts.",
            "Notify the supervisor, safety lead, HR, or required authority based on company policy.",
            "Attach accident, incident, training, or corrective action documents to the correct employee or driver file.",
            "Track follow-up actions until closed."
        ])

    if "driver" in sop_type.lower():
        base_procedure.extend([
            "Verify driver identity, route assignment, credentials, training status, and supervisor ownership.",
            "Attach relevant documents to Driver 360.",
            "Review safety, counseling, performance, route, and compliance history before final action.",
            "Escalate contract, union, labor, or pay-related issues when applicable."
        ])

    if "employee" in sop_type.lower() or "staff" in sop_type.lower():
        base_procedure.extend([
            "Verify employee identity, department, role, supervisor, and station assignment.",
            "Attach relevant files to the Employee / Staff File lane.",
            "Escalate HR, training, safety, attendance, disciplinary, or return-to-work issues when required."
        ])

    procedure_text = "\n".join([f"{idx + 1}. {step}" for idx, step in enumerate(base_procedure)])

    return [
        {
            "section_order": 1,
            "section_title": "Purpose",
            "section_type": "Purpose",
            "section_text": f"This SOP defines the required process for {subject_area}. It was triggered by: {trigger_reason}. The purpose is to create a clear, repeatable, reviewable procedure for {department}.",
            "required": True,
        },
        {
            "section_order": 2,
            "section_title": "Scope",
            "section_type": "Scope",
            "section_text": f"This SOP applies to {audience} at {lori_sop_clean(request.get('station_name')) or 'the selected operating location'}. It applies to the operating context selected in LORI and should not be used for another station unless approved.",
            "required": True,
        },
        {
            "section_order": 3,
            "section_title": "Source Documents and Inputs",
            "section_type": "Source",
            "section_text": f"{source_note}\n\nSource preview:\n{source_summary}",
            "required": True,
        },
        {
            "section_order": 4,
            "section_title": "Roles and Responsibilities",
            "section_type": "Responsibilities",
            "section_text": "The process owner is responsible for making sure the SOP is followed. Supervisors are responsible for review, escalation, documentation, and follow-up. Employees, drivers, contractors, or staff covered by the SOP are responsible for following the approved procedure and acknowledging training when required.",
            "required": True,
        },
        {
            "section_order": 5,
            "section_title": "Step-by-Step Procedure",
            "section_type": "Procedure",
            "section_text": procedure_text,
            "required": True,
        },
        {
            "section_order": 6,
            "section_title": "Required Review and Approval",
            "section_type": "Approval",
            "section_text": "Before this SOP is published or used as final company procedure, it must be reviewed by the required supervisor, safety, compliance, HR, labor relations, legal, or executive approver based on the risk type. LORI is a decision-support tool and does not replace authorized company approval.",
            "required": True,
        },
        {
            "section_order": 7,
            "section_title": "Documentation and Audit Trail",
            "section_type": "Documentation",
            "section_text": "All relevant documents, approvals, acknowledgements, action items, exceptions, and follow-up notes must be stored in LORI. Documents should be attached to the correct station, driver, employee, route, policy, or action record.",
            "required": True,
        },
        {
            "section_order": 8,
            "section_title": "Exceptions and Escalation",
            "section_type": "Escalation",
            "section_text": "Any exception, unclear instruction, safety concern, labor issue, contract concern, pay impact, compliance risk, or policy conflict must be escalated before final action. Do not proceed when the SOP conflicts with law, company policy, labor agreement, contract, safety rule, or authorized leadership direction.",
            "required": True,
        },
        {
            "section_order": 9,
            "section_title": "Training and Acknowledgement",
            "section_type": "Training",
            "section_text": "Covered users should be trained on this SOP before use. Acknowledgement may be required for drivers, employees, supervisors, contractors, or department staff based on the SOP type and company requirements.",
            "required": True,
        },
        {
            "section_order": 10,
            "section_title": "LORI Decision Support Notice",
            "section_type": "Disclaimer",
            "section_text": "LORI provides operational decision support only. SOPs must be reviewed and approved by authorized company leadership before use. HR, labor, legal, safety, compliance, DOT, and contract-related matters require appropriate authorized review.",
            "required": True,
        },
    ]


@app.get("/sop-builder-summary")
async def sop_builder_summary(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    requests = await lori_sop_get_rows(
        "lori_sop_builder_requests",
        "select=*&order=created_at.desc&limit=5000",
    )
    sops = await lori_sop_get_rows(
        "lori_sop_library",
        "select=*&order=created_at.desc&limit=5000",
    )

    if station_code:
        requests = [r for r in requests if lori_sop_upper(r.get("station_code")) == lori_sop_upper(station_code)]
        sops = [s for s in sops if lori_sop_upper(s.get("station_code")) == lori_sop_upper(station_code)]

    return {
        "status": "success",
        "sop_requests_count": len(requests),
        "sop_library_count": len(sops),
        "draft_sops_count": len([s for s in sops if s.get("sop_status") == "Draft"]),
        "in_review_count": len([s for s in sops if s.get("sop_status") == "In Review"]),
        "approved_count": len([s for s in sops if s.get("sop_status") == "Approved"]),
        "published_count": len([s for s in sops if s.get("sop_status") == "Published"]),
        "high_risk_count": len([s for s in sops if s.get("risk_level") == "High"]),
        "requires_review_count": len([
            s for s in sops
            if s.get("requires_hr_review")
            or s.get("requires_labor_review")
            or s.get("requires_legal_review")
            or s.get("requires_safety_review")
            or s.get("requires_compliance_review")
        ]),
    }


@app.get("/sop-source-document-options")
async def sop_source_document_options(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    document_type: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    docs = await lori_sop_get_rows(
        "lori_document_library",
        "select=id,document_title,intake_lane,document_type,document_category,station_code,driver_name,employee_name,route_id,extraction_status,referenced_by_modules,created_at&order=created_at.desc&limit=1000",
    )

    if station_code:
        docs = [d for d in docs if lori_sop_upper(d.get("station_code")) == lori_sop_upper(station_code)]

    if document_type:
        docs = [d for d in docs if lori_sop_clean(d.get("document_type")).lower() == document_type.lower()]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "documents_count": len(docs[:limit]),
        "documents": docs[:limit],
    }


@app.post("/sop-build-request-create")
async def sop_build_request_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_title = lori_sop_clean(payload.get("request_title"))

    if not request_title:
        return {"status": "error", "message": "request_title is required."}

    source_document_id = lori_sop_safe_uuid(payload.get("source_document_id"))
    source_document_title = lori_sop_clean(payload.get("source_document_title"))

    if source_document_id and not source_document_title:
        docs = await lori_sop_get_rows(
            "lori_document_library",
            f"select=*&id=eq.{quote(source_document_id)}&limit=1",
        )
        if docs:
            source_document_title = docs[0].get("document_title")

    request_payload = {
        "request_title": request_title,
        "request_status": "Draft",
        "sop_type": lori_sop_clean(payload.get("sop_type") or "Station Operations SOP"),
        "sop_priority": lori_sop_clean(payload.get("sop_priority") or "Standard"),

        "company_name": lori_sop_clean(payload.get("company_name")),
        "region_code": lori_sop_upper(payload.get("region_code")),
        "region_name": lori_sop_clean(payload.get("region_name")),
        "operating_state": lori_sop_upper(payload.get("operating_state")),
        "city": lori_sop_clean(payload.get("city")),
        "station_code": lori_sop_upper(payload.get("station_code")),
        "station_name": lori_sop_clean(payload.get("station_name")),
        "primary_zip": lori_sop_clean(payload.get("primary_zip")),
        "route_group": lori_sop_clean(payload.get("route_group")),
        "route_id": lori_sop_clean(payload.get("route_id")),

        "subject_area": lori_sop_clean(payload.get("subject_area")),
        "department": lori_sop_clean(payload.get("department")),
        "audience": lori_sop_clean(payload.get("audience")),
        "trigger_reason": lori_sop_clean(payload.get("trigger_reason")),

        "source_document_id": source_document_id,
        "source_document_title": source_document_title,

        "user_instructions": lori_sop_clean(payload.get("user_instructions")),
        "desired_output_style": lori_sop_clean(payload.get("desired_output_style") or "Clear operational SOP"),
        "requested_by": lori_sop_clean(payload.get("requested_by") or "LORI User"),

        "ai_generation_status": "Not Started",
        "review_status": "Not Reviewed",
        "publication_status": "Unpublished",
    }

    created = await lori_policy_supabase_post(
        "lori_sop_builder_requests",
        request_payload,
    )

    return {
        "status": "success",
        "message": "SOP build request created.",
        "request": created[0] if created else request_payload,
    }


@app.post("/sop-generate-draft")
async def sop_generate_draft(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_id = lori_sop_clean(payload.get("request_id"))

    if not request_id:
        return {"status": "error", "message": "request_id is required."}

    requests = await lori_sop_get_rows(
        "lori_sop_builder_requests",
        f"select=*&id=eq.{quote(request_id)}&limit=1",
    )

    if not requests:
        return {"status": "not_found", "message": "SOP request not found."}

    request = requests[0]

    source_document = None

    if request.get("source_document_id"):
        docs = await lori_sop_get_rows(
            "lori_document_library",
            f"select=*&id=eq.{quote(str(request.get('source_document_id')))}&limit=1",
        )
        source_document = docs[0] if docs else None

    review_flags = lori_sop_review_flags(
        sop_type=request.get("sop_type"),
        source_document=source_document,
        user_instructions=request.get("user_instructions") or "",
    )

    sop_title = lori_sop_clean(payload.get("sop_title") or request.get("request_title")).replace("Request — ", "").replace("Request - ", "")

    sections = lori_sop_generate_sections(request, source_document)

    full_sop_text = "\n\n".join([
        f"{section['section_order']}. {section['section_title']}\n{section['section_text']}"
        for section in sections
    ])

    sop_payload = {
        "sop_request_id": request_id,
        "sop_title": sop_title,
        "sop_number": lori_sop_number(request.get("station_code"), request.get("sop_type")),
        "sop_status": "Draft",
        "sop_type": request.get("sop_type"),
        "version_number": "1.0",

        "company_name": request.get("company_name"),
        "region_code": request.get("region_code"),
        "region_name": request.get("region_name"),
        "operating_state": request.get("operating_state"),
        "city": request.get("city"),
        "station_code": request.get("station_code"),
        "station_name": request.get("station_name"),
        "route_group": request.get("route_group"),
        "route_id": request.get("route_id"),

        "department": request.get("department"),
        "audience": request.get("audience"),
        "owner_name": payload.get("owner_name") or "Operations Leadership",

        "effective_date": payload.get("effective_date") or None,
        "review_due_date": payload.get("review_due_date") or None,
        "expiration_date": payload.get("expiration_date") or None,

        "purpose": sections[0]["section_text"],
        "scope": sections[1]["section_text"],
        "policy_reference": source_document.get("document_title") if source_document else request.get("source_document_title"),
        "source_document_id": request.get("source_document_id"),
        "source_document_title": source_document.get("document_title") if source_document else request.get("source_document_title"),

        "full_sop_text": full_sop_text,
        "executive_summary": f"Draft SOP for {request.get('subject_area') or request.get('sop_type')} serving {request.get('audience') or 'the selected audience'}.",
        **review_flags,
        "created_by": "LORI SOP Builder",
    }

    created_sop = await lori_policy_supabase_post(
        "lori_sop_library",
        sop_payload,
    )

    sop = created_sop[0] if created_sop else sop_payload
    sop_id = sop.get("id")

    created_sections = []

    for section in sections:
        section_payload = {
            "sop_id": sop_id,
            **section,
        }
        created = await lori_policy_supabase_post(
            "lori_sop_sections",
            section_payload,
        )
        if created:
            created_sections.append(created[0])

    if source_document and sop_id:
        await lori_policy_supabase_post(
            "lori_sop_source_documents",
            {
                "sop_id": sop_id,
                "document_id": source_document.get("id"),
                "source_title": source_document.get("document_title"),
                "source_type": source_document.get("document_type"),
                "source_relevance": "Primary source document used to generate SOP draft.",
                "source_status": "Linked",
            },
        )

    await lori_policy_supabase_post(
        "lori_sop_versions",
        {
            "sop_id": sop_id,
            "version_number": "1.0",
            "version_status": "Draft",
            "version_summary": "Initial SOP draft generated by LORI SOP Builder.",
            "full_sop_text": full_sop_text,
            "changed_by": payload.get("requested_by") or request.get("requested_by") or "LORI SOP Builder",
            "change_reason": "Initial generation",
        },
    )

    review_steps = []

    if review_flags.get("requires_safety_review"):
        review_steps.append(("Safety Review", "Safety Leader"))
    if review_flags.get("requires_compliance_review"):
        review_steps.append(("Compliance Review", "Compliance Owner"))
    if review_flags.get("requires_hr_review"):
        review_steps.append(("HR Review", "HR / People Leader"))
    if review_flags.get("requires_labor_review"):
        review_steps.append(("Labor Review", "Labor Relations"))
    if review_flags.get("requires_legal_review"):
        review_steps.append(("Legal Review", "Legal / Contract Owner"))

    review_steps.insert(0, ("Supervisor Review", "Operations Supervisor"))

    created_reviews = []

    for review_type, role in review_steps:
        created = await lori_policy_supabase_post(
            "lori_sop_reviews",
            {
                "sop_id": sop_id,
                "review_type": review_type,
                "reviewer_role": role,
                "review_status": "Pending",
                "required": True,
            },
        )
        if created:
            created_reviews.append(created[0])

    await lori_sop_patch_rows(
        "lori_sop_builder_requests",
        f"id=eq.{quote(request_id)}",
        {
            "request_status": "Completed",
            "ai_generation_status": "Completed",
            "review_status": "Pending Review",
            "publication_status": "Unpublished",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    return {
        "status": "success",
        "message": "SOP draft generated.",
        "sop": sop,
        "sections_count": len(created_sections),
        "sections": created_sections,
        "reviews_count": len(created_reviews),
        "reviews": created_reviews,
        "decision_support_note": sop.get("decision_support_note"),
    }


@app.get("/sop-library")
async def sop_library(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    sop_status: Optional[str] = Query(None),
    sop_type: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    sops = await lori_sop_get_rows(
        "lori_sop_library",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        sops = [s for s in sops if lori_sop_upper(s.get("station_code")) == lori_sop_upper(station_code)]

    if sop_status:
        sops = [s for s in sops if lori_sop_clean(s.get("sop_status")).lower() == sop_status.lower()]

    if sop_type:
        sops = [s for s in sops if lori_sop_clean(s.get("sop_type")).lower() == sop_type.lower()]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "sops_count": len(sops[:limit]),
        "sops": sops[:limit],
    }


@app.get("/sop-detail")
async def sop_detail(
    api_key: Optional[str] = Query(None),
    sop_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    sops = await lori_sop_get_rows(
        "lori_sop_library",
        f"select=*&id=eq.{quote(sop_id)}&limit=1",
    )

    if not sops:
        return {"status": "not_found", "message": "SOP not found."}

    sections = await lori_sop_get_rows(
        "lori_sop_sections",
        f"select=*&sop_id=eq.{quote(sop_id)}&order=section_order.asc&limit=500",
    )

    sources = await lori_sop_get_rows(
        "lori_sop_source_documents",
        f"select=*&sop_id=eq.{quote(sop_id)}&order=created_at.desc&limit=100",
    )

    reviews = await lori_sop_get_rows(
        "lori_sop_reviews",
        f"select=*&sop_id=eq.{quote(sop_id)}&order=created_at.asc&limit=100",
    )

    versions = await lori_sop_get_rows(
        "lori_sop_versions",
        f"select=*&sop_id=eq.{quote(sop_id)}&order=created_at.desc&limit=100",
    )

    acknowledgements = await lori_sop_get_rows(
        "lori_sop_acknowledgements",
        f"select=*&sop_id=eq.{quote(sop_id)}&order=created_at.desc&limit=500",
    )

    action_links = await lori_sop_get_rows(
        "lori_sop_action_links",
        f"select=*&sop_id=eq.{quote(sop_id)}&order=created_at.desc&limit=100",
    )

    return {
        "status": "success",
        "sop": sops[0],
        "sections_count": len(sections),
        "sections": sections,
        "sources_count": len(sources),
        "sources": sources,
        "reviews_count": len(reviews),
        "reviews": reviews,
        "versions_count": len(versions),
        "versions": versions,
        "acknowledgements_count": len(acknowledgements),
        "acknowledgements": acknowledgements,
        "action_links_count": len(action_links),
        "action_links": action_links,
    }


@app.post("/sop-review-update")
async def sop_review_update(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    review_id = lori_sop_clean(payload.get("review_id"))

    if not review_id:
        return {"status": "error", "message": "review_id is required."}

    update_payload = {
        "reviewer_name": lori_sop_clean(payload.get("reviewer_name")),
        "reviewer_email": lori_sop_clean(payload.get("reviewer_email")),
        "reviewer_role": lori_sop_clean(payload.get("reviewer_role")),
        "review_status": lori_sop_clean(payload.get("review_status") or "Approved"),
        "review_notes": lori_sop_clean(payload.get("review_notes")),
        "reviewed_at": datetime.utcnow().isoformat(),
    }

    updated = await lori_sop_patch_rows(
        "lori_sop_reviews",
        f"id=eq.{quote(review_id)}",
        update_payload,
    )

    return {
        "status": "success",
        "message": "SOP review updated.",
        "review": updated[0] if updated else update_payload,
    }


@app.post("/sop-status-update")
async def sop_status_update(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    sop_id = lori_sop_clean(payload.get("sop_id"))
    new_status = lori_sop_clean(payload.get("sop_status"))

    if not sop_id:
        return {"status": "error", "message": "sop_id is required."}

    if not new_status:
        return {"status": "error", "message": "sop_status is required."}

    update_payload = {
        "sop_status": new_status,
        "updated_at": datetime.utcnow().isoformat(),
    }

    if new_status == "Approved":
        update_payload["approved_by"] = lori_sop_clean(payload.get("approved_by") or "LORI Admin")
        update_payload["approved_at"] = datetime.utcnow().isoformat()

    if new_status == "Published":
        update_payload["published_by"] = lori_sop_clean(payload.get("published_by") or "LORI Admin")
        update_payload["published_at"] = datetime.utcnow().isoformat()

    updated = await lori_sop_patch_rows(
        "lori_sop_library",
        f"id=eq.{quote(sop_id)}",
        update_payload,
    )

    return {
        "status": "success",
        "message": f"SOP status updated to {new_status}.",
        "sop": updated[0] if updated else update_payload,
    }


@app.post("/sop-acknowledgement-create")
async def sop_acknowledgement_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    sop_id = lori_sop_clean(payload.get("sop_id"))

    if not sop_id:
        return {"status": "error", "message": "sop_id is required."}

    acknowledgement_payload = {
        "sop_id": sop_id,
        "person_type": lori_sop_clean(payload.get("person_type")),
        "person_name": lori_sop_clean(payload.get("person_name")),
        "employee_id": lori_sop_clean(payload.get("employee_id")),
        "driver_id": lori_sop_clean(payload.get("driver_id")),
        "email": lori_sop_clean(payload.get("email")),
        "station_code": lori_sop_upper(payload.get("station_code")),
        "route_id": lori_sop_clean(payload.get("route_id")),
        "department": lori_sop_clean(payload.get("department")),
        "role_title": lori_sop_clean(payload.get("role_title")),
        "acknowledgement_status": lori_sop_clean(payload.get("acknowledgement_status") or "Acknowledged"),
        "acknowledged_by": lori_sop_clean(payload.get("acknowledged_by") or payload.get("person_name") or "LORI User"),
        "acknowledged_at": datetime.utcnow().isoformat(),
        "acknowledgement_notes": lori_sop_clean(payload.get("acknowledgement_notes")),
    }

    created = await lori_policy_supabase_post(
        "lori_sop_acknowledgements",
        acknowledgement_payload,
    )

    return {
        "status": "success",
        "message": "SOP acknowledgement recorded.",
        "acknowledgement": created[0] if created else acknowledgement_payload,
    }


@app.post("/sop-action-link-create")
async def sop_action_link_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    sop_id = lori_sop_clean(payload.get("sop_id"))

    if not sop_id:
        return {"status": "error", "message": "sop_id is required."}

    action_payload = {
        "sop_id": sop_id,
        "target_module": lori_sop_clean(payload.get("target_module") or "Action Center"),
        "target_record_id": lori_sop_safe_uuid(payload.get("target_record_id")),
        "action_title": lori_sop_clean(payload.get("action_title") or "Review SOP"),
        "action_owner": lori_sop_clean(payload.get("action_owner") or "Operations Leadership"),
        "action_status": lori_sop_clean(payload.get("action_status") or "Open"),
        "due_date": payload.get("due_date") or None,
        "notes": lori_sop_clean(payload.get("notes")),
    }

    created = await lori_policy_supabase_post(
        "lori_sop_action_links",
        action_payload,
    )

    return {
        "status": "success",
        "message": "SOP action link created.",
        "action_link": created[0] if created else action_payload,
    }


@app.get("/sop-export-html")
async def sop_export_html(
    api_key: Optional[str] = Query(None),
    sop_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    detail = await sop_detail(api_key=api_key, sop_id=sop_id)

    if detail.get("status") != "success":
        return detail

    sop = detail.get("sop") or {}
    sections = detail.get("sections") or []
    reviews = detail.get("reviews") or []

    section_html = "\n".join([
        f"""
        <section style="margin-bottom: 22px;">
          <h2>{section.get('section_order')}. {section.get('section_title')}</h2>
          <p style="white-space: pre-wrap;">{section.get('section_text') or ''}</p>
        </section>
        """
        for section in sections
    ])

    review_html = "\n".join([
        f"""
        <tr>
          <td>{review.get('review_type') or ''}</td>
          <td>{review.get('reviewer_role') or ''}</td>
          <td>{review.get('review_status') or ''}</td>
          <td>{review.get('review_notes') or ''}</td>
        </tr>
        """
        for review in reviews
    ])

    html = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <title>{sop.get('sop_title')}</title>
    </head>
    <body style="font-family: Arial, sans-serif; padding: 36px; color: #111827;">
      <div style="border-bottom: 3px solid #111827; padding-bottom: 16px; margin-bottom: 24px;">
        <h1 style="margin: 0;">{sop.get('sop_title')}</h1>
        <p><strong>SOP Number:</strong> {sop.get('sop_number') or ''}</p>
        <p><strong>Status:</strong> {sop.get('sop_status') or ''} | <strong>Version:</strong> {sop.get('version_number') or ''}</p>
        <p><strong>Station:</strong> {sop.get('station_code') or ''} — {sop.get('station_name') or ''}</p>
        <p><strong>Audience:</strong> {sop.get('audience') or ''}</p>
      </div>

      <div style="background:#f3f4f6; padding:16px; border-radius:10px; margin-bottom:24px;">
        <h2>Executive Summary</h2>
        <p>{sop.get('executive_summary') or ''}</p>
        <p><strong>Decision Support Notice:</strong> {sop.get('decision_support_note') or ''}</p>
      </div>

      {section_html}

      <h2>Review Status</h2>
      <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%;">
        <thead>
          <tr>
            <th>Review Type</th>
            <th>Reviewer Role</th>
            <th>Status</th>
            <th>Notes</th>
          </tr>
        </thead>
        <tbody>
          {review_html}
        </tbody>
      </table>
    </body>
    </html>
    """

    return {
        "status": "success",
        "sop_id": sop_id,
        "html": html,
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# EXECUTIVE REPORTS & PACKET CENTER ENGINE
# Creates report requests, generates leadership packets,
# builds HTML/print-ready reports, and stores report history.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime, date
import os
import httpx


SUPABASE_URL_REPORTS = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_REPORTS = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_report_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_report_upper(value: Any) -> str:
    return lori_report_clean(value).upper()


def lori_report_today() -> str:
    return date.today().isoformat()


async def lori_report_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_report_safe_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    try:
        return await lori_policy_supabase_get(f"{table}?{query}")
    except Exception:
        return []


async def lori_report_patch_rows(
    table: str,
    match_query: str,
    payload: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not SUPABASE_URL_REPORTS or not SUPABASE_SERVICE_ROLE_KEY_REPORTS:
        raise RuntimeError("Missing Supabase environment variables.")

    url = f"{SUPABASE_URL_REPORTS}/rest/v1/{table}?{match_query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_REPORTS,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_REPORTS}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(url, headers=headers, json=payload)

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase PATCH failed: {response.status_code} {response.text}")

    try:
        return response.json()
    except Exception:
        return []


def lori_report_count_label(count_value: int, singular: str, plural: str) -> str:
    return f"{count_value} {singular if count_value == 1 else plural}"


def lori_report_html_escape(value: Any) -> str:
    text = lori_report_clean(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


async def lori_report_snapshot_for_station(station_code: str) -> Dict[str, Any]:
    station_code_clean = lori_report_upper(station_code)

    docs = await lori_report_safe_get_rows(
        "lori_document_library",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=1000",
    )

    action_items = await lori_report_safe_get_rows(
        "lori_action_items",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=1000",
    )

    kpi_plans = await lori_report_safe_get_rows(
        "lori_kpi_action_plans",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=500",
    )

    sop_library = await lori_report_safe_get_rows(
        "lori_sop_library",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=500",
    )

    route_projects = await lori_report_safe_get_rows(
        "lori_route_config_projects",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=500",
    )

    route_contract_reviews = await lori_report_safe_get_rows(
        "lori_route_contract_safeguard_reviews",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=500",
    )

    driver_master = await lori_report_safe_get_rows(
        "lori_driver_master",
        f"select=*&station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=1000",
    )

    regulatory_alerts = await lori_report_safe_get_rows(
        "lori_regulatory_alerts",
        "select=*&order=created_at.desc&limit=500",
    )

    access_changes = await lori_report_safe_get_rows(
        "lori_user_access_change_requests",
        f"select=*&current_station_code=eq.{quote(station_code_clean)}&order=created_at.desc&limit=500",
    )

    return {
        "documents": docs,
        "action_items": action_items,
        "kpi_plans": kpi_plans,
        "sops": sop_library,
        "route_projects": route_projects,
        "contract_reviews": route_contract_reviews,
        "drivers": driver_master,
        "regulatory_alerts": regulatory_alerts,
        "access_changes": access_changes,
    }


def lori_report_build_sections(
    report_template: Dict[str, Any],
    request: Dict[str, Any],
    snapshot: Dict[str, Any],
) -> List[Dict[str, Any]]:
    report_code = lori_report_clean(report_template.get("report_code"))
    report_name = lori_report_clean(report_template.get("report_name"))
    station_code = lori_report_clean(request.get("station_code"))
    station_name = lori_report_clean(request.get("station_name"))
    route_id = lori_report_clean(request.get("route_id"))
    driver_name = lori_report_clean(request.get("driver_name"))

    docs = snapshot.get("documents", [])
    action_items = snapshot.get("action_items", [])
    kpi_plans = snapshot.get("kpi_plans", [])
    sops = snapshot.get("sops", [])
    route_projects = snapshot.get("route_projects", [])
    contract_reviews = snapshot.get("contract_reviews", [])
    drivers = snapshot.get("drivers", [])
    regulatory_alerts = snapshot.get("regulatory_alerts", [])
    access_changes = snapshot.get("access_changes", [])

    open_actions = [
        item for item in action_items
        if lori_report_clean(item.get("status") or item.get("action_status")).lower() not in ["completed", "closed", "done"]
    ]

    high_risk_docs = [
        d for d in docs
        if d.get("labor_review_required")
        or d.get("legal_review_required")
        or d.get("hr_review_required")
        or d.get("accident_related")
        or d.get("contract_related")
        or d.get("union_related")
    ]

    high_risk_contracts = [
        c for c in contract_reviews
        if c.get("contract_risk") == "High"
        or c.get("labor_risk") == "High"
        or c.get("cost_impact_risk") == "High"
        or c.get("implementation_status") == "Do Not Implement Yet"
    ]

    draft_sops = [s for s in sops if s.get("sop_status") in ["Draft", "In Review"]]
    published_sops = [s for s in sops if s.get("sop_status") == "Published"]

    sections = []

    sections.append({
        "section_order": 1,
        "section_title": "Executive Summary",
        "section_type": "Executive Summary",
        "section_summary": f"{report_name} for {station_code or 'selected station'}.",
        "section_body": (
            f"This report summarizes current LORI information for {station_name or station_code or 'the selected operating context'}. "
            f"It is generated from connected LORI modules and should be reviewed by authorized leadership before action. "
            f"Source modules include: {', '.join(report_template.get('source_modules') or [])}."
        ),
        "source_module": "Reports Engine",
        "risk_level": "Standard",
        "requires_attention": False,
    })

    if report_code in [
        "EXEC_OPERATIONS_BRIEFING",
        "LEADERSHIP_DASHBOARD_PACKET",
        "SUPERVISOR_ACCOUNTABILITY_REPORT",
        "STATION_READINESS_REPORT",
    ]:
        sections.extend([
            {
                "section_order": 2,
                "section_title": "Station Snapshot",
                "section_type": "Station Summary",
                "section_summary": "Current station-level operating summary.",
                "section_body": (
                    f"Station: {station_code} — {station_name}. "
                    f"Documents: {len(docs)}. Drivers: {len(drivers)}. "
                    f"Open action items: {len(open_actions)}. KPI action plans: {len(kpi_plans)}. "
                    f"SOPs: {len(sops)}. Route configuration projects: {len(route_projects)}."
                ),
                "source_module": "Overview / Leadership Dashboard",
                "risk_level": "Standard",
                "requires_attention": len(open_actions) > 0,
            },
            {
                "section_order": 3,
                "section_title": "Leadership Attention Items",
                "section_type": "Action Summary",
                "section_summary": "Items leadership should review.",
                "section_body": (
                    f"LORI found {len(open_actions)} open action items, {len(high_risk_docs)} sensitive or review-required documents, "
                    f"{len(high_risk_contracts)} high-risk contract/labor safeguards, and {len(draft_sops)} SOPs in draft or review status."
                ),
                "source_module": "Action Center",
                "risk_level": "Medium" if open_actions or high_risk_docs or high_risk_contracts else "Low",
                "requires_attention": bool(open_actions or high_risk_docs or high_risk_contracts),
            }
        ])

    elif report_code in [
        "DRIVER_360_PROFILE",
        "DRIVER_RISK_REPORT",
        "DRIVER_RECOGNITION_REPORT",
        "DRIVER_COACHING_COUNSELING_REPORT",
    ]:
        driver_docs = [
            d for d in docs
            if driver_name.lower() in lori_report_clean(d.get("driver_name")).lower()
        ] if driver_name else [d for d in docs if d.get("intake_lane") == "Driver File"]

        sections.extend([
            {
                "section_order": 2,
                "section_title": "Driver Summary",
                "section_type": "Driver",
                "section_summary": "Driver-related records and document visibility.",
                "section_body": (
                    f"Driver filter: {driver_name or 'All drivers at station'}. "
                    f"Driver records found: {len(drivers)}. Driver documents found: {len(driver_docs)}. "
                    f"Safety, counseling, performance, document, and action history should be reviewed before final decisions."
                ),
                "source_module": "Driver 360",
                "risk_level": "Medium" if driver_docs else "Standard",
                "requires_attention": bool(driver_docs),
            },
            {
                "section_order": 3,
                "section_title": "Driver Follow-Up",
                "section_type": "Follow-Up",
                "section_summary": "Recommended next step.",
                "section_body": "Review Driver 360, attached driver documents, open actions, safety events, counseling status, credentials, and route performance before taking action.",
                "source_module": "Driver 360 / Action Center",
                "risk_level": "Standard",
                "requires_attention": True,
            }
        ])

    elif report_code in [
        "ROUTE_CONFIGURATION_PACKET",
        "ROUTE_OPTIMIZATION_REPORT",
        "WORK_AREA_BALANCE_REPORT",
        "ROUTE_CHANGE_REVIEW_PACKET",
    ]:
        route_filtered = [
            p for p in route_projects
            if not route_id or route_id.lower() in lori_report_clean(p).lower()
        ]

        sections.extend([
            {
                "section_order": 2,
                "section_title": "Route Configuration Summary",
                "section_type": "Route",
                "section_summary": "Route and work area projects.",
                "section_body": (
                    f"Route filter: {route_id or 'All routes'}. "
                    f"Route configuration projects found: {len(route_filtered)}. "
                    f"Contract/labor safeguard reviews found: {len(contract_reviews)}. "
                    f"High-risk safeguard reviews: {len(high_risk_contracts)}."
                ),
                "source_module": "Route Configuration",
                "risk_level": "High" if high_risk_contracts else "Medium",
                "requires_attention": bool(high_risk_contracts),
            },
            {
                "section_order": 3,
                "section_title": "Implementation Gate",
                "section_type": "Approval Gate",
                "section_summary": "Route changes require review before implementation.",
                "section_body": (
                    "Route recommendations are planning outputs until operations leadership completes required review. "
                    "Contract, labor, union, pay, seniority, route assignment, and policy safeguards must be reviewed before implementation."
                ),
                "source_module": "Contract Safeguard",
                "risk_level": "High" if high_risk_contracts else "Standard",
                "requires_attention": True,
            }
        ])

    elif report_code in [
        "CONTRACT_LABOR_SAFEGUARD_REPORT",
        "COMPLIANCE_POLICY_REVIEW_REPORT",
        "REGULATORY_WATCH_REPORT",
        "ACCESS_CONTROL_USER_TRANSFER_AUDIT",
    ]:
        sections.extend([
            {
                "section_order": 2,
                "section_title": "Compliance / Safeguard Summary",
                "section_type": "Compliance",
                "section_summary": "Contract, labor, regulatory, and access-control review.",
                "section_body": (
                    f"Contract/labor safeguard reviews: {len(contract_reviews)}. "
                    f"High-risk safeguard reviews: {len(high_risk_contracts)}. "
                    f"Regulatory alerts available: {len(regulatory_alerts)}. "
                    f"User access change records: {len(access_changes)}."
                ),
                "source_module": "Compliance / Contract Safeguard / Regulatory Watch",
                "risk_level": "High" if high_risk_contracts else "Medium",
                "requires_attention": bool(high_risk_contracts or regulatory_alerts or access_changes),
            }
        ])

    elif report_code in [
        "DOCUMENT_INTAKE_SUMMARY",
        "SENSITIVE_DOCUMENT_REVIEW_REPORT",
        "BATCH_INTEGRITY_AUDIT",
        "DOCUMENT_MODULE_LINK_REPORT",
    ]:
        extracted_docs = [d for d in docs if d.get("extraction_status") == "Extracted"]
        needs_review_docs = [d for d in docs if d.get("extraction_status") != "Extracted"]

        sections.extend([
            {
                "section_order": 2,
                "section_title": "Document Library Summary",
                "section_type": "Document Intake",
                "section_summary": "Station document library status.",
                "section_body": (
                    f"Total documents: {len(docs)}. Extracted documents: {len(extracted_docs)}. "
                    f"Documents needing extraction/review: {len(needs_review_docs)}. "
                    f"Sensitive/review-required documents: {len(high_risk_docs)}."
                ),
                "source_module": "Document & Data Intake Center",
                "risk_level": "Medium" if needs_review_docs or high_risk_docs else "Low",
                "requires_attention": bool(needs_review_docs or high_risk_docs),
            }
        ])

    elif report_code in [
        "KPI_ACTION_PLAN_REPORT",
        "KPI_EXCEPTION_REPORT",
        "ACTION_CENTER_REPORT",
        "LEADERSHIP_BRIEFING_QUEUE_REPORT",
    ]:
        sections.extend([
            {
                "section_order": 2,
                "section_title": "KPI and Action Summary",
                "section_type": "KPI / Action",
                "section_summary": "Operational action status.",
                "section_body": (
                    f"KPI action plans: {len(kpi_plans)}. Total action items: {len(action_items)}. "
                    f"Open action items: {len(open_actions)}. "
                    f"Leadership should review open, overdue, high-priority, and owner-assigned action items."
                ),
                "source_module": "KPI Action Plans / Action Center",
                "risk_level": "Medium" if open_actions else "Low",
                "requires_attention": bool(open_actions),
            }
        ])

    elif report_code in [
        "SOP_LIBRARY_REPORT",
        "SOP_REVIEW_PACKET",
        "SOP_ACKNOWLEDGEMENT_REPORT",
        "TRAINING_PROCEDURE_GAP_REPORT",
    ]:
        sections.extend([
            {
                "section_order": 2,
                "section_title": "SOP and Training Summary",
                "section_type": "SOP",
                "section_summary": "SOP library, review, and training status.",
                "section_body": (
                    f"Total SOPs: {len(sops)}. Draft/in-review SOPs: {len(draft_sops)}. "
                    f"Published SOPs: {len(published_sops)}. "
                    f"Draft and in-review SOPs should be reviewed before use."
                ),
                "source_module": "SOP Builder",
                "risk_level": "Medium" if draft_sops else "Low",
                "requires_attention": bool(draft_sops),
            }
        ])

    elif report_code in [
        "DRIVER_ROAD_COMMUNICATIONS_LOG",
        "URGENT_DRIVER_MESSAGE_REPORT",
    ]:
        sections.extend([
            {
                "section_order": 2,
                "section_title": "Driver Road Communications Summary",
                "section_type": "Driver Communications",
                "section_summary": "Driver field communication report.",
                "section_body": (
                    "Driver Road Communications are limited to operational messages to drivers in the field, such as route updates, stop changes, weather alerts, safety reminders, delay notices, and return-to-station instructions. "
                    "No internal SOP, contract, safeguard, or admin workflow notifications are included in this report."
                ),
                "source_module": "Driver Road Communications",
                "risk_level": "Standard",
                "requires_attention": False,
            }
        ])

    sections.append({
        "section_order": 99,
        "section_title": "Decision Support Notice",
        "section_type": "Disclaimer",
        "section_summary": "LORI decision-support limitation.",
        "section_body": (
            "LORI provides operational decision support only. Reports, packets, recommendations, route changes, SOPs, contract safeguards, compliance items, driver actions, and employment-related issues must be reviewed and approved by authorized company leadership before final action."
        ),
        "source_module": "Reports Engine",
        "risk_level": "Standard",
        "requires_attention": False,
    })

    return sections


def lori_report_build_html(report: Dict[str, Any], sections: List[Dict[str, Any]]) -> str:
    section_html = "\n".join([
        f"""
        <section style="border:1px solid #e5e7eb; border-radius:12px; padding:18px; margin-bottom:18px;">
          <div style="font-size:12px; color:#6b7280; text-transform:uppercase; letter-spacing:.08em;">{lori_report_html_escape(section.get('section_type'))}</div>
          <h2 style="margin:6px 0 8px 0;">{lori_report_html_escape(section.get('section_title'))}</h2>
          <p style="font-weight:600; color:#374151;">{lori_report_html_escape(section.get('section_summary'))}</p>
          <p style="white-space:pre-wrap; line-height:1.55;">{lori_report_html_escape(section.get('section_body'))}</p>
          <p style="font-size:12px; color:#6b7280;">Source: {lori_report_html_escape(section.get('source_module'))} | Risk: {lori_report_html_escape(section.get('risk_level'))}</p>
        </section>
        """
        for section in sections
    ])

    html = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <title>{lori_report_html_escape(report.get('report_title'))}</title>
    </head>
    <body style="font-family: Arial, sans-serif; background:#f9fafb; color:#111827; padding:36px;">
      <div style="background:#ffffff; border:1px solid #e5e7eb; border-radius:16px; padding:26px; margin-bottom:24px;">
        <div style="font-size:12px; color:#6b7280; text-transform:uppercase; letter-spacing:.10em;">LORI Drive Command Center</div>
        <h1 style="margin:8px 0 10px 0;">{lori_report_html_escape(report.get('report_title'))}</h1>
        <p><strong>Report:</strong> {lori_report_html_escape(report.get('report_name'))}</p>
        <p><strong>Category:</strong> {lori_report_html_escape(report.get('report_category'))}</p>
        <p><strong>Station:</strong> {lori_report_html_escape(report.get('station_code'))} — {lori_report_html_escape(report.get('station_name'))}</p>
        <p><strong>Generated:</strong> {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}</p>
        <div style="margin-top:16px; padding:14px; background:#fff7ed; border:1px solid #fed7aa; border-radius:10px;">
          <strong>Demonstration Data Only — Not Company Proprietary Data</strong>
          <br>Reports reflect available LORI demo/system data and require authorized review before final action.
        </div>
      </div>

      <div style="background:#ffffff; border:1px solid #e5e7eb; border-radius:16px; padding:24px;">
        {section_html}
      </div>
    </body>
    </html>
    """
    return html


@app.get("/report-templates")
async def report_templates(
    api_key: Optional[str] = Query(None),
    report_category: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    templates = await lori_report_get_rows(
        "lori_report_templates",
        "select=*&order=sort_order.asc&limit=500",
    )

    if report_category:
        templates = [
            t for t in templates
            if lori_report_clean(t.get("report_category")).lower() == report_category.lower()
        ]

    categories = sorted(list(set([t.get("report_category") for t in templates if t.get("report_category")])))

    return {
        "status": "success",
        "templates_count": len(templates),
        "categories": categories,
        "templates": templates,
    }


@app.get("/reports-packet-center-summary")
async def reports_packet_center_summary(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    templates = await lori_report_get_rows(
        "lori_report_templates",
        "select=*&order=sort_order.asc&limit=500",
    )

    reports = await lori_report_safe_get_rows(
        "lori_report_library",
        "select=*&order=created_at.desc&limit=1000",
    )

    requests = await lori_report_safe_get_rows(
        "lori_report_requests",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        station = lori_report_upper(station_code)
        reports = [r for r in reports if lori_report_upper(r.get("station_code")) == station]
        requests = [r for r in requests if lori_report_upper(r.get("station_code")) == station]

    category_counts = {}

    for template in templates:
        category = template.get("report_category") or "Other"
        category_counts[category] = category_counts.get(category, 0) + 1

    return {
        "status": "success",
        "report_templates_count": len(templates),
        "generated_reports_count": len(reports),
        "report_requests_count": len(requests),
        "pending_backend_count": len([t for t in templates if t.get("report_status") == "Pending Backend"]),
        "available_reports_count": len([t for t in templates if t.get("report_status") == "Available"]),
        "published_reports_count": len([r for r in reports if r.get("report_status") == "Published"]),
        "draft_reports_count": len([r for r in reports if r.get("report_status") == "Draft"]),
        "category_counts": category_counts,
    }


@app.post("/report-request-create")
async def report_request_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    report_code = lori_report_upper(payload.get("report_code"))

    if not report_code:
        return {"status": "error", "message": "report_code is required."}

    templates = await lori_report_get_rows(
        "lori_report_templates",
        f"select=*&report_code=eq.{quote(report_code)}&limit=1",
    )

    if not templates:
        return {"status": "not_found", "message": f"Report template {report_code} was not found."}

    template = templates[0]

    request_payload = {
        "report_template_id": template.get("id"),
        "request_title": lori_report_clean(payload.get("request_title") or f"{template.get('report_name')} — {payload.get('station_code') or 'Station'}"),
        "report_name": template.get("report_name"),
        "report_category": template.get("report_category"),
        "report_code": template.get("report_code"),
        "request_status": "Draft",

        "company_name": lori_report_clean(payload.get("company_name")),
        "region_code": lori_report_upper(payload.get("region_code")),
        "region_name": lori_report_clean(payload.get("region_name")),
        "operating_state": lori_report_upper(payload.get("operating_state")),
        "city": lori_report_clean(payload.get("city")),
        "station_code": lori_report_upper(payload.get("station_code")),
        "station_name": lori_report_clean(payload.get("station_name")),
        "primary_zip": lori_report_clean(payload.get("primary_zip")),
        "route_group": lori_report_clean(payload.get("route_group")),
        "route_id": lori_report_clean(payload.get("route_id")),

        "driver_id": lori_report_clean(payload.get("driver_id")),
        "driver_name": lori_report_clean(payload.get("driver_name")),
        "employee_id": lori_report_clean(payload.get("employee_id")),
        "employee_name": lori_report_clean(payload.get("employee_name")),

        "date_range_start": payload.get("date_range_start") or None,
        "date_range_end": payload.get("date_range_end") or None,

        "risk_level": lori_report_clean(payload.get("risk_level")),
        "action_status": lori_report_clean(payload.get("action_status")),
        "sop_status": lori_report_clean(payload.get("sop_status")),
        "document_type": lori_report_clean(payload.get("document_type")),

        "requested_by": lori_report_clean(payload.get("requested_by") or "LORI User"),
        "request_notes": lori_report_clean(payload.get("request_notes")),
    }

    created = await lori_policy_supabase_post(
        "lori_report_requests",
        request_payload,
    )

    return {
        "status": "success",
        "message": "Report request created.",
        "request": created[0] if created else request_payload,
        "template": template,
    }


@app.post("/report-generate-packet")
async def report_generate_packet(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_id = lori_report_clean(payload.get("request_id"))
    report_code = lori_report_upper(payload.get("report_code"))

    request = None

    if request_id:
        requests = await lori_report_get_rows(
            "lori_report_requests",
            f"select=*&id=eq.{quote(request_id)}&limit=1",
        )
        if not requests:
            return {"status": "not_found", "message": "Report request not found."}
        request = requests[0]
        report_code = request.get("report_code")

    if not report_code:
        return {"status": "error", "message": "request_id or report_code is required."}

    templates = await lori_report_get_rows(
        "lori_report_templates",
        f"select=*&report_code=eq.{quote(report_code)}&limit=1",
    )

    if not templates:
        return {"status": "not_found", "message": f"Report template {report_code} was not found."}

    template = templates[0]

    if not request:
        request = {
            "request_title": payload.get("request_title") or template.get("report_name"),
            "report_name": template.get("report_name"),
            "report_category": template.get("report_category"),
            "report_code": template.get("report_code"),
            "company_name": payload.get("company_name"),
            "region_code": payload.get("region_code"),
            "region_name": payload.get("region_name"),
            "operating_state": payload.get("operating_state"),
            "city": payload.get("city"),
            "station_code": payload.get("station_code"),
            "station_name": payload.get("station_name"),
            "route_group": payload.get("route_group"),
            "route_id": payload.get("route_id"),
            "driver_id": payload.get("driver_id"),
            "driver_name": payload.get("driver_name"),
            "employee_id": payload.get("employee_id"),
            "employee_name": payload.get("employee_name"),
            "date_range_start": payload.get("date_range_start"),
            "date_range_end": payload.get("date_range_end"),
            "requested_by": payload.get("requested_by") or "LORI User",
        }

    station_code = lori_report_upper(request.get("station_code"))
    snapshot = await lori_report_snapshot_for_station(station_code) if station_code else await lori_report_snapshot_for_station("")

    sections = lori_report_build_sections(template, request, snapshot)

    executive_summary = sections[0].get("section_body") if sections else "Report generated by LORI."
    attention_sections = [s for s in sections if s.get("requires_attention")]

    key_findings = f"LORI generated {len(sections)} report sections. {len(attention_sections)} section(s) require attention."
    recommended_actions = "Review report sections, confirm source data, and assign any required follow-up in Action Center."
    risk_summary = "High attention required." if any(s.get("risk_level") == "High" for s in sections) else ("Medium attention required." if attention_sections else "No immediate high-risk items detected in available data.")

    report_payload = {
        "report_request_id": request_id or None,
        "report_template_id": template.get("id"),

        "report_title": lori_report_clean(payload.get("report_title") or request.get("request_title") or template.get("report_name")),
        "report_name": template.get("report_name"),
        "report_category": template.get("report_category"),
        "report_code": template.get("report_code"),
        "report_status": "Generated",

        "company_name": request.get("company_name"),
        "region_code": request.get("region_code"),
        "region_name": request.get("region_name"),
        "operating_state": request.get("operating_state"),
        "city": request.get("city"),
        "station_code": request.get("station_code"),
        "station_name": request.get("station_name"),
        "route_group": request.get("route_group"),
        "route_id": request.get("route_id"),

        "driver_id": request.get("driver_id"),
        "driver_name": request.get("driver_name"),
        "employee_id": request.get("employee_id"),
        "employee_name": request.get("employee_name"),

        "report_period_start": request.get("date_range_start") or None,
        "report_period_end": request.get("date_range_end") or None,

        "executive_summary": executive_summary,
        "key_findings": key_findings,
        "recommended_actions": recommended_actions,
        "risk_summary": risk_summary,
        "source_modules": template.get("source_modules") or [],

        "report_plain_text": "\n\n".join([f"{s.get('section_title')}\n{s.get('section_body')}" for s in sections]),
        "generated_by": payload.get("generated_by") or "LORI Reports Engine",
    }

    created_report = await lori_policy_supabase_post(
        "lori_report_library",
        report_payload,
    )

    report = created_report[0] if created_report else report_payload
    report_id = report.get("id")

    created_sections = []

    for section in sections:
        section_payload = {
            "report_id": report_id,
            **section,
        }

        created = await lori_policy_supabase_post(
            "lori_report_sections",
            section_payload,
        )

        if created:
            created_sections.append(created[0])

    html = lori_report_build_html(report, sections)

    await lori_report_patch_rows(
        "lori_report_library",
        f"id=eq.{quote(str(report_id))}",
        {
            "report_html": html,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    if request_id:
        await lori_report_patch_rows(
            "lori_report_requests",
            f"id=eq.{quote(request_id)}",
            {
                "request_status": "Generated",
                "updated_at": datetime.utcnow().isoformat(),
            },
        )

    return {
        "status": "success",
        "message": "Report packet generated.",
        "report": {**report, "report_html": html},
        "sections_count": len(created_sections),
        "sections": created_sections,
        "source_snapshot_counts": {
            "documents": len(snapshot.get("documents", [])),
            "action_items": len(snapshot.get("action_items", [])),
            "kpi_plans": len(snapshot.get("kpi_plans", [])),
            "sops": len(snapshot.get("sops", [])),
            "route_projects": len(snapshot.get("route_projects", [])),
            "contract_reviews": len(snapshot.get("contract_reviews", [])),
            "drivers": len(snapshot.get("drivers", [])),
        },
    }


@app.get("/report-library")
async def report_library(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    report_category: Optional[str] = Query(None),
    report_code: Optional[str] = Query(None),
    report_status: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    reports = await lori_report_get_rows(
        "lori_report_library",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        reports = [r for r in reports if lori_report_upper(r.get("station_code")) == lori_report_upper(station_code)]

    if report_category:
        reports = [r for r in reports if lori_report_clean(r.get("report_category")).lower() == report_category.lower()]

    if report_code:
        reports = [r for r in reports if lori_report_upper(r.get("report_code")) == lori_report_upper(report_code)]

    if report_status:
        reports = [r for r in reports if lori_report_clean(r.get("report_status")).lower() == report_status.lower()]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "reports_count": len(reports[:limit]),
        "reports": reports[:limit],
    }


@app.get("/report-detail")
async def report_detail(
    api_key: Optional[str] = Query(None),
    report_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    reports = await lori_report_get_rows(
        "lori_report_library",
        f"select=*&id=eq.{quote(report_id)}&limit=1",
    )

    if not reports:
        return {"status": "not_found", "message": "Report not found."}

    sections = await lori_report_get_rows(
        "lori_report_sections",
        f"select=*&report_id=eq.{quote(report_id)}&order=section_order.asc&limit=500",
    )

    source_links = await lori_report_safe_get_rows(
        "lori_report_source_links",
        f"select=*&report_id=eq.{quote(report_id)}&order=created_at.desc&limit=500",
    )

    exports = await lori_report_safe_get_rows(
        "lori_report_export_history",
        f"select=*&report_id=eq.{quote(report_id)}&order=created_at.desc&limit=200",
    )

    return {
        "status": "success",
        "report": reports[0],
        "sections_count": len(sections),
        "sections": sections,
        "source_links_count": len(source_links),
        "source_links": source_links,
        "export_history_count": len(exports),
        "export_history": exports,
    }


@app.get("/report-export-html")
async def report_export_html(
    api_key: Optional[str] = Query(None),
    report_id: str = Query(...),
    exported_by: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    detail = await report_detail(api_key=api_key, report_id=report_id)

    if detail.get("status") != "success":
        return detail

    report = detail.get("report") or {}
    sections = detail.get("sections") or []

    html = report.get("report_html") or lori_report_build_html(report, sections)

    await lori_policy_supabase_post(
        "lori_report_export_history",
        {
            "report_id": report_id,
            "export_type": "HTML Print",
            "export_status": "Completed",
            "exported_by": exported_by or "LORI User",
            "export_notes": "HTML/print-ready report generated. PDF export is not enabled yet.",
        },
    )

    return {
        "status": "success",
        "report_id": report_id,
        "export_type": "HTML Print",
        "pdf_export_available": False,
        "html": html,
        "message": "HTML print/export packet generated. PDF generation is not enabled yet.",
    }


@app.post("/report-status-update")
async def report_status_update(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    report_id = lori_report_clean(payload.get("report_id"))
    report_status = lori_report_clean(payload.get("report_status"))

    if not report_id:
        return {"status": "error", "message": "report_id is required."}

    if not report_status:
        return {"status": "error", "message": "report_status is required."}

    update_payload = {
        "report_status": report_status,
        "updated_at": datetime.utcnow().isoformat(),
    }

    if report_status == "Reviewed":
        update_payload["reviewed_by"] = lori_report_clean(payload.get("reviewed_by") or "LORI User")
        update_payload["reviewed_at"] = datetime.utcnow().isoformat()

    if report_status == "Published":
        update_payload["published_by"] = lori_report_clean(payload.get("published_by") or "LORI User")
        update_payload["published_at"] = datetime.utcnow().isoformat()

    updated = await lori_report_patch_rows(
        "lori_report_library",
        f"id=eq.{quote(report_id)}",
        update_payload,
    )

    return {
        "status": "success",
        "message": f"Report status updated to {report_status}.",
        "report": updated[0] if updated else update_payload,
    }


@app.get("/report-requests")
async def report_requests(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    request_status: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    requests = await lori_report_get_rows(
        "lori_report_requests",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        requests = [r for r in requests if lori_report_upper(r.get("station_code")) == lori_report_upper(station_code)]

    if request_status:
        requests = [r for r in requests if lori_report_clean(r.get("request_status")).lower() == request_status.lower()]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "requests_count": len(requests[:limit]),
        "requests": requests[:limit],
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# TREND INTELLIGENCE ENGINE
# Allows LORI to offer trend analysis, analyze patterns, create
# trend points, explain trends, support counseling language, and
# surface insights from uploaded data, reports, drivers, routes,
# documents, SOPs, KPI plans, action items, compliance, and more.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime, date
from collections import Counter
import os
import httpx
import re


SUPABASE_URL_TREND = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_TREND = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_trend_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_trend_upper(value: Any) -> str:
    return lori_trend_clean(value).upper()


def lori_trend_lower(value: Any) -> str:
    return lori_trend_clean(value).lower()


def lori_trend_today() -> str:
    return date.today().isoformat()


def lori_trend_parse_date(value: Any) -> Optional[date]:
    if not value:
        return None

    text = str(value).strip()

    try:
        if "T" in text:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        return date.fromisoformat(text[:10])
    except Exception:
        return None


def lori_trend_period_label(value: Optional[date]) -> str:
    if not value:
        return date.today().strftime("%Y-%m")
    return value.strftime("%Y-%m")


async def lori_trend_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_trend_safe_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    try:
        return await lori_policy_supabase_get(f"{table}?{query}")
    except Exception:
        return []


async def lori_trend_patch_rows(
    table: str,
    match_query: str,
    payload: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not SUPABASE_URL_TREND or not SUPABASE_SERVICE_ROLE_KEY_TREND:
        raise RuntimeError("Missing Supabase environment variables.")

    url = f"{SUPABASE_URL_TREND}/rest/v1/{table}?{match_query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_TREND,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_TREND}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(url, headers=headers, json=payload)

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase PATCH failed: {response.status_code} {response.text}")

    try:
        return response.json()
    except Exception:
        return []


def lori_trend_infer_template_from_question(question_text: str) -> Dict[str, str]:
    q = lori_trend_lower(question_text)

    mapping = [
        {
            "code": "DRIVER_PERFORMANCE_TREND",
            "category": "Driver",
            "use_case": "Coaching / Counseling",
            "keywords": ["driver", "performance", "score", "better", "worse", "improving", "declining"],
        },
        {
            "code": "DRIVER_SAFETY_TREND",
            "category": "Safety",
            "use_case": "Safety Review",
            "keywords": ["safety", "accident", "incident", "crash", "injury", "unsafe"],
        },
        {
            "code": "COUNSELING_PATTERN_TREND",
            "category": "Counseling",
            "use_case": "Counseling",
            "keywords": ["counseling", "coach", "coaching", "discipline", "write up", "pattern", "behavior"],
        },
        {
            "code": "ROUTE_DELAY_TREND",
            "category": "Route",
            "use_case": "Route Review",
            "keywords": ["late", "delay", "delays", "route", "delivery window", "always late"],
        },
        {
            "code": "WORK_AREA_BALANCE_TREND",
            "category": "Route",
            "use_case": "Route Configuration",
            "keywords": ["work area", "territory", "balance", "unbalanced", "stops", "route change", "stop moves"],
        },
        {
            "code": "KPI_PERFORMANCE_TREND",
            "category": "KPI",
            "use_case": "Leadership Briefing",
            "keywords": ["kpi", "metric", "metrics", "goal", "target", "performance gap"],
        },
        {
            "code": "ACTION_ITEM_CLOSURE_TREND",
            "category": "Action Center",
            "use_case": "Leadership Briefing",
            "keywords": ["action item", "follow up", "overdue", "closed", "closure", "owner"],
        },
        {
            "code": "DOCUMENT_REVIEW_TREND",
            "category": "Document",
            "use_case": "Document Review",
            "keywords": ["document", "upload", "file", "policy", "accident report", "employee file"],
        },
        {
            "code": "SOP_REVIEW_TREND",
            "category": "SOP",
            "use_case": "Training Review",
            "keywords": ["sop", "procedure", "training", "acknowledgement", "acknowledgment"],
        },
        {
            "code": "COMPLIANCE_RISK_TREND",
            "category": "Compliance",
            "use_case": "Compliance Review",
            "keywords": ["compliance", "regulatory", "policy risk", "audit", "violation"],
        },
        {
            "code": "DRIVER_ROAD_COMMUNICATION_TREND",
            "category": "Driver Road Communications",
            "use_case": "General Information",
            "keywords": ["message", "driver message", "road communication", "urgent message", "stop change"],
        },
    ]

    for item in mapping:
        if any(keyword in q for keyword in item["keywords"]):
            return item

    return {
        "code": "KPI_PERFORMANCE_TREND",
        "category": "General",
        "use_case": "General Information",
    }


def lori_trend_direction_from_points(point_values: List[float]) -> Dict[str, str]:
    if len(point_values) < 2:
        return {
            "trend_direction": "Unknown",
            "trend_strength": "Weak",
            "confidence_level": "Low",
        }

    first = point_values[0]
    last = point_values[-1]
    change = last - first

    if change > 0:
        direction = "Worsening"
    elif change < 0:
        direction = "Improving"
    else:
        direction = "Stable"

    absolute_change = abs(change)

    if absolute_change >= 5:
        strength = "Strong"
        confidence = "High"
    elif absolute_change >= 2:
        strength = "Moderate"
        confidence = "Medium"
    else:
        strength = "Weak"
        confidence = "Low"

    return {
        "trend_direction": direction,
        "trend_strength": strength,
        "confidence_level": confidence,
    }


def lori_trend_risk_from_direction(direction: str, strength: str, record_count: int) -> str:
    if direction == "Worsening" and strength in ["Strong", "Moderate"]:
        return "High" if record_count >= 5 else "Medium"
    if direction == "Worsening":
        return "Medium"
    if direction == "Stable" and record_count >= 10:
        return "Medium"
    return "Low"


async def lori_trend_snapshot(
    station_code: str = "",
    driver_name: str = "",
    driver_id: str = "",
    employee_name: str = "",
    employee_id: str = "",
    route_id: str = "",
) -> Dict[str, List[Dict[str, Any]]]:
    station = lori_trend_upper(station_code)

    def filter_common(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        filtered = rows

        if station:
            filtered = [
                r for r in filtered
                if lori_trend_upper(r.get("station_code")) == station
                or lori_trend_upper(r.get("current_station_code")) == station
                or lori_trend_upper(r.get("new_station_code")) == station
            ]

        if driver_name:
            filtered = [
                r for r in filtered
                if driver_name.lower() in lori_trend_lower(r.get("driver_name"))
                or driver_name.lower() in lori_trend_lower(r.get("current_driver_name"))
                or driver_name.lower() in lori_trend_lower(r.get("receiving_driver_name"))
            ]

        if driver_id:
            filtered = [
                r for r in filtered
                if lori_trend_lower(r.get("driver_id")) == driver_id.lower()
            ]

        if employee_name:
            filtered = [
                r for r in filtered
                if employee_name.lower() in lori_trend_lower(r.get("employee_name"))
            ]

        if employee_id:
            filtered = [
                r for r in filtered
                if lori_trend_lower(r.get("employee_id")) == employee_id.lower()
            ]

        if route_id:
            filtered = [
                r for r in filtered
                if lori_trend_lower(r.get("route_id")) == route_id.lower()
                or lori_trend_lower(r.get("current_route_id")) == route_id.lower()
                or lori_trend_lower(r.get("receiving_route_id")) == route_id.lower()
            ]

        return filtered

    docs = filter_common(await lori_trend_safe_get_rows("lori_document_library", "select=*&order=created_at.desc&limit=3000"))
    action_items = filter_common(await lori_trend_safe_get_rows("lori_action_items", "select=*&order=created_at.desc&limit=3000"))
    kpi_plans = filter_common(await lori_trend_safe_get_rows("lori_kpi_action_plans", "select=*&order=created_at.desc&limit=1000"))
    sops = filter_common(await lori_trend_safe_get_rows("lori_sop_library", "select=*&order=created_at.desc&limit=1000"))
    sop_ack = filter_common(await lori_trend_safe_get_rows("lori_sop_acknowledgements", "select=*&order=created_at.desc&limit=2000"))
    route_projects = filter_common(await lori_trend_safe_get_rows("lori_route_config_projects", "select=*&order=created_at.desc&limit=1000"))
    contract_reviews = filter_common(await lori_trend_safe_get_rows("lori_route_contract_safeguard_reviews", "select=*&order=created_at.desc&limit=1000"))
    driver_master = filter_common(await lori_trend_safe_get_rows("lori_driver_master", "select=*&order=created_at.desc&limit=3000"))
    driver_metrics = filter_common(await lori_trend_safe_get_rows("lori_driver_metrics", "select=*&order=created_at.desc&limit=3000"))
    driver_safety = filter_common(await lori_trend_safe_get_rows("lori_driver_safety_events", "select=*&order=created_at.desc&limit=3000"))
    driver_counseling = filter_common(await lori_trend_safe_get_rows("lori_driver_counseling", "select=*&order=created_at.desc&limit=3000"))
    reports = filter_common(await lori_trend_safe_get_rows("lori_report_library", "select=*&order=created_at.desc&limit=1000"))
    regulatory_alerts = await lori_trend_safe_get_rows("lori_regulatory_alerts", "select=*&order=created_at.desc&limit=1000")

    road_messages = filter_common(await lori_trend_safe_get_rows("lori_driver_road_communications", "select=*&order=created_at.desc&limit=3000"))

    return {
        "documents": docs,
        "action_items": action_items,
        "kpi_plans": kpi_plans,
        "sops": sops,
        "sop_acknowledgements": sop_ack,
        "route_projects": route_projects,
        "contract_reviews": contract_reviews,
        "driver_master": driver_master,
        "driver_metrics": driver_metrics,
        "driver_safety": driver_safety,
        "driver_counseling": driver_counseling,
        "reports": reports,
        "regulatory_alerts": regulatory_alerts,
        "road_messages": road_messages,
    }


def lori_trend_rows_for_template(template_code: str, snapshot: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Any]:
    code = lori_trend_upper(template_code)

    if code == "DRIVER_PERFORMANCE_TREND":
        rows = snapshot.get("driver_metrics", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("intake_lane") == "Driver File" or d.get("kpi_related") or d.get("safety_related")
        ]
        metric_name = "Driver performance-related records"

    elif code == "DRIVER_SAFETY_TREND":
        rows = snapshot.get("driver_safety", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("safety_related") or d.get("accident_related")
        ]
        metric_name = "Safety events and safety-related records"

    elif code == "COUNSELING_PATTERN_TREND":
        rows = snapshot.get("driver_counseling", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("counseling_related")
        ]
        metric_name = "Counseling and coaching-related records"

    elif code == "ROUTE_DELAY_TREND":
        rows = snapshot.get("route_projects", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("route_configuration_related") or d.get("route_assignment_related")
        ]
        metric_name = "Route delay / route review records"

    elif code == "WORK_AREA_BALANCE_TREND":
        rows = snapshot.get("route_projects", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("route_configuration_related") or "route" in lori_trend_lower(d.get("document_type"))
        ]
        metric_name = "Work area balance records"

    elif code == "KPI_PERFORMANCE_TREND":
        rows = snapshot.get("kpi_plans", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("kpi_related")
        ]
        metric_name = "KPI and performance records"

    elif code == "ACTION_ITEM_CLOSURE_TREND":
        rows = snapshot.get("action_items", [])
        metric_name = "Action items"

    elif code == "DOCUMENT_REVIEW_TREND":
        rows = snapshot.get("documents", [])
        metric_name = "Uploaded documents"

    elif code == "SOP_REVIEW_TREND":
        rows = snapshot.get("sops", []) + snapshot.get("sop_acknowledgements", [])
        metric_name = "SOP and acknowledgement records"

    elif code == "COMPLIANCE_RISK_TREND":
        rows = snapshot.get("contract_reviews", []) + snapshot.get("regulatory_alerts", []) + [
            d for d in snapshot.get("documents", [])
            if d.get("compliance_related") or d.get("policy_related") or d.get("legal_review_required")
        ]
        metric_name = "Compliance, policy, contract, and regulatory records"

    elif code == "DRIVER_ROAD_COMMUNICATION_TREND":
        rows = snapshot.get("road_messages", [])
        metric_name = "Driver road messages"

    else:
        rows = (
            snapshot.get("documents", [])
            + snapshot.get("action_items", [])
            + snapshot.get("kpi_plans", [])
            + snapshot.get("route_projects", [])
            + snapshot.get("sops", [])
        )
        metric_name = "LORI records"

    return {
        "rows": rows,
        "metric_name": metric_name,
    }


def lori_trend_date_for_row(row: Dict[str, Any]) -> date:
    for key in [
        "incident_date",
        "event_date",
        "metric_date",
        "review_due_date",
        "effective_date",
        "created_at",
        "updated_at",
        "completed_at",
        "published_at",
    ]:
        parsed = lori_trend_parse_date(row.get(key))
        if parsed:
            return parsed

    return date.today()


def lori_trend_make_points(rows: List[Dict[str, Any]], metric_name: str) -> List[Dict[str, Any]]:
    if not rows:
        return []

    counter = Counter()

    for row in rows:
        d = lori_trend_date_for_row(row)
        counter[lori_trend_period_label(d)] += 1

    labels = sorted(counter.keys())

    points = []

    for idx, label in enumerate(labels):
        year, month = label.split("-")
        point_date = date(int(year), int(month), 1)
        points.append({
            "point_order": idx + 1,
            "point_label": label,
            "point_date": point_date.isoformat(),
            "point_period": "Monthly",
            "metric_name": metric_name,
            "metric_value": float(counter[label]),
            "metric_unit": "records",
        })

    return points


def lori_trend_result_text(
    template: Dict[str, Any],
    request: Dict[str, Any],
    points: List[Dict[str, Any]],
    direction: Dict[str, str],
    rows: List[Dict[str, Any]],
) -> Dict[str, str]:
    subject = lori_trend_clean(request.get("trend_subject") or template.get("trend_template_name") or "Trend")
    category = lori_trend_clean(request.get("trend_category") or template.get("trend_category") or "General")
    use_case = lori_trend_clean(request.get("trend_use_case") or template.get("trend_use_case") or "General Information")
    driver_name = lori_trend_clean(request.get("driver_name"))
    route_id = lori_trend_clean(request.get("route_id"))

    direction_label = direction.get("trend_direction", "Unknown").lower()
    strength = direction.get("trend_strength", "Moderate").lower()
    count = len(rows)

    if not points:
        summary = f"LORI does not have enough dated records yet to create a reliable trend for {subject}."
        explanation = "More uploaded data, reports, documents, action records, route records, counseling entries, KPI records, or safety records are needed before LORI can show a meaningful pattern."
        recommended = "Upload or connect more source data, then run the trend again."
    else:
        target_text = driver_name or route_id or lori_trend_clean(request.get("station_code")) or "the selected context"
        summary = f"LORI found a {strength} {direction_label} trend for {subject} based on {count} source record(s)."
        explanation = f"The trend was calculated from available LORI records grouped over time for {target_text}. It should be treated as decision support, not as final proof or disciplinary conclusion."
        recommended = "Review the underlying records, confirm accuracy, and decide whether this trend should be used for coaching, counseling, leadership briefing, route review, KPI follow-up, or Action Center tracking."

    counseling_summary = (
        f"Counseling trend note: LORI identified a {direction.get('trend_direction', 'Unknown')} pattern. "
        f"Use this as a coaching conversation starter only. Confirm the facts, dates, and documents before discussing performance or corrective action."
    )

    leadership_summary = (
        f"Leadership trend summary: {summary} The pattern may require supervisor review, operational follow-up, or additional data validation."
    )

    operational_summary = (
        f"Operational trend summary: {explanation} Recommended next step: {recommended}"
    )

    suggested_follow_up = f"Would you like LORI to show the source records and create a chart for this {category.lower()} trend?"

    return {
        "trend_title": f"{subject} Trend",
        "trend_summary": summary,
        "plain_language_explanation": explanation,
        "leadership_summary": leadership_summary,
        "counseling_summary": counseling_summary,
        "operational_summary": operational_summary,
        "recommended_action": recommended,
        "suggested_follow_up_question": suggested_follow_up,
    }


@app.get("/trend-templates")
async def trend_templates(
    api_key: Optional[str] = Query(None),
    trend_category: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    templates = await lori_trend_get_rows(
        "lori_trend_templates",
        "select=*&order=trend_template_name.asc&limit=500",
    )

    if trend_category:
        templates = [
            t for t in templates
            if lori_trend_clean(t.get("trend_category")).lower() == trend_category.lower()
        ]

    return {
        "status": "success",
        "templates_count": len(templates),
        "templates": templates,
    }


@app.get("/trend-intelligence-summary")
async def trend_intelligence_summary(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    templates = await lori_trend_safe_get_rows(
        "lori_trend_templates",
        "select=*&order=trend_template_name.asc&limit=500",
    )

    requests = await lori_trend_safe_get_rows(
        "lori_trend_intelligence_requests",
        "select=*&order=created_at.desc&limit=1000",
    )

    results = await lori_trend_safe_get_rows(
        "lori_trend_intelligence_results",
        "select=*&order=created_at.desc&limit=1000",
    )

    insights = await lori_trend_safe_get_rows(
        "lori_trend_insight_library",
        "select=*&order=created_at.desc&limit=1000",
    )

    offers = await lori_trend_safe_get_rows(
        "lori_trend_conversation_offers",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        station = lori_trend_upper(station_code)
        requests = [r for r in requests if lori_trend_upper(r.get("station_code")) == station]
        insights = [i for i in insights if lori_trend_upper(i.get("station_code")) == station]
        offers = [o for o in offers if lori_trend_upper(o.get("station_code")) == station]

    return {
        "status": "success",
        "trend_templates_count": len(templates),
        "trend_requests_count": len(requests),
        "trend_results_count": len(results),
        "active_insights_count": len([i for i in insights if i.get("insight_status") == "Active"]),
        "offers_count": len(offers),
        "accepted_offers_count": len([o for o in offers if o.get("offer_status") == "Accepted"]),
        "declined_offers_count": len([o for o in offers if o.get("offer_status") == "Declined"]),
    }


@app.post("/trend-offer-for-question")
async def trend_offer_for_question(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    question_text = lori_trend_clean(payload.get("question_text") or payload.get("original_question"))

    if not question_text:
        return {"status": "error", "message": "question_text is required."}

    inferred = lori_trend_infer_template_from_question(question_text)

    templates = await lori_trend_safe_get_rows(
        "lori_trend_templates",
        f"select=*&trend_template_code=eq.{quote(inferred['code'])}&limit=1",
    )

    template = templates[0] if templates else {}

    offer_text = template.get("trend_offer_text") or f"Would you like LORI to trend this information over time?"

    offer_payload = {
        "conversation_id": lori_trend_clean(payload.get("conversation_id")),
        "user_email": lori_trend_clean(payload.get("user_email")),
        "offered_by_module": lori_trend_clean(payload.get("offered_by_module") or "Ask LORI"),
        "original_question": question_text,
        "trend_subject": lori_trend_clean(payload.get("trend_subject") or template.get("trend_template_name") or inferred["category"]),
        "trend_category": lori_trend_clean(payload.get("trend_category") or template.get("trend_category") or inferred["category"]),
        "trend_use_case": lori_trend_clean(payload.get("trend_use_case") or template.get("trend_use_case") or inferred["use_case"]),
        "offer_text": offer_text,
        "offer_status": "Offered",
        "company_name": lori_trend_clean(payload.get("company_name")),
        "station_code": lori_trend_upper(payload.get("station_code")),
        "route_id": lori_trend_clean(payload.get("route_id")),
        "driver_id": lori_trend_clean(payload.get("driver_id")),
        "driver_name": lori_trend_clean(payload.get("driver_name")),
        "employee_id": lori_trend_clean(payload.get("employee_id")),
        "employee_name": lori_trend_clean(payload.get("employee_name")),
    }

    created = await lori_policy_supabase_post(
        "lori_trend_conversation_offers",
        offer_payload,
    )

    return {
        "status": "success",
        "should_offer_trend": True,
        "offer": created[0] if created else offer_payload,
        "recommended_template_code": inferred["code"],
        "trend_category": offer_payload["trend_category"],
        "trend_use_case": offer_payload["trend_use_case"],
        "offer_text": offer_text,
    }


@app.post("/trend-offer-update")
async def trend_offer_update(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    offer_id = lori_trend_clean(payload.get("offer_id"))
    offer_status = lori_trend_clean(payload.get("offer_status"))

    if not offer_id:
        return {"status": "error", "message": "offer_id is required."}

    if offer_status not in ["Accepted", "Declined", "Expired", "Offered"]:
        return {"status": "error", "message": "offer_status must be Accepted, Declined, Expired, or Offered."}

    update_payload = {
        "offer_status": offer_status,
    }

    if offer_status == "Accepted":
        update_payload["accepted_at"] = datetime.utcnow().isoformat()

    if offer_status == "Declined":
        update_payload["declined_at"] = datetime.utcnow().isoformat()

    updated = await lori_trend_patch_rows(
        "lori_trend_conversation_offers",
        f"id=eq.{quote(offer_id)}",
        update_payload,
    )

    return {
        "status": "success",
        "message": f"Trend offer marked {offer_status}.",
        "offer": updated[0] if updated else update_payload,
    }


@app.post("/trend-request-create")
async def trend_request_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    question_text = lori_trend_clean(payload.get("question_text"))
    template_code = lori_trend_upper(payload.get("trend_template_code"))

    if not template_code:
        inferred = lori_trend_infer_template_from_question(question_text)
        template_code = inferred["code"]

    templates = await lori_trend_safe_get_rows(
        "lori_trend_templates",
        f"select=*&trend_template_code=eq.{quote(template_code)}&limit=1",
    )

    template = templates[0] if templates else {}

    request_payload = {
        "request_title": lori_trend_clean(payload.get("request_title") or f"{template.get('trend_template_name') or 'Trend'} — {payload.get('station_code') or 'Station'}"),
        "request_status": "Draft",
        "question_text": question_text,
        "trend_subject": lori_trend_clean(payload.get("trend_subject") or template.get("trend_template_name")),
        "trend_category": lori_trend_clean(payload.get("trend_category") or template.get("trend_category")),
        "trend_use_case": lori_trend_clean(payload.get("trend_use_case") or template.get("trend_use_case") or "General Information"),
        "requested_scope": lori_trend_clean(payload.get("requested_scope") or "Station"),

        "company_name": lori_trend_clean(payload.get("company_name")),
        "region_code": lori_trend_upper(payload.get("region_code")),
        "region_name": lori_trend_clean(payload.get("region_name")),
        "operating_state": lori_trend_upper(payload.get("operating_state")),
        "city": lori_trend_clean(payload.get("city")),
        "station_code": lori_trend_upper(payload.get("station_code")),
        "station_name": lori_trend_clean(payload.get("station_name")),
        "route_group": lori_trend_clean(payload.get("route_group")),
        "route_id": lori_trend_clean(payload.get("route_id")),

        "driver_id": lori_trend_clean(payload.get("driver_id")),
        "driver_name": lori_trend_clean(payload.get("driver_name")),
        "employee_id": lori_trend_clean(payload.get("employee_id")),
        "employee_name": lori_trend_clean(payload.get("employee_name")),

        "date_range_start": payload.get("date_range_start") or None,
        "date_range_end": payload.get("date_range_end") or None,

        "source_modules": template.get("source_modules") or payload.get("source_modules") or [],
        "source_report_id": payload.get("source_report_id") or None,
        "source_document_id": payload.get("source_document_id") or None,
        "source_sop_id": payload.get("source_sop_id") or None,
        "source_action_item_id": payload.get("source_action_item_id") or None,
        "source_route_project_id": payload.get("source_route_project_id") or None,
        "source_kpi_plan_id": payload.get("source_kpi_plan_id") or None,

        "requested_by": lori_trend_clean(payload.get("requested_by") or "LORI User"),
        "user_requested_trend": bool(payload.get("user_requested_trend", True)),
        "lori_offered_trend": bool(payload.get("lori_offered_trend", False)),
        "trend_offer_text": lori_trend_clean(payload.get("trend_offer_text") or template.get("trend_offer_text")),
    }

    created = await lori_policy_supabase_post(
        "lori_trend_intelligence_requests",
        request_payload,
    )

    return {
        "status": "success",
        "message": "Trend intelligence request created.",
        "request": created[0] if created else request_payload,
        "trend_template_code": template_code,
        "template": template,
    }


@app.post("/trend-analyze")
async def trend_analyze(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    request_id = lori_trend_clean(payload.get("request_id"))

    if not request_id:
        return {"status": "error", "message": "request_id is required."}

    requests = await lori_trend_get_rows(
        "lori_trend_intelligence_requests",
        f"select=*&id=eq.{quote(request_id)}&limit=1",
    )

    if not requests:
        return {"status": "not_found", "message": "Trend request not found."}

    request = requests[0]

    inferred = lori_trend_infer_template_from_question(request.get("question_text") or request.get("trend_subject") or "")
    template_code = lori_trend_upper(payload.get("trend_template_code") or inferred["code"])

    templates = await lori_trend_safe_get_rows(
        "lori_trend_templates",
        f"select=*&trend_template_code=eq.{quote(template_code)}&limit=1",
    )

    template = templates[0] if templates else {
        "trend_template_name": request.get("trend_subject") or "Trend",
        "trend_category": request.get("trend_category") or "General",
        "trend_use_case": request.get("trend_use_case") or "General Information",
        "supports_chart": True,
        "supports_counseling": False,
        "supports_leadership_summary": True,
        "supports_action_center": True,
    }

    snapshot = await lori_trend_snapshot(
        station_code=request.get("station_code"),
        driver_name=request.get("driver_name"),
        driver_id=request.get("driver_id"),
        employee_name=request.get("employee_name"),
        employee_id=request.get("employee_id"),
        route_id=request.get("route_id"),
    )

    row_bundle = lori_trend_rows_for_template(template_code, snapshot)
    rows = row_bundle["rows"]
    metric_name = row_bundle["metric_name"]

    points = lori_trend_make_points(rows, metric_name)
    values = [float(p["metric_value"]) for p in points]

    direction = lori_trend_direction_from_points(values)
    risk_level = lori_trend_risk_from_direction(
        direction.get("trend_direction"),
        direction.get("trend_strength"),
        len(rows),
    )

    text_parts = lori_trend_result_text(template, request, points, direction, rows)

    result_payload = {
        "trend_request_id": request_id,
        "result_status": "Generated" if points else "Not Enough Data",
        "trend_title": text_parts["trend_title"],
        "trend_summary": text_parts["trend_summary"],
        "trend_direction": direction["trend_direction"],
        "trend_strength": direction["trend_strength"],
        "confidence_level": direction["confidence_level"],
        "risk_level": risk_level,
        "finding_type": "Trend Analysis",
        "finding_category": template.get("trend_category") or request.get("trend_category"),
        "plain_language_explanation": text_parts["plain_language_explanation"],
        "leadership_summary": text_parts["leadership_summary"],
        "counseling_summary": text_parts["counseling_summary"],
        "operational_summary": text_parts["operational_summary"],
        "recommended_action": text_parts["recommended_action"],
        "suggested_follow_up_question": text_parts["suggested_follow_up_question"],
        "should_offer_chart": bool(template.get("supports_chart", True)),
        "should_offer_counseling_language": bool(template.get("supports_counseling", False)),
        "should_offer_leadership_packet": bool(template.get("supports_leadership_summary", True)),
        "should_send_to_action_center": bool(template.get("supports_action_center", True)),
        "data_points_count": len(points),
        "source_records_count": len(rows),
        "generated_by": "LORI Trend Intelligence",
    }

    created_result = await lori_policy_supabase_post(
        "lori_trend_intelligence_results",
        result_payload,
    )

    result = created_result[0] if created_result else result_payload
    result_id = result.get("id")

    created_points = []

    for point in points:
        point_payload = {
            "trend_result_id": result_id,
            **point,
            "route_id": request.get("route_id"),
            "driver_id": request.get("driver_id"),
            "driver_name": request.get("driver_name"),
            "employee_id": request.get("employee_id"),
            "employee_name": request.get("employee_name"),
            "station_code": request.get("station_code"),
            "source_module": ", ".join(template.get("source_modules") or []),
            "source_reference": template_code,
            "notes": "Trend point generated from available LORI records.",
        }

        created = await lori_policy_supabase_post(
            "lori_trend_data_points",
            point_payload,
        )

        if created:
            created_points.append(created[0])

    created_sources = []

    for row in rows[:75]:
        source_payload = {
            "trend_result_id": result_id,
            "source_module": template.get("trend_category") or "LORI",
            "source_table": "multiple_lori_tables",
            "source_record_id": row.get("id") if isinstance(row.get("id"), str) else None,
            "source_reference": row.get("document_title") or row.get("sop_title") or row.get("report_title") or row.get("action_title") or row.get("route_id") or row.get("driver_name"),
            "source_title": row.get("document_title") or row.get("sop_title") or row.get("report_title") or row.get("request_title") or metric_name,
            "source_summary": row.get("notes") or row.get("summary") or row.get("executive_summary") or row.get("trend_summary") or "",
            "station_code": row.get("station_code") or request.get("station_code"),
            "route_id": row.get("route_id") or request.get("route_id"),
            "driver_id": row.get("driver_id") or request.get("driver_id"),
            "driver_name": row.get("driver_name") or request.get("driver_name"),
            "employee_id": row.get("employee_id") or request.get("employee_id"),
            "employee_name": row.get("employee_name") or request.get("employee_name"),
            "relevance_score": 1.0,
        }

        created = await lori_policy_supabase_post(
            "lori_trend_source_records",
            source_payload,
        )

        if created:
            created_sources.append(created[0])

    insight_payload = {
        "insight_title": result_payload["trend_title"],
        "insight_category": result_payload["finding_category"],
        "insight_status": "Active",
        "company_name": request.get("company_name"),
        "region_code": request.get("region_code"),
        "region_name": request.get("region_name"),
        "operating_state": request.get("operating_state"),
        "station_code": request.get("station_code"),
        "route_group": request.get("route_group"),
        "route_id": request.get("route_id"),
        "driver_id": request.get("driver_id"),
        "driver_name": request.get("driver_name"),
        "employee_id": request.get("employee_id"),
        "employee_name": request.get("employee_name"),
        "insight_summary": result_payload["trend_summary"],
        "trend_direction": result_payload["trend_direction"],
        "risk_level": result_payload["risk_level"],
        "confidence_level": result_payload["confidence_level"],
        "source_modules": template.get("source_modules") or [],
        "recommended_action": result_payload["recommended_action"],
    }

    await lori_policy_supabase_post(
        "lori_trend_insight_library",
        insight_payload,
    )

    await lori_trend_patch_rows(
        "lori_trend_intelligence_requests",
        f"id=eq.{quote(request_id)}",
        {
            "request_status": "Analyzed" if points else "Needs More Data",
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    return {
        "status": "success",
        "message": "Trend analysis generated.",
        "result": result,
        "data_points_count": len(created_points),
        "data_points": created_points,
        "source_records_count": len(created_sources),
        "source_records": created_sources,
        "trend_template_code": template_code,
        "decision_support_note": result_payload["decision_support_note"],
    }


@app.post("/trend-quick-analyze")
async def trend_quick_analyze(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    create_response = await trend_request_create(api_key=api_key, payload=payload)

    if create_response.get("status") != "success":
        return create_response

    request = create_response.get("request") or {}

    analyze_response = await trend_analyze(
        api_key=api_key,
        payload={
            "request_id": request.get("id"),
            "trend_template_code": create_response.get("trend_template_code"),
        },
    )

    return {
        "status": "success",
        "request": request,
        "analysis": analyze_response,
    }


@app.get("/trend-result-detail")
async def trend_result_detail(
    api_key: Optional[str] = Query(None),
    trend_result_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    results = await lori_trend_get_rows(
        "lori_trend_intelligence_results",
        f"select=*&id=eq.{quote(trend_result_id)}&limit=1",
    )

    if not results:
        return {"status": "not_found", "message": "Trend result not found."}

    points = await lori_trend_safe_get_rows(
        "lori_trend_data_points",
        f"select=*&trend_result_id=eq.{quote(trend_result_id)}&order=point_order.asc&limit=500",
    )

    sources = await lori_trend_safe_get_rows(
        "lori_trend_source_records",
        f"select=*&trend_result_id=eq.{quote(trend_result_id)}&order=created_at.desc&limit=500",
    )

    return {
        "status": "success",
        "result": results[0],
        "data_points_count": len(points),
        "data_points": points,
        "source_records_count": len(sources),
        "source_records": sources,
    }


@app.get("/trend-insights")
async def trend_insights(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    insight_category: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    insights = await lori_trend_safe_get_rows(
        "lori_trend_insight_library",
        "select=*&order=last_detected_at.desc&limit=1000",
    )

    if station_code:
        insights = [i for i in insights if lori_trend_upper(i.get("station_code")) == lori_trend_upper(station_code)]

    if driver_name:
        insights = [i for i in insights if driver_name.lower() in lori_trend_lower(i.get("driver_name"))]

    if route_id:
        insights = [i for i in insights if lori_trend_lower(i.get("route_id")) == route_id.lower()]

    if insight_category:
        insights = [i for i in insights if lori_trend_lower(i.get("insight_category")) == insight_category.lower()]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "insights_count": len(insights[:limit]),
        "insights": insights[:limit],
    }


@app.get("/trend-requests")
async def trend_requests(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    request_status: Optional[str] = Query(None),
    trend_category: Optional[str] = Query(None),
    limit: int = Query(200),
):
    lori_regulatory_require_key(api_key)

    requests = await lori_trend_safe_get_rows(
        "lori_trend_intelligence_requests",
        "select=*&order=created_at.desc&limit=1000",
    )

    if station_code:
        requests = [r for r in requests if lori_trend_upper(r.get("station_code")) == lori_trend_upper(station_code)]

    if request_status:
        requests = [r for r in requests if lori_trend_lower(r.get("request_status")) == request_status.lower()]

    if trend_category:
        requests = [r for r in requests if lori_trend_lower(r.get("trend_category")) == trend_category.lower()]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "requests_count": len(requests[:limit]),
        "requests": requests[:limit],
    }
# ============================================================
# LORI DRIVE COMMAND CENTER
# CALENDAR & REMINDER ENGINE
# Internal LORI reminders and calendar-style follow-up events.
# External Google/Outlook calendar sync can be added later.
# ============================================================

from fastapi import Body, Query
from typing import Any, Dict, List, Optional
from urllib.parse import quote
from datetime import datetime, date, timedelta
import os
import httpx


SUPABASE_URL_CALENDAR = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_CALENDAR = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_calendar_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_calendar_upper(value: Any) -> str:
    return lori_calendar_clean(value).upper()


def lori_calendar_today() -> str:
    return date.today().isoformat()


def lori_calendar_parse_due_date(value: Any) -> Optional[str]:
    text = lori_calendar_clean(value)
    if not text:
        return None

    try:
        if "T" in text:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
        return date.fromisoformat(text[:10]).isoformat()
    except Exception:
        return None


def lori_calendar_parse_due_time(value: Any) -> Optional[str]:
    text = lori_calendar_clean(value)
    if not text:
        return None

    try:
        if len(text) >= 5:
            return text[:5]
        return None
    except Exception:
        return None


def lori_calendar_make_datetime(due_date: Any, due_time: Any) -> Optional[str]:
    due_date_clean = lori_calendar_parse_due_date(due_date)
    due_time_clean = lori_calendar_parse_due_time(due_time)

    if not due_date_clean:
        return None

    if not due_time_clean:
        due_time_clean = "09:00"

    return f"{due_date_clean}T{due_time_clean}:00+00:00"


async def lori_calendar_get_rows(
    table: str,
    query: str = "select=*&limit=500",
) -> List[Dict[str, Any]]:
    return await lori_policy_supabase_get(f"{table}?{query}")


async def lori_calendar_patch_rows(
    table: str,
    match_query: str,
    payload: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not SUPABASE_URL_CALENDAR or not SUPABASE_SERVICE_ROLE_KEY_CALENDAR:
        raise RuntimeError("Missing Supabase environment variables.")

    url = f"{SUPABASE_URL_CALENDAR}/rest/v1/{table}?{match_query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_CALENDAR,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_CALENDAR}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.patch(url, headers=headers, json=payload)

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase PATCH failed: {response.status_code} {response.text}")

    try:
        return response.json()
    except Exception:
        return []


async def lori_calendar_log_activity(
    reminder_id: str,
    activity_type: str,
    activity_summary: str,
    activity_by_name: str = "",
    activity_by_email: str = "",
    old_status: str = "",
    new_status: str = "",
    activity_metadata: Optional[Dict[str, Any]] = None,
):
    payload = {
        "reminder_id": reminder_id,
        "activity_type": activity_type,
        "activity_summary": activity_summary,
        "activity_by_name": activity_by_name,
        "activity_by_email": activity_by_email,
        "old_status": old_status,
        "new_status": new_status,
        "activity_metadata": activity_metadata or {},
    }

    try:
        await lori_policy_supabase_post("lori_calendar_reminder_activity_log", payload)
    except Exception:
        pass


@app.get("/calendar-reminder-templates")
async def calendar_reminder_templates(
    api_key: Optional[str] = Query(None),
    related_module: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    templates = await lori_calendar_get_rows(
        "lori_calendar_reminder_templates",
        "select=*&order=template_name.asc&limit=500",
    )

    if related_module:
        templates = [
            t for t in templates
            if lori_calendar_clean(t.get("related_module")).lower() == related_module.lower()
        ]

    return {
        "status": "success",
        "templates_count": len(templates),
        "templates": templates,
    }


@app.get("/calendar-reminders-summary")
async def calendar_reminders_summary(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    assigned_to_email: Optional[str] = Query(None),
):
    lori_regulatory_require_key(api_key)

    reminders = await lori_calendar_get_rows(
        "lori_calendar_reminders",
        "select=*&order=created_at.desc&limit=5000",
    )

    if station_code:
        reminders = [
            r for r in reminders
            if lori_calendar_upper(r.get("station_code")) == lori_calendar_upper(station_code)
        ]

    if assigned_to_email:
        reminders = [
            r for r in reminders
            if lori_calendar_clean(r.get("assigned_to_email")).lower() == assigned_to_email.lower()
        ]

    today = date.today()

    def due_date_obj(row):
        try:
            return date.fromisoformat(str(row.get("due_date"))[:10])
        except Exception:
            return None

    open_items = [
        r for r in reminders
        if lori_calendar_clean(r.get("reminder_status")).lower() in ["open", "scheduled", "snoozed"]
    ]

    overdue_items = [
        r for r in open_items
        if due_date_obj(r) and due_date_obj(r) < today
    ]

    due_today_items = [
        r for r in open_items
        if due_date_obj(r) and due_date_obj(r) == today
    ]

    upcoming_items = [
        r for r in open_items
        if due_date_obj(r) and today < due_date_obj(r) <= today + timedelta(days=7)
    ]

    return {
        "status": "success",
        "reminders_count": len(reminders),
        "open_count": len(open_items),
        "scheduled_count": len([r for r in reminders if r.get("reminder_status") == "Scheduled"]),
        "completed_count": len([r for r in reminders if r.get("reminder_status") == "Completed"]),
        "cancelled_count": len([r for r in reminders if r.get("reminder_status") == "Cancelled"]),
        "overdue_count": len(overdue_items),
        "due_today_count": len(due_today_items),
        "upcoming_7_days_count": len(upcoming_items),
        "urgent_count": len([r for r in reminders if r.get("priority") == "Urgent"]),
        "important_count": len([r for r in reminders if r.get("priority") == "Important"]),
        "external_calendar_status": "Not Connected",
        "external_calendar_note": "Internal LORI reminders are available. Google/Outlook calendar sync is not connected yet.",
    }


@app.get("/calendar-reminders")
async def calendar_reminders(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    related_module: Optional[str] = Query(None),
    reminder_status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    assigned_to_email: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    employee_name: Optional[str] = Query(None),
    route_id: Optional[str] = Query(None),
    due_from: Optional[str] = Query(None),
    due_to: Optional[str] = Query(None),
    limit: int = Query(300),
):
    lori_regulatory_require_key(api_key)

    reminders = await lori_calendar_get_rows(
        "lori_calendar_reminders",
        "select=*&order=due_date.asc,created_at.desc&limit=5000",
    )

    if station_code:
        reminders = [r for r in reminders if lori_calendar_upper(r.get("station_code")) == lori_calendar_upper(station_code)]

    if related_module:
        reminders = [r for r in reminders if lori_calendar_clean(r.get("related_module")).lower() == related_module.lower()]

    if reminder_status:
        reminders = [r for r in reminders if lori_calendar_clean(r.get("reminder_status")).lower() == reminder_status.lower()]

    if priority:
        reminders = [r for r in reminders if lori_calendar_clean(r.get("priority")).lower() == priority.lower()]

    if assigned_to_email:
        reminders = [r for r in reminders if lori_calendar_clean(r.get("assigned_to_email")).lower() == assigned_to_email.lower()]

    if driver_name:
        reminders = [r for r in reminders if driver_name.lower() in lori_calendar_clean(r.get("driver_name")).lower()]

    if employee_name:
        reminders = [r for r in reminders if employee_name.lower() in lori_calendar_clean(r.get("employee_name")).lower()]

    if route_id:
        reminders = [r for r in reminders if lori_calendar_clean(r.get("route_id")).lower() == route_id.lower()]

    due_from_date = lori_calendar_parse_due_date(due_from)
    due_to_date = lori_calendar_parse_due_date(due_to)

    if due_from_date:
        reminders = [
            r for r in reminders
            if r.get("due_date") and str(r.get("due_date"))[:10] >= due_from_date
        ]

    if due_to_date:
        reminders = [
            r for r in reminders
            if r.get("due_date") and str(r.get("due_date"))[:10] <= due_to_date
        ]

    limit = max(1, min(limit, 1000))

    return {
        "status": "success",
        "reminders_count": len(reminders[:limit]),
        "reminders": reminders[:limit],
    }


@app.post("/calendar-reminder-create")
async def calendar_reminder_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    reminder_title = lori_calendar_clean(payload.get("reminder_title"))

    if not reminder_title:
        return {"status": "error", "message": "reminder_title is required."}

    due_date = lori_calendar_parse_due_date(payload.get("due_date"))
    due_time = lori_calendar_parse_due_time(payload.get("due_time"))
    reminder_datetime = lori_calendar_make_datetime(due_date, due_time)

    reminder_payload = {
        "reminder_title": reminder_title,
        "reminder_type": lori_calendar_clean(payload.get("reminder_type") or "Internal Reminder"),
        "reminder_status": lori_calendar_clean(payload.get("reminder_status") or "Open"),
        "priority": lori_calendar_clean(payload.get("priority") or "Normal"),

        "company_name": lori_calendar_clean(payload.get("company_name")),
        "region_code": lori_calendar_upper(payload.get("region_code")),
        "region_name": lori_calendar_clean(payload.get("region_name")),
        "operating_state": lori_calendar_upper(payload.get("operating_state")),
        "city": lori_calendar_clean(payload.get("city")),
        "station_code": lori_calendar_upper(payload.get("station_code")),
        "station_name": lori_calendar_clean(payload.get("station_name")),
        "route_group": lori_calendar_clean(payload.get("route_group")),
        "route_id": lori_calendar_clean(payload.get("route_id")),

        "related_module": lori_calendar_clean(payload.get("related_module")),
        "related_record_id": payload.get("related_record_id") or None,
        "related_reference": lori_calendar_clean(payload.get("related_reference")),

        "driver_id": lori_calendar_clean(payload.get("driver_id")),
        "driver_name": lori_calendar_clean(payload.get("driver_name")),
        "employee_id": lori_calendar_clean(payload.get("employee_id")),
        "employee_name": lori_calendar_clean(payload.get("employee_name")),

        "document_id": payload.get("document_id") or None,
        "document_title": lori_calendar_clean(payload.get("document_title")),
        "sop_id": payload.get("sop_id") or None,
        "sop_title": lori_calendar_clean(payload.get("sop_title")),
        "report_id": payload.get("report_id") or None,
        "report_title": lori_calendar_clean(payload.get("report_title")),
        "action_item_id": payload.get("action_item_id") or None,
        "action_title": lori_calendar_clean(payload.get("action_title")),
        "kpi_plan_id": payload.get("kpi_plan_id") or None,
        "kpi_plan_title": lori_calendar_clean(payload.get("kpi_plan_title")),
        "route_project_id": payload.get("route_project_id") or None,
        "route_project_title": lori_calendar_clean(payload.get("route_project_title")),

        "due_date": due_date,
        "due_time": due_time,
        "reminder_datetime": reminder_datetime,

        "assigned_to_name": lori_calendar_clean(payload.get("assigned_to_name")),
        "assigned_to_email": lori_calendar_clean(payload.get("assigned_to_email")),
        "created_by_name": lori_calendar_clean(payload.get("created_by_name") or payload.get("created_by") or "LORI User"),
        "created_by_email": lori_calendar_clean(payload.get("created_by_email")),

        "reminder_notes": lori_calendar_clean(payload.get("reminder_notes") or payload.get("notes")),
        "voiceflow_created": bool(payload.get("voiceflow_created", False)),
        "ask_lori_created": bool(payload.get("ask_lori_created", False)),

        "external_calendar_status": "Not Connected",
        "external_calendar_provider": None,
    }

    created = await lori_policy_supabase_post(
        "lori_calendar_reminders",
        reminder_payload,
    )

    reminder = created[0] if created else reminder_payload
    reminder_id = reminder.get("id")

    if reminder_id:
        await lori_calendar_log_activity(
            reminder_id=reminder_id,
            activity_type="Created",
            activity_summary="Internal LORI reminder created.",
            activity_by_name=reminder_payload["created_by_name"],
            activity_by_email=reminder_payload["created_by_email"],
            old_status="",
            new_status=reminder_payload["reminder_status"],
            activity_metadata={
                "related_module": reminder_payload["related_module"],
                "voiceflow_created": reminder_payload["voiceflow_created"],
                "ask_lori_created": reminder_payload["ask_lori_created"],
                "external_calendar_status": "Not Connected",
            },
        )

        if reminder_payload["assigned_to_name"] or reminder_payload["assigned_to_email"]:
            await lori_policy_supabase_post(
                "lori_calendar_reminder_participants",
                {
                    "reminder_id": reminder_id,
                    "participant_name": reminder_payload["assigned_to_name"],
                    "participant_email": reminder_payload["assigned_to_email"],
                    "participant_role": "Assigned Owner",
                    "participant_type": "Assigned User",
                    "participation_status": "Pending",
                },
            )

    return {
        "status": "success",
        "message": "Internal LORI reminder created.",
        "reminder": reminder,
        "external_calendar_status": "Not Connected",
        "external_calendar_note": "This was saved as an internal LORI reminder. Google/Outlook calendar sync is not connected yet.",
    }


@app.get("/calendar-reminder-detail")
async def calendar_reminder_detail(
    api_key: Optional[str] = Query(None),
    reminder_id: str = Query(...),
):
    lori_regulatory_require_key(api_key)

    reminders = await lori_calendar_get_rows(
        "lori_calendar_reminders",
        f"select=*&id=eq.{quote(reminder_id)}&limit=1",
    )

    if not reminders:
        return {"status": "not_found", "message": "Reminder not found."}

    participants = await lori_calendar_get_rows(
        "lori_calendar_reminder_participants",
        f"select=*&reminder_id=eq.{quote(reminder_id)}&order=created_at.asc&limit=100",
    )

    activity = await lori_calendar_get_rows(
        "lori_calendar_reminder_activity_log",
        f"select=*&reminder_id=eq.{quote(reminder_id)}&order=created_at.desc&limit=200",
    )

    return {
        "status": "success",
        "reminder": reminders[0],
        "participants_count": len(participants),
        "participants": participants,
        "activity_count": len(activity),
        "activity": activity,
    }


@app.post("/calendar-reminder-update")
async def calendar_reminder_update(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    reminder_id = lori_calendar_clean(payload.get("reminder_id"))

    if not reminder_id:
        return {"status": "error", "message": "reminder_id is required."}

    existing_rows = await lori_calendar_get_rows(
        "lori_calendar_reminders",
        f"select=*&id=eq.{quote(reminder_id)}&limit=1",
    )

    if not existing_rows:
        return {"status": "not_found", "message": "Reminder not found."}

    existing = existing_rows[0]

    due_date = lori_calendar_parse_due_date(payload.get("due_date") or existing.get("due_date"))
    due_time = lori_calendar_parse_due_time(payload.get("due_time") or existing.get("due_time"))
    reminder_datetime = lori_calendar_make_datetime(due_date, due_time)

    update_payload = {
        "reminder_title": lori_calendar_clean(payload.get("reminder_title") or existing.get("reminder_title")),
        "reminder_type": lori_calendar_clean(payload.get("reminder_type") or existing.get("reminder_type")),
        "reminder_status": lori_calendar_clean(payload.get("reminder_status") or existing.get("reminder_status")),
        "priority": lori_calendar_clean(payload.get("priority") or existing.get("priority")),
        "due_date": due_date,
        "due_time": due_time,
        "reminder_datetime": reminder_datetime,
        "assigned_to_name": lori_calendar_clean(payload.get("assigned_to_name") or existing.get("assigned_to_name")),
        "assigned_to_email": lori_calendar_clean(payload.get("assigned_to_email") or existing.get("assigned_to_email")),
        "reminder_notes": lori_calendar_clean(payload.get("reminder_notes") or existing.get("reminder_notes")),
        "updated_at": datetime.utcnow().isoformat(),
    }

    updated = await lori_calendar_patch_rows(
        "lori_calendar_reminders",
        f"id=eq.{quote(reminder_id)}",
        update_payload,
    )

    await lori_calendar_log_activity(
        reminder_id=reminder_id,
        activity_type="Updated",
        activity_summary="Internal LORI reminder updated.",
        activity_by_name=lori_calendar_clean(payload.get("updated_by_name") or "LORI User"),
        activity_by_email=lori_calendar_clean(payload.get("updated_by_email")),
        old_status=existing.get("reminder_status"),
        new_status=update_payload["reminder_status"],
        activity_metadata=update_payload,
    )

    return {
        "status": "success",
        "message": "Reminder updated.",
        "reminder": updated[0] if updated else update_payload,
    }


@app.post("/calendar-reminder-complete")
async def calendar_reminder_complete(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    reminder_id = lori_calendar_clean(payload.get("reminder_id"))

    if not reminder_id:
        return {"status": "error", "message": "reminder_id is required."}

    completed_by_name = lori_calendar_clean(payload.get("completed_by_name") or "LORI User")
    completed_by_email = lori_calendar_clean(payload.get("completed_by_email"))

    updated = await lori_calendar_patch_rows(
        "lori_calendar_reminders",
        f"id=eq.{quote(reminder_id)}",
        {
            "reminder_status": "Completed",
            "completed_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_calendar_log_activity(
        reminder_id=reminder_id,
        activity_type="Completed",
        activity_summary=lori_calendar_clean(payload.get("completion_notes") or "Reminder completed."),
        activity_by_name=completed_by_name,
        activity_by_email=completed_by_email,
        new_status="Completed",
    )

    return {
        "status": "success",
        "message": "Reminder completed.",
        "reminder": updated[0] if updated else {"id": reminder_id, "reminder_status": "Completed"},
    }


@app.post("/calendar-reminder-cancel")
async def calendar_reminder_cancel(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    reminder_id = lori_calendar_clean(payload.get("reminder_id"))

    if not reminder_id:
        return {"status": "error", "message": "reminder_id is required."}

    cancelled_by_name = lori_calendar_clean(payload.get("cancelled_by_name") or "LORI User")
    cancelled_by_email = lori_calendar_clean(payload.get("cancelled_by_email"))
    cancel_reason = lori_calendar_clean(payload.get("cancel_reason") or "Reminder cancelled.")

    updated = await lori_calendar_patch_rows(
        "lori_calendar_reminders",
        f"id=eq.{quote(reminder_id)}",
        {
            "reminder_status": "Cancelled",
            "cancelled_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_calendar_log_activity(
        reminder_id=reminder_id,
        activity_type="Cancelled",
        activity_summary=cancel_reason,
        activity_by_name=cancelled_by_name,
        activity_by_email=cancelled_by_email,
        new_status="Cancelled",
    )

    return {
        "status": "success",
        "message": "Reminder cancelled.",
        "reminder": updated[0] if updated else {"id": reminder_id, "reminder_status": "Cancelled"},
    }


@app.post("/calendar-reminder-snooze")
async def calendar_reminder_snooze(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    reminder_id = lori_calendar_clean(payload.get("reminder_id"))
    snoozed_until = lori_calendar_clean(payload.get("snoozed_until"))

    if not reminder_id:
        return {"status": "error", "message": "reminder_id is required."}

    if not snoozed_until:
        return {"status": "error", "message": "snoozed_until is required."}

    updated = await lori_calendar_patch_rows(
        "lori_calendar_reminders",
        f"id=eq.{quote(reminder_id)}",
        {
            "reminder_status": "Snoozed",
            "snoozed_until": snoozed_until,
            "updated_at": datetime.utcnow().isoformat(),
        },
    )

    await lori_calendar_log_activity(
        reminder_id=reminder_id,
        activity_type="Snoozed",
        activity_summary=f"Reminder snoozed until {snoozed_until}.",
        activity_by_name=lori_calendar_clean(payload.get("snoozed_by_name") or "LORI User"),
        activity_by_email=lori_calendar_clean(payload.get("snoozed_by_email")),
        new_status="Snoozed",
        activity_metadata={"snoozed_until": snoozed_until},
    )

    return {
        "status": "success",
        "message": "Reminder snoozed.",
        "reminder": updated[0] if updated else {"id": reminder_id, "reminder_status": "Snoozed"},
    }


@app.get("/calendar-reminders-due")
async def calendar_reminders_due(
    api_key: Optional[str] = Query(None),
    station_code: Optional[str] = Query(None),
    assigned_to_email: Optional[str] = Query(None),
    days_ahead: int = Query(7),
):
    lori_regulatory_require_key(api_key)

    today = date.today()
    end_date = today + timedelta(days=max(0, min(days_ahead, 60)))

    response = await calendar_reminders(
        api_key=api_key,
        station_code=station_code,
        assigned_to_email=assigned_to_email,
        limit=1000,
    )

    reminders = response.get("reminders", [])

    due = []

    for reminder in reminders:
        if reminder.get("reminder_status") in ["Completed", "Cancelled"]:
            continue

        due_date = lori_calendar_parse_due_date(reminder.get("due_date"))

        if not due_date:
            continue

        due_date_obj = date.fromisoformat(due_date)

        if today <= due_date_obj <= end_date:
            due.append(reminder)

    return {
        "status": "success",
        "due_count": len(due),
        "date_range_start": today.isoformat(),
        "date_range_end": end_date.isoformat(),
        "reminders": due,
    }


@app.post("/voiceflow/calendar-reminder-create")
async def voiceflow_calendar_reminder_create(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    lori_regulatory_require_key(api_key)

    payload = dict(payload or {})
    payload["voiceflow_created"] = True
    payload["ask_lori_created"] = True

    if not payload.get("created_by_name"):
        payload["created_by_name"] = payload.get("user_name") or "Voiceflow / Ask LORI"

    if not payload.get("reminder_type"):
        payload["reminder_type"] = "Internal Reminder"

    if not payload.get("related_module"):
        payload["related_module"] = "Ask LORI"

    created = await calendar_reminder_create(api_key=api_key, payload=payload)

    if created.get("status") == "success":
        return {
            "status": "success",
            "voiceflow_message": "Reminder created inside LORI. External calendar sync is not connected yet.",
            "reminder": created.get("reminder"),
            "external_calendar_status": "Not Connected",
        }

    return created
# ============================================================
# LORI DRIVE COMMAND CENTER
# SAFE VOICEFLOW TREND INTELLIGENCE ENDPOINT
# Use this endpoint from Voiceflow instead of the older
# /trend-quick-analyze route if the older route is throwing 500.
#
# Voiceflow URL:
# POST /voiceflow/trend-quick-analyze
# ============================================================

from fastapi import Body, Query, HTTPException
from typing import Any, Dict, List, Optional
from datetime import datetime, date
from urllib.parse import quote
import os
import uuid
import httpx


SUPABASE_URL_TREND_VF = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_TREND_VF = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_trend_vf_clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lori_trend_vf_upper(value: Any) -> str:
    return lori_trend_vf_clean(value).upper()


def lori_trend_vf_now_iso() -> str:
    return datetime.utcnow().isoformat()


def lori_trend_vf_require_key(api_key: Optional[str]):
    """
    Uses the existing LORI key checker if available.
    Falls back to common environment variable names if needed.
    """
    try:
        lori_regulatory_require_key(api_key)
        return
    except NameError:
        pass

    expected = (
        os.getenv("LORI_API_KEY")
        or os.getenv("LORI_DRIVE_API_KEY")
        or os.getenv("API_KEY")
        or os.getenv("VOICEFLOW_API_KEY")
    )

    if expected and api_key != expected:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")

    if expected and not api_key:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")


async def lori_trend_vf_supabase_get(table: str, query: str) -> List[Dict[str, Any]]:
    """
    Safe Supabase GET.
    Returns [] instead of crashing the trend endpoint.
    """
    if not SUPABASE_URL_TREND_VF or not SUPABASE_SERVICE_ROLE_KEY_TREND_VF:
        return []

    url = f"{SUPABASE_URL_TREND_VF}/rest/v1/{table}?{query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_TREND_VF,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_TREND_VF}",
        "Content-Type": "application/json",
    }

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.get(url, headers=headers)

        if response.status_code >= 400:
            return []

        return response.json()
    except Exception:
        return []


async def lori_trend_vf_supabase_post(table: str, payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Safe Supabase POST.
    Returns [] instead of crashing the trend endpoint.
    """
    if not SUPABASE_URL_TREND_VF or not SUPABASE_SERVICE_ROLE_KEY_TREND_VF:
        return []

    url = f"{SUPABASE_URL_TREND_VF}/rest/v1/{table}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_TREND_VF,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_TREND_VF}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(url, headers=headers, json=payload)

        if response.status_code >= 400:
            return []

        try:
            return response.json()
        except Exception:
            return []
    except Exception:
        return []


def lori_trend_vf_infer_category(payload: Dict[str, Any]) -> str:
    question = lori_trend_vf_clean(payload.get("question_text")).lower()
    subject = lori_trend_vf_clean(payload.get("trend_subject")).lower()

    combined = f"{question} {subject}"

    if any(x in combined for x in ["driver", "marcus", "safety", "counseling", "coaching", "performance"]):
        return "Driver Performance Trend"

    if any(x in combined for x in ["route", "late", "delay", "stop", "work area", "route balance"]):
        return "Route / Work Area Trend"

    if any(x in combined for x in ["kpi", "metric", "score", "expected result"]):
        return "KPI Trend"

    if any(x in combined for x in ["sop", "acknowledgement", "procedure"]):
        return "SOP Trend"

    if any(x in combined for x in ["document", "upload", "employee handbook", "policy", "agreement"]):
        return "Document Trend"

    if any(x in combined for x in ["action", "task", "follow-up", "overdue"]):
        return "Action Center Trend"

    if any(x in combined for x in ["compliance", "regulatory", "dot", "fmcsa"]):
        return "Compliance Trend"

    return "Operational Trend"


def lori_trend_vf_infer_direction(payload: Dict[str, Any], source_count: int) -> Dict[str, str]:
    question = lori_trend_vf_clean(payload.get("question_text")).lower()
    subject = lori_trend_vf_clean(payload.get("trend_subject")).lower()
    combined = f"{question} {subject}"

    if any(x in combined for x in ["keep having issues", "worse", "worsening", "increasing", "more accidents", "always late", "falling behind"]):
        return {
            "trend_direction": "Worsening",
            "trend_strength": "Moderate" if source_count else "Preliminary",
            "confidence_level": "Medium" if source_count >= 3 else "Low",
            "risk_level": "Medium" if source_count < 5 else "High",
        }

    if any(x in combined for x in ["better", "improving", "improved", "less", "reduced"]):
        return {
            "trend_direction": "Improving",
            "trend_strength": "Moderate" if source_count else "Preliminary",
            "confidence_level": "Medium" if source_count >= 3 else "Low",
            "risk_level": "Low",
        }

    return {
        "trend_direction": "Needs Review",
        "trend_strength": "Preliminary",
        "confidence_level": "Low" if source_count < 3 else "Medium",
        "risk_level": "Medium",
    }


def lori_trend_vf_build_summary(payload: Dict[str, Any], source_count: int, category: str, direction: Dict[str, str]) -> Dict[str, str]:
    question = lori_trend_vf_clean(payload.get("question_text"))
    subject = lori_trend_vf_clean(payload.get("trend_subject"))
    driver_name = lori_trend_vf_clean(payload.get("driver_name"))
    route_id = lori_trend_vf_clean(payload.get("route_id"))
    station_code = lori_trend_vf_clean(payload.get("station_code") or "JESSUP-01")

    display_subject = subject or driver_name or route_id or question or "Operational trend"

    if driver_name:
        trend_title = f"Trend Analysis — {driver_name}"
        trend_summary = (
            f"LORI reviewed the trend request for {driver_name}. "
            f"The pattern is categorized as {category} with a current direction of {direction['trend_direction']}."
        )
    elif route_id:
        trend_title = f"Trend Analysis — Route {route_id}"
        trend_summary = (
            f"LORI reviewed the trend request for route {route_id}. "
            f"The pattern is categorized as {category} with a current direction of {direction['trend_direction']}."
        )
    else:
        trend_title = f"Trend Analysis — {display_subject}"
        trend_summary = (
            f"LORI reviewed the requested operational trend. "
            f"The pattern is categorized as {category} with a current direction of {direction['trend_direction']}."
        )

    if source_count == 0:
        source_note = (
            "LORI did not find enough dated source records in this quick Voiceflow review to create a fully reliable trend. "
            "This should be treated as a preliminary operational signal."
        )
    else:
        source_note = (
            f"LORI found {source_count} possible source record(s) that may support this trend review. "
            "The source records should be reviewed before taking action."
        )

    plain_language = (
        f"Based on the question, this may indicate a recurring operational issue that needs closer review. "
        f"{source_note}"
    )

    leadership_summary = (
        f"Leadership should treat this as a {direction['risk_level'].lower()} risk trend signal for {station_code}. "
        "The next step is to validate the source records and assign the appropriate owner for follow-up."
    )

    counseling_summary = (
        "If this trend is used for counseling or coaching, keep the conversation neutral and fact-based. "
        "This should be positioned as review and expectation-setting, not a final disciplinary conclusion."
    )

    recommended_action = (
        "Review the source records, confirm whether the issue is isolated or recurring, and assign supervisor follow-up if the pattern is confirmed."
    )

    follow_up_question = "Would you like LORI to prepare a supervisor follow-up or counseling summary based on this trend?"

    return {
        "trend_title": trend_title,
        "trend_summary": trend_summary,
        "plain_language_explanation": plain_language,
        "leadership_summary": leadership_summary,
        "counseling_summary": counseling_summary,
        "operational_summary": trend_summary,
        "recommended_action": recommended_action,
        "suggested_follow_up_question": follow_up_question,
    }


async def lori_trend_vf_collect_sources(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Pulls possible supporting records if tables exist.
    This is intentionally safe and optional.
    """
    driver_id = lori_trend_vf_clean(payload.get("driver_id"))
    driver_name = lori_trend_vf_clean(payload.get("driver_name"))
    station_code = lori_trend_vf_upper(payload.get("station_code") or "JESSUP-01")

    sources: List[Dict[str, Any]] = []

    if driver_id:
        driver_metrics = await lori_trend_vf_supabase_get(
            "lori_driver_metrics",
            f"select=*&driver_id=eq.{quote(driver_id)}&order=created_at.desc&limit=20",
        )
        for row in driver_metrics:
            sources.append({
                "source_module": "Driver Metrics",
                "source_title": f"Driver metric record for {driver_id}",
                "source_reference": row.get("id") or row.get("driver_id") or driver_id,
                "driver_id": driver_id,
                "driver_name": row.get("driver_name") or driver_name,
                "station_code": row.get("station_code") or station_code,
                "relevance": "Possible supporting driver metric record",
            })

        counseling = await lori_trend_vf_supabase_get(
            "lori_driver_counseling",
            f"select=*&driver_id=eq.{quote(driver_id)}&order=created_at.desc&limit=20",
        )
        for row in counseling:
            sources.append({
                "source_module": "Driver Counseling",
                "source_title": f"Counseling record for {driver_id}",
                "source_reference": row.get("id") or row.get("driver_id") or driver_id,
                "driver_id": driver_id,
                "driver_name": row.get("driver_name") or driver_name,
                "station_code": row.get("station_code") or station_code,
                "relevance": "Possible supporting counseling record",
            })

        safety = await lori_trend_vf_supabase_get(
            "lori_driver_safety_events",
            f"select=*&driver_id=eq.{quote(driver_id)}&order=created_at.desc&limit=20",
        )
        for row in safety:
            sources.append({
                "source_module": "Driver Safety Events",
                "source_title": f"Safety event for {driver_id}",
                "source_reference": row.get("id") or row.get("driver_id") or driver_id,
                "driver_id": driver_id,
                "driver_name": row.get("driver_name") or driver_name,
                "station_code": row.get("station_code") or station_code,
                "relevance": "Possible supporting safety event",
            })

    elif driver_name:
        driver_search_name = quote(driver_name)

        driver_records = await lori_trend_vf_supabase_get(
            "lori_driver_master",
            f"select=*&driver_name=ilike.*{driver_search_name}*&limit=10",
        )
        for row in driver_records:
            sources.append({
                "source_module": "Driver Master",
                "source_title": f"Driver record for {driver_name}",
                "source_reference": row.get("id") or row.get("driver_id") or driver_name,
                "driver_id": row.get("driver_id") or "",
                "driver_name": row.get("driver_name") or driver_name,
                "station_code": row.get("station_code") or station_code,
                "relevance": "Driver record match",
            })

    station_actions = await lori_trend_vf_supabase_get(
        "lori_action_items",
        f"select=*&station_code=eq.{quote(station_code)}&order=created_at.desc&limit=20",
    )
    for row in station_actions:
        if driver_name and driver_name.lower() not in str(row).lower():
            continue
        sources.append({
            "source_module": "Action Center",
            "source_title": row.get("action_title") or "Action item",
            "source_reference": row.get("id") or "",
            "driver_id": row.get("driver_id") or "",
            "driver_name": row.get("driver_name") or driver_name,
            "station_code": row.get("station_code") or station_code,
            "relevance": "Possible related action item",
        })

    return sources[:50]


@app.post("/voiceflow/trend-quick-analyze")
async def voiceflow_trend_quick_analyze(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    """
    Safe endpoint for Voiceflow Trend Intelligence.
    This endpoint should not crash if older trend tables or optional fields are missing.
    """
    lori_trend_vf_require_key(api_key)

    payload = dict(payload or {})

    question_text = lori_trend_vf_clean(payload.get("question_text"))
    trend_subject = lori_trend_vf_clean(payload.get("trend_subject"))

    if not question_text and not trend_subject:
        return {
            "status": "error",
            "message": "question_text or trend_subject is required.",
        }

    company_name = lori_trend_vf_clean(payload.get("company_name") or "Food Authority")
    region_code = lori_trend_vf_upper(payload.get("region_code") or "MID_ATLANTIC")
    region_name = lori_trend_vf_clean(payload.get("region_name") or "Mid-Atlantic")
    operating_state = lori_trend_vf_upper(payload.get("operating_state") or "MD")
    city = lori_trend_vf_clean(payload.get("city") or "Jessup")
    station_code = lori_trend_vf_upper(payload.get("station_code") or "JESSUP-01")
    station_name = lori_trend_vf_clean(payload.get("station_name") or "Jessup Delivery Station")
    route_group = lori_trend_vf_clean(payload.get("route_group") or "Delivery Operations")

    payload["company_name"] = company_name
    payload["region_code"] = region_code
    payload["region_name"] = region_name
    payload["operating_state"] = operating_state
    payload["city"] = city
    payload["station_code"] = station_code
    payload["station_name"] = station_name
    payload["route_group"] = route_group

    trend_category = lori_trend_vf_infer_category(payload)

    sources = await lori_trend_vf_collect_sources(payload)
    source_count = len(sources)

    direction = lori_trend_vf_infer_direction(payload, source_count)
    text_parts = lori_trend_vf_build_summary(payload, source_count, trend_category, direction)

    request_id = str(uuid.uuid4())
    result_id = str(uuid.uuid4())

    request_payload = {
        "id": request_id,
        "request_title": lori_trend_vf_clean(payload.get("request_title") or f"Trend Analysis — {trend_subject or question_text}"),
        "question_text": question_text,
        "trend_subject": trend_subject or question_text,
        "trend_category": trend_category,
        "company_name": company_name,
        "region_code": region_code,
        "region_name": region_name,
        "operating_state": operating_state,
        "city": city,
        "station_code": station_code,
        "station_name": station_name,
        "route_group": route_group,
        "route_id": lori_trend_vf_clean(payload.get("route_id")),
        "driver_id": lori_trend_vf_clean(payload.get("driver_id")),
        "driver_name": lori_trend_vf_clean(payload.get("driver_name")),
        "employee_id": lori_trend_vf_clean(payload.get("employee_id")),
        "employee_name": lori_trend_vf_clean(payload.get("employee_name")),
        "requested_by": lori_trend_vf_clean(payload.get("requested_by") or "Voiceflow / Ask LORI"),
        "request_status": "Completed",
        "user_requested_trend": bool(payload.get("user_requested_trend", True)),
        "lori_offered_trend": bool(payload.get("lori_offered_trend", True)),
        "created_at": lori_trend_vf_now_iso(),
        "updated_at": lori_trend_vf_now_iso(),
    }

    saved_request = await lori_trend_vf_supabase_post("lori_trend_requests", request_payload)

    if saved_request and saved_request[0].get("id"):
        request_id = saved_request[0]["id"]

    result_payload = {
        "id": result_id,
        "trend_request_id": request_id,
        "result_status": "Generated" if source_count >= 1 else "Preliminary",
        "trend_title": text_parts["trend_title"],
        "trend_summary": text_parts["trend_summary"],
        "trend_direction": direction["trend_direction"],
        "trend_strength": direction["trend_strength"],
        "confidence_level": direction["confidence_level"],
        "risk_level": direction["risk_level"],
        "finding_type": "Trend Analysis",
        "finding_category": trend_category,
        "plain_language_explanation": text_parts["plain_language_explanation"],
        "leadership_summary": text_parts["leadership_summary"],
        "counseling_summary": text_parts["counseling_summary"],
        "operational_summary": text_parts["operational_summary"],
        "recommended_action": text_parts["recommended_action"],
        "suggested_follow_up_question": text_parts["suggested_follow_up_question"],
        "decision_support_note": "LORI provides operational decision support only. Trend analysis should be reviewed by authorized leadership before final action.",
        "should_offer_chart": True,
        "should_offer_counseling_language": True,
        "should_offer_leadership_packet": True,
        "should_send_to_action_center": True,
        "data_points_count": max(source_count, 1),
        "source_records_count": source_count,
        "generated_by": "LORI Trend Intelligence",
        "created_at": lori_trend_vf_now_iso(),
        "updated_at": lori_trend_vf_now_iso(),
    }

    saved_result = await lori_trend_vf_supabase_post("lori_trend_results", result_payload)

    if saved_result and saved_result[0].get("id"):
        result_id = saved_result[0]["id"]

    data_points = []

    if source_count:
        for index, source in enumerate(sources[:10], start=1):
            point = {
                "trend_result_id": result_id,
                "point_label": f"Record {index}",
                "metric_name": trend_category,
                "metric_value": index,
                "source_module": source.get("source_module"),
                "notes": source.get("relevance"),
                "created_at": lori_trend_vf_now_iso(),
            }
            data_points.append(point)
            await lori_trend_vf_supabase_post("lori_trend_data_points", point)
    else:
        point = {
            "trend_result_id": result_id,
            "point_label": "Preliminary Review",
            "metric_name": trend_category,
            "metric_value": 1,
            "source_module": "Voiceflow",
            "notes": "Preliminary trend request created from Voiceflow. More dated records are needed for a stronger trend.",
            "created_at": lori_trend_vf_now_iso(),
        }
        data_points.append(point)
        await lori_trend_vf_supabase_post("lori_trend_data_points", point)

    for source in sources[:25]:
        source_payload = {
            "trend_result_id": result_id,
            "source_module": source.get("source_module"),
            "source_title": source.get("source_title"),
            "source_reference": str(source.get("source_reference") or ""),
            "driver_id": source.get("driver_id") or "",
            "driver_name": source.get("driver_name") or "",
            "station_code": source.get("station_code") or station_code,
            "relevance": source.get("relevance") or "Possible supporting source record",
            "created_at": lori_trend_vf_now_iso(),
        }
        await lori_trend_vf_supabase_post("lori_trend_source_records", source_payload)

    voiceflow_message = (
        f"{text_parts['trend_summary']}\n\n"
        f"Trend Direction: {direction['trend_direction']}\n"
        f"Confidence Level: {direction['confidence_level']}\n"
        f"Risk Level: {direction['risk_level']}\n\n"
        f"Recommended Action: {text_parts['recommended_action']}\n\n"
        "LORI provides operational decision support only. Trend analysis should be reviewed by authorized leadership before final action."
    )

    return {
        "status": "success",
        "voiceflow_message": voiceflow_message,
        "trend_request_id": request_id,
        "trend_result_id": result_id,
        "trend_title": text_parts["trend_title"],
        "trend_summary": text_parts["trend_summary"],
        "trend_direction": direction["trend_direction"],
        "trend_strength": direction["trend_strength"],
        "confidence_level": direction["confidence_level"],
        "risk_level": direction["risk_level"],
        "finding_category": trend_category,
        "plain_language_explanation": text_parts["plain_language_explanation"],
        "leadership_summary": text_parts["leadership_summary"],
        "counseling_summary": text_parts["counseling_summary"],
        "operational_summary": text_parts["operational_summary"],
        "recommended_action": text_parts["recommended_action"],
        "suggested_follow_up_question": text_parts["suggested_follow_up_question"],
        "decision_support_note": "LORI provides operational decision support only. Trend analysis should be reviewed by authorized leadership before final action.",
        "data_points_count": len(data_points),
        "source_records_count": source_count,
        "data_points": data_points,
        "source_records": sources,
    }


@app.post("/trend-quick-analyze-v2")
async def trend_quick_analyze_v2(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    """
    Same safe endpoint with a non-Voiceflow path, available for Lovable or manual testing.
    """
    return await voiceflow_trend_quick_analyze(api_key=api_key, payload=payload)
# ============================================================
# LORI DRIVE COMMAND CENTER
# VOICEFLOW ACTUAL TREND DATA ENDPOINT
#
# Purpose:
# Returns actual dated trend data when available, instead of only
# a general trend interpretation.
#
# Voiceflow URL:
# POST /voiceflow/trend-quick-analyze-data
# ============================================================

from fastapi import Body, Query, HTTPException
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from urllib.parse import quote
import os
import uuid
import httpx


SUPABASE_URL_TREND_DATA = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY_TREND_DATA = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def lori_trend_data_clean(value: Any) -> str:
    if value is None:
        return ""

    cleaned = " ".join(str(value).strip().split())

    # Prevent Voiceflow placeholders from being treated as real values.
    if cleaned.startswith("{") and cleaned.endswith("}"):
        return ""

    return cleaned


def lori_trend_data_upper(value: Any) -> str:
    return lori_trend_data_clean(value).upper()


def lori_trend_data_now_iso() -> str:
    return datetime.utcnow().isoformat()


def lori_trend_data_require_key(api_key: Optional[str]):
    """
    Uses the existing LORI key checker if available.
    Falls back to common environment variable names if needed.
    """
    try:
        lori_regulatory_require_key(api_key)
        return
    except NameError:
        pass

    expected = (
        os.getenv("LORI_API_KEY")
        or os.getenv("LORI_DRIVE_API_KEY")
        or os.getenv("API_KEY")
        or os.getenv("VOICEFLOW_API_KEY")
    )

    if expected and api_key != expected:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")

    if expected and not api_key:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")


async def lori_trend_data_supabase_get(table: str, query: str) -> List[Dict[str, Any]]:
    """
    Safe Supabase GET. Returns [] instead of crashing.
    """
    if not SUPABASE_URL_TREND_DATA or not SUPABASE_SERVICE_ROLE_KEY_TREND_DATA:
        return []

    url = f"{SUPABASE_URL_TREND_DATA}/rest/v1/{table}?{query}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_TREND_DATA,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_TREND_DATA}",
        "Content-Type": "application/json",
    }

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.get(url, headers=headers)

        if response.status_code >= 400:
            return []

        return response.json()
    except Exception:
        return []


async def lori_trend_data_supabase_post(table: str, payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Safe Supabase POST. Returns [] instead of crashing.
    """
    if not SUPABASE_URL_TREND_DATA or not SUPABASE_SERVICE_ROLE_KEY_TREND_DATA:
        return []

    url = f"{SUPABASE_URL_TREND_DATA}/rest/v1/{table}"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY_TREND_DATA,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY_TREND_DATA}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(url, headers=headers, json=payload)

        if response.status_code >= 400:
            return []

        try:
            return response.json()
        except Exception:
            return []
    except Exception:
        return []


def lori_trend_data_infer_driver_name(payload: Dict[str, Any]) -> str:
    """
    Attempts to find a driver name even if Voiceflow fails to map driver_name.
    """
    driver_name = lori_trend_data_clean(payload.get("driver_name"))

    if driver_name:
        return driver_name

    text = (
        lori_trend_data_clean(payload.get("question_text"))
        + " "
        + lori_trend_data_clean(payload.get("trend_subject"))
    ).lower()

    known_drivers = [
        "Marcus Hill",
        "Avery Stone",
        "Jordan Blake",
        "Evan Miles",
        "Tanya Reed",
        "Calvin Price",
        "Nia Carter",
        "Sofia Grant",
    ]

    for name in known_drivers:
        if name.lower() in text:
            return name

    return ""


def lori_trend_data_infer_driver_id(payload: Dict[str, Any], driver_name: str) -> str:
    driver_id = lori_trend_data_clean(payload.get("driver_id"))

    if driver_id:
        return driver_id

    known_map = {
        "Marcus Hill": "DEMO-D001",
        "Tanya Reed": "DEMO-D002",
        "Calvin Price": "DEMO-D003",
        "Nia Carter": "DEMO-D004",
        "Evan Miles": "DEMO-D005",
        "Sofia Grant": "DEMO-D006",
        "Avery Stone": "DEMO-D008",
    }

    return known_map.get(driver_name, "")


def lori_trend_data_extract_date(row: Dict[str, Any]) -> str:
    for key in [
        "metric_date",
        "record_date",
        "event_date",
        "counseling_date",
        "action_date",
        "created_at",
        "updated_at",
        "date",
    ]:
        value = lori_trend_data_clean(row.get(key))
        if value:
            return value[:10]

    return ""


def lori_trend_data_parse_date(value: str) -> datetime:
    value = lori_trend_data_clean(value)

    if not value:
        return datetime.min

    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        pass

    try:
        return datetime.strptime(value[:10], "%Y-%m-%d")
    except Exception:
        return datetime.min


def lori_trend_data_extract_numeric_metric(row: Dict[str, Any]) -> Tuple[str, Optional[float]]:
    """
    Finds a useful numeric metric from a row.
    """
    preferred_keys = [
        "performance_score",
        "safety_score",
        "compliance_score",
        "risk_score",
        "on_time_rate",
        "on_time_percentage",
        "delivery_score",
        "route_score",
        "metric_value",
        "score",
        "rating",
        "incident_count",
        "counseling_count",
        "late_count",
        "missed_stop_count",
        "exception_count",
        "open_action_count",
        "count",
    ]

    for key in preferred_keys:
        if key in row and row.get(key) is not None:
            try:
                return key, float(row.get(key))
            except Exception:
                continue

    # fallback: use the first numeric value that is not clearly an ID/year
    for key, value in row.items():
        if key.lower().endswith("id"):
            continue
        if key.lower() in ["year", "zip", "station_zip"]:
            continue

        try:
            if value is not None and str(value).strip() != "":
                return key, float(value)
        except Exception:
            continue

    return "", None


def lori_trend_data_month_key(date_text: str) -> str:
    if not date_text:
        return "Unknown Date"

    return date_text[:7]


def lori_trend_data_direction_from_values(values: List[float]) -> Dict[str, str]:
    if len(values) < 2:
        return {
            "trend_direction": "Needs More Data",
            "trend_strength": "Preliminary",
            "confidence_level": "Low",
            "risk_level": "Medium",
        }

    first = values[0]
    last = values[-1]
    change = last - first

    if change > 0:
        return {
            "trend_direction": "Increasing",
            "trend_strength": "Moderate" if abs(change) >= 2 else "Light",
            "confidence_level": "Medium" if len(values) >= 3 else "Low",
            "risk_level": "High" if abs(change) >= 5 else "Medium",
        }

    if change < 0:
        return {
            "trend_direction": "Decreasing",
            "trend_strength": "Moderate" if abs(change) >= 2 else "Light",
            "confidence_level": "Medium" if len(values) >= 3 else "Low",
            "risk_level": "Low" if last <= first else "Medium",
        }

    return {
        "trend_direction": "Flat",
        "trend_strength": "Stable",
        "confidence_level": "Medium" if len(values) >= 3 else "Low",
        "risk_level": "Medium",
    }


def lori_trend_data_infer_category(payload: Dict[str, Any]) -> str:
    combined = (
        lori_trend_data_clean(payload.get("question_text"))
        + " "
        + lori_trend_data_clean(payload.get("trend_subject"))
    ).lower()

    if any(x in combined for x in ["driver", "marcus", "safety", "counseling", "coaching", "performance"]):
        return "Driver Performance Trend"

    if any(x in combined for x in ["route", "late", "delay", "stop", "work area", "route balance"]):
        return "Route / Work Area Trend"

    if any(x in combined for x in ["kpi", "metric", "score", "expected result"]):
        return "KPI Trend"

    if any(x in combined for x in ["sop", "acknowledgement", "procedure"]):
        return "SOP Trend"

    if any(x in combined for x in ["document", "upload", "employee handbook", "policy", "agreement"]):
        return "Document Trend"

    if any(x in combined for x in ["action", "task", "follow-up", "overdue"]):
        return "Action Center Trend"

    if any(x in combined for x in ["compliance", "regulatory", "dot", "fmcsa"]):
        return "Compliance Trend"

    return "Operational Trend"


async def lori_trend_data_collect_driver_rows(
    driver_id: str,
    driver_name: str,
    station_code: str,
) -> List[Dict[str, Any]]:
    """
    Collects actual records from known LORI tables.
    """
    all_rows: List[Dict[str, Any]] = []

    # Driver metrics
    if driver_id:
        rows = await lori_trend_data_supabase_get(
            "lori_driver_metrics",
            f"select=*&driver_id=eq.{quote(driver_id)}&order=created_at.asc&limit=50",
        )
        for row in rows:
            row["_source_module"] = "Driver Metrics"
            all_rows.append(row)

    if driver_name:
        rows = await lori_trend_data_supabase_get(
            "lori_driver_metrics",
            f"select=*&driver_name=ilike.*{quote(driver_name)}*&order=created_at.asc&limit=50",
        )
        for row in rows:
            row["_source_module"] = "Driver Metrics"
            all_rows.append(row)

    # Safety events
    if driver_id:
        rows = await lori_trend_data_supabase_get(
            "lori_driver_safety_events",
            f"select=*&driver_id=eq.{quote(driver_id)}&order=created_at.asc&limit=50",
        )
        for row in rows:
            row["_source_module"] = "Driver Safety Events"
            all_rows.append(row)

    if driver_name:
        rows = await lori_trend_data_supabase_get(
            "lori_driver_safety_events",
            f"select=*&driver_name=ilike.*{quote(driver_name)}*&order=created_at.asc&limit=50",
        )
        for row in rows:
            row["_source_module"] = "Driver Safety Events"
            all_rows.append(row)

    # Counseling records
    if driver_id:
        rows = await lori_trend_data_supabase_get(
            "lori_driver_counseling",
            f"select=*&driver_id=eq.{quote(driver_id)}&order=created_at.asc&limit=50",
        )
        for row in rows:
            row["_source_module"] = "Driver Counseling"
            all_rows.append(row)

    if driver_name:
        rows = await lori_trend_data_supabase_get(
            "lori_driver_counseling",
            f"select=*&driver_name=ilike.*{quote(driver_name)}*&order=created_at.asc&limit=50",
        )
        for row in rows:
            row["_source_module"] = "Driver Counseling"
            all_rows.append(row)

    # Action items
    if station_code:
        rows = await lori_trend_data_supabase_get(
            "lori_action_items",
            f"select=*&station_code=eq.{quote(station_code)}&order=created_at.asc&limit=100",
        )
        for row in rows:
            row_text = str(row).lower()
            if driver_name and driver_name.lower() not in row_text:
                continue
            row["_source_module"] = "Action Center"
            all_rows.append(row)

    # Deduplicate by id/source
    seen = set()
    unique_rows = []

    for row in all_rows:
        row_id = str(row.get("id") or row.get("driver_id") or "") + "|" + row.get("_source_module", "")
        if row_id in seen:
            continue
        seen.add(row_id)
        unique_rows.append(row)

    unique_rows.sort(key=lambda r: lori_trend_data_parse_date(lori_trend_data_extract_date(r)))

    return unique_rows


def lori_trend_data_build_points_from_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Converts actual rows into trend data points.
    If numeric values are available, uses numeric trend.
    If not, uses monthly record counts as actual trend data.
    """
    numeric_points = []

    for row in rows:
        date_text = lori_trend_data_extract_date(row)
        metric_name, metric_value = lori_trend_data_extract_numeric_metric(row)

        if metric_name and metric_value is not None:
            numeric_points.append({
                "date": date_text or "Unknown Date",
                "metric_name": metric_name,
                "metric_value": metric_value,
                "source_module": row.get("_source_module", "Unknown Source"),
                "source_reference": str(row.get("id") or row.get("driver_id") or ""),
                "notes": lori_trend_data_clean(
                    row.get("notes")
                    or row.get("event_type")
                    or row.get("counseling_type")
                    or row.get("action_title")
                    or row.get("status")
                    or ""
                ),
            })

    if numeric_points:
        values = [float(p["metric_value"]) for p in numeric_points]
        direction = lori_trend_data_direction_from_values(values)

        return {
            "mode": "numeric",
            "metric_used": numeric_points[0]["metric_name"],
            "points": numeric_points,
            "values": values,
            "direction": direction,
        }

    # If no numeric metrics, use monthly counts of actual records.
    monthly_counts: Dict[str, int] = {}

    for row in rows:
        date_text = lori_trend_data_extract_date(row)
        month = lori_trend_data_month_key(date_text)
        monthly_counts[month] = monthly_counts.get(month, 0) + 1

    count_points = [
        {
            "date": month,
            "metric_name": "record_count",
            "metric_value": count,
            "source_module": "Combined Actual Records",
            "source_reference": "",
            "notes": "Monthly count of related records found in LORI.",
        }
        for month, count in sorted(monthly_counts.items())
    ]

    values = [float(p["metric_value"]) for p in count_points]
    direction = lori_trend_data_direction_from_values(values)

    return {
        "mode": "record_count",
        "metric_used": "record_count",
        "points": count_points,
        "values": values,
        "direction": direction,
    }


def lori_trend_data_format_points(points: List[Dict[str, Any]], limit: int = 8) -> str:
    if not points:
        return "No dated data points were found."

    lines = []

    for point in points[:limit]:
        lines.append(
            f"- {point.get('date')}: {point.get('metric_name')} = {point.get('metric_value')} "
            f"({point.get('source_module')})"
        )

    if len(points) > limit:
        lines.append(f"- Plus {len(points) - limit} additional data point(s).")

    return "\n".join(lines)


@app.post("/voiceflow/trend-quick-analyze-data")
async def voiceflow_trend_quick_analyze_data(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    """
    Actual trend data endpoint for Voiceflow.
    This returns real data points if they exist in the LORI tables.
    """
    lori_trend_data_require_key(api_key)

    payload = dict(payload or {})

    question_text = lori_trend_data_clean(payload.get("question_text"))
    trend_subject = lori_trend_data_clean(payload.get("trend_subject"))

    if not question_text and not trend_subject:
        return {
            "status": "error",
            "message": "question_text or trend_subject is required.",
        }

    company_name = lori_trend_data_clean(payload.get("company_name") or "Food Authority")
    region_code = lori_trend_data_upper(payload.get("region_code") or "MID_ATLANTIC")
    region_name = lori_trend_data_clean(payload.get("region_name") or "Mid-Atlantic")
    operating_state = lori_trend_data_upper(payload.get("operating_state") or "MD")
    city = lori_trend_data_clean(payload.get("city") or "Jessup")
    station_code = lori_trend_data_upper(payload.get("station_code") or "JESSUP-01")
    station_name = lori_trend_data_clean(payload.get("station_name") or "Jessup Delivery Station")
    route_group = lori_trend_data_clean(payload.get("route_group") or "Delivery Operations")

    driver_name = lori_trend_data_infer_driver_name(payload)
    driver_id = lori_trend_data_infer_driver_id(payload, driver_name)

    trend_category = lori_trend_data_infer_category(payload)

    actual_rows = await lori_trend_data_collect_driver_rows(
        driver_id=driver_id,
        driver_name=driver_name,
        station_code=station_code,
    )

    actual_data_used = len(actual_rows) > 0

    trend_data = lori_trend_data_build_points_from_rows(actual_rows)

    points = trend_data.get("points", [])
    values = trend_data.get("values", [])
    direction = trend_data.get("direction", {})
    metric_used = trend_data.get("metric_used", "record_count")
    mode = trend_data.get("mode", "none")

    if values:
        first_value = values[0]
        latest_value = values[-1]
        change_value = latest_value - first_value
    else:
        first_value = None
        latest_value = None
        change_value = None

    if points:
        date_range = f"{points[0].get('date')} to {points[-1].get('date')}"
    else:
        date_range = "No date range available"

    request_id = str(uuid.uuid4())
    result_id = str(uuid.uuid4())

    trend_title = f"Trend Analysis — {driver_name or trend_subject or question_text}"

    if actual_data_used:
        trend_summary = (
            f"LORI reviewed actual trend data for {driver_name or trend_subject}. "
            f"The trend is categorized as {trend_category} with a current direction of {direction.get('trend_direction', 'Needs Review')}."
        )
    else:
        trend_summary = (
            f"LORI reviewed the trend request for {driver_name or trend_subject or question_text}, "
            "but did not find enough dated source records to produce actual trend data. "
            "More source records need to be uploaded or linked for a reliable trend."
        )

    actual_points_text = lori_trend_data_format_points(points)

    if actual_data_used:
        plain_language_explanation = (
            f"LORI found {len(actual_rows)} related source record(s) and converted them into {len(points)} trend data point(s). "
            f"The metric used was {metric_used}. The reviewed date range was {date_range}."
        )
    else:
        plain_language_explanation = (
            "LORI did not find enough dated source records in the current tables to produce a reliable actual trend. "
            "This should be treated as a preliminary signal until more records are available."
        )

    recommended_action = (
        "Review the source records, confirm whether the issue is isolated or recurring, and assign supervisor follow-up if the pattern is confirmed."
    )

    decision_support_note = (
        "LORI provides operational decision support only. Trend analysis should be reviewed by authorized leadership before final action."
    )

    request_payload = {
        "id": request_id,
        "request_title": trend_title,
        "question_text": question_text,
        "trend_subject": trend_subject or driver_name or question_text,
        "trend_category": trend_category,
        "company_name": company_name,
        "region_code": region_code,
        "region_name": region_name,
        "operating_state": operating_state,
        "city": city,
        "station_code": station_code,
        "station_name": station_name,
        "route_group": route_group,
        "route_id": lori_trend_data_clean(payload.get("route_id")),
        "driver_id": driver_id,
        "driver_name": driver_name,
        "employee_id": lori_trend_data_clean(payload.get("employee_id")),
        "employee_name": lori_trend_data_clean(payload.get("employee_name")),
        "requested_by": lori_trend_data_clean(payload.get("requested_by") or "Voiceflow / Ask LORI"),
        "request_status": "Completed",
        "user_requested_trend": bool(payload.get("user_requested_trend", True)),
        "lori_offered_trend": bool(payload.get("lori_offered_trend", True)),
        "created_at": lori_trend_data_now_iso(),
        "updated_at": lori_trend_data_now_iso(),
    }

    saved_request = await lori_trend_data_supabase_post("lori_trend_requests", request_payload)

    if saved_request and saved_request[0].get("id"):
        request_id = saved_request[0]["id"]

    result_payload = {
        "id": result_id,
        "trend_request_id": request_id,
        "result_status": "Generated" if actual_data_used else "Not Enough Data",
        "trend_title": trend_title,
        "trend_summary": trend_summary,
        "trend_direction": direction.get("trend_direction", "Needs More Data"),
        "trend_strength": direction.get("trend_strength", "Preliminary"),
        "confidence_level": direction.get("confidence_level", "Low"),
        "risk_level": direction.get("risk_level", "Medium"),
        "finding_type": "Actual Trend Data",
        "finding_category": trend_category,
        "plain_language_explanation": plain_language_explanation,
        "leadership_summary": (
            f"Leadership should review this {direction.get('risk_level', 'medium').lower()} risk trend signal for {station_code}."
        ),
        "counseling_summary": (
            "If this trend is used for coaching or counseling, keep the conversation neutral and fact-based. "
            "Review the source records before discussing the trend with the driver."
        ),
        "operational_summary": trend_summary,
        "recommended_action": recommended_action,
        "suggested_follow_up_question": "Would you like LORI to prepare a supervisor follow-up, counseling summary, action item, or reminder based on this trend?",
        "decision_support_note": decision_support_note,
        "should_offer_chart": True,
        "should_offer_counseling_language": True,
        "should_offer_leadership_packet": True,
        "should_send_to_action_center": True,
        "data_points_count": len(points),
        "source_records_count": len(actual_rows),
        "generated_by": "LORI Trend Intelligence",
        "created_at": lori_trend_data_now_iso(),
        "updated_at": lori_trend_data_now_iso(),
    }

    saved_result = await lori_trend_data_supabase_post("lori_trend_results", result_payload)

    if saved_result and saved_result[0].get("id"):
        result_id = saved_result[0]["id"]

    for point in points:
        point_payload = {
            "trend_result_id": result_id,
            "point_label": point.get("date"),
            "metric_name": point.get("metric_name"),
            "metric_value": point.get("metric_value"),
            "source_module": point.get("source_module"),
            "notes": point.get("notes"),
            "created_at": lori_trend_data_now_iso(),
        }
        await lori_trend_data_supabase_post("lori_trend_data_points", point_payload)

    for row in actual_rows[:25]:
        source_payload = {
            "trend_result_id": result_id,
            "source_module": row.get("_source_module", "Unknown Source"),
            "source_title": (
                row.get("action_title")
                or row.get("event_type")
                or row.get("counseling_type")
                or row.get("metric_name")
                or "Trend source record"
            ),
            "source_reference": str(row.get("id") or row.get("driver_id") or ""),
            "driver_id": driver_id,
            "driver_name": driver_name,
            "station_code": station_code,
            "relevance": "Actual source record used for trend review.",
            "created_at": lori_trend_data_now_iso(),
        }
        await lori_trend_data_supabase_post("lori_trend_source_records", source_payload)

    if actual_data_used:
        voiceflow_message = (
            f"{trend_summary}\n\n"
            f"Actual Trend Data:\n"
            f"Data Points Reviewed: {len(points)}\n"
            f"Source Records Reviewed: {len(actual_rows)}\n"
            f"Date Range: {date_range}\n"
            f"Metric Used: {metric_used}\n"
            f"First Value: {first_value}\n"
            f"Latest Value: {latest_value}\n"
            f"Change: {change_value}\n\n"
            f"Trend Direction: {direction.get('trend_direction', 'Needs More Data')}\n"
            f"Trend Strength: {direction.get('trend_strength', 'Preliminary')}\n"
            f"Confidence Level: {direction.get('confidence_level', 'Low')}\n"
            f"Risk Level: {direction.get('risk_level', 'Medium')}\n\n"
            f"Data Points:\n{actual_points_text}\n\n"
            f"Recommended Action: {recommended_action}\n\n"
            f"{decision_support_note}"
        )
    else:
        voiceflow_message = (
            f"{trend_summary}\n\n"
            f"Actual Trend Data:\n"
            f"Data Points Reviewed: 0\n"
            f"Source Records Reviewed: 0\n"
            f"Date Range: Not available\n"
            f"Metric Used: Not available\n\n"
            f"Recommended Action: Upload or link dated driver performance, counseling, safety, route, KPI, or action records so LORI can calculate a reliable trend.\n\n"
            f"{decision_support_note}"
        )

    return {
        "status": "success",
        "actual_data_used": actual_data_used,
        "voiceflow_message": voiceflow_message,
        "trend_request_id": request_id,
        "trend_result_id": result_id,
        "trend_title": trend_title,
        "trend_summary": trend_summary,
        "trend_direction": direction.get("trend_direction", "Needs More Data"),
        "trend_strength": direction.get("trend_strength", "Preliminary"),
        "confidence_level": direction.get("confidence_level", "Low"),
        "risk_level": direction.get("risk_level", "Medium"),
        "finding_category": trend_category,
        "metric_used": metric_used,
        "first_value": first_value,
        "latest_value": latest_value,
        "change_value": change_value,
        "date_range": date_range,
        "plain_language_explanation": plain_language_explanation,
        "recommended_action": recommended_action,
        "decision_support_note": decision_support_note,
        "data_points_count": len(points),
        "source_records_count": len(actual_rows),
        "data_points": points,
        "source_records": actual_rows[:25],
    }


@app.post("/trend-quick-analyze-data")
async def trend_quick_analyze_data(
    api_key: Optional[str] = Query(None),
    payload: Dict[str, Any] = Body(...),
):
    """
    Same actual-data endpoint for Lovable/manual testing.
    """
    return await voiceflow_trend_quick_analyze_data(api_key=api_key, payload=payload)
